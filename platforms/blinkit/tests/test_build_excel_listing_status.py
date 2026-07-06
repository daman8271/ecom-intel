import datetime
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
BUILDER = ROOT / "platforms" / "blinkit" / "build_excel.py"


def row(status, sku, canonical, pincode, in_stock, source, sale, prid):
    return {
        "city": "Delhi",
        "pincode": pincode,
        "locality": f"Locality {pincode}",
        "store_id": f"store-{pincode}",
        "store_name": f"Store {pincode}",
        "sku_raw": sku,
        "canonical": canonical,
        "pack": "1 l",
        "vol_ml": 1000,
        "sale": sale,
        "base_sale": sale,
        "offer_sale": None,
        "mrp": sale + 100,
        "discount_pct": round(100 * 100 / (sale + 100), 1),
        "per_litre": sale,
        "eta_min": 10,
        "in_stock": in_stock,
        "listing_status": status,
        "stock_source": source,
        "price_source": source,
        "pdp_checked": 1 if source in ("pdp", "pdp_probe") else 0,
        "prid": prid,
        "listing_url": f"https://blinkit.com/prn/{canonical}/prid/{prid}",
    }


def test_listing_status_distinguishes_not_listed_from_oos():
    sku_seen = row(
        "listed_in_stock",
        "Jivo Test Listed Oil",
        "jivo-test-listed-oil-1l",
        "110001",
        1,
        "pdp_probe",
        100,
        "100001",
    )
    sku_oos = row(
        "listed_out_of_stock",
        "Jivo Test Missing Oil",
        "jivo-test-missing-oil-1l",
        "110001",
        0,
        "pdp",
        200,
        "100002",
    )
    sku_other_pin = row(
        "listed_in_stock",
        "Jivo Test Listed Oil",
        "jivo-test-listed-oil-1l",
        "110002",
        1,
        "pdp_probe",
        101,
        "100001",
    )
    result = {
        "summary": {
            "captured_at": "2099-01-01T04:00:00.000Z",
            "pincodes_total": 2,
            "pincodes_with_jivo": 2,
            "unique_skus": 2,
            "total_rows": 3,
            "wall_s": 1,
        },
        "allRows": [sku_seen, sku_oos, sku_other_pin],
        "perPin": [
            {
                "city": "Delhi",
                "pincode": "110001",
                "locality": "Locality 110001",
                "resolved": True,
                "store_name": "Store 110001",
                "rows": [sku_seen, sku_oos],
            },
            {
                "city": "Delhi",
                "pincode": "110002",
                "locality": "Locality 110002",
                "resolved": True,
                "store_name": "Store 110002",
                "rows": [sku_other_pin],
            },
        ],
    }

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        (tmp_path / "result.json").write_text(json.dumps(result), encoding="utf-8")
        subprocess.run([sys.executable, str(BUILDER)], cwd=tmp_path, check=True, capture_output=True, text=True)
        workbook = tmp_path / f"Jivo-{tmp_path.name.title()}-Live-Report-{datetime.date.today()}.xlsx"
        assert workbook.exists()

        wb = load_workbook(workbook, read_only=True, data_only=True)
        ws = wb["Listing Status"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        records = [dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]

        oos = [
            r for r in records
            if str(r["Pincode"]) == "110001" and r["SKU"] == "Jivo Test Missing Oil 1 L"
        ][0]
        not_listed = [
            r for r in records
            if str(r["Pincode"]) == "110002" and r["SKU"] == "Jivo Test Missing Oil 1 L"
        ][0]

        assert oos["Product status"] == "Listed - Out of stock"
        assert oos["In stock"] == "No"
        assert oos["Source"] == "pdp"
        assert not_listed["Product status"] == "Not listed"
        assert not_listed["In stock"] is None
        assert not_listed["Source"] == "search_absent"


if __name__ == "__main__":
    test_listing_status_distinguishes_not_listed_from_oos()
    print("PASS blinkit workbook listing-status regression")
