import json, datetime, statistics, os, re
from pathlib import Path
from collections import defaultdict, OrderedDict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

d = json.load(open('result.json'))
rows = d['allRows']
per = d['perPin']
summary = d['summary']

IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))


def clean_name(s):
    """Repair a scrape-truncated listing title that ends in a dangling '(' —
    e.g. 'Jivo Cold Pressed Canola Oil (' -> 'Jivo Cold Pressed Canola Oil'."""
    s = str(s or "").strip()
    while s.endswith("("):
        s = s[:-1].rstrip()
    return s


def captured_ist(cap):
    """Format an ISO-8601 UTC captured_at as real IST ('YYYY-MM-DD HH:MM IST'),
    not the UTC clock mislabelled IST (matches tools/report_dashboard.py)."""
    try:
        dt = datetime.datetime.fromisoformat(str(cap).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        return dt.astimezone(IST).strftime("%Y-%m-%d %H:%M IST")
    except Exception:
        return f"{str(cap)[:16].replace('T', ' ')} IST"

# platform name derived from the folder this runs in (blinkit, zepto, ...)
PLATFORM = os.path.basename(os.getcwd()).replace('-', ' ').title()

JIVO_GREEN = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO_GREEN)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=JIVO_GREEN, size=18)
SUB_FONT = Font(italic=True, color="555555", size=10)
RED = PatternFill("solid", fgColor="F4CCCC")
GREEN = PatternFill("solid", fgColor="D9EAD3")
YEL = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

wb = Workbook()

# pretty SKU label from canonical
def label(canon):
    parts = canon.rsplit('-', 1)
    name = parts[0].replace('-', ' ').title()
    pack = parts[1].upper().replace('ML', ' ml').replace('L', ' L') if len(parts) > 1 else ''
    return f"{name} {pack}".strip()

def parse_vol_ml(pack):
    if not pack:
        return None
    s = str(pack).lower()
    def to_ml(n, u):
        if u in ('ml', 'g'):
            return n
        if u in ('l', 'ltr', 'litre', 'kg'):
            return n * 1000
        return None
    m = re.search(r'([\d.]+)\s*(ml|l|ltr|litre|kg|g)\b\s*[x×]\s*([\d.]+)', s)
    if m:
        base = to_ml(float(m.group(1)), m.group(2))
        return base * float(m.group(3)) if base is not None else None
    m = re.search(r'([\d.]+)\s*[x×]\s*([\d.]+)\s*(ml|l|ltr|litre|kg|g)\b', s)
    if m:
        base = to_ml(float(m.group(2)), m.group(3))
        return float(m.group(1)) * base if base is not None else None
    m = re.search(r'([\d.]+)\s*(ml|l|ltr|litre|kg|g)', s)
    if not m:
        return None
    return to_ml(float(m.group(1)), m.group(2))

def canonical_py(name, pack):
    base = re.sub(r'\(.*?\)', '', str(name or '').lower())
    base = re.sub(r'[^a-z0-9 ]', '', base)
    base = re.sub(r'\s+', ' ', base).strip().replace(' ', '-')
    vol = parse_vol_ml(pack)
    if vol:
        vol_tag = f"{vol / 1000:g}l" if vol >= 1000 else f"{vol:g}ml"
    else:
        vol_tag = "na"
    return re.sub(r'--+', '-', f"{base}-{vol_tag}")

def pack_from_name(name):
    m = re.search(r'(\d+(?:\.\d+)?)\s*(ml|l|ltr|litre|kg|g)\b', str(name or ''), re.I)
    if not m:
        return ""
    unit = m.group(2).lower()
    if unit == "ltr":
        unit = "l"
    return f"{m.group(1)} {unit}"

def load_expected_sku_meta(rows):
    meta = {}
    by_prid = {}
    for r in rows:
        canon = r.get('canonical')
        if not canon:
            continue
        meta.setdefault(canon, {"canonical": canon, "label": label(canon), "prid": r.get('prid', ''), "url": r.get('listing_url', '')})
        if r.get('prid'):
            by_prid[str(r.get('prid'))] = r

    root = Path(__file__).resolve().parents[2]
    map_path = root / "tools" / "pricematch" / "sku_map.json"
    try:
        sku_map = json.load(open(map_path, encoding="utf-8")).get("skus") or {}
    except Exception:
        sku_map = {}

    for master_name, rec in sku_map.items():
        b = (rec.get("platforms") or {}).get("blinkit") or {}
        if not b:
            continue
        prid = str(b.get("id") or "").strip()
        live = by_prid.get(prid)
        if live and live.get('canonical'):
            canon = live['canonical']
        else:
            canon = canonical_py(clean_name(b.get("title") or master_name), pack_from_name(master_name))
        meta.setdefault(canon, {
            "canonical": canon,
            "label": label(canon),
            "prid": prid,
            "url": b.get("url", ""),
        })
        if prid and not meta[canon].get("prid"):
            meta[canon]["prid"] = prid
        if b.get("url") and not meta[canon].get("url"):
            meta[canon]["url"] = b.get("url")
    return meta

sku_meta = load_expected_sku_meta(rows)
skus = sorted(sku_meta)
cities_with = sorted(set(r['city'] for r in rows))
all_cities = OrderedDict()
for p in per:
    all_cities.setdefault(p['city'], 0)
    all_cities[p['city']] += len(p['rows'])
cities_without = [c for c, n in all_cities.items() if n == 0]

def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDR_FONT; cell.alignment = CEN; cell.border = BORDER

def autosize(ws, maxw=42):
    for col in ws.columns:
        L = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[L].width = min(maxw, max(10, w + 2))

def pct_fraction(v):
    if v is None:
        return None
    return round(float(v) / 100.0, 4)

def product_status(x):
    s = x.get('listing_status')
    if s == 'listed_in_stock' or x.get('in_stock') == 1:
        return "Listed - In stock"
    if s == 'listed_out_of_stock' or x.get('in_stock') == 0:
        return "Listed - Out of stock"
    return str(s or "Listed").replace("_", " ").title()

def row_preference(x):
    """Pick the clearest row when multiple stores survive for one pincode/SKU."""
    return (
        1 if x.get('in_stock') == 1 else 0,
        1 if x.get('stock_source') in ('pdp', 'pdp_probe') or x.get('pdp_checked') else 0,
        1 if x.get('stock_probe') else 0,
    )

not_listed_rows = []

def collect_not_listed_row(p, sku):
    meta = sku_meta.get(sku) or {}
    return {
        "city": p['city'],
        "pincode": p['pincode'],
        "locality": p.get('locality', ''),
        "store_name": p.get('store_name', ''),
        "sku": meta.get("label") or label(sku),
        "canonical": sku,
        "prid": meta.get("prid", ""),
        "listing_url": meta.get("url", ""),
        "source": "search_absent",
        "note": "SKU was not listed for this resolved Blinkit pincode/store; this is not an out-of-stock row.",
    }

# ---------- Sheet 1: Executive Summary ----------
ws = wb.active; ws.title = "Summary"
ws["A1"] = f"Jivo x {PLATFORM} - Live Pricing Intelligence"; ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")
ws["A2"] = f"Captured {captured_ist(summary['captured_at'])}  -  {summary['pincodes_with_jivo']}/{summary['pincodes_total']} pincodes carry Jivo  -  {summary['unique_skus']} unique SKUs  -  {summary['total_rows']} datapoints  -  scrape {summary['wall_s']}s"
ws["A2"].font = SUB_FONT; ws.merge_cells("A2:G2")

# KPI cards
kpis = [("Unique SKUs", summary['unique_skus']), ("Pincodes w/ Jivo", f"{summary['pincodes_with_jivo']}/{summary['pincodes_total']}"),
        ("Datapoints", summary['total_rows']), ("Cities w/ ZERO Jivo", len(cities_without))]
r = 4
for i, (k, v) in enumerate(kpis):
    c = 1 + i * 2
    ws.cell(row=r, column=c, value=k).font = Font(bold=True, size=10, color="555555")
    cell = ws.cell(row=r + 1, column=c, value=v); cell.font = Font(bold=True, size=20, color=JIVO_GREEN)

# cheapest per SKU
ws.cell(row=7, column=1, value="Cheapest pincode per SKU (in-stock, sale price)").font = Font(bold=True, size=12)
hdr = ["SKU", "City", "Pincode", "Store", "Sale Rs", "MRP Rs", "Disc %"]
for j, h in enumerate(hdr, 1):
    ws.cell(row=8, column=j, value=h)
style_header(ws, 8, len(hdr))
rr = 9
for s in skus:
    cand = [x for x in rows if x['canonical'] == s and x['in_stock'] == 1]
    if not cand: cand = [x for x in rows if x['canonical'] == s]
    if not cand: continue
    b = min(cand, key=lambda x: x['sale'])
    for j, v in enumerate([label(s), b['city'], b['pincode'], b['store_name'], b['sale'], b['mrp'], pct_fraction(b['discount_pct'])], 1):
        cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
        if j >= 5: cell.alignment = CEN
        if j == 7: cell.number_format = '0.0%'
    rr += 1
# distribution gaps
rr += 1
ws.cell(row=rr, column=1, value=f"WHITESPACE - Cities with ZERO Jivo on {PLATFORM}:").font = Font(bold=True, size=11, color="CC0000")
rr += 1
ws.cell(row=rr, column=1, value=", ".join(cities_without) if cities_without else "None - full coverage").font = Font(size=11)
ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
autosize(ws)

# ---------- Sheet 2: Master Data ----------
ws = wb.create_sheet("Master Data")
cols = ["City", "Pincode", "Locality", "Store", "SKU", "Pack", "Vol (ml)", "Sale Rs", "Base Sale Rs", "Offer Rs", "MRP Rs", "Disc %", "Rs/L", "ETA min", "In stock", "Product status", "Stock source", "Price source"]
ws.append(cols)
for x in sorted(rows, key=lambda r: (r['city'], r['pincode'], r['canonical'])):
    ws.append([x['city'], x['pincode'], x.get('locality',''), x['store_name'], clean_name(x['sku_raw']), x['pack'], x['vol_ml'],
               x['sale'], x.get('base_sale'), x.get('offer_sale'), x['mrp'], pct_fraction(x['discount_pct']), x['per_litre'], x['eta_min'], "Yes" if x['in_stock'] else "No",
               product_status(x), x.get('stock_source',''), x.get('price_source','')])
style_header(ws)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
        if cell.column in (8, 9, 10, 11): cell.number_format = '"Rs"#,##0'
        if cell.column == 12: cell.number_format = '0.0%'
    sc = row[7].value
    if row[14].value == "No": row[14].fill = RED
    if isinstance(row[11].value, (int, float)) and row[11].value and row[11].value >= 0.40: row[11].fill = GREEN
autosize(ws)

# ---------- Sheet 2b: Listing Status ----------
ws = wb.create_sheet("Listing Status")
cols = ["City", "Pincode", "Locality", "SKU", "Product status", "In stock", "Sale Rs", "Base Sale Rs", "Offer Rs", "MRP Rs", "Store", "PRID", "Source"]
ws.append(cols)
for p in sorted(per, key=lambda r: (r['city'], r['pincode'])):
    if not p.get('resolved'):
        continue
    by_sku = {}
    for x in (p.get('rows') or []):
        key = x.get('canonical')
        if not key:
            continue
        if key not in by_sku or row_preference(x) > row_preference(by_sku[key]):
            by_sku[key] = x
    for s in skus:
        x = by_sku.get(s)
        if x:
            ws.append([p['city'], p['pincode'], p.get('locality',''), label(s), product_status(x),
                       "Yes" if x.get('in_stock') else "No", x.get('sale'), x.get('base_sale'), x.get('offer_sale'), x.get('mrp'),
                       x.get('store_name',''), x.get('prid',''), x.get('stock_source','')])
        else:
            not_listed_rows.append(collect_not_listed_row(p, s))
            ws.append([p['city'], p['pincode'], p.get('locality',''), label(s), "Not listed",
                       "", None, None, None, None, p.get('store_name',''), "", "search_absent"])
style_header(ws)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
        if cell.column in (7, 8, 9, 10): cell.number_format = '"Rs"#,##0'
    status = row[4].value
    if status == "Listed - In stock":
        row[4].fill = GREEN
    elif status == "Listed - Out of stock":
        row[4].fill = RED
    elif status == "Not listed":
        row[4].fill = YEL
autosize(ws)

# ---------- Sheet 2c: Not Listed Pincodes ----------
ws = wb.create_sheet("Not Listed Pincodes")
not_listed_cols = ["City", "Pincode", "Locality", "Store", "SKU", "Canonical", "PRID", "Listing URL", "Source", "Note"]
ws.append(not_listed_cols)
for x in sorted(not_listed_rows, key=lambda r: (r["city"], r["pincode"], r["sku"])):
    ws.append([x["city"], x["pincode"], x["locality"], x["store_name"], x["sku"], x["canonical"], x["prid"], x["listing_url"], x["source"], x["note"]])
style_header(ws)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(not_listed_cols))}{max(ws.max_row, 1)}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
    row[8].fill = YEL
autosize(ws)

# ---------- Matrix builder ----------
def matrix(sheet_name, valfn, fmt=None, scale=False, scale_rev=False):
    ws = wb.create_sheet(sheet_name)
    cols = ["SKU"] + cities_with
    ws.append(cols)
    for s in skus:
        rowvals = [label(s)]
        for c in cities_with:
            cand = [x for x in rows if x['canonical'] == s and x['city'] == c]
            rowvals.append(valfn(cand) if cand else None)
        ws.append(rowvals)
    style_header(ws)
    ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            if cell.column > 1:
                cell.alignment = CEN
                if fmt and isinstance(cell.value, (int, float)): cell.number_format = fmt
                if cell.value is None: cell.value = "-"
    if scale:
        last = get_column_letter(ws.max_column)
        lo, hi = ("F4CCCC", "D9EAD3") if scale_rev else ("D9EAD3", "F4CCCC")
        ws.conditional_formatting.add(f"B2:{last}{ws.max_row}",
            ColorScaleRule(start_type="min", start_color=lo, end_type="max", end_color=hi))
    autosize(ws, maxw=14)
    return ws

# Sheet 3: Pricing Matrix (modal sale per city) - green cheap -> red expensive
def modal_price(cands):
    prices = [x['sale'] for x in cands if x['sale'] is not None]
    if not prices: return None
    cnt = Counter(prices); top = max(cnt.values())
    return round(min(p for p, n in cnt.items() if n == top))
matrix("Pricing Matrix", modal_price, '"Rs"#,##0', scale=True)
# Sheet 4: Stock Status (% in stock)
def stock_cell(c):
    pct = round(sum(x['in_stock'] for x in c) / len(c), 4)
    return pct
wsS = matrix("Stock Status", stock_cell, '0%')
for row in wsS.iter_rows(min_row=2):
    for cell in row:
        if cell.column > 1 and isinstance(cell.value, (int, float)):
            cell.fill = GREEN if cell.value == 1 else (RED if cell.value == 0 else YEL)
# Sheet 5: Discount Analysis (modal disc per city) - higher = greener
def modal_disc(cands):
    vals = [x['discount_pct'] for x in cands if x['discount_pct'] is not None]
    if not vals: return None
    cnt = Counter(vals); top = max(cnt.values())
    return pct_fraction(round(min(v for v, n in cnt.items() if n == top), 1))
matrix("Discount Analysis", modal_disc, '0.0%', scale=True, scale_rev=True)

# ---------- Sheet 6: Coverage / Gaps ----------
ws = wb.create_sheet("Coverage & Gaps")
ws.append(["City", "Pincode", "Locality", "Store assigned", "Jivo SKUs found"])
for p in per:
    ws.append([p['city'], p['pincode'], p.get('locality',''), p['store_name'], len(p['rows'])])
style_header(ws); ws.freeze_panes = "A2"
for row in ws.iter_rows(min_row=2):
    for cell in row: cell.border = BORDER
    if row[4].value == 0: row[4].fill = RED
    else: row[4].fill = GREEN
autosize(ws)

fname = f"Jivo-{PLATFORM.replace(' ', '')}-Live-Report-{datetime.date.today()}.xlsx"
wb.save(fname)
print("SAVED:", fname)
print("Sheets:", wb.sheetnames)

nl_wb = Workbook()
nl_ws = nl_wb.active
nl_ws.title = "Not Listed Pincodes"
nl_ws.append(not_listed_cols)
for x in sorted(not_listed_rows, key=lambda r: (r["city"], r["pincode"], r["sku"])):
    nl_ws.append([x["city"], x["pincode"], x["locality"], x["store_name"], x["sku"], x["canonical"], x["prid"], x["listing_url"], x["source"], x["note"]])
style_header(nl_ws)
nl_ws.freeze_panes = "A2"
nl_ws.auto_filter.ref = f"A1:{get_column_letter(len(not_listed_cols))}{max(nl_ws.max_row, 1)}"
for row in nl_ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
    row[8].fill = YEL
autosize(nl_ws)
nl_name = f"Jivo-{PLATFORM.replace(' ', '')}-Not-Listed-Pincodes-{datetime.date.today()}.xlsx"
nl_wb.save(nl_name)
print("SAVED:", nl_name)
