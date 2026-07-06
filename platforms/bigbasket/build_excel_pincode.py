"""BigBasket PINCODE-WISE report — proper, matching the other platforms' standard.
Sheets: Summary (KPI cards + verdict + cheapest-per-SKU + whitespace), Master Data,
Pincode Coverage, Pricing Matrix, Stock Status, Discount Analysis. Fixed column widths
(no more ####), conditional formatting, Jivo-only. Reads result_pincode.json."""
import json, datetime, statistics, os
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

HERE = os.path.dirname(os.path.abspath(__file__))
d = json.load(open(os.path.join(HERE, os.environ.get('PIN_RESULT', 'result_pincode.json'))))
rows, per, summary = d['allRows'], d['perPin'], d['summary']
MIN_REPORT_PINS = int(os.environ.get('BB_PINCODE_MIN_REPORT_PINS', '20'))
ALLOW_PARTIAL_REPORT = os.environ.get('BB_ALLOW_PARTIAL_PINCODE_REPORT', '').lower() in ('1', 'true', 'yes')
target_pins = summary.get('pincodes_target')
target_pins = int(target_pins) if target_pins is not None else None
required_pins = min(target_pins, MIN_REPORT_PINS) if target_pins is not None else MIN_REPORT_PINS
if not ALLOW_PARTIAL_REPORT and int(summary.get('pincodes_total') or len(per) or 0) < required_pins:
    raise SystemExit(
        f"Refusing under-covered BigBasket pincode report: "
        f"{summary.get('pincodes_total', len(per))}/{target_pins or '?'} pincodes; "
        f"minimum required {required_pins}. Set BB_ALLOW_PARTIAL_PINCODE_REPORT=1 only for diagnostics."
    )
PRICEMATCH = {'110092', '560006'}  # the active price-match pins — always highlighted

GREEN_H = "008B3A"
HDR = PatternFill("solid", fgColor=GREEN_H); HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE = Font(bold=True, color=GREEN_H, size=18); SUB = Font(italic=True, color="555555", size=10)
RED = PatternFill("solid", fgColor="F4CCCC"); GREEN = PatternFill("solid", fgColor="D9EAD3")
YEL = PatternFill("solid", fgColor="FFF2CC"); BLUE = PatternFill("solid", fgColor="CFE2F3")
thin = Side(style="thin", color="D0D0D0"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")
RS = '"₹"#,##0.00'; PCT = '0.0%'
wb = Workbook()


def label(c):
    p = c.rsplit('-', 1); n = p[0].replace('-', ' ').title()
    pk = p[1].upper().replace('ML', ' ml').replace('L', ' L') if len(p) > 1 else ''
    return f"{n} {pk}".strip()


def style_header(ws, row=1, n=None):
    for c in range(1, (n or ws.max_column) + 1):
        cell = ws.cell(row=row, column=c); cell.fill = HDR; cell.font = HDR_FONT; cell.alignment = CEN; cell.border = BORDER


def widths(ws, spec):
    """spec: dict {col_index: width}; everything else gets a sensible autosize with buffer
    that accounts for ₹ number formatting (the #### fix)."""
    for col in ws.columns:
        i = col[0].column; L = get_column_letter(i)
        if i in spec:
            ws.column_dimensions[L].width = spec[i]; continue
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[L].width = min(46, max(11, w + 3))


def pct_fraction(v):
    if v is None:
        return None
    return round(float(v) / 100.0, 4)


skus = sorted(set(r['canonical'] for r in rows))
pins_with = [p for p in per if p['rows']]
pins_zero = [p for p in per if not p['rows']]
plabel = lambda p: f"{p['pincode']} {p['city']}"


def location_issue(p):
    status = str(p.get('set_status') or '')
    pin = str(p.get('pincode') or '')
    rpin = str(p.get('resolved_pincode') or '')
    city = str(p.get('city') or p.get('requested_city') or '')
    rcity = str(p.get('resolved_city') or '')
    issues = []
    if status and status != 'ok':
        issues.append(status)
    if rpin and rpin != pin:
        issues.append(f"resolved pin {rpin}")
    if rcity and city and rcity.lower() != city.lower():
        issues.append(f"resolved city {rcity}")
    if not p.get('rows'):
        issues.append("zero Jivo rows")
    return "; ".join(issues)


loc_issues = [p for p in per if location_issue(p)]

# ---------------- Sheet 1: Summary ----------------
ws = wb.active; ws.title = "Summary"
ws["A1"] = "Jivo × BigBasket — PINCODE-WISE Pricing Intelligence"; ws["A1"].font = TITLE; ws.merge_cells("A1:H1")
surveyed = (f"{summary['pincodes_total']}/{summary['pincodes_target']} pincodes surveyed   |   "
            if summary.get('pincodes_target') and summary.get('pincodes_target') != summary.get('pincodes_total') else "")
ws["A2"] = (f"Captured {summary.get('captured_at','')[:16].replace('T',' ')} IST   |   "
            f"{surveyed}"
            f"{summary['pincodes_with_jivo']}/{summary['pincodes_total']} pincodes carry Jivo   |   "
            f"{summary['unique_skus']} SKUs   |   {summary['total_rows']} datapoints   |   source: {summary.get('source','')}")
ws["A2"].font = SUB; ws.merge_cells("A2:H2")
ws["A3"] = "VERDICT: " + str(summary.get('verdict', '')); ws["A3"].font = Font(bold=True, color=GREEN_H, size=12); ws.merge_cells("A3:H3")
avg_disc = round(statistics.mean([r['discount_pct'] for r in rows if r['discount_pct']]), 1) if any(r['discount_pct'] for r in rows) else 0
kpis = [("Pincodes carrying Jivo", f"{summary['pincodes_with_jivo']}/{summary['pincodes_total']}"), ("Unique SKUs", summary['unique_skus']),
        ("Datapoints", summary['total_rows']), ("Avg discount", f"{avg_disc}%")]
for i, (k, v) in enumerate(kpis):
    c = 1 + i * 2
    ws.cell(row=5, column=c, value=k).font = Font(bold=True, size=10, color="555555")
    ws.cell(row=6, column=c, value=v).font = Font(bold=True, size=20, color=GREEN_H)
# price-match pin callout
pm = []
for pin in sorted(PRICEMATCH):
    pe = next((p for p in per if str(p['pincode']) == pin), None)
    pm.append(f"{pin} ({pe['city']}): {len(pe['rows'])} Jivo SKUs" if pe else f"{pin}: not surveyed")
ws["A8"] = "Price-match pins: " + "   |   ".join(pm); ws["A8"].font = Font(bold=True, color="1155CC", size=10); ws.merge_cells("A8:H8")
ws["A9"] = f"Location audit: {len(loc_issues)} pincodes have a zero-row, failed, or resolved-location warning"; ws["A9"].font = Font(bold=True, color="CC0000" if loc_issues else GREEN_H, size=10); ws.merge_cells("A9:H9")
# cheapest pincode per SKU
ws.cell(row=10, column=1, value="Cheapest pincode per SKU (in-stock)").font = Font(bold=True, size=12)
for j, h in enumerate(["SKU", "Pincode", "City", "Sale", "MRP", "Disc %"], 1):
    ws.cell(row=11, column=j, value=h)
style_header(ws, 11, 6); rr = 12
for s in skus:
    cand = [x for x in rows if x['canonical'] == s and x['in_stock']] or [x for x in rows if x['canonical'] == s]
    if not cand:
        continue
    b = min(cand, key=lambda x: x['sale'])
    for j, v in enumerate([label(s), b['pincode'], b['city'], b['sale'], b['mrp'], pct_fraction(b['discount_pct'])], 1):
        cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
        if j == 4 or j == 5:
            cell.number_format = RS
        if j == 6:
            cell.number_format = PCT
        if j >= 4:
            cell.alignment = CEN
    rr += 1
rr += 1
ws.cell(row=rr, column=1, value=f"WHITESPACE — {len(pins_zero)} pincodes with ZERO Jivo (skip these in the daily cron):").font = Font(bold=True, color="CC0000", size=11)
ws.cell(row=rr + 1, column=1, value=", ".join(f"{p['pincode']}({p['city']})" for p in pins_zero[:60]) + (" …" if len(pins_zero) > 60 else "")).font = Font(size=9)
ws.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 1, end_column=8)
widths(ws, {1: 34, 2: 11, 3: 13, 4: 12, 5: 12, 6: 9, 7: 10, 8: 10})

# ---------------- Sheet 2: Master Data ----------------
ws = wb.create_sheet("Master Data")
cols = ["City", "Pincode", "Resolved Pin", "Locality", "SKU", "SKU Code", "Pack", "Vol (ml)", "Sale", "MRP", "Disc %", "₹/L", "In stock", "PM pin"]
ws.append(cols)
for x in sorted(rows, key=lambda r: (r['city'], r['pincode'], r['canonical'])):
    ws.append([x['city'], x['pincode'], x.get('resolved_pincode', ''), x.get('locality', ''), x['sku_raw'], str(x.get('sku_id') or ''), x['pack'], x['vol_ml'], x['sale'], x['mrp'],
               pct_fraction(x['discount_pct']), x['per_litre'], "Yes" if x['in_stock'] else "No", "★" if str(x['pincode']) in PRICEMATCH else ""])
style_header(ws); ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
    row[8].number_format = RS; row[9].number_format = RS; row[10].number_format = PCT; row[11].number_format = RS
    if row[12].value == "No":
        row[12].fill = RED
    if isinstance(row[10].value, (int, float)) and row[10].value >= 0.40:
        row[10].fill = GREEN
    if row[13].value:
        row[1].fill = BLUE
    if row[2].value and str(row[2].value) != str(row[1].value):
        row[2].fill = YEL
widths(ws, {1: 13, 2: 11, 3: 12, 4: 22, 5: 40, 6: 11, 8: 9, 9: 13, 10: 13, 11: 9, 12: 11, 13: 9, 14: 7})

# ---------------- Sheet 3: Pincode Coverage ----------------
ws = wb.create_sheet("Pincode Coverage")
ws.append(["City", "Pincode", "Resolved City", "Resolved Pin", "Locality", "PM pin", "Jivo SKUs", "In-stock", "OOS", "Avg Sale", "Min Sale", "Issue"])
for p in sorted(per, key=lambda p: (-len(p['rows']), p['city'], p['pincode'])):
    pr = p['rows']; ins = sum(1 for r in pr if r['in_stock']); sales = [r['sale'] for r in pr if r.get('sale') is not None]
    ws.append([p['city'], p['pincode'], p.get('resolved_city', ''), p.get('resolved_pincode', ''),
               p.get('locality', ''), "★" if str(p['pincode']) in PRICEMATCH else "",
               len(pr), ins, len(pr) - ins, round(statistics.mean(sales)) if sales else None, min(sales) if sales else None,
               location_issue(p)])
style_header(ws); ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:L{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
    row[6].fill = RED if row[6].value == 0 else GREEN
    if isinstance(row[8].value, int) and row[8].value > 0:
        row[8].fill = YEL
    row[9].number_format = RS; row[10].number_format = RS
    if row[5].value:
        row[1].fill = BLUE
    if row[11].value:
        row[11].fill = YEL
widths(ws, {1: 13, 2: 11, 3: 14, 4: 12, 5: 22, 6: 7, 7: 10, 8: 9, 9: 6, 10: 12, 11: 12, 12: 34})

# ---------------- Sheet 4: Location Audit ----------------
ws = wb.create_sheet("Location Audit")
ws.append(["City", "Pincode", "Locality", "Resolved City", "Resolved Pin", "Resolved Locality", "Status", "Jivo SKUs", "Store/SA", "Issue"])
for p in sorted(per, key=lambda p: (p['city'], p['pincode'])):
    issue = location_issue(p)
    ws.append([
        p.get('city', ''), p.get('pincode', ''), p.get('locality', ''),
        p.get('resolved_city', ''), p.get('resolved_pincode', ''), p.get('resolved_locality', ''),
        p.get('set_status', ''), len(p.get('rows') or []), p.get('serving_sa') or p.get('store_id', ''), issue,
    ])
style_header(ws); ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:J{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
    if row[9].value:
        row[9].fill = YEL
    if row[6].value != 'ok':
        row[6].fill = RED
    elif row[7].value:
        row[7].fill = GREEN
widths(ws, {1: 14, 2: 11, 3: 24, 4: 14, 5: 12, 6: 24, 7: 20, 8: 10, 9: 22, 10: 42})

# ---------------- Sheet 5: SKU City Coverage ----------------
ws = wb.create_sheet("SKU City Coverage")
ws.append([
    "SKU", "Pack", "Alive Cities", "Alive City List", "Alive Pincodes",
    "Listed Cities", "Listed Pincodes", "In-stock Rows", "OOS Rows",
])
sku_city_rows = []
for s in skus:
    sku_rows = [r for r in rows if r['canonical'] == s]
    alive_rows = [r for r in sku_rows if r.get('in_stock')]
    alive_cities = sorted({str(r.get('city', '')) for r in alive_rows if r.get('city')})
    listed_cities = sorted({str(r.get('city', '')) for r in sku_rows if r.get('city')})
    alive_pins = sorted({str(r.get('pincode', '')) for r in alive_rows if r.get('pincode')})
    listed_pins = sorted({str(r.get('pincode', '')) for r in sku_rows if r.get('pincode')})
    sample = sku_rows[0] if sku_rows else {}
    sku_city_rows.append((
        len(alive_cities), len(alive_pins), label(s), sample.get('pack', ''),
        ", ".join(alive_cities), len(listed_cities), len(listed_pins),
        len(alive_rows), len(sku_rows) - len(alive_rows),
    ))
for alive_city_n, alive_pin_n, sku_label, pack, alive_city_list, listed_city_n, listed_pin_n, in_rows, oos_rows in sorted(
        sku_city_rows, key=lambda x: (-x[0], -x[1], x[2])):
    ws.append([sku_label, pack, alive_city_n, alive_city_list, alive_pin_n, listed_city_n, listed_pin_n, in_rows, oos_rows])
style_header(ws); ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:I{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
    row[2].fill = GREEN if row[2].value else RED
    row[4].fill = GREEN if row[4].value else RED
widths(ws, {1: 42, 2: 14, 3: 12, 4: 46, 5: 14, 6: 13, 7: 14, 8: 13, 9: 10})


# ---------------- Matrix helper (pincode rows x SKU cols) ----------------
def matrix(name, valfn, fmt=None, scale=False, scale_rev=False, only_with=True):
    ws = wb.create_sheet(name)
    src = pins_with if only_with else per
    ws.append(["Pincode / City"] + [label(s) for s in skus])
    for p in sorted(src, key=lambda p: (p['city'], p['pincode'])):
        byc = {}
        for r in p['rows']:
            byc.setdefault(r['canonical'], []).append(r)
        ws.append([plabel(p)] + [valfn(byc[s]) if s in byc else None for s in skus])
    style_header(ws); ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            if cell.column > 1:
                cell.alignment = CEN
                if fmt and isinstance(cell.value, (int, float)):
                    cell.number_format = fmt
                elif cell.value is None:
                    cell.value = "-"
    if scale and ws.max_row > 1:
        lo, hi = ("F4CCCC", "D9EAD3") if scale_rev else ("D9EAD3", "F4CCCC")
        ws.conditional_formatting.add(f"B2:{get_column_letter(ws.max_column)}{ws.max_row}",
                                      ColorScaleRule(start_type="min", start_color=lo, end_type="max", end_color=hi))
    widths(ws, {1: 20, **{c: 12 for c in range(2, ws.max_column + 1)}})
    return ws


# Sheet 4: Pricing Matrix (avg sale per pincode) — green cheap -> red dear
matrix("Pricing Matrix", lambda c: round(statistics.mean([x['sale'] for x in c]), 2), RS, scale=True)
# Sheet 5: Stock Status (% in-stock per pincode/SKU)
wsS = matrix("Stock Status", lambda c: round(sum(x['in_stock'] for x in c) / len(c), 4), '0%')
for row in wsS.iter_rows(min_row=2):
    for cell in row:
        if cell.column > 1 and isinstance(cell.value, (int, float)):
            cell.fill = GREEN if cell.value == 1 else (RED if cell.value == 0 else YEL)
# Sheet 6: Discount Analysis (avg disc per pincode/SKU) — higher = greener
matrix("Discount Analysis", lambda c: pct_fraction(round(statistics.mean([x['discount_pct'] for x in c]), 1)), PCT, scale=True, scale_rev=True)

fname = os.path.join(HERE, f"Jivo-BigBasket-Pincode-Report-{datetime.date.today()}.xlsx")
wb.save(fname)
print("SAVED:", fname)
print("Sheets:", wb.sheetnames)
