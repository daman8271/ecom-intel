import json, datetime, statistics, os, re
from collections import defaultdict, OrderedDict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

d = json.load(open('result.json'))
rows = d['allRows']
per = d['perPin']
summary = d['summary']

# Fail before touching an existing workbook when the adapter output is incomplete.
# The ingest adapter has its own gate; this is independent defense for manual runs.
if len(rows) < 10:
    raise SystemExit(f"QUALITY FAIL: only {len(rows)} Instamart rows; refusing empty report")
if summary.get('total_rows') != len(rows):
    raise SystemExit(
        f"QUALITY FAIL: summary total_rows={summary.get('total_rows')} but allRows={len(rows)}"
    )
if len(per) < 648 or summary.get('pincodes_total') != len(per):
    raise SystemExit(
        f"QUALITY FAIL: incomplete pincode coverage ({len(per)} perPin, "
        f"summary={summary.get('pincodes_total')})"
    )
per_by_pin = {str(p.get('pincode', '')): p for p in per}
if len(per_by_pin) != len(per):
    raise SystemExit("QUALITY FAIL: duplicate pincode entries in perPin")
for required_pin in ('110059', '110064'):
    required = per_by_pin.get(required_pin)
    if required is None:
        raise SystemExit(f"QUALITY FAIL: required direct pincode {required_pin} is absent")
    if not required.get('rows') and required.get('collection_status') != 'failed':
        raise SystemExit(
            f"QUALITY FAIL: {required_pin} has neither data nor an explicit failure marker"
        )

# platform name derived from the folder this runs in (blinkit, zepto, ...)
PLATFORM = os.path.basename(os.getcwd()).replace('-', ' ').title()

JIVO_GREEN = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO_GREEN)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=JIVO_GREEN, size=18)
SUB_FONT = Font(italic=True, color="555555", size=10)
RED = PatternFill("solid", fgColor="F4CCCC")
GREEN = PatternFill("solid", fgColor="D9EAD3")
YEL = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

wb = Workbook()

# ── FIX E: SKU label hygiene helper ──────────────────────────────────────────
def clean_name(s):
    """
    Tidy a raw SKU/canonical display string. Never throws; always returns
    a non-empty string that still identifies the product.
    Rules:
      (a) Replace non-breaking spaces (\xa0) with ordinary space; collapse runs.
      (b) Remove a redundant standalone 'oil' word when 'oil' appears later.
          e.g. "Cold Pressed oil Groundnut Oil" -> "Cold Pressed Groundnut Oil"
      (c) Remove a trailing duplicated size token: "Mustard Oil 5L 5 L" -> "Mustard Oil 5L"
      (d) Normalise combo-pack notation: "2 LX2" / "1lx2" -> "1 L × 2"
      (e) Title-case with sensible unit casing (L, ml, kg).
    """
    if not s:
        return s
    try:
        # (a) non-breaking space and whitespace collapse
        s = s.replace('\xa0', ' ')
        s = re.sub(r'  +', ' ', s).strip()

        # (d) combo-pack: detect patterns like "2LX2", "2 LX2", "1lx2", "2l x2", "500mlx2", etc.
        # Normalise to "<unit> × <count>"  e.g. "1 L × 2"
        def _fmt_combo(m):
            qty   = m.group(1).strip()   # numeric quantity e.g. "2" or "500"
            unit  = m.group(2).strip()   # unit e.g. "L" or "ml"
            count = m.group(3).strip()   # pack count e.g. "2"
            # unit normalisation
            ul = unit.lower()
            if ul in ('l', 'ltr', 'litre', 'liter'):
                unit_clean = 'L'
            elif ul in ('ml', 'millilitre', 'milliliter'):
                unit_clean = 'ml'
            elif ul in ('kg', 'kilogram'):
                unit_clean = 'kg'
            elif ul in ('g', 'gm', 'gram'):
                unit_clean = 'g'
            else:
                unit_clean = unit
            return f"{qty} {unit_clean} × {count}"

        # Pattern: <number> <unit> x/X <count>   (with optional spaces around x)
        s = re.sub(
            r'\b(\d+(?:\.\d+)?)\s*(l|L|ltr|ml|ML|kg|KG|g|G)\s*[xX×]\s*(\d+)\b',
            _fmt_combo, s
        )

        # (e) Title-case the whole string first
        s = s.title()

        # (e) Fix unit casing that title() breaks: "5L" stays "5L", "500Ml"->"500ml", "1Kg"->"1kg"
        s = re.sub(r'(\d+(?:\.\d+)?)\s*Ml\b', lambda m: m.group(1) + ' ml', s)
        s = re.sub(r'(\d+(?:\.\d+)?)\s*Kg\b', lambda m: m.group(1) + ' kg', s)
        # Keep "L" uppercase after a digit (litre abbreviation)
        # title() already gives "5L" -> "5L" but "5 L" -> "5 L" (fine)

        # (b) Remove a redundant standalone 'oil' word when 'Oil' (title-cased) appears later.
        # e.g. "Cold Pressed Oil Groundnut Oil" -> "Cold Pressed Groundnut Oil"
        # Only strip the FIRST occurrence if a later one exists.
        parts = s.split()
        if parts.count('Oil') >= 2:
            first_idx = parts.index('Oil')
            parts.pop(first_idx)
            s = ' '.join(parts)

        # (c) Remove a trailing duplicated size token.
        # After title-case: "Mustard Oil 5L 5 L" or "Ricebran Oil 1L 1 L"
        # Pattern: a size token (number+unit) at end, preceded by same value in a compact form.
        # Match trailing "  <n> L" or "<n>L" pair:
        s = re.sub(
            r'(\b\d+(?:\.\d+)?\s*(?:L|ml|kg|g)\b)([\s,]+)(\d+(?:\.\d+)?)\s*(L|ml|kg|g)\s*$',
            lambda m: m.group(1) if m.group(1).replace(' ', '').lower() ==
                       (m.group(3) + m.group(4)).lower() else m.group(0),
            s
        )
        # Simpler fallback for "5L 5 L" -> strip the space-separated duplicate at end
        s = re.sub(
            r'\b(\d+(?:\.\d+)?)(L|ml|kg|g)\s+\1\s+(L|ml|kg|g)\s*$',
            lambda m: m.group(1) + m.group(2) if m.group(2).lower() == m.group(3).lower() else m.group(0),
            s
        )

        return s.strip()
    except Exception:
        # Safety net: never crash, return original
        return s

# pretty SKU label from canonical (FIX E: apply clean_name)
def label(canon):
    parts = canon.rsplit('-', 1)
    name = parts[0].replace('-', ' ').title()
    pack = parts[1].upper().replace('ML', ' ml').replace('L', ' L') if len(parts) > 1 else ''
    raw = f"{name} {pack}".strip()
    return clean_name(raw)

skus = sorted(set(r['canonical'] for r in rows))
cities_with = sorted(set(r['city'] for r in rows))
all_cities = OrderedDict()
for p in per:
    all_cities.setdefault(p['city'], 0)
    all_cities[p['city']] += len(p['rows'])
cities_without = [c for c, n in all_cities.items() if n == 0]

def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDR_FONT; cell.alignment = CEN; cell.border = BORDER

def autosize(ws, maxw=42):
    for col in ws.columns:
        L = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[L].width = min(maxw, max(10, w + 2))

def pct_fraction(v):
    """Convert source percentage points (36.3) to Excel percent fraction (0.363)."""
    if v is None:
        return None
    return round(float(v) / 100.0, 4)

# ---------- Sheet 1: Executive Summary ----------
ws = wb.active; ws.title = "Summary"
ws["A1"] = f"Jivo x {PLATFORM} - Live Pricing Intelligence"; ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")
# FIX C: parse captured_at as UTC and convert to IST (+05:30) before display
_cap_raw = summary['captured_at']
_cap_utc = datetime.datetime.fromisoformat(_cap_raw.replace('Z', '+00:00'))
_cap_ist = _cap_utc + datetime.timedelta(hours=5, minutes=30)
_cap_str = _cap_ist.strftime('%Y-%m-%d %H:%M')
ws["A2"] = f"Captured {_cap_str} IST  -  {summary['pincodes_with_jivo']}/{summary['pincodes_total']} pincodes carry Jivo  -  {summary['unique_skus']} unique SKUs  -  {summary['total_rows']} datapoints  -  scrape {summary['wall_s']}s"
ws["A2"].font = SUB_FONT; ws.merge_cells("A2:G2")

# KPI cards
kpis = [("Unique SKUs", summary['unique_skus']), ("Pincodes w/ Jivo", f"{summary['pincodes_with_jivo']}/{summary['pincodes_total']}"),
        ("Datapoints", summary['total_rows']), ("Cities w/ ZERO Jivo", len(cities_without))]
r = 4
for i, (k, v) in enumerate(kpis):
    c = 1 + i * 2
    ws.cell(row=r, column=c, value=k).font = Font(bold=True, size=10, color="555555")
    cell = ws.cell(row=r + 1, column=c, value=v); cell.font = Font(bold=True, size=20, color=JIVO_GREEN)

# FIX B: cheapest per SKU — pick min sale among IN-STOCK rows only.
# If a SKU is OOS everywhere, annotate the row "(OOS everywhere)" and skip pricing.
ws.cell(row=7, column=1, value="Cheapest pincode per SKU (in-stock, sale price)").font = Font(bold=True, size=12)
hdr = ["SKU", "City", "Pincode", "Store", "Sale Rs", "MRP Rs", "Disc %"]
for j, h in enumerate(hdr, 1):
    ws.cell(row=8, column=j, value=h)
style_header(ws, 8, len(hdr))
rr = 9
for s in skus:
    # Only consider in-stock rows (FIX B)
    cand = [x for x in rows if x['canonical'] == s and x['in_stock'] == 1]
    if not cand:
        # OOS everywhere: show annotation, no price data
        ws.cell(row=rr, column=1, value=f"{label(s)} (OOS everywhere)").border = BORDER
        for j in range(2, len(hdr) + 1):
            ws.cell(row=rr, column=j, value="-").border = BORDER
        rr += 1
        continue
    b = min(cand, key=lambda x: x['sale'])
    for j, v in enumerate([label(s), b['city'], b['pincode'], b['store_name'], b['sale'], b['mrp'], pct_fraction(b['discount_pct'])], 1):
        cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
        if j >= 5: cell.alignment = CEN
        if j == 7: cell.number_format = '0.0%'
    rr += 1
# distribution gaps
rr += 1
ws.cell(row=rr, column=1, value=f"WHITESPACE - Cities with ZERO Jivo on {PLATFORM}:").font = Font(bold=True, size=11, color="CC0000")
rr += 1
ws.cell(row=rr, column=1, value=", ".join(cities_without) if cities_without else "None - full coverage").font = Font(size=11)
ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
autosize(ws)

# ---------- Sheet 2: Master Data ----------
ws = wb.create_sheet("Master Data")
cols = ["City", "Pincode", "Locality", "Store", "SKU", "Pack", "Vol (ml)", "Sale Rs", "MRP Rs", "Disc %", "Rs/L", "ETA min", "In stock"]
ws.append(cols)
for x in sorted(rows, key=lambda r: (r['city'], r['pincode'], r['canonical'])):
    # FIX D: suppress Rs/L for ghee (weight-based product; litre metric is meaningless)
    is_ghee = 'ghee' in (x.get('sku_raw') or '').lower() or 'ghee' in (x.get('canonical') or '').lower()
    per_litre_val = None if is_ghee else x['per_litre']
    # FIX E: display clean SKU name in Master Data
    ws.append([x['city'], x['pincode'], x['locality'], x['store_name'], clean_name(x['sku_raw']), x['pack'], x['vol_ml'],
               x['sale'], x['mrp'], pct_fraction(x['discount_pct']), per_litre_val, x['eta_min'], "Yes" if x['in_stock'] else "No"])
# Keep explicitly failed collections visible instead of silently dropping their
# pincodes from Master Data. These marker rows are excluded from all SKU metrics.
for p in sorted((p for p in per if p.get('collection_status') == 'failed'), key=lambda p: (p['city'], p['pincode'])):
    ws.append([
        p['city'], p['pincode'], p['locality'], p['store_name'],
        "DATA UNAVAILABLE - COLLECTION FAILED", "", None, None, None, None,
        None, None, "Unknown",
    ])
style_header(ws)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
        if cell.column in (8, 9): cell.number_format = '"Rs"#,##0'
        if cell.column == 10: cell.number_format = '0.0%'
    sc = row[7].value
    if row[12].value == "No": row[12].fill = RED
    if row[4].value == "DATA UNAVAILABLE - COLLECTION FAILED":
        for cell in row:
            cell.fill = RED
            cell.font = Font(italic=True, color="9C0006")
    if isinstance(row[9].value, (int, float)) and row[9].value and row[9].value >= 0.40: row[9].fill = GREEN
autosize(ws)

# ---------- Matrix builder ----------
def matrix(sheet_name, valfn, fmt=None, scale=False, scale_rev=False):
    ws = wb.create_sheet(sheet_name)
    cols = ["SKU"] + cities_with
    ws.append(cols)
    for s in skus:
        rowvals = [label(s)]
        for c in cities_with:
            cand = [x for x in rows if x['canonical'] == s and x['city'] == c]
            rowvals.append(valfn(cand) if cand else None)
        ws.append(rowvals)
    style_header(ws)
    ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            if cell.column > 1:
                cell.alignment = CEN
                if fmt and isinstance(cell.value, (int, float)): cell.number_format = fmt
                if cell.value is None: cell.value = "-"
    if scale and ws.max_row >= 2 and ws.max_column >= 2:
        last = get_column_letter(ws.max_column)
        lo, hi = ("F4CCCC", "D9EAD3") if scale_rev else ("D9EAD3", "F4CCCC")
        ws.conditional_formatting.add(f"B2:{last}{ws.max_row}",
            ColorScaleRule(start_type="min", start_color=lo, end_type="max", end_color=hi))
    autosize(ws, maxw=14)
    return ws

# Sheet 3: Pricing Matrix (modal sale price per city) - green cheap -> red expensive.
# We show the MOST COMMON real shelf price in each city, NOT the mean: averaging a
# SKU sold at 485 in most pincodes and 520 in a few yields 490, a phantom price no
# store actually charges. The mode is always a price a customer can really see; ties
# break to the cheaper price (consistent with the report's cheapest-first framing).
# The full per-pincode spread (e.g. the 3 Delhi stores at 520) stays in Master Data.
#
# FIX A: only consider IN-STOCK rows for the modal price.  If a (SKU, city) has no
# in-stock rows at all, return None so the matrix renders "-" (not an OOS price).
def modal_price(cands):
    # Filter to in-stock rows only (FIX A)
    live = [x for x in cands if x['in_stock'] == 1]
    prices = [x['sale'] for x in live if x['sale'] is not None]
    if not prices: return None
    cnt = Counter(prices); top = max(cnt.values())
    return round(min(p for p, n in cnt.items() if n == top))
matrix("Pricing Matrix", modal_price, '"Rs"#,##0', scale=True)
# Sheet 4: Stock Status (% in stock) - intentionally uses ALL rows (not just in-stock)
def stock_cell(c):
    pct = round(sum(x['in_stock'] for x in c) / len(c), 4)
    return pct
wsS = matrix("Stock Status", stock_cell, '0%')
for row in wsS.iter_rows(min_row=2):
    for cell in row:
        if cell.column > 1 and isinstance(cell.value, (int, float)):
            cell.fill = GREEN if cell.value == 1 else (RED if cell.value == 0 else YEL)
# Sheet 5: Discount Analysis (modal disc per city, in-stock only) - higher = greener
def modal_disc(cands):
    live = [x for x in cands if x['in_stock'] == 1]
    vals = [x['discount_pct'] for x in live if x['discount_pct'] is not None]
    if not vals: return None
    cnt = Counter(vals); top = max(cnt.values())
    return pct_fraction(round(min(v for v, n in cnt.items() if n == top), 1))
matrix("Discount Analysis", modal_disc, '0.0%', scale=True, scale_rev=True)

# ---------- Sheet 6: Coverage / Gaps ----------
ws = wb.create_sheet("Coverage & Gaps")
ws.append(["City", "Pincode", "Locality", "Store assigned", "Jivo SKUs found"])
for p in per:
    ws.append([p['city'], p['pincode'], p['locality'], p['store_name'], len(p['rows'])])
style_header(ws); ws.freeze_panes = "A2"
for row in ws.iter_rows(min_row=2):
    for cell in row: cell.border = BORDER
    if row[4].value == 0: row[4].fill = RED
    else: row[4].fill = GREEN
autosize(ws)

fname = f"Jivo-{PLATFORM.replace(' ', '')}-Live-Report-{datetime.date.today()}.xlsx"
wb.save(fname)
print("SAVED:", fname)
print("Sheets:", wb.sheetnames)
