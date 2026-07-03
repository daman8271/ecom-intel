import json, datetime, statistics, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

d = json.load(open('result.json'))
rows = d['allRows']
per = d.get('perPin', [])
summary = d['summary']

PLATFORM = os.path.basename(os.getcwd()).replace('-', ' ').title()
HAS_SELLER = any('seller' in r for r in rows)


def gv(r, k, default=''):
    v = r.get(k, default)
    return default if v is None else v


def ist_str(cap):
    """UTC ISO timestamp -> 'YYYY-MM-DD HH:MM IST'. captured_at is stored in UTC;
    the old report sliced the raw string and mislabelled it IST."""
    try:
        dt = datetime.datetime.fromisoformat(str(cap).replace('Z', '+00:00'))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        ist = dt.astimezone(datetime.timezone(datetime.timedelta(hours=5, minutes=30)))
        return ist.strftime('%Y-%m-%d %H:%M IST')
    except Exception:
        return str(cap)


CAPTURED_IST = ist_str(summary.get('captured_at'))

# A SKU is a combo (multi-pack / bundle) if: the pack field encodes a multipack
# ("1+1 LTR", "5+1 LTR"); the text says combo/bundle; or "pack/set of N" with N>=2.
# Guarded so "Pack of 1" stays a SINGLE, and title marketing '+' isn't counted.
_COMBO_KW = re.compile(r'\b(?:combo|bundle)\b', re.I)
_COMBO_PACKN = re.compile(r'\b(?:pack|set|combo)\s+of\s+(\d+)', re.I)


def is_combo(r):
    pack = str(gv(r, 'pack'))
    both = "%s %s" % (pack, gv(r, 'sku_raw'))
    if '+' in pack or _COMBO_KW.search(both):
        return True
    m = _COMBO_PACKN.search(both)
    return bool(m and int(m.group(1)) >= 2)


def pack_type(r):
    return "Combo" if is_combo(r) else "Single"


def med(xs):
    xs = [x for x in xs if x is not None]
    return statistics.median(xs) if xs else None


def pct_fraction(v):
    if v is None:
        return None
    return round(float(v) / 100.0, 4)


def label(canon):
    parts = str(canon).rsplit('-', 1)
    name = parts[0].replace('-', ' ').title()
    pack = parts[1].upper().replace('ML', ' ml').replace('L', ' L') if len(parts) > 1 else ''
    return ("%s %s" % (name, pack)).strip()


# ---------- palette / shared styles ----------
JIVO_GREEN = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO_GREEN)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=JIVO_GREEN, size=18)
SUB_FONT = Font(italic=True, color="555555", size=10)
SECTION_FONT = Font(bold=True, color=JIVO_GREEN, size=13)
NOTE_FONT = Font(italic=True, color="777777", size=9)
KPI_LBL = Font(bold=True, size=10, color="666666")
KPI_VAL = Font(bold=True, size=22, color=JIVO_GREEN)
BOLD = Font(bold=True)
BODY = Font(size=11)
RED = PatternFill("solid", fgColor="F4CCCC")
GREEN = PatternFill("solid", fgColor="D9EAD3")
YEL = PatternFill("solid", fgColor="FFF2CC")
BAND = PatternFill("solid", fgColor="F2F8F4")
COMBO_FILL = PatternFill("solid", fgColor="DDEBF7")
KPI_BG = PatternFill("solid", fgColor="EAF4EE")
_thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
MONEY = '"₹"#,##0'
PCT = '0.0%'
PCT0 = '0%'


def title_block(ws, title, subtitle, span=8):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.row_dimensions[1].height = 26


def header_at(ws, headers, row):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = HDR; c.font = HDR_FONT; c.alignment = CEN; c.border = BORDER
    ws.row_dimensions[row].height = 26


def autosize(ws, maxw=46, minw=10):
    for col in ws.columns:
        try:
            L = get_column_letter(col[0].column)
        except Exception:
            continue
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[L].width = min(maxw, max(minw, w + 2))


wb = Workbook()

if HAS_SELLER:
    # ---------------- aggregates ----------------
    def _missing(x):
        return gv(x, 'availability_text') in ("NOT FOUND", "BLOCKED")

    n_total = len(rows)
    n_instk = sum(1 for x in rows if x['in_stock'])
    n_miss = sum(1 for x in rows if _missing(x))
    n_oos = n_total - n_instk - n_miss
    n_sellers = len(set(gv(x, 'seller') for x in rows if gv(x, 'seller')))
    combos = [x for x in rows if is_combo(x)]
    singles = [x for x in rows if not is_combo(x)]
    instk_rows = [x for x in rows if x['in_stock'] and x.get('sale') is not None]
    disc_vals = [x['discount_pct'] for x in instk_rows if x.get('discount_pct') is not None]
    avg_disc = round(statistics.mean(disc_vals), 1) if disc_vals else 0
    asins_tracked = summary.get('asins_targeted', n_total)

    # ---------------- Sheet 1: Summary (dashboard) ----------------
    ws = wb.active
    ws.title = "Summary"
    title_block(ws, "Jivo × Amazon — Live Pricing Intelligence",
                "Captured %s    ·    %s ASINs tracked    ·    national marketplace pricing    ·    scrape %ss"
                % (CAPTURED_IST, asins_tracked, summary.get('wall_s', '?')), span=8)

    kpis = [("ASINs Tracked", asins_tracked), ("In Stock", n_instk),
            ("Out of Stock", n_oos), ("Not Found / Blocked", n_miss),
            ("Distinct Retailers", n_sellers), ("Combo SKUs", len(combos)),
            ("Single SKUs", len(singles)), ("Avg Discount (in-stock)", "%s%%" % avg_disc)]
    for idx, (k, v) in enumerate(kpis):
        col = 1 + (idx % 4) * 2
        lr = 4 + (idx // 4) * 3
        lc = ws.cell(row=lr, column=col, value=k); lc.font = KPI_LBL; lc.alignment = LEFT
        vc = ws.cell(row=lr + 1, column=col, value=v); vc.font = KPI_VAL; vc.alignment = LEFT
        for cc in (col, col + 1):
            ws.cell(row=lr, column=cc).fill = KPI_BG
            ws.cell(row=lr + 1, column=cc).fill = KPI_BG

    # Highlights
    ws.cell(row=11, column=1, value="Highlights").font = SECTION_FONT
    bullets = []
    priced_instk = [x for x in instk_rows if (x.get('per_litre') or 0) > 0]
    if priced_instk:
        c = min(priced_instk, key=lambda x: x['per_litre'])
        bullets.append("Cheapest in stock: %s — ₹%d / L (%s)"
                       % (c['sku_raw'][:58], round(c['per_litre']), gv(c, 'seller') or '—'))
    if instk_rows:
        dp = max(instk_rows, key=lambda x: (x.get('discount_pct') or 0))
        if (dp.get('discount_pct') or 0) > 0:
            bullets.append("Deepest discount: %s — %.0f%% off (₹%s)"
                           % (dp['sku_raw'][:58], dp['discount_pct'],
                              ('%d' % dp['sale']) if dp.get('sale') is not None else '—'))
    bullets.append("Stock health: %d of %d in stock (%.0f%%)  ·  %d out of stock  ·  %d not found/blocked"
                   % (n_instk, n_total, (100 * n_instk / n_total) if n_total else 0, n_oos, n_miss))
    bullets.append("Pack mix: %d combo SKUs  ·  %d single SKUs" % (len(combos), len(singles)))
    by_seller_ct = defaultdict(int)
    for x in rows:
        if gv(x, 'seller'):
            by_seller_ct[gv(x, 'seller')] += 1
    if by_seller_ct:
        ts = max(by_seller_ct.items(), key=lambda kv: kv[1])
        bullets.append("Top retailer: %s (%d SKUs)" % (ts[0], ts[1]))
    r = 12
    for b in bullets:
        cell = ws.cell(row=r, column=1, value="•  " + b); cell.font = BODY; cell.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1

    # Top 10 deals
    r += 1
    ws.cell(row=r, column=1, value="Top 10 Deals (in stock, by discount)").font = SECTION_FONT
    r += 1
    header_at(ws, ["SKU / Product", "Retailer (Sold by)", "Sale ₹", "MRP ₹", "Disc %"], r)
    r += 1
    for i, x in enumerate(sorted(instk_rows, key=lambda x: -(x.get('discount_pct') or 0))[:10]):
        vals = [x['sku_raw'], gv(x, 'seller') or "—", x.get('sale'), x.get('mrp'), pct_fraction(x.get('discount_pct'))]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=j, value=v); cell.border = BORDER
            cell.alignment = LEFT if j in (1, 2) else CEN
            if j in (3, 4) and isinstance(v, (int, float)):
                cell.number_format = MONEY
            if j == 5 and isinstance(v, (int, float)):
                cell.number_format = PCT
            if i % 2:
                cell.fill = BAND
        dcell = ws.cell(row=r, column=5)
        if isinstance(dcell.value, (int, float)) and dcell.value >= 0.40:
            dcell.fill = GREEN
        r += 1
    autosize(ws)
    ws.column_dimensions['A'].width = 46

    # ---------------- Sheet 2: Master Data ----------------
    ws = wb.create_sheet("Master Data")
    title_block(ws, "Master Data — All Tracked ASINs",
                "%d ASINs  ·  captured %s  ·  sort & filter enabled (headers in row 3)"
                % (n_total, CAPTURED_IST), span=14)
    cols = ["SKU / Product", "ASIN", "Category", "Pack Type", "Retailer (Sold by)", "Ships From",
            "In Stock?", "Availability", "Units Left", "Sale Price (₹)", "MRP (₹)", "Discount %",
            "Pack Size", "Price / Litre (₹)"]
    HROW = 3
    header_at(ws, cols, HROW)
    ordered = sorted(rows, key=lambda r: (not r['in_stock'], gv(r, 'category'),
                                          (r.get('sale') is None, r.get('sale') or 0)))
    rr = HROW + 1
    for i, x in enumerate(ordered):
        pt = pack_type(x)
        instk = "Yes" if x['in_stock'] else "No"
        seller = gv(x, 'seller') or "—"
        vals = [x['sku_raw'], gv(x, 'asin'), gv(x, 'category'), pt, seller, gv(x, 'ships_from'),
                instk, gv(x, 'availability_text'), x.get('units_left'),
                x.get('sale'), x.get('mrp'), pct_fraction(x.get('discount_pct')), gv(x, 'pack'), x.get('per_litre')]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
            cell.alignment = LEFT if j in (1, 5, 8, 13) else CEN
            if j in (10, 11, 14) and isinstance(v, (int, float)):
                cell.number_format = MONEY
            if j == 12 and isinstance(v, (int, float)):
                cell.number_format = PCT
            if i % 2:
                cell.fill = BAND
        # status colours override zebra
        if instk == "No":
            ws.cell(row=rr, column=7).fill = RED
            ws.cell(row=rr, column=8).fill = RED if _missing(x) else YEL
        else:
            ws.cell(row=rr, column=7).fill = GREEN
        if pt == "Combo":
            ws.cell(row=rr, column=4).fill = COMBO_FILL
        if seller == "—":
            ws.cell(row=rr, column=5).fill = YEL
        dcell = ws.cell(row=rr, column=12)
        if isinstance(dcell.value, (int, float)) and dcell.value >= 0.40:
            dcell.fill = GREEN
        rr += 1
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:%s%d" % (get_column_letter(len(cols)), rr - 1)
    autosize(ws)
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['H'].width = 30

    # ---------------- Sheet 3: Stock Status ----------------
    ws = wb.create_sheet("Stock Status")
    title_block(ws, "Stock Status — Availability Breakdown",
                "By category and pack type  ·  captured %s" % CAPTURED_IST, span=6)
    ws.cell(row=3, column=1, value="By Category (worst availability first)").font = SECTION_FONT
    header_at(ws, ["Category", "Total ASINs", "In Stock", "Out of Stock", "Not Found / Blocked", "% In Stock"], 4)
    by_cat = defaultdict(list)
    for x in rows:
        by_cat[gv(x, 'category') or "(uncategorised)"].append(x)
    cat_rows = []
    for cat, grp in by_cat.items():
        ins = sum(1 for x in grp if x['in_stock'])
        mis = sum(1 for x in grp if _missing(x))
        cat_rows.append((cat, len(grp), ins, len(grp) - ins - mis, mis,
                         round(ins / len(grp), 4) if grp else 0))
    cat_rows.sort(key=lambda t: t[5])
    rr = 5
    for row_vals in cat_rows:
        for j, v in enumerate(row_vals, 1):
            cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
            cell.alignment = LEFT if j == 1 else CEN
            if j == 6:
                cell.number_format = PCT0
        pc = ws.cell(row=rr, column=6)
        pc.fill = GREEN if row_vals[5] >= 0.80 else (RED if row_vals[5] <= 0.20 else YEL)
        rr += 1
    first_data, last_data = 5, rr - 1
    tot = ["ALL CATEGORIES", n_total, n_instk, n_oos, n_miss, round(n_instk / n_total, 4) if n_total else 0]
    for j, v in enumerate(tot, 1):
        cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER; cell.font = BOLD
        cell.alignment = LEFT if j == 1 else CEN
        if j == 6:
            cell.number_format = PCT0
    try:
        ws.conditional_formatting.add("F%d:F%d" % (first_data, last_data),
                                      DataBarRule(start_type="num", start_value=0,
                                                  end_type="num", end_value=1, color="63BE7B"))
    except Exception:
        pass
    rr += 2

    ws.cell(row=rr, column=1, value="By Pack Type").font = SECTION_FONT
    rr += 1
    header_at(ws, ["Pack Type", "Total ASINs", "In Stock", "Out of Stock", "% In Stock"], rr)
    rr += 1
    for nm, grp in (("Combo", combos), ("Single", singles)):
        ins = sum(1 for x in grp if x['in_stock'])
        mis = sum(1 for x in grp if _missing(x))
        pct = round(ins / len(grp), 4) if grp else 0
        vals = [nm, len(grp), ins, len(grp) - ins - mis, pct]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
            cell.alignment = LEFT if j == 1 else CEN
            if j == 5:
                cell.number_format = PCT0
        ws.cell(row=rr, column=5).fill = GREEN if pct >= 0.80 else (RED if pct <= 0.20 else YEL)
        rr += 1
    rr += 1
    if cat_rows:
        w = cat_rows[0]
        ws.cell(row=rr, column=1,
                value="Note: lowest availability is %s — %d%% in stock (%d of %d ASINs)."
                % (w[0], round(w[5] * 100), w[2], w[1])).font = NOTE_FONT
    autosize(ws)

    # ---------------- Sheet 4: Retailers ----------------
    ws = wb.create_sheet("Retailers")
    title_block(ws, "Retailers — Who Sells Jivo on Amazon",
                "Buy-box seller breakdown  ·  captured %s" % CAPTURED_IST, span=6)
    header_at(ws, ["Retailer (Sold by)", "SKUs Sold", "In-Stock SKUs", "Min ₹", "Median ₹", "Max ₹"], 3)
    by_seller = defaultdict(list)
    for x in rows:
        by_seller[gv(x, 'seller') or "(no buy-box / unavailable)"].append(x)
    rr = 4
    for sname in sorted(by_seller, key=lambda s: (-len(by_seller[s]), s)):
        grp = by_seller[sname]
        sales = sorted(x['sale'] for x in grp if x.get('sale') is not None)
        vals = [sname, len(grp), sum(1 for x in grp if x['in_stock']),
                sales[0] if sales else None,
                round(statistics.median(sales)) if sales else None,
                sales[-1] if sales else None]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
            cell.alignment = LEFT if j == 1 else CEN
            if j >= 4 and isinstance(v, (int, float)):
                cell.number_format = MONEY
        if sname.startswith("(no buy-box"):
            ws.cell(row=rr, column=1).fill = YEL
        rr += 1
    ws.freeze_panes = "A4"
    autosize(ws)
    ws.column_dimensions['A'].width = 36

    # ---------------- Sheet 5: Deals & Discounts ----------------
    ws = wb.create_sheet("Deals & Discounts")
    title_block(ws, "Deals & Discounts — In-Stock SKUs by Discount",
                "Ranked by discount off listed MRP  ·  captured %s" % CAPTURED_IST, span=7)
    header_at(ws, ["SKU / Product", "Category", "Retailer (Sold by)", "Sale ₹", "MRP ₹", "Disc %", "₹ / L"], 3)
    rr = 4
    for i, x in enumerate(sorted(instk_rows, key=lambda x: (-(x.get('discount_pct') or 0), x.get('sale') or 0))):
        vals = [x['sku_raw'], gv(x, 'category'), gv(x, 'seller') or "—",
                x.get('sale'), x.get('mrp'), pct_fraction(x.get('discount_pct')), x.get('per_litre')]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
            cell.alignment = LEFT if j in (1, 2, 3) else CEN
            if j in (4, 5, 7) and isinstance(v, (int, float)):
                cell.number_format = MONEY
            if j == 6 and isinstance(v, (int, float)):
                cell.number_format = PCT
            if i % 2:
                cell.fill = BAND
        dcell = ws.cell(row=rr, column=6)
        if isinstance(dcell.value, (int, float)):
            dcell.fill = GREEN if dcell.value >= 0.40 else (RED if dcell.value <= 0.05 else YEL)
        rr += 1
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:G%d" % (rr - 1)
    autosize(ws)
    ws.column_dimensions['A'].width = 46

    # ---------------- Sheet 6: Combos vs Singles ----------------
    ws = wb.create_sheet("Combos vs Singles")
    title_block(ws, "Combos vs Singles — Pack-Type Breakdown",
                "Combo = bundle / multi-pack SKUs (pack contains '+', 'bundle', 'combo'); "
                "Single = standalone pack  ·  captured %s" % CAPTURED_IST, span=7)
    ws.cell(row=3, column=1, value="Overview").font = SECTION_FONT
    header_at(ws, ["Pack Type", "# ASINs", "In Stock", "% In Stock", "Median Sale ₹", "Median Disc %"], 4)

    def grp_stats(grp):
        ins = sum(1 for x in grp if x['in_stock'])
        pct = round(ins / len(grp), 4) if grp else 0
        msale = med([x['sale'] for x in grp if x.get('sale') is not None])
        mdisc = med([x['discount_pct'] for x in grp if x['in_stock'] and x.get('discount_pct') is not None])
        return ins, pct, msale, mdisc

    rr = 5
    for nm, grp in (("Combo", combos), ("Single", singles), ("ALL", rows)):
        ins, pct, msale, mdisc = grp_stats(grp)
        vals = [nm, len(grp), ins, pct,
                round(msale) if msale is not None else None,
                pct_fraction(round(mdisc, 1)) if mdisc is not None else None]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
            cell.alignment = LEFT if j == 1 else CEN
            if j == 4:
                cell.number_format = PCT0
            if j == 5 and isinstance(v, (int, float)):
                cell.number_format = MONEY
            if j == 6 and isinstance(v, (int, float)):
                cell.number_format = PCT
            if nm == "ALL":
                cell.font = BOLD
        rr += 1
    rr += 1

    def listing(start, heading, grp):
        ws.cell(row=start, column=1, value=heading).font = SECTION_FONT
        header_at(ws, ["SKU / Product", "Pack Size", "Retailer (Sold by)", "Sale ₹", "MRP ₹", "Disc %", "In Stock?"],
                  start + 1)
        rw = start + 2
        for i, x in enumerate(sorted(grp, key=lambda r: (not r['in_stock'], r.get('sale') or 1e9))):
            instk = "Yes" if x['in_stock'] else "No"
            vals = [x['sku_raw'], gv(x, 'pack') or "—", gv(x, 'seller') or "—",
                    x.get('sale'), x.get('mrp'), pct_fraction(x.get('discount_pct')), instk]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=rw, column=j, value=v); cell.border = BORDER
                cell.alignment = LEFT if j in (1, 2, 3) else CEN
                if j in (4, 5) and isinstance(v, (int, float)):
                    cell.number_format = MONEY
                if j == 6 and isinstance(v, (int, float)):
                    cell.number_format = PCT
                if i % 2:
                    cell.fill = BAND
            ws.cell(row=rw, column=7).fill = GREEN if instk == "Yes" else RED
            rw += 1
        return rw + 1

    nxt = listing(rr, "Combo SKUs (%d)" % len(combos), combos)
    listing(nxt, "Single SKUs (%d)" % len(singles), singles)
    autosize(ws)
    ws.column_dimensions['A'].width = 46

else:
    # ---------------- generic fallback (per-pincode platforms) ----------------
    # This file is Amazon's copy (HAS_SELLER is always true above); this branch is a
    # defensive minimum so the script never crashes if run without seller data.
    ws = wb.active
    ws.title = "Summary"
    title_block(ws, "Jivo × %s — Live Pricing Intelligence" % PLATFORM,
                "Captured %s  ·  %d rows" % (CAPTURED_IST, len(rows)), span=8)
    ws2 = wb.create_sheet("Master Data")
    header_at(ws2, ["City", "Pincode", "Store", "SKU / Product", "Pack", "Sale ₹", "MRP ₹",
                    "Disc %", "₹ / L", "In Stock?"], 1)
    for x in sorted(rows, key=lambda r: (gv(r, 'city'), gv(r, 'pincode'), gv(r, 'canonical'))):
        ws2.append([gv(x, 'city'), gv(x, 'pincode'), gv(x, 'store_name'), gv(x, 'sku_raw'), gv(x, 'pack'),
                    x.get('sale'), x.get('mrp'), pct_fraction(x.get('discount_pct')), x.get('per_litre'),
                    "Yes" if x.get('in_stock') else "No"])
    ws2.freeze_panes = "A2"
    autosize(ws2)

fname = "Jivo-%s-Live-Report-%s.xlsx" % (PLATFORM.replace(' ', ''), datetime.date.today())
wb.save(fname)
print("SAVED:", fname)
print("Sheets:", wb.sheetnames)
