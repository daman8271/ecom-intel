import json, datetime, statistics, os
from collections import defaultdict, OrderedDict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

d = json.load(open('result.json'))
rows = d['allRows']
per = d['perPin']
summary = d['summary']

# platform name derived from the folder this runs in (blinkit, zepto, ...)
PLATFORM = os.path.basename(os.getcwd()).replace('-', ' ').title()

# --- Amazon Now (ctnow surface) -----------------------------------------------
# Every published row is a GENUINE Amazon Now offer (the scraper keeps only cards
# carrying the blue Now badge). The speed-tier data (now_eta/now_slot) stays in
# result.json but is NOT displayed: owner order 2026-06-06 — "that thing is not
# our thing, kindly remove from all the sheets".

# Catalog coverage: classify the full Jivo catalog (the core Amazon scraper's
# products.json, the authoritative 314-ASIN list) into PRESENT / OUT OF STOCK /
# NOT ON NOW. Matched by ASIN (the stable key across surfaces). Best-effort: if the
# catalog file is unavailable the Catalog Coverage sheet is simply skipped.
CATALOG = {}
for cat_path in ('../amazon/products.json', 'products.json'):
    try:
        for p in json.load(open(cat_path)):
            if p.get('asin'):
                CATALOG[p['asin']] = p
        if CATALOG:
            break
    except Exception:
        continue

now_by_asin = defaultdict(list)
for r in rows:
    if r.get('asin'):
        now_by_asin[r['asin']].append(r)

def coverage_status(asin):
    rs = now_by_asin.get(asin)
    if not rs:
        return "NOT ON NOW"
    return "PRESENT" if any(x.get('in_stock') for x in rs) else "OUT OF STOCK"

cov_counts = OrderedDict([("PRESENT", 0), ("OUT OF STOCK", 0), ("NOT ON NOW", 0)])
for asin in CATALOG:
    cov_counts[coverage_status(asin)] += 1

JIVO_GREEN = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO_GREEN)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=JIVO_GREEN, size=18)
SUB_FONT = Font(italic=True, color="555555", size=10)
RED = PatternFill("solid", fgColor="F4CCCC")
GREEN = PatternFill("solid", fgColor="D9EAD3")
YEL = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

wb = Workbook()

# Display label = the REAL listing title (longest seen for the SKU), cleaned.
# Re-titling the canonical slug garbled names ("1+1"→"11 Litres", truncated
# "Mustard D" + a literal "NA" pack) and doubled the pack suffix ("1 Litre 1 L")
# — fresh-eyes 2026-06-06 MUST-3. Slug titling stays only as a fallback.
RAW_BY_CANON = {}
for _r in rows:
    _raw = (_r.get('sku_raw') or '').strip().rstrip(',-– ').strip()
    if _raw and len(_raw) > len(RAW_BY_CANON.get(_r['canonical'], '')):
        RAW_BY_CANON[_r['canonical']] = _raw

def label(canon, maxlen=60):
    raw = RAW_BY_CANON.get(canon)
    if raw:
        return raw if len(raw) <= maxlen else raw[:maxlen - 1].rstrip() + "…"
    parts = canon.rsplit('-', 1)
    name = parts[0].replace('-', ' ').title()
    pack = parts[1].upper().replace('ML', ' ml').replace('L', ' L') if len(parts) > 1 else ''
    return f"{name} {pack}".strip()

skus = sorted(set(r['canonical'] for r in rows))
cities_with = sorted(set(r['city'] for r in rows))
all_cities = OrderedDict()
for p in per:
    all_cities.setdefault(p['city'], 0)
    all_cities[p['city']] += len(p['rows'])
cities_without = [c for c, n in all_cities.items() if n == 0]

def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDR_FONT; cell.alignment = CEN; cell.border = BORDER

def autosize(ws, maxw=42):
    for col in ws.columns:
        L = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[L].width = min(maxw, max(10, w + 2))

# ---------- Sheet 1: Executive Summary ----------
ws = wb.active; ws.title = "Summary"
ws["A1"] = f"Jivo x {PLATFORM} - Live Pricing Intelligence"; ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")
# ONE honest IST timestamp (captured_at is UTC "Z"; the old slice printed UTC
# labelled IST — the 07:24-vs-12:54 confusion). No scrape-seconds (internal).
_utc = datetime.datetime.fromisoformat(summary['captured_at'].replace('Z', '+00:00'))
CAPTURED_IST = (_utc + datetime.timedelta(hours=5, minutes=30)).strftime('%Y-%m-%d %H:%M')
ws["A2"] = f"Captured {CAPTURED_IST} IST  -  {summary['pincodes_with_jivo']}/{summary['pincodes_total']} pincodes carry Jivo  -  {summary['unique_skus']} unique SKUs  -  {summary['total_rows']} datapoints"
ws["A2"].font = SUB_FONT; ws.merge_cells("A2:G2")
# Source caption: plain words, no URL parameters (fresh-eyes MUST-6).
ws["A3"] = "Source: Amazon Now storefront."
ws["A3"].font = SUB_FONT; ws.merge_cells("A3:G3")

# KPI cards
kpis = [("Unique SKUs", summary['unique_skus']), ("Pincodes w/ Jivo", f"{summary['pincodes_with_jivo']}/{summary['pincodes_total']}"),
        ("Datapoints", summary['total_rows']), ("Cities w/ ZERO Jivo", len(cities_without))]
r = 4
for i, (k, v) in enumerate(kpis):
    c = 1 + i * 2
    ws.cell(row=r, column=c, value=k).font = Font(bold=True, size=10, color="555555")
    cell = ws.cell(row=r + 1, column=c, value=v); cell.font = Font(bold=True, size=20, color=JIVO_GREEN)
if cities_without:
    nz = ws.cell(row=6, column=1, value="Zero-Jivo cities: " + ", ".join(cities_without))
    nz.font = Font(italic=True, size=9, color="CC0000")
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=8)

# cheapest per SKU
ws.cell(row=7, column=1, value="Cheapest pincode per SKU (in-stock, sale price)").font = Font(bold=True, size=12)
hdr = ["SKU", "City", "Pincode", "Store", "Sale ₹", "MRP ₹", "Disc %"]
for j, h in enumerate(hdr, 1):
    ws.cell(row=8, column=j, value=h)
style_header(ws, 8, len(hdr))
rr = 9
for s in skus:
    cand = [x for x in rows if x['canonical'] == s and x['in_stock'] == 1]
    if not cand: cand = [x for x in rows if x['canonical'] == s]
    if not cand: continue
    cand = [x for x in cand if x.get('sale') is not None]
    if not cand: continue
    b = min(cand, key=lambda x: x['sale'])
    disc = b['discount_pct'] / 100 if b.get('discount_pct') is not None else None
    for j, v in enumerate([label(s), b['city'], b['pincode'], b['store_name'], b['sale'], b['mrp'], disc], 1):
        cell = ws.cell(row=rr, column=j, value=v); cell.border = BORDER
        if j >= 5: cell.alignment = CEN
        if j in (5, 6): cell.number_format = '"₹"#,##0'
        if j == 7: cell.number_format = '0.0%'
    rr += 1
# distribution gaps
rr += 1
ws.cell(row=rr, column=1, value=f"WHITESPACE - Cities with ZERO Jivo on {PLATFORM}:").font = Font(bold=True, size=11, color="CC0000")
rr += 1
ws.cell(row=rr, column=1, value=", ".join(cities_without) if cities_without else "None - full coverage").font = Font(size=11)
ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
autosize(ws)

# ---------- Sheet 2: Master Data ----------
# Amazon Now-specific: show ASIN (the stable cross-surface key). No slot/speed-tier
# column (owner order 2026-06-06).
ws = wb.create_sheet("Master Data")
cols = ["City", "Pincode", "Locality", "ASIN", "SKU", "Pack", "Vol (ml)", "Sale ₹", "MRP ₹", "Disc %", "₹/L", "In stock"]
ws.append(cols)
for x in sorted(rows, key=lambda r: (r['city'], r['pincode'], r['canonical'])):
    ws.append([x['city'], x['pincode'], x['locality'], x.get('asin'), x['sku_raw'], x['pack'], x['vol_ml'],
               x['sale'], x['mrp'],
               (x['discount_pct'] / 100 if x.get('discount_pct') is not None else None),
               x['per_litre'], "Yes" if x['in_stock'] else "No"])
style_header(ws)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = BORDER
        if cell.column in (8, 9, 11): cell.number_format = '"₹"#,##0'
        if cell.column == 10: cell.number_format = '0.0%'  # fraction + true % (never literal-% on a ×100 value)
    if row[11].value == "No": row[11].fill = RED
    if isinstance(row[9].value, (int, float)) and row[9].value and row[9].value >= 0.4: row[9].fill = GREEN
autosize(ws)

# ---------- Matrix builder ----------
def matrix(sheet_name, valfn, fmt=None, scale=False, scale_rev=False):
    ws = wb.create_sheet(sheet_name)
    cols = ["SKU"] + cities_with
    ws.append(cols)
    for s in skus:
        rowvals = [label(s)]
        for c in cities_with:
            cand = [x for x in rows if x['canonical'] == s and x['city'] == c]
            rowvals.append(valfn(cand) if cand else None)
        ws.append(rowvals)
    style_header(ws)
    ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            if cell.column > 1:
                cell.alignment = CEN
                if fmt and isinstance(cell.value, (int, float)): cell.number_format = fmt
                if cell.value is None: cell.value = "-"
    if scale:
        last = get_column_letter(ws.max_column)
        lo, hi = ("F4CCCC", "D9EAD3") if scale_rev else ("D9EAD3", "F4CCCC")
        ws.conditional_formatting.add(f"B2:{last}{ws.max_row}",
            ColorScaleRule(start_type="min", start_color=lo, end_type="max", end_color=hi))
    autosize(ws, maxw=14)
    ws.column_dimensions["A"].width = 46   # full product names readable (fresh-eyes MUST-2)
    return ws

# Sheet 3: Pricing Matrix (modal sale per city) - green cheap -> red expensive
def modal_price(cands):
    prices = [x['sale'] for x in cands if x['sale'] is not None]
    if not prices: return None
    cnt = Counter(prices); top = max(cnt.values())
    return round(min(p for p, n in cnt.items() if n == top))
matrix("Pricing Matrix", modal_price, '"₹"#,##0', scale=True)
# Sheet 4: Stock Status (share in stock; FRACTION + true % format)
def stock_cell(c):
    return round(sum(x['in_stock'] for x in c) / len(c), 2)
wsS = matrix("Stock Status", stock_cell, '0%')
n_full = 0; n_cells = 0
for row in wsS.iter_rows(min_row=2):
    for cell in row:
        if cell.column > 1 and isinstance(cell.value, (int, float)):
            n_cells += 1
            if cell.value == 1: n_full += 1
            cell.fill = GREEN if cell.value == 1 else (RED if cell.value == 0 else YEL)
# one-line takeaway above the grid (fresh-eyes S12: the wall of green is one bit)
wsS.insert_rows(1)
wsS["A1"] = (f"All {len(skus)} SKUs 100% in stock in every covered city."
             if n_full == n_cells else
             f"{n_full} of {n_cells} SKU×city cells fully in stock — yellow/red cells below need attention.")
wsS["A1"].font = Font(bold=True, size=11, color=JIVO_GREEN if n_full == n_cells else "CC0000")
wsS.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(7, wsS.max_column))
wsS.freeze_panes = "B3"
# Sheet 5: Discount Analysis (avg disc per city) — deep discount is a BRAND
# COMPLIANCE RISK, not a win: deepest discount = RED (fresh-eyes S14).
matrix("Discount Analysis", lambda c: round(statistics.mean([x['discount_pct'] / 100 for x in c if x['discount_pct'] is not None]), 3) if any(x['discount_pct'] is not None for x in c) else None, '0.0%', scale=True)

# ---------- Sheet 6: Now Serviceability & Coverage ----------
# For Amazon Now the key per-pincode facts are: is Now serviceable here at all, and
# does Jivo appear (vs only competitors). Three states: green=Jivo on Now,
# yellow=Now serviceable but NO Jivo (competitor-only whitespace), red=no Now here.
ws = wb.create_sheet("Now Serviceability")
n_svc = sum(1 for p in per if p.get('serviceable', len(p['rows']) > 0))
ws["A1"] = f"{n_svc} of {len(per)} pincodes serviceable on Now · {summary['pincodes_with_jivo']} carry Jivo"
ws["A1"].font = Font(bold=True, size=11, color=JIVO_GREEN)
ws.merge_cells("A1:E1")
ws.append(["City", "Pincode", "Locality", "Now serviceable", "Jivo SKUs on Now"])
# serviceable first (fresh-eyes: alphabetical put the red wall on screen one)
for p in sorted(per, key=lambda p: (not p.get('serviceable', len(p['rows']) > 0),
                                    -len(p['rows']), p['city'], p['pincode'])):
    svc = p.get('serviceable', len(p['rows']) > 0)
    ws.append([p['city'], p['pincode'], p['locality'], "Yes" if svc else "No", len(p['rows'])])
style_header(ws, 2); ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:E{ws.max_row}"
for row in ws.iter_rows(min_row=3):
    for cell in row: cell.border = BORDER
    njivo = row[4].value
    svc = row[3].value == "Yes"
    row[4].fill = GREEN if njivo else (YEL if svc else RED)
    row[3].fill = GREEN if svc else RED
autosize(ws)

# (the former "Now Speed Tiers" sheet was REMOVED — owner order 2026-06-06)

# ---------- Sheet 7: Catalog Coverage (PRESENT / OUT OF STOCK / NOT ON NOW) ----------
# The honest answer to "how much of the Jivo catalog is actually on Amazon Now?".
# Classifies the full 314-SKU core catalog by ASIN. NOT ON NOW is by Amazon's design
# (the Now storefront indexes only a subset), not a scrape gap.
if CATALOG:
    ws = wb.create_sheet("Catalog Coverage")
    # title + header + data — no blank-band sandwich, header styled LAST so no
    # data fill can bleed into it (fresh-eyes S11).
    ws["A1"] = f"Jivo catalog on Amazon Now — {cov_counts['PRESENT']} present · {cov_counts['OUT OF STOCK']} out of stock · {cov_counts['NOT ON NOW']} not on Now (of {len(CATALOG)} catalog SKUs)"
    ws["A1"].font = Font(bold=True, size=11, color=JIVO_GREEN); ws.merge_cells("A1:E1")
    hdr_row = 2
    ws.append(["ASIN", "Catalog name", "Category", "Status", "Now sale ₹ (min)"])
    STATUS_ORDER = {"PRESENT": 0, "OUT OF STOCK": 1, "NOT ON NOW": 2}
    for asin in sorted(CATALOG, key=lambda a: (STATUS_ORDER[coverage_status(a)], CATALOG[a].get('category') or '', CATALOG[a].get('name') or '')):
        p = CATALOG[asin]
        st = coverage_status(asin)
        rs = now_by_asin.get(asin, [])
        sales = [x['sale'] for x in rs if x.get('sale') is not None]
        ws.append([asin, (p.get('name') or '')[:60], p.get('category') or p.get('item') or '',
                   st, (min(sales) if sales else None)])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:E{ws.max_row}"
    for row in ws.iter_rows(min_row=hdr_row + 1):
        for cell in row:
            cell.border = BORDER
            if cell.column in (4, 5):
                cell.alignment = CEN
            if cell.column == 5: cell.number_format = '"₹"#,##0'
        sc = row[3].value
        row[3].fill = GREEN if sc == "PRESENT" else (YEL if sc == "OUT OF STOCK" else RED)
    style_header(ws, hdr_row, 5)
    autosize(ws)

fname = f"Jivo-{PLATFORM.replace(' ', '')}-Live-Report-{datetime.date.today()}.xlsx"
wb.save(fname)
print("SAVED:", fname)
print("Sheets:", wb.sheetnames)
