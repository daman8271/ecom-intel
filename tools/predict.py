#!/usr/bin/env python3
"""
predict.py <platform> <xlsx_path>

Standalone, platform-agnostic "Predictions" page generator. Opens an
already-built workbook and APPENDS a worksheet named "Predictions",
generated DETERMINISTICALLY (no LLM, no network) from:
    data/<platform>/history.csv      (per-SKU-per-location run history)
    platforms/<platform>/result.json (current run; perPin or allRows shape)

Sections produced:
    - WHAT TO WATCH        : top data-driven call-outs
    - STOCK-OUT RISK       : current OOS + (with history) trending-to-stockout
    - PRICE & DISCOUNT MOVE : per-SKU change vs previous run(s); big-move flags
    - COVERAGE TREND        : pincodes/stores carrying Jivo over time + gaps

Handles thin history (1 run) gracefully -> current-state insights + a note that
trend predictions sharpen as history accumulates. Never crashes: a missing or
short history.csv still yields a useful current-state sheet.

Python 3 stdlib + openpyxl only. Run AFTER build_excel.py, BEFORE delivery.

Usage:
    python3 tools/predict.py blinkit platforms/blinkit/Jivo-Blinkit-...xlsx
"""

import sys, os, csv, json, datetime
from collections import defaultdict, OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styling
# Mirror build_excel.py so the appended sheet matches the rest of the book.
JIVO_GREEN = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO_GREEN)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=JIVO_GREEN, size=18)
SUB_FONT = Font(italic=True, color="555555", size=10)
SECTION_FONT = Font(bold=True, color=JIVO_GREEN, size=13)
NOTE_FONT = Font(italic=True, color="777777", size=9)
RED = PatternFill("solid", fgColor="F4CCCC")
GREEN = PatternFill("solid", fgColor="D9EAD3")
YEL = PatternFill("solid", fgColor="FFF2CC")
_thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Thresholds for "big move" flagging.
PRICE_PCT_BIG = 5.0      # >=5% price change is notable
PRICE_ABS_BIG = 10.0     # ... or >=Rs.10 absolute
DISC_PT_BIG = 5.0        # >=5 percentage-point discount change is notable
LOW_STOCK_FRAC = 0.5     # SKU in stock in <50% of carrying stores = "low"


# ---------------------------------------------------------------- helpers
def _f(v):
    """Best-effort float; None/'' /junk -> None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def label(canon):
    """Pretty SKU label from a canonical slug (matches build_excel.py)."""
    if not canon:
        return "?"
    parts = str(canon).rsplit("-", 1)
    name = parts[0].replace("-", " ").title()
    pack = ""
    if len(parts) > 1:
        pack = parts[1].upper().replace("ML", " ml").replace("L", " L")
    out = f"{name} {pack}".strip()
    return out if out else str(canon)


def in_stock_of(row):
    """
    Robustly read stock state from a result.json row across platform shapes
    and future extra fields. Returns True / False / None (unknown).
    Recognised: in_stock, availability, available, units_left/stock_qty, eta.
    """
    if "in_stock" in row and row["in_stock"] is not None:
        v = row["in_stock"]
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)):
            return v > 0
        s = str(v).strip().lower()
        if s in ("1", "true", "yes", "in stock", "instock", "available"):
            return True
        if s in ("0", "false", "no", "out of stock", "oos", "unavailable"):
            return False
    # availability / available string or bool (Amazon may add this)
    for key in ("availability", "available"):
        if key in row and row[key] is not None:
            v = row[key]
            if isinstance(v, bool):
                return v
            s = str(v).strip().lower()
            if s in ("in stock", "instock", "available", "yes", "true", "1"):
                return True
            if s in ("out of stock", "oos", "unavailable", "no", "false", "0"):
                return False
    # explicit unit counts
    for key in ("units_left", "stock_qty", "qty", "stock"):
        if key in row and row[key] is not None:
            n = _f(row[key])
            if n is not None:
                return n > 0
    return None


def rows_from_result(d):
    """Flatten result.json to a row list, supporting allRows and perPin shapes."""
    if not isinstance(d, dict):
        return []
    if d.get("allRows"):
        return list(d["allRows"])
    out = []
    for p in d.get("perPin", []) or []:
        for r in p.get("rows", []) or []:
            out.append(r)
    return out


def loc_key(row):
    """Stable location key for a result/history row."""
    pin = str(row.get("pincode") or "").strip()
    city = str(row.get("city") or "").strip()
    if pin and pin != "-":
        return pin
    return city or "-"


def fmt_money(v):
    if v is None:
        return "-"
    return f"Rs.{v:,.0f}" if abs(v - round(v)) < 0.05 else f"Rs.{v:,.2f}"


def fmt_signed_money(v):
    if v is None:
        return "-"
    sign = "+" if v >= 0 else "-"
    return f"{sign}Rs.{abs(v):,.0f}"


def fmt_pct(v, signed=False):
    if v is None:
        return "-"
    if signed:
        return f"{'+' if v >= 0 else ''}{v:.1f}%"
    return f"{v:.1f}%"


# ---------------------------------------------------------------- load data
def load_history(path):
    """
    Read history.csv -> list of dict rows (typed). Returns (rows, runs_sorted).
    runs_sorted: list of run_ids in chronological order (sorted by run_id;
    run_id format YYYY-MM-DD-HHMM sorts correctly as a string).
    Tolerates missing file / malformed rows.
    """
    if not os.path.exists(path):
        return [], []
    rows = []
    try:
        with open(path, newline="", encoding="utf-8") as fh:
            for r in csv.DictReader(fh):
                rows.append(r)
    except Exception:
        return [], []
    runs = sorted({r.get("run_id", "") for r in rows if r.get("run_id")})
    return rows, runs


def load_result(path):
    if not os.path.exists(path):
        return {}, []
    try:
        d = json.load(open(path, encoding="utf-8"))
    except Exception:
        return {}, []
    return d, rows_from_result(d)


# ---------------------------------------------------------------- sheet writer
class Sheet:
    """Thin cursor over a worksheet with section/table/note helpers."""

    def __init__(self, ws):
        self.ws = ws
        self.r = 1
        self.maxcol = 1

    def title(self, text, sub=None):
        c = self.ws.cell(row=self.r, column=1, value=text)
        c.font = TITLE_FONT
        self.r += 1
        if sub:
            s = self.ws.cell(row=self.r, column=1, value=sub)
            s.font = SUB_FONT
            self.r += 1
        self.r += 1  # blank spacer

    def section(self, text, sub=None):
        c = self.ws.cell(row=self.r, column=1, value=text)
        c.font = SECTION_FONT
        self.r += 1
        if sub:
            s = self.ws.cell(row=self.r, column=1, value=sub)
            s.font = NOTE_FONT
            self.r += 1

    def note(self, text):
        c = self.ws.cell(row=self.r, column=1, value=text)
        c.font = NOTE_FONT
        c.alignment = LEFT
        self.r += 1

    def bullet(self, text, fill=None):
        c = self.ws.cell(row=self.r, column=1, value=u"• " + text)
        c.font = Font(size=11)
        c.alignment = LEFT
        if fill:
            c.fill = fill
        self.r += 1

    def blank(self, n=1):
        self.r += n

    def table(self, headers, data_rows, fills=None):
        """
        headers: list[str]; data_rows: list[list]; fills: optional list (one
        PatternFill-or-None per data row) to tint the whole row.
        """
        ncol = len(headers)
        self.maxcol = max(self.maxcol, ncol)
        for j, h in enumerate(headers, start=1):
            cell = self.ws.cell(row=self.r, column=j, value=h)
            cell.fill = HDR
            cell.font = HDR_FONT
            cell.alignment = CEN
            cell.border = BORDER
        self.r += 1
        for i, drow in enumerate(data_rows):
            rowfill = fills[i] if (fills and i < len(fills)) else None
            for j in range(ncol):
                val = drow[j] if j < len(drow) else ""
                cell = self.ws.cell(row=self.r, column=j + 1, value=val)
                cell.border = BORDER
                cell.alignment = CEN if j > 0 else LEFT
                if rowfill:
                    cell.fill = rowfill
            self.r += 1
        self.r += 1  # spacer after table

    def autosize(self, maxw=46):
        ws = self.ws
        for col in ws.columns:
            try:
                L = get_column_letter(col[0].column)
            except Exception:
                continue
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[L].width = min(maxw, max(12, w + 2))
        # keep the first (label/bullet) column generous
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 40)


# ---------------------------------------------------------------- analysis
def analyze_current(result_rows):
    """
    Current-run aggregates per SKU. Returns dict:
      sku -> {label, n_loc, n_in, n_oos, n_unknown, oos_cities, prices, discs,
              min_price, max_price, med_price, med_disc}
    plus a top-level 'oos_locations' list (loc-level OOS, capped).
    """
    per_sku = OrderedDict()
    oos_loc_events = []  # (sku, city, pincode)
    for row in result_rows:
        sku = row.get("canonical") or row.get("sku_raw") or "?"
        e = per_sku.setdefault(sku, {
            "label": label(sku), "n_loc": 0, "n_in": 0, "n_oos": 0,
            "n_unknown": 0, "oos_cities": set(), "prices": [], "discs": [],
        })
        e["n_loc"] += 1
        st = in_stock_of(row)
        price = _f(row.get("sale"))
        if price is None:
            price = _f(row.get("price"))
        disc = _f(row.get("discount_pct"))
        if st is True:
            e["n_in"] += 1
            if price is not None:
                e["prices"].append(price)
            if disc is not None:
                e["discs"].append(disc)
        elif st is False:
            e["n_oos"] += 1
            city = row.get("city") or "?"
            e["oos_cities"].add(city)
            oos_loc_events.append((sku, city, str(row.get("pincode") or "")))
        else:
            e["n_unknown"] += 1
            # unknown stock: still use price/disc for outlier detection
            if price is not None:
                e["prices"].append(price)
            if disc is not None:
                e["discs"].append(disc)
    # derive stats
    import statistics
    for e in per_sku.values():
        ps = e["prices"]
        e["min_price"] = min(ps) if ps else None
        e["max_price"] = max(ps) if ps else None
        e["med_price"] = statistics.median(ps) if ps else None
        e["med_disc"] = statistics.median(e["discs"]) if e["discs"] else None
    return per_sku, oos_loc_events


def run_date(history_rows, run_id):
    for r in history_rows:
        if r.get("run_id") == run_id:
            return r.get("date_ist") or run_id
    return run_id


def per_run_sku_stats(history_rows):
    """
    Build {run_id: {sku: {n, n_in, n_oos, prices[], discs[]}}} from history.
    """
    out = defaultdict(lambda: defaultdict(lambda: {
        "n": 0, "n_in": 0, "n_oos": 0, "prices": [], "discs": [], "discs_in": []}))
    for r in history_rows:
        run = r.get("run_id")
        sku = r.get("canonical_sku")
        if not run or not sku:
            continue
        s = out[run][sku]
        s["n"] += 1
        st = _f(r.get("in_stock"))
        in_stk = st is not None and st > 0
        if st is not None:
            if st > 0:
                s["n_in"] += 1
            else:
                s["n_oos"] += 1
        p = _f(r.get("price"))
        if p is not None and p > 0:
            s["prices"].append(p)
        dsc = _f(r.get("discount_pct"))
        if dsc is not None:
            s["discs"].append(dsc)
            if in_stk:  # in-stock-only discount, to match the current-run KPI basis
                s["discs_in"].append(dsc)
    return out


def median(xs):
    import statistics
    return statistics.median(xs) if xs else None


# ---------------------------------------------------------------- dashboard
def build_dashboard(wb, platform, when, result_rows, per_sku_cur,
                    runstats, runs, history_rows, watch):
    """
    Insert a chart-driven 'Leadership View' as the FIRST sheet: the handful of
    things the e-com head needs at a glance and nothing else --
        - catalog availability (in stock vs out)        : donut
        - discount depth / price-erosion exposure        : bar
        - buy-box control (who sells Jivo)               : bar  (Amazon)
          -> falls back to "where Jivo is live by city"        (per-pincode)
        - availability & discount trend over recent runs : line
    plus a KPI strip and the top things to act on.

    Chart source data is written into a clearly-labelled block lower in the same
    sheet; the charts reference it. Platform-agnostic and defensive -- the caller
    wraps this so a hiccup never costs us the Predictions sheet.
    """
    from openpyxl.chart import DoughnutChart, BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties

    if not result_rows:
        return

    pname = platform.replace("-", " ").title()
    HAS_SELLER = any("seller" in r for r in result_rows)
    GREEN, AMBER, REDC = "008B3A", "E69138", "CC0000"

    # ---- availability breakdown (in / out / not-found-or-blocked) ----
    n_in = n_oos = n_nf = 0
    for r in result_rows:
        at = str(r.get("availability_text") or "").upper()
        if at in ("NOT FOUND", "BLOCKED"):
            n_nf += 1
            continue
        st = in_stock_of(r)
        if st is True:
            n_in += 1
        elif st is False:
            n_oos += 1
        else:
            n_nf += 1
    n_total = len(result_rows)
    instock_pct = (100.0 * n_in / (n_in + n_oos)) if (n_in + n_oos) else 0.0

    # ---- discount distribution (in-stock, priced SKUs) ----
    bands = [("0-20% off", 0, 20), ("20-40% off", 20, 40),
             ("40-60% off", 40, 60), ("60%+ off", 60, 1e9)]
    band_counts = [0] * len(bands)
    discs = []
    for r in result_rows:
        if in_stock_of(r) is not True:
            continue
        d = _f(r.get("discount_pct"))
        if d is None:
            continue
        discs.append(d)
        for i, (_, lo, hi) in enumerate(bands):
            if lo <= d < hi:
                band_counts[i] += 1
                break
    avg_disc = (sum(discs) / len(discs)) if discs else 0.0
    deep = sum(1 for d in discs if d >= 50)
    n_risk = sum(1 for e in per_sku_cur.values() if e["n_oos"] > 0)

    # ---- control chart: buy-box (Amazon) or coverage by city ----
    if HAS_SELLER:
        sc = {}
        for r in result_rows:
            s = (r.get("seller") or "").strip() or "No buy-box / unavailable"
            sc[s] = sc.get(s, 0) + 1
        ranked = sorted(sc.items(), key=lambda kv: -kv[1])
        control_title = "Who controls the buy-box (SKUs sold by)"
        control_rows = ranked[:6]
        if len(ranked) > 6:
            control_rows = control_rows + [("Other", sum(n for _, n in ranked[6:]))]
    else:
        cc = {}
        for r in result_rows:
            if in_stock_of(r) is False:
                continue
            city = (str(r.get("city") or "?").strip()) or "?"
            cc[city] = cc.get(city, 0) + 1
        ranked = sorted(cc.items(), key=lambda kv: -kv[1])
        control_title = "Where Jivo is live (in-stock datapoints by city)"
        control_rows = ranked[:8]

    # ---- availability & discount trend over the recent runs ----
    WIN = 14
    window = runs[-WIN:] if runs else []
    trend_rows = []
    for run in window:
        rs = runstats.get(run, {})
        tin = tk = 0
        dd = []
        for s in rs.values():
            tin += s["n_in"]
            tk += s["n_in"] + s["n_oos"]
            dd += s["discs_in"]   # in-stock-only, matches the headline KPI basis
        ip = (100.0 * tin / tk) if tk else None
        ad = (sum(dd) / len(dd)) if dd else None
        rid = str(run)            # run_id format: YYYY-MM-DD-HHMM
        try:                      # -> 'MM-DD HH:MM' so the 3 daily runs stay distinct
            dlabel = "%s %s:%s" % (rid[5:10], rid[11:13], rid[13:15])
        except Exception:
            dlabel = rid
        trend_rows.append([dlabel,
                           round(ip, 1) if ip is not None else None,
                           round(ad, 1) if ad is not None else None])
    have_trend = len([t for t in trend_rows if t[1] is not None]) >= 2

    # =========================== write the sheet ===========================
    ws = wb.create_sheet("Leadership View", 0)
    ws.sheet_view.showGridLines = False

    t = ws.cell(1, 1, "Jivo x %s - Leadership View" % pname)
    t.font = Font(bold=True, color=GREEN, size=20)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    headline = ("%.0f%% of catalog in stock    .    avg %.0f%% off    .    "
                "%d SKU(s) at stock-out risk    .    %s"
                % (instock_pct, avg_disc, n_risk, when))
    h = ws.cell(2, 1, headline)
    h.font = Font(italic=True, color="555555", size=11)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    # KPI strip
    kpis = [("Catalog SKUs", n_total),
            ("In Stock", "%.0f%%" % instock_pct),
            ("Avg Discount", "%.0f%%" % avg_disc),
            ("Deep Discounts (>=50%)", deep),
            ("At Stock-out Risk", n_risk)]
    for i, (k, v) in enumerate(kpis):
        c = 1 + i * 2
        ws.cell(4, c, k).font = Font(bold=True, size=9, color="666666")
        ws.cell(5, c, v).font = Font(bold=True, size=22, color=GREEN)

    # what to act on (top 5 from the prioritised watch list)
    ws.cell(7, 1, "What to act on").font = Font(bold=True, color=GREEN, size=13)
    best = {}
    for prio, msg in watch:
        if msg not in best or prio > best[msg]:
            best[msg] = prio
    ordered = [m for m, _ in sorted(best.items(), key=lambda kv: -kv[1])[:5]]
    rr = 8
    if ordered:
        for msg in ordered:
            cell = ws.cell(rr, 1, u"• " + msg)
            cell.font = Font(size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            rr += 1
    else:
        ws.cell(rr, 1, u"• Stock, pricing and coverage all steady this run.").font = Font(size=11)

    # ---- chart source-data block (charts reference these cells) ----
    DATA0 = 46
    ws.cell(DATA0 - 1, 1, "Chart source data (auto-generated)").font = NOTE_FONT

    def put(top, headers, drows):
        for j, hh in enumerate(headers, 1):
            ws.cell(top, j, hh).font = Font(bold=True, size=9, color="888888")
        for i, rw in enumerate(drows):
            for j, v in enumerate(rw, 1):
                ws.cell(top + 1 + i, j, v)
        return top, top + len(drows)

    stk_rows = [["In Stock", n_in], ["Out of Stock", n_oos]]
    if n_nf > 0 or HAS_SELLER:
        stk_rows.append(["Not Found / Blocked", n_nf])
    s_hdr, s_last = put(DATA0, ["Availability", "SKUs"], stk_rows)

    b_hdr, b_last = put(s_last + 2, ["Discount band", "SKUs"],
                        [[bands[i][0], band_counts[i]] for i in range(len(bands))])

    c_hdr, c_last = put(b_last + 2, ["Label", "SKUs"],
                        [[k, v] for k, v in control_rows])

    if have_trend:
        t_hdr, t_last = put(c_last + 2, ["Run", "In-stock %", "Avg disc %"], trend_rows)

    def solid(color):
        return GraphicalProperties(solidFill=color)

    # 1) availability donut
    d = DoughnutChart()
    d.title = "Catalog availability"
    d.add_data(Reference(ws, min_col=2, min_row=s_hdr, max_row=s_last),
               titles_from_data=True)
    d.set_categories(Reference(ws, min_col=1, min_row=s_hdr + 1, max_row=s_last))
    d.dataLabels = DataLabelList()
    d.dataLabels.showPercent = True
    palette = [GREEN, REDC, AMBER]
    ser = d.series[0]
    for i in range(s_last - s_hdr):
        pt = DataPoint(idx=i)
        pt.graphicalProperties = solid(palette[i % len(palette)])
        ser.data_points.append(pt)
    d.height, d.width = 7.2, 11.5
    ws.add_chart(d, "A14")

    # 2) discount-depth bar
    b = BarChart()
    b.type = "col"
    b.title = "Discount depth (in-stock SKUs)"
    b.add_data(Reference(ws, min_col=2, min_row=b_hdr, max_row=b_last),
               titles_from_data=True)
    b.set_categories(Reference(ws, min_col=1, min_row=b_hdr + 1, max_row=b_last))
    b.series[0].graphicalProperties = solid(GREEN)
    b.legend = None
    b.dataLabels = DataLabelList()
    b.dataLabels.showVal = True
    b.height, b.width = 7.2, 11.5
    ws.add_chart(b, "G14")

    # 3) buy-box / coverage bar (horizontal -> long labels read well)
    c = BarChart()
    c.type = "bar"
    c.title = control_title
    c.add_data(Reference(ws, min_col=2, min_row=c_hdr, max_row=c_last),
               titles_from_data=True)
    c.set_categories(Reference(ws, min_col=1, min_row=c_hdr + 1, max_row=c_last))
    c.series[0].graphicalProperties = solid(GREEN)
    c.legend = None
    c.dataLabels = DataLabelList()
    c.dataLabels.showVal = True
    c.height, c.width = 7.2, 11.5
    ws.add_chart(c, "A29")

    # 4) availability & discount trend line
    if have_trend:
        ln = LineChart()
        ln.title = "Availability & discount trend (recent runs)"
        ln.add_data(Reference(ws, min_col=2, max_col=3, min_row=t_hdr, max_row=t_last),
                    titles_from_data=True)
        ln.set_categories(Reference(ws, min_col=1, min_row=t_hdr + 1, max_row=t_last))
        line_cols = [GREEN, AMBER]
        for i, s in enumerate(ln.series):
            gp = GraphicalProperties()
            gp.line = LineProperties(solidFill=line_cols[i % 2], w=28000)
            s.graphicalProperties = gp
            s.marker = Marker(symbol="circle", size=5)
            s.smooth = False
        ln.y_axis.title = "%"
        ln.height, ln.width = 7.2, 11.5
        ws.add_chart(ln, "G29")

    # cosmetics
    ws.column_dimensions["A"].width = 26
    for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J"):
        ws.column_dimensions[col].width = 13
    try:
        wb.active = wb.sheetnames.index("Leadership View")
    except Exception:
        pass


# ---------------------------------------------------------------- build sheet
def build(platform, xlsx_path):
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    hist_path = os.path.join(root, "data", platform, "history.csv")
    result_path = os.path.join(root, "platforms", platform, "result.json")

    history_rows, runs = load_history(hist_path)
    result, result_rows = load_result(result_path)

    per_sku_cur, oos_loc_events = analyze_current(result_rows)
    runstats = per_run_sku_stats(history_rows)

    cur_run = runs[-1] if runs else None
    prev_run = runs[-2] if len(runs) >= 2 else None
    n_runs = len(runs)

    # National platforms (Amazon marketplace) scrape a single national location;
    # per-location framing then reads as noise ("OOS locations: 1, Cities: All India").
    # Detect that and reshape STOCK-OUT and TREND sections below.
    NATIONAL = bool(result_rows) and len(
        {(r.get("city") or "", r.get("pincode") or "") for r in result_rows}
    ) <= 1

    # captured_at -> IST string for the subtitle
    cap = (result.get("summary") or {}).get("captured_at")
    gen_when = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    when = gen_when
    try:
        dt = datetime.datetime.fromisoformat(str(cap).replace("Z", "+00:00"))
        ist = dt.astimezone(datetime.timezone(datetime.timedelta(hours=5, minutes=30)))
        when = ist.strftime("%Y-%m-%d %H:%M IST")
    except Exception:
        pass

    wb = load_workbook(xlsx_path)
    if "Predictions" in wb.sheetnames:
        del wb["Predictions"]
    ws = wb.create_sheet("Predictions")
    S = Sheet(ws)

    pname = platform.replace("-", " ").title()
    S.title(
        f"Jivo x {pname} - Predictions & Early Warnings",
        sub=(f"Generated {when}  |  history: {n_runs} run(s)"
             + (f", latest {cur_run}" if cur_run else "")
             + "  |  deterministic, no-LLM"),
    )

    if n_runs <= 1:
        S.note("THIN HISTORY: only one run on record (or none). This page shows "
               "current-state insights now; trend & stock-out-trajectory "
               "predictions sharpen automatically as more runs accumulate.")
        S.blank()

    # ----- gather watch call-outs as we compute the sections -----
    watch = []

    # ============================================================ STOCK-OUT RISK
    S.section("1. STOCK-OUT RISK",
              "Currently out-of-stock ASINs." if NATIONAL
              else "Currently out-of-stock / low coverage, plus SKUs trending toward stockout.")

    # current OOS by SKU
    OOS_TABLE_CAP = 40
    oos_skus = [(sku, e) for sku, e in per_sku_cur.items() if e["n_oos"] > 0]
    oos_skus.sort(key=lambda x: (-x[1]["n_oos"], x[1]["label"]))
    if oos_skus:
        if NATIONAL:
            # National marketplace: each SKU is one row, so "OOS locations" / "Cities
            # affected" are always 1 — drop them; show ASIN + raw availability text.
            sku_first = {}
            for rr in result_rows:
                k = rr.get("canonical") or rr.get("sku_raw") or "?"
                sku_first.setdefault(k, rr)
            data, fills = [], []
            for sku, e in oos_skus[:OOS_TABLE_CAP]:
                rr = sku_first.get(sku, {})
                data.append([e["label"], rr.get("asin") or "-",
                             rr.get("availability_text") or "out of stock"])
                fills.append(RED)
            S.table(["SKU", "ASIN", "Availability"], data, fills)
            if len(oos_skus) > OOS_TABLE_CAP:
                S.note(f"Showing top {OOS_TABLE_CAP} of {len(oos_skus)} OOS ASINs.")
                S.blank()
            if len(oos_skus) >= 6:
                watch.append((len(oos_skus) + 60,
                              f"{len(oos_skus)} ASINs currently out of stock"))
            for sku, e in oos_skus[:5]:
                watch.append((10, f"{e['label']} currently out of stock"))
        else:
            data, fills = [], []
            for sku, e in oos_skus[:OOS_TABLE_CAP]:
                cities = sorted(e["oos_cities"])
                shown = ", ".join(cities[:6]) + (f" +{len(cities) - 6} more" if len(cities) > 6 else "")
                data.append([e["label"], e["n_oos"], e["n_loc"], len(cities), shown])
                frac = e["n_oos"] / e["n_loc"] if e["n_loc"] else 0
                fills.append(RED if frac >= LOW_STOCK_FRAC else YEL)
            S.table(["SKU", "OOS locations", "Total locations", "Cities affected",
                     "Sample cities"], data, fills)
            if len(oos_skus) > OOS_TABLE_CAP:
                S.note(f"Showing top {OOS_TABLE_CAP} of {len(oos_skus)} OOS SKUs "
                       "(sorted by number of locations affected).")
                S.blank()
            multi = [(sku, e) for sku, e in oos_skus if e["n_oos"] >= 2]
            if len(oos_skus) >= 6:
                watch.append((len(oos_skus) + 60,
                              f"{len(oos_skus)} SKUs currently out of stock in >=1 location"))
            for sku, e in (multi if multi else oos_skus[:5]):
                cities = sorted(e["oos_cities"])
                watch.append((e["n_oos"] + (10 if e["n_oos"] >= 2 else 0),
                              f"{e['label']} out of stock in {e['n_oos']} location(s)"
                              f" across {len(cities)} city/cities"))
    else:
        if result_rows:
            S.bullet("No out-of-stock SKUs detected in the current run.", GREEN)
        else:
            S.note("No current rows available to assess stock.")
        S.blank()

    # low-coverage (in stock in <50% of locations where it appears, but not all OOS).
    # Needs >=4 locations per SKU, which never holds on a national single-location feed.
    if not NATIONAL:
        low = []
        for sku, e in per_sku_cur.items():
            known = e["n_in"] + e["n_oos"]
            if known >= 4 and e["n_oos"] > 0:
                frac_in = e["n_in"] / known
                if frac_in < LOW_STOCK_FRAC:
                    low.append([e["label"], e["n_in"], known, fmt_pct(100 * frac_in)])
        if low:
            S.section("Low availability (in stock in <50% of carrying locations)")
            S.table(["SKU", "In stock", "Known locations", "In-stock rate"], low,
                    [YEL] * len(low))

    # trending toward stockout (needs >=2 runs).
    # National: single row per (run, sku), so use flip detection (in-stock <-> OOS).
    # Per-pincode: keep the in-stock-rate decline analysis.
    if prev_run and runstats:
        if NATIONAL:
            flips_oos, flips_back = [], []
            for sku in sorted(set(runstats[cur_run]) | set(runstats[prev_run])):
                c = runstats[cur_run].get(sku); p = runstats[prev_run].get(sku)
                if not c or not p:
                    continue
                c_in = c["n_in"] > 0; p_in = p["n_in"] > 0
                if p_in and not c_in:
                    flips_oos.append(label(sku))
                elif (not p_in) and c_in:
                    flips_back.append(label(sku))
            if flips_oos:
                S.section("Flipped to OUT OF STOCK since last run")
                S.table(["SKU", "Status change"],
                        [[s, f"in stock ({prev_run}) -> OOS ({cur_run})"] for s in flips_oos[:30]],
                        [RED] * min(30, len(flips_oos)))
                watch.append((50 + len(flips_oos),
                              f"{len(flips_oos)} ASIN(s) flipped to OUT OF STOCK since last run"))
            if flips_back:
                S.section("Came BACK in stock since last run")
                S.table(["SKU", "Status change"],
                        [[s, f"OOS ({prev_run}) -> in stock ({cur_run})"] for s in flips_back[:30]],
                        [GREEN] * min(30, len(flips_back)))
                watch.append((30 + len(flips_back),
                              f"{len(flips_back)} ASIN(s) came back in stock since last run"))
            if not flips_oos and not flips_back:
                S.note("No stock-status flips since the previous run.")
                S.blank()
        else:
            trend = []
            for sku in sorted(set(runstats[cur_run]) | set(runstats[prev_run])):
                c = runstats[cur_run].get(sku)
                p = runstats[prev_run].get(sku)
                if not c or not p:
                    continue
                ck = c["n_in"] + c["n_oos"]
                pk = p["n_in"] + p["n_oos"]
                if ck < 3 or pk < 3:
                    continue
                cr = c["n_in"] / ck
                pr = p["n_in"] / pk
                drop = pr - cr
                if drop >= 0.10:  # in-stock rate fell by >=10 pts
                    trend.append([label(sku), fmt_pct(100 * pr), fmt_pct(100 * cr),
                                  fmt_pct(-100 * drop, signed=True)])
                    watch.append((int(drop * 100) + 50,
                                  f"{label(sku)} in-stock rate falling: "
                                  f"{fmt_pct(100 * pr)} -> {fmt_pct(100 * cr)} vs last run"))
            if trend:
                S.section("Trending toward stockout (in-stock rate dropping vs last run)")
                S.table(["SKU", f"Prev ({prev_run})", f"Now ({cur_run})", "Change"],
                        trend, [RED] * len(trend))
    elif n_runs <= 1:
        S.note("Stockout-trajectory analysis needs >=2 runs; shows here once more history exists.")
        S.blank()

    # ============================================================ PRICE & DISCOUNT
    S.section("2. PRICE & DISCOUNT MOVEMENT",
              "Median sale price & discount per SKU vs the previous run; big moves flagged.")
    if prev_run and runstats:
        moves = []
        for sku in sorted(set(runstats[cur_run]) | set(runstats[prev_run])):
            c = runstats[cur_run].get(sku)
            p = runstats[prev_run].get(sku)
            if not c or not p:
                continue
            cp, pp = median(c["prices"]), median(p["prices"])
            cd, pd = median(c["discs"]), median(p["discs"])
            d_abs = (cp - pp) if (cp is not None and pp is not None) else None
            d_pct = (100 * d_abs / pp) if (d_abs is not None and pp) else None
            d_disc = (cd - pd) if (cd is not None and pd is not None) else None
            big = ((d_pct is not None and abs(d_pct) >= PRICE_PCT_BIG) or
                   (d_abs is not None and abs(d_abs) >= PRICE_ABS_BIG) or
                   (d_disc is not None and abs(d_disc) >= DISC_PT_BIG))
            moves.append({
                "sku": sku, "lbl": label(sku), "pp": pp, "cp": cp,
                "d_abs": d_abs, "d_pct": d_pct, "pd": pd, "cd": cd,
                "d_disc": d_disc, "big": big,
                "rank": abs(d_pct) if d_pct is not None else 0,
            })
        moves.sort(key=lambda m: (-1 if m["big"] else 0, -m["rank"], m["lbl"]))
        data, fills = [], []
        for m in moves:
            data.append([
                m["lbl"], fmt_money(m["pp"]), fmt_money(m["cp"]),
                fmt_signed_money(m["d_abs"]), fmt_pct(m["d_pct"], signed=True),
                fmt_pct(m["pd"]) if m["pd"] is not None else "-",
                fmt_pct(m["cd"]) if m["cd"] is not None else "-",
                fmt_pct(m["d_disc"], signed=True) if m["d_disc"] is not None else "-",
            ])
            if m["big"]:
                up = (m["d_abs"] or 0) > 0
                fills.append(RED if up else GREEN)  # price up = red (worse), down = green
                if m["d_pct"] is not None and abs(m["d_pct"]) >= PRICE_PCT_BIG:
                    watch.append((int(abs(m["d_pct"])) + 40,
                                  f"{m['lbl']} price {'up' if up else 'down'} "
                                  f"{fmt_pct(m['d_pct'], signed=True)} "
                                  f"({fmt_money(m['pp'])} -> {fmt_money(m['cp'])}) vs last run"))
                elif m["d_disc"] is not None and abs(m["d_disc"]) >= DISC_PT_BIG:
                    watch.append((int(abs(m["d_disc"])) + 30,
                                  f"{m['lbl']} discount {fmt_pct(m['d_disc'], signed=True)} pts vs last run"))
            else:
                fills.append(None)
        if data:
            S.table(["SKU", f"Prev price ({prev_run})", f"Now price ({cur_run})",
                     "delta Rs.", "delta %", "Prev disc", "Now disc", "delta disc pts"],
                    data, fills)
            S.note("Red = sale price rose vs last run; Green = sale price fell. "
                   "Flagged moves: >=5% / >=Rs.10 price, or >=5 pts discount.")
            S.blank()
        else:
            S.note("No SKUs shared between the last two runs to compare.")
            S.blank()
    else:
        # thin history -> current price/discount & cross-location-spread outliers.
        # Run-over-run movement isn't available yet, so surface the most useful
        # current-state signal: SKUs whose price varies across locations (a
        # cross-location anomaly worth watching) and the discount extremes.
        S.note("Only one run on record - showing current cross-location price "
               "spread & discount outliers instead of run-over-run movement.")
        TOPN = 25
        priced = [(sku, e) for sku, e in per_sku_cur.items() if e["med_price"] is not None]

        def spread_of(e):
            if e["max_price"] is None or e["min_price"] is None:
                return 0.0
            return e["max_price"] - e["min_price"]

        any_spread = any(spread_of(e) > 0.5 for _, e in priced)

        if any_spread:
            # per-pincode platform on its first run: rank by cross-location spread
            ranked = sorted(priced, key=lambda kv: -spread_of(kv[1]))
            shown = [(s, e) for s, e in ranked if spread_of(e) > 0.5][:TOPN]
            data = []
            for sku, e in shown:
                spread = spread_of(e)
                data.append([
                    e["label"], fmt_money(e["min_price"]), fmt_money(e["med_price"]),
                    fmt_money(e["max_price"]), fmt_money(spread),
                    fmt_pct(e["med_disc"]) if e["med_disc"] is not None else "-",
                ])
                if e["med_price"] and spread / e["med_price"] >= 0.15:
                    watch.append((int(100 * spread / e["med_price"]),
                                  f"{e['label']} price varies {fmt_money(e['min_price'])}-"
                                  f"{fmt_money(e['max_price'])} across locations this run"))
            S.section("Widest cross-location price spread (top by Rs. spread)")
            S.table(["SKU", "Min price", "Median price", "Max price",
                     "Spread", "Median discount"], data)
            if len(ranked) > len(shown):
                S.note(f"Showing top {len(shown)} of {len(priced)} priced SKUs by spread.")
                S.blank()
        else:
            # national platform (single price per SKU): no spread -> show discount
            # extremes, which is the actionable current-state signal.
            by_disc = sorted([(s, e) for s, e in priced if e["med_disc"] is not None],
                             key=lambda kv: kv[1]["med_disc"])
            if by_disc:
                hi = list(reversed(by_disc[-TOPN:]))   # biggest discounts
                lo = by_disc[:10]                       # smallest discounts
                S.section(f"Highest discounts right now (top {len(hi)})")
                S.table(["SKU", "Price", "MRP-implied discount", "Locations"],
                        [[e["label"], fmt_money(e["med_price"]),
                          fmt_pct(e["med_disc"]), e["n_loc"]] for _, e in hi])
                top = hi[0][1]
                watch.append((int(top["med_disc"]),
                              f"Deepest discount this run: {top['label']} at "
                              f"{fmt_pct(top['med_disc'])} off"))
                S.section(f"Lowest / no discount right now (bottom {len(lo)})")
                S.table(["SKU", "Price", "MRP-implied discount", "Locations"],
                        [[e["label"], fmt_money(e["med_price"]),
                          fmt_pct(e["med_disc"]), e["n_loc"]] for _, e in lo],
                        [YEL] * len(lo))
                S.note(f"{len(priced)} priced SKUs this run; full per-SKU pricing "
                       "is on the other sheets. Discount = vs listed MRP.")
                S.blank()
            else:
                # priced but no discount info: compact price-only list, capped
                data = [[e["label"], fmt_money(e["med_price"]), e["n_loc"]]
                        for _, e in sorted(priced, key=lambda kv: -kv[1]["med_price"])[:TOPN]]
                if data:
                    S.section(f"Highest-priced SKUs right now (top {len(data)})")
                    S.table(["SKU", "Price", "Locations"], data)
                    S.blank()
                else:
                    S.note("No priced rows available in the current run.")
                    S.blank()

    # ============================================================ COVERAGE / CATALOGUE TREND
    if NATIONAL:
        S.section("3. CATALOGUE TREND",
                  "ASINs scraped and in-stock count per run.")
        if runstats:
            cov = []
            recent = runs[-12:]
            prev_in_n = None
            for run in recent:
                rs = runstats[run]
                tot_sku = len(rs)
                tot_in = sum(1 for s in rs.values() if s["n_in"] > 0)
                tot_oos = sum(1 for s in rs.values() if s["n_oos"] > 0)
                delta = (tot_in - prev_in_n) if prev_in_n is not None else None
                arrow = ""
                if delta is not None:
                    arrow = "up" if delta > 0 else ("down" if delta < 0 else "flat")
                cov.append([run_date(history_rows, run), run, tot_sku, tot_in, tot_oos,
                            (f"{'+' if delta and delta > 0 else ''}{delta}" if delta is not None else "-"),
                            arrow])
                prev_in_n = tot_in
            if len(runs) > len(recent):
                S.note(f"Showing last {len(recent)} of {len(runs)} runs.")
            S.table(["Date", "Run", "ASINs scraped", "In stock", "Out of stock",
                     "delta in-stock", "Trend"], cov)
            if len(cov) >= 2:
                last_in = cov[-1][3]; prev_in_val = cov[-2][3]
                d = last_in - prev_in_val
                if d <= -5:
                    watch.append((30 + abs(d),
                                  f"In-stock ASIN count fell by {-d} since last run "
                                  f"({prev_in_val} -> {last_in})"))
                elif d >= 5:
                    watch.append((20 + d,
                                  f"In-stock ASIN count grew by {d} since last run "
                                  f"({prev_in_val} -> {last_in})"))
        else:
            S.note("No history yet for a catalogue trend.")
            S.blank()
        cur_in = sum(1 for r in result_rows if in_stock_of(r) is True)
        cur_oos = sum(1 for r in result_rows if in_stock_of(r) is False)
        cur_unk = len(result_rows) - cur_in - cur_oos
        S.section("Current catalogue snapshot")
        S.bullet(f"{len(result_rows)} ASINs tracked this run.")
        S.bullet(f"{cur_in} in stock  ·  {cur_oos} out of stock"
                 + (f"  ·  {cur_unk} not found / blocked" if cur_unk else "") + ".")
        if len(result_rows):
            S.bullet(f"In-stock rate: {fmt_pct(100 * cur_in / len(result_rows))}.")
        S.blank()
    else:
        S.section("3. COVERAGE TREND",
                  "Locations carrying Jivo over time, and notable gaps.")

        # per-run distinct location count (from history)
        if runstats:
            cov = []
            run_loc = {}  # run -> set of locations carrying any Jivo
            # recompute distinct carrying locations per run from raw history
            carry = defaultdict(set)
            for r in history_rows:
                run = r.get("run_id")
                st = _f(r.get("in_stock"))
                # a location "carries" the SKU if it appears in history (it was found)
                key = (r.get("city") or "") + "|" + (r.get("pincode") or "")
                if run:
                    carry[run].add(key)
            for run in runs:
                run_loc[run] = carry.get(run, set())
            prev_n = None
            for run in runs:
                n = len(run_loc[run])
                delta = (n - prev_n) if prev_n is not None else None
                arrow = ""
                if delta is not None:
                    arrow = "up" if delta > 0 else ("down" if delta < 0 else "flat")
                cov.append([run_date(history_rows, run), run, n,
                            (f"{'+' if delta and delta > 0 else ''}{delta}" if delta is not None else "-"),
                            arrow])
                prev_n = n
            S.table(["Date", "Run", "Locations carrying Jivo", "delta vs prev", "Trend"], cov)

            # coverage drop call-out + lost locations vs previous run
            if prev_run:
                now_set = run_loc.get(cur_run, set())
                prev_set = run_loc.get(prev_run, set())
                lost = prev_set - now_set
                gained = now_set - prev_set
                if lost:
                    # group lost by city
                    lost_cities = defaultdict(int)
                    for k in lost:
                        city = k.split("|", 1)[0] or "?"
                        lost_cities[city] += 1
                    S.section("Locations that LOST coverage since last run")
                    rows_lost = [[c, n] for c, n in sorted(lost_cities.items(),
                                                           key=lambda x: -x[1])]
                    S.table(["City", "Locations lost"], rows_lost, [YEL] * len(rows_lost))
                    top_city = max(lost_cities.items(), key=lambda x: x[1])
                    watch.append((len(lost) + 20,
                                  f"Coverage dropped: {len(lost)} location(s) lost Jivo vs last run"
                                  f" (most in {top_city[0]})"))
                if gained:
                    watch.append((len(gained),
                                  f"Coverage grew: {len(gained)} new location(s) now carry Jivo"))
        else:
            S.note("No history yet for a coverage trend; current coverage shown below.")
            S.blank()

        # current coverage snapshot from result.json summary + gaps
        summ = result.get("summary") or {}
        pin_with = summ.get("pincodes_with_jivo")
        pin_tot = summ.get("pincodes_total")
        if pin_with is not None and pin_tot:
            S.section("Current coverage snapshot")
            S.bullet(f"{pin_with} of {pin_tot} locations probed currently carry Jivo "
                     f"({fmt_pct(100 * pin_with / pin_tot)}).")
            if pin_tot - pin_with > 0:
                watch.append((1, f"{pin_tot - pin_with} of {pin_tot} probed locations have NO Jivo"))

        # cities with zero Jivo (gap), derived from perPin if available
        cities_total = OrderedDict()
        for p in (result.get("perPin") or []):
            c = p.get("city") or "?"
            cities_total.setdefault(c, 0)
            cities_total[c] += len(p.get("rows") or [])
        zero_cities = [c for c, n in cities_total.items() if n == 0]
        if zero_cities:
            S.bullet("Cities with ZERO Jivo this run: " + ", ".join(sorted(zero_cities)), RED)
            watch.append((len(zero_cities) + 10,
                          f"{len(zero_cities)} city/cities have ZERO Jivo this run: "
                          + ", ".join(sorted(zero_cities)[:5])
                          + (" ..." if len(zero_cities) > 5 else "")))
        S.blank()

    # ============================================================ WHAT TO WATCH
    # (rendered near the top would need pre-pass; instead put it as a clearly
    # labelled lead section by writing it first. We compute after, so place it
    # at the end but make it visually prominent.)
    S.section("4. WHAT TO WATCH (top call-outs)",
              "Highest-signal, data-driven items from the sections above.")
    if watch:
        # dedupe by message, keep highest priority
        best = {}
        for prio, msg in watch:
            if msg not in best or prio > best[msg]:
                best[msg] = prio
        ordered = sorted(best.items(), key=lambda kv: -kv[1])[:10]
        for msg, _ in ordered:
            S.bullet(msg)
    else:
        S.bullet("Nothing notable: stock, pricing and coverage all look steady this run.", GREEN)
    if n_runs <= 1:
        S.blank()
        S.note("As more runs accumulate, this list will surface run-over-run "
               "price moves, stockout trajectories and coverage shifts.")

    S.autosize()
    ws.freeze_panes = "A4"

    # Chart-driven 'Leadership View' front sheet for the e-com head. Wrapped so a
    # dashboard hiccup never costs us the Predictions sheet we just built.
    try:
        build_dashboard(wb, platform, when, result_rows, per_sku_cur,
                        runstats, runs, history_rows, watch)
    except Exception as e:
        import traceback
        sys.stderr.write("predict.py: Leadership View skipped (non-fatal): %s\n" % e)
        traceback.print_exc(file=sys.stderr)

    wb.save(xlsx_path)
    return xlsx_path, n_runs, len(per_sku_cur), len(watch)


def main(argv):
    if len(argv) < 3:
        sys.stderr.write("usage: predict.py <platform> <xlsx_path>\n")
        return 2
    platform, xlsx_path = argv[1], argv[2]
    if not os.path.exists(xlsx_path):
        sys.stderr.write(f"predict.py: xlsx not found: {xlsx_path}\n")
        return 1
    try:
        path, n_runs, n_sku, n_watch = build(platform, xlsx_path)
        sys.stderr.write(f"predict.py: appended 'Predictions' to {path} "
                         f"(history runs={n_runs}, skus={n_sku}, watch items={n_watch})\n")
        return 0
    except Exception as e:
        # Never break the pipeline: log and exit non-zero so run.sh's `|| true`
        # swallows it without losing the already-built workbook.
        import traceback
        sys.stderr.write("predict.py: FAILED (non-fatal): %s\n" % e)
        traceback.print_exc(file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
