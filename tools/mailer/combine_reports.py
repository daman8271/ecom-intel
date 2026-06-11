#!/usr/bin/env python3
"""Combine the day's report workbooks into ONE .xlsx with all sheets.

Each source sheet is copied cell-by-cell (values, styles, number formats,
merged cells, column widths, row heights, freeze panes, tab colors,
auto-filters, conditional formatting) under a platform-prefixed sheet name,
e.g. "FRESH·Pricing Matrix". Content is not altered.

Known limitation: openpyxl drops embedded charts/images on load, so chart
objects in Leadership View sheets do not survive the merge — all underlying
data does.

Usage: combine_reports.py --out combined.xlsx report1.xlsx report2.xlsx ...
"""
import argparse
import re
import sys
from copy import copy

import openpyxl

PREFIX = {
    "Amazon": "AMZ",
    "AmazonFresh": "FRESH",
    "AmazonNow": "NOW",
    "Bigbasket": "BB",
    "Blinkit": "BLK",
    "Flipkart": "FK",
    "FlipkartMinutes": "FKM",
    "Zepto": "ZEPTO",
    "Price-Match": "PM",
}

BAD_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def platform_of(path):
    name = path.rsplit("/", 1)[-1]
    m = re.match(r"Jivo-(.+?)-Live-Report-\d{4}-\d{2}-\d{2}\.xlsx$", name)
    if m:
        return m.group(1)
    m = re.match(r"Jivo-(.+?)-\d{4}-\d{2}-\d{2}\.xlsx$", name)
    if m:
        return m.group(1)
    return name.rsplit(".", 1)[0]


def sheet_name(prefix, original, used):
    name = BAD_SHEET_CHARS.sub(" ", f"{prefix}·{original}")[:31]
    base, n = name, 2
    while name in used:
        suffix = f"~{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def copy_sheet(src, dst):
    for row in src.iter_rows():
        for c in row:
            nc = dst.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                nc.font = copy(c.font)
                nc.border = copy(c.border)
                nc.fill = copy(c.fill)
                nc.number_format = c.number_format
                nc.protection = copy(c.protection)
                nc.alignment = copy(c.alignment)
            if c.hyperlink:
                nc.hyperlink = copy(c.hyperlink)
            if c.comment:
                nc.comment = copy(c.comment)
    for rng in src.merged_cells.ranges:
        dst.merge_cells(str(rng))
    for key, dim in src.column_dimensions.items():
        d = dst.column_dimensions[key]
        d.width, d.hidden = dim.width, dim.hidden
    for key, dim in src.row_dimensions.items():
        d = dst.row_dimensions[key]
        d.height, d.hidden = dim.height, dim.hidden
    dst.freeze_panes = src.freeze_panes
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    dst.sheet_view.zoomScale = src.sheet_view.zoomScale
    dst.sheet_properties.tabColor = src.sheet_properties.tabColor
    try:
        dst.sheet_format = copy(src.sheet_format)
    except Exception:
        pass
    if src.auto_filter.ref:
        dst.auto_filter.ref = src.auto_filter.ref
    try:
        for cf in src.conditional_formatting:
            for rule in cf.rules:
                dst.conditional_formatting.add(str(cf.sqref), rule)
    except Exception as e:
        print(f"WARN: conditional formatting not copied for {dst.title}: {e}",
              file=sys.stderr)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    ap.add_argument("files", nargs="+")
    args = ap.parse_args()

    out = openpyxl.Workbook()
    out.remove(out.active)
    used = set()
    copied = 0
    for path in args.files:
        prefix = PREFIX.get(platform_of(path), platform_of(path)[:6])
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            print(f"WARN: cannot open {path}: {e}", file=sys.stderr)
            continue
        for src in wb.worksheets:
            dst = out.create_sheet(sheet_name(prefix, src.title, used))
            copy_sheet(src, dst)
            copied += 1
        wb.close()

    if not copied:
        print("ERROR: no sheets copied", file=sys.stderr)
        return 1
    out.active = 0
    out.save(args.out)
    print(f"OK: {copied} sheets from {len(args.files)} workbooks -> {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
