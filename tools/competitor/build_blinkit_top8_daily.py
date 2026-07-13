#!/usr/bin/env python3
"""Merge the daily Blinkit top-8 device run and build its Excel workbook."""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import importlib.util
import json
import os
import shutil
import subprocess
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path("/opt/ecom-intel")
IST = dt.timezone(dt.timedelta(hours=5, minutes=30))
TARGET_BRANDS = [
    "Fortune",
    "Saffola",
    "Borges",
    "Tata",
    "Del Monte",
    "Figaro",
    "Sundrop",
    "Gulab",
]
ALLOWED_BRANDS = {brand.casefold() for brand in TARGET_BRANDS} | {"jivo", "sano"}
EXPECTED_SHEETS = [
    "Summary",
    "City-Pin-SKU Prices",
    "Run Scope",
    "Anchor Watch",
    "Master Data",
]
CITY_PRICE_BRAND_ORDER = [
    "Fortune",
    "Borges",
    "Del Monte",
    "Figaro",
    "Tata",
    "Saffola",
    "Sundrop",
    "Gulab",
]


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def norm_brand(value: Any) -> str:
    return " ".join(str(value or "").split()).casefold()


def pin(value: Any) -> str:
    return str(value or "").strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def captured_ist_date(value: Any) -> str:
    parsed = dt.datetime.fromisoformat(str(value or "").replace("Z", "+00:00"))
    return parsed.astimezone(IST).date().isoformat()


def validate_and_merge(
    run_dir: Path, run_meta: dict[str, Any], date_ist: str
) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, str], dict[str, Any]]:
    expected = read_json(run_dir / "pincodes.all.json")
    expected_pins = {pin(row.get("pincode")) for row in expected}
    city_counts = Counter(str(row.get("city") or "").strip() for row in expected)
    if len(expected) != 75 or len(expected_pins) != 75:
        raise SystemExit("scope gate failed: expected 75 unique pincodes")
    if len(city_counts) != 25 or set(city_counts.values()) != {3}:
        raise SystemExit(f"city scope gate failed: {dict(city_counts)}")

    workers = run_meta.get("workers") or []
    if not workers:
        raise SystemExit("run metadata has no workers")
    local_scraper_sha = sha256(ROOT / "platforms/blinkit/scrape.js")
    progress: dict[str, dict[str, Any]] = {}
    pin_device: dict[str, str] = {}
    summaries: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []

    for worker in workers:
        worker_id = str(worker["id"])
        label = str(worker["label"])
        worker_pins = read_json(run_dir / str(worker["pincodes_file"]))
        worker_expected = {pin(row.get("pincode")) for row in worker_pins}
        worker_progress = read_json(run_dir / str(worker["progress_file"]))
        capture = read_json(run_dir / str(worker["capture_file"]))
        if set(worker_progress) != worker_expected:
            raise SystemExit(
                f"{label} coverage mismatch: expected={len(worker_expected)} "
                f"actual={len(worker_progress)}"
            )
        duplicate_pins = sorted(set(progress) & set(worker_progress))
        if duplicate_pins:
            raise SystemExit(f"duplicate worker pincodes: {duplicate_pins}")
        for pincode, item in worker_progress.items():
            progress[pincode] = item
            pin_device[pincode] = label

        worker_rows = capture.get("allRows") or []
        progress_rows = sum(
            len(item.get("rows") or []) for item in worker_progress.values()
        )
        if progress_rows != len(worker_rows):
            raise SystemExit(
                f"{label} row gate failed: progress={progress_rows}, "
                f"capture={len(worker_rows)}"
            )
        summary = capture.get("summary") or {}
        if captured_ist_date(summary.get("captured_at")) != date_ist:
            raise SystemExit(
                f"{label} capture is not from {date_ist}: {summary.get('captured_at')}"
            )
        if summary.get("scraper_sha256") != local_scraper_sha:
            raise SystemExit(f"{label} scraper hash does not match the reviewed VPS file")
        summaries.append(summary)
        rows.extend(worker_rows)
        if worker_id not in {"windows", "macpro"}:
            raise SystemExit(f"unsupported worker id: {worker_id}")

    if set(progress) != expected_pins:
        raise SystemExit(
            f"combined coverage mismatch: missing={sorted(expected_pins - set(progress))} "
            f"extra={sorted(set(progress) - expected_pins)}"
        )
    unresolved = sorted(p for p, item in progress.items() if not item.get("resolved"))
    blocked = sorted(
        p
        for p, item in progress.items()
        if item.get("blocked") or item.get("partial_block")
    )
    auth_rejected = sorted(
        p for p, item in progress.items() if item.get("auth_accepted") != 1
    )
    empty = sorted(p for p, item in progress.items() if not (item.get("rows") or []))
    if unresolved or blocked or auth_rejected or empty:
        raise SystemExit(
            "live gate failed: "
            f"unresolved={unresolved} blocked={blocked} "
            f"auth_rejected={auth_rejected} empty={empty}"
        )

    bad_rows: list[int] = []
    for index, row in enumerate(rows):
        brand = norm_brand(row.get("brand"))
        if (
            pin(row.get("pincode")) not in expected_pins
            or str(row.get("platform") or "").casefold() != "blinkit"
            or brand not in ALLOWED_BRANDS
            or not isinstance(row.get("sale"), (int, float))
            or row.get("sale") <= 0
            or row.get("in_stock") not in (0, 1)
        ):
            bad_rows.append(index)
    if bad_rows:
        raise SystemExit(f"row quality gate failed: bad row indexes={bad_rows[:20]}")

    deduped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = (
            pin(row.get("pincode")),
            str(row.get("store_id") or ""),
            norm_brand(row.get("brand")),
            str(row.get("canonical") or ""),
        )
        deduped.setdefault(key, row)
    rows = list(deduped.values())
    found = Counter(norm_brand(row.get("brand")) for row in rows)

    summary = {
        "mode": "competitor",
        "platform": "blinkit",
        "date_ist": date_ist,
        "pincodes_total": 75,
        "pincodes_resolved": 75,
        "pincodes_unresolved": 0,
        "pincodes_blocked": 0,
        "pincodes_with_rows": 75,
        "total_rows": len(rows),
        "unique_skus": len({str(row.get("canonical") or "") for row in rows}),
        "wall_s": max(int(item.get("wall_s") or 0) for item in summaries),
        "partial": False,
        "auth_session": 1,
        "auth_required": 1,
        "auth_verified": 1,
        "auth_verified_pincodes": 75,
        "captured_at": max(str(item.get("captured_at") or "") for item in summaries),
        "scope": {
            "cities": 25,
            "pincodes_per_city": 3,
            "devices": [
                str(worker.get("display_name") or worker["label"])
                for worker in workers
            ],
            "competitors": TARGET_BRANDS,
            "selection_source": str(run_meta.get("selection_source") or ""),
        },
    }
    merged = {"summary": summary, "allRows": rows}
    audit = {
        "date": date_ist,
        "run_id": run_meta.get("run_id"),
        "summary": summary,
        "rows_by_brand": {
            brand: found.get(brand.casefold(), 0) for brand in TARGET_BRANDS
        }
        | {"Jivo": found.get("jivo", 0), "Sano": found.get("sano", 0)},
        "rows_by_city": dict(
            sorted(Counter(str(row.get("city") or "") for row in rows).items())
        ),
        "pin_device": pin_device,
    }
    return expected, merged, pin_device, audit


def add_scope_sheet(
    workbook_path: Path,
    expected: list[dict[str, Any]],
    progress: dict[str, dict[str, Any]],
    pin_device: dict[str, str],
    summary: dict[str, Any],
    run_meta: dict[str, Any],
    date_ist: str,
) -> None:
    workbook = load_workbook(workbook_path)
    if "Run Scope" in workbook.sheetnames:
        del workbook["Run Scope"]
    sheet = workbook.create_sheet("Run Scope", 1)
    green = "008B3A"
    header_fill = PatternFill("solid", fgColor=green)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    display_date = dt.date.fromisoformat(date_ist).strftime("%d %b %Y").lstrip("0")

    sheet["A1"] = "Blinkit Competitor Run Scope"
    sheet["A1"].font = Font(size=18, bold=True, color=green)
    sheet.merge_cells("A1:H1")
    sheet["A2"] = (
        f"{display_date} | 25 cities | 3 pincodes per city | 75/75 resolved | "
        f"{summary['total_rows']} datapoints"
    )
    sheet["A2"].font = Font(italic=True, color="555555")
    sheet.merge_cells("A2:H2")
    sheet["A4"] = "Top 8 competitors"
    sheet["B4"] = ", ".join(TARGET_BRANDS)
    sheet.merge_cells("B4:H4")
    sheet["A5"] = "Devices"
    concurrencies = {int(worker.get("concurrency", 2)) for worker in run_meta["workers"]}
    labels = " + ".join(
        str(worker.get("display_name") or worker["label"])
        for worker in run_meta["workers"]
    )
    if len(concurrencies) == 1:
        worker_text = f"{labels} ({concurrencies.pop()} workers each)"
    else:
        worker_text = " + ".join(
            f"{worker.get('display_name') or worker['label']} ({worker.get('concurrency', 2)} workers)"
            for worker in run_meta["workers"]
        )
    sheet["B5"] = worker_text
    sheet.merge_cells("B5:H5")

    headers = [
        "City",
        "Pincode",
        "Locality",
        "Device",
        "Resolved",
        "Authenticated",
        "Rows",
        "Brands captured",
    ]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(row=7, column=column, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    ordered = sorted(
        expected,
        key=lambda row: (
            str(row.get("city") or ""),
            int(row.get("competitor_sample_index") or 0),
        ),
    )
    for row_number, item in enumerate(ordered, 8):
        pincode = pin(item.get("pincode"))
        result = progress[pincode]
        brands = sorted(
            {
                " ".join(str(row.get("brand") or "").split())
                for row in result.get("rows") or []
                if row.get("brand")
            },
            key=str.casefold,
        )
        values = [
            item.get("city"),
            pincode,
            item.get("locality"),
            pin_device[pincode],
            "Yes" if result.get("resolved") else "No",
            "Yes" if result.get("auth_accepted") == 1 else "No",
            len(result.get("rows") or []),
            ", ".join(brands),
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A7:H{sheet.max_row}"
    for column, width in enumerate([20, 12, 32, 12, 12, 15, 10, 70], 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    workbook.save(workbook_path)


def load_report_builder() -> Any:
    path = ROOT / "tools/competitor/build_competitor_report.py"
    spec = importlib.util.spec_from_file_location("competitor_report_builder", path)
    if spec is None or spec.loader is None:
        raise SystemExit(f"cannot load report builder: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def add_city_pin_sku_sheet(
    workbook_path: Path, rows: list[dict[str, Any]], date_ist: str
) -> None:
    report = load_report_builder()
    annotated = report.annotate([dict(row) for row in rows], "blinkit")
    matched_brands = {
        row.get("_brand")
        for row in annotated
        if any(report.rival_matches(row, anchor) for anchor in report.ANCHORS)
    }
    price_brands = [brand for brand in CITY_PRICE_BRAND_ORDER if brand in matched_brands]
    unmatched_brands = sorted(set(TARGET_BRANDS) - set(price_brands))

    workbook = load_workbook(workbook_path)
    if "City-Pin-SKU Prices" in workbook.sheetnames:
        del workbook["City-Pin-SKU Prices"]
    sheet = workbook.create_sheet("City-Pin-SKU Prices", 1)
    green = "008B3A"
    header_fill = PatternFill("solid", fgColor=green)
    header_font = Font(color="FFFFFF", bold=True)
    stripe_fill = PatternFill("solid", fgColor="F0F6F1")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    columns = ["City", "Pincode", "SKU (JIVO anchor)", "JIVO Rs"]
    columns += [f"{brand} Rs" for brand in price_brands]
    columns += ["Cheapest Rival", "Gap Rs", "Verdict"]
    last_column = get_column_letter(len(columns))

    sheet["A1"] = "City × Pincode × SKU — Competitor Prices (Blinkit)"
    sheet["A1"].font = Font(size=18, bold=True, color=green)
    sheet.merge_cells(f"A1:{last_column}1")
    note = (
        f"{date_ist}  -  listed sale price Rs of the matched pack at that pincode "
        "(rival packs = same size as the JIVO SKU, so prices compare directly)  -  "
        "red = listed but OUT OF STOCK at capture  -  blank = not listed at that "
        "pincode"
    )
    if unmatched_brands:
        note += "  -  no name-matched rival SKU found for: " + ", ".join(
            unmatched_brands
        )
    sheet["A2"] = note
    sheet["A2"].font = Font(italic=True, color="555555", size=10)
    sheet.merge_cells(f"A2:{last_column}2")
    for column, value in enumerate(columns, 1):
        cell = sheet.cell(row=4, column=column, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border

    rows_by_pin: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in annotated:
        key = (str(row.get("city") or ""), pin(row.get("pincode")))
        rows_by_pin.setdefault(key, []).append(row)
    striped_cities = {
        city
        for index, city in enumerate(sorted({key[0] for key in rows_by_pin}))
        if index % 2 == 0
    }

    output_row = 5
    for (city, pincode), pin_rows in sorted(
        rows_by_pin.items(), key=lambda item: (item[0][0], item[0][1])
    ):
        for anchor in report.ANCHORS:
            jivo_rows = [
                row
                for row in pin_rows
                if row.get("_jivo") and report.jivo_matches(row, anchor)
            ]
            rival_rows = [
                row for row in pin_rows if report.rival_matches(row, anchor)
            ]
            if not jivo_rows and not rival_rows:
                continue

            jivo_row = min(
                jivo_rows,
                key=lambda row: (
                    not bool(row.get("in_stock")),
                    row.get("sale") is None,
                    row.get("sale") if row.get("sale") is not None else float("inf"),
                ),
                default=None,
            )
            brand_rows: dict[str, dict[str, Any]] = {}
            for row in rival_rows:
                brand = str(row.get("_brand") or "")
                current = brand_rows.get(brand)
                if current is None or (
                    not bool(row.get("in_stock")),
                    float(row["sale"]),
                ) < (
                    not bool(current.get("in_stock")),
                    float(current["sale"]),
                ):
                    brand_rows[brand] = row
            cheapest = min(
                brand_rows.values(), key=lambda row: float(row["sale"]), default=None
            )
            jivo_sale = jivo_row.get("sale") if jivo_row else None
            cheapest_sale = cheapest.get("sale") if cheapest else None
            cheapest_brand = cheapest.get("_brand") if cheapest else None
            gap = (
                cheapest_sale - jivo_sale
                if cheapest_sale is not None and jivo_sale is not None
                else None
            )
            if jivo_sale is None:
                verdict = "Rival only (JIVO missing)"
            elif cheapest_sale is None:
                verdict = "JIVO only listed"
            else:
                verdict, _, _ = report.verdict_for(cheapest_sale, jivo_sale)
                if verdict == "THREAT - rival cheaper":
                    verdict = "THREAT - rival cheaper"
                elif verdict == "JIVO cheaper":
                    verdict = "JIVO cheaper"
                else:
                    verdict = "Level"

            values = [
                city,
                pincode,
                anchor.get("jivo_example", anchor["key"]),
                jivo_sale,
            ]
            values += [
                brand_rows.get(brand, {}).get("sale") for brand in price_brands
            ]
            values += [cheapest_brand, gap, verdict]
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row=output_row, column=column, value=value)
                cell.border = border
                if city in striped_cities:
                    cell.fill = stripe_fill
                if column >= 4 and column <= 3 + len(price_brands) + 1:
                    cell.number_format = "#,##0"
            if jivo_row and not jivo_row.get("in_stock"):
                sheet.cell(row=output_row, column=4).font = Font(
                    bold=True, color="C00000"
                )
            for offset, brand in enumerate(price_brands, 5):
                if brand in brand_rows and not brand_rows[brand].get("in_stock"):
                    sheet.cell(row=output_row, column=offset).font = Font(
                        bold=True, color="C00000"
                    )
            gap_column = len(columns) - 1
            verdict_column = len(columns)
            sheet.cell(row=output_row, column=gap_column).number_format = "+#,##0;-#,##0;0"
            if verdict == "THREAT - rival cheaper":
                sheet.cell(row=output_row, column=verdict_column).font = Font(
                    bold=True, color="C00000"
                )
            elif verdict == "JIVO cheaper":
                sheet.cell(row=output_row, column=verdict_column).font = Font(
                    color=green
                )
            elif verdict == "Level":
                sheet.cell(row=output_row, column=verdict_column).font = Font(
                    color="B45309"
                )
            elif verdict == "Rival only (JIVO missing)":
                sheet.cell(row=output_row, column=verdict_column).font = Font(
                    bold=True, color="B45309"
                )
            output_row += 1

    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{last_column}{sheet.max_row}"
    widths = [15, 10, 44, 9] + [11] * len(price_brands) + [15, 9, 22]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    workbook.save(workbook_path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    parser.add_argument("--run-dir", type=Path, required=True)
    args = parser.parse_args()
    run_dir = args.run_dir.resolve()
    run_meta = read_json(run_dir / "run.json")
    if run_meta.get("date") != args.date:
        raise SystemExit("run metadata date does not match --date")

    expected, merged, pin_device, audit = validate_and_merge(
        run_dir, run_meta, args.date
    )
    progress: dict[str, dict[str, Any]] = {}
    for worker in run_meta["workers"]:
        progress.update(read_json(run_dir / str(worker["progress_file"])))

    data_dir = Path(
        os.environ.get("COMPETITOR_DATA_DIR", ROOT / "tools/competitor/data")
    )
    output_dir = Path(os.environ.get("COMPETITOR_OUT_DIR", ROOT / "output"))
    capture_path = data_dir / f"blinkit_competitor_{args.date}.json"
    report_path = output_dir / f"Competitor-Price-Watch-Blinkit-{args.date}.xlsx"
    capture_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    if capture_path.exists():
        backup = run_dir / f"preexisting-{capture_path.name}"
        if not backup.exists():
            shutil.copy2(capture_path, backup)
    temporary = capture_path.with_suffix(".json.tmp")
    temporary.write_text(
        json.dumps(merged, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    os.replace(temporary, capture_path)

    env = os.environ.copy()
    env["COMPETITOR_INPUT"] = str(capture_path)
    env["COMPETITOR_DATA_DIR"] = str(data_dir)
    env["COMPETITOR_OUT_DIR"] = str(output_dir)
    subprocess.run(
        [
            "python3",
            str(ROOT / "tools/competitor/build_competitor_report.py"),
            "blinkit",
            args.date,
        ],
        check=True,
        env=env,
    )
    add_scope_sheet(
        report_path,
        expected,
        progress,
        pin_device,
        merged["summary"],
        run_meta,
        args.date,
    )
    add_city_pin_sku_sheet(report_path, merged["allRows"], args.date)
    workbook = load_workbook(report_path, read_only=True, data_only=False)
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise SystemExit(f"workbook sheet gate failed: {workbook.sheetnames}")
    if workbook["Run Scope"].max_row != 82:
        raise SystemExit("Run Scope must contain 75 pincode rows")
    workbook.close()

    audit["capture_path"] = str(capture_path)
    audit["capture_sha256"] = sha256(capture_path)
    audit["workbook_path"] = str(report_path)
    audit["workbook_sha256"] = sha256(report_path)
    audit_text = json.dumps(audit, indent=2, ensure_ascii=False) + "\n"
    (run_dir / "merge-audit.json").write_text(audit_text, encoding="utf-8")
    audit_dir = Path(os.environ.get("BLINKIT_TOP8_AUDIT_DIR", ROOT / "logs"))
    audit_log = audit_dir / f"blinkit-top8-{args.date}.audit.json"
    audit_log.parent.mkdir(parents=True, exist_ok=True)
    audit_log.write_text(audit_text, encoding="utf-8")
    print(json.dumps(audit["summary"], sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
