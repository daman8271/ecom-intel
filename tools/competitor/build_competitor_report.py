#!/usr/bin/env python3
"""Anchor-based competitor price watch (V2, quick-commerce).

V2 fixes a real accuracy bug: oil type is now derived from the PRODUCT NAME,
never from the search query the row came back on. e.g. "Patanjali Virgin Sesame
Oil (Gingelly/Til Oil)" surfaced on the *canola oil* search but is name-derived
as **sesame** and therefore never appears under a canola anchor.

Pipeline
  1. Read the env-gated competitor capture for one platform
     (tools/competitor/data/<platform>_competitor_<date>.json, shape
     {"summary":{...},"allRows":[...]}).
  2. AUTHORITATIVELY re-derive oil_type + grade from each row's NAME using
     oil_classifier.json (overrides the scraper's query-based `category`).
  3. Attach competitors to the 9 JIVO anchors in competitor_match_map.json ONLY
     when name-derived oil_type EQUALS the anchor's oil_type AND grade is in the
     anchor's grades_ok AND brand is allowed AND no blend/exclude token appears
     AND pack volume is within ~15% (grams->ml at density 0.916). Sesame /
     mustard / sunflower / groundnut can NEVER match canola; canola never matches
     anything else. Olive grades (extra_virgin | extra_light | pomace) are never
     crossed; mustard kachi_ghani/cold_pressed is never crossed with refined.
  4. Emit, per anchor, JIVO price / MRP / Discount % / per-litre vs each MATCHED
     rival, modal across the captured pincodes. Where no like-for-like rival
     exists, say "no direct competitor" honestly - never force a wrong oil.

Outputs (Competitor- prefix, NEVER the mailer-globbed Jivo-*):
  output/Competitor-Price-Watch-<Platform>-<date>.xlsx   (the run's platform)
  output/Competitor-Price-Watch-AllQcomm-<date>.xlsx     (all platforms for <date>)

House style matches platforms/<p>/build_excel.py: JIVO green #008B3A headers,
freeze panes, autofilter, MODAL aggregation (most common value, ties -> lowest).

GUARDRAILS honoured:
  - reads ONLY tools/competitor/data/*_competitor_<date>.json (the combined book
    globs by <date>, so a TESTFIX run can never read a live YYYY-MM-DD capture).
  - writes ONLY to output/ with the "Competitor-Price-Watch-" prefix.
  - read-only on the config JSONs; never touches the price-match master.

Usage:
  build_competitor_report.py <platform> [date]
    <platform>  e.g. blinkit | zepto | flipkart-minutes
    [date]      YYYY-MM-DD (default: today, IST); any tag (e.g. TESTFIX) works.
Env overrides (sane defaults; used for offline smoke tests):
  COMPETITOR_DATA_DIR  default /root/ecom-intel/tools/competitor/data
  COMPETITOR_OUT_DIR   default /root/ecom-intel/output
  COMPETITOR_INPUT     explicit path to the run platform's capture json
"""
import json, os, re, sys, glob, datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..'))

# ---- palette / styling (house build_excel.py) ----
JIVO_GREEN = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO_GREEN)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=JIVO_GREEN, size=18)
SUB_FONT = Font(italic=True, color="555555", size=10)
SECTION_FILL = PatternFill("solid", fgColor="E2EFDA")   # pale green section band
SECTION_FONT = Font(bold=True, color="006100", size=11)
JIVO_FILL = PatternFill("solid", fgColor="D9EAD3")       # JIVO anchor row
RED = PatternFill("solid", fgColor="F4CCCC")             # rival CHEAPER than JIVO = threat
GREEN = PatternFill("solid", fgColor="D9EAD3")           # JIVO cheaper = good
YEL = PatternFill("solid", fgColor="FFF2CC")             # ~level
GREY = PatternFill("solid", fgColor="F2F2F2")            # no direct competitor
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
EPS_PCT = 0.02  # within 2% per-litre counts as "level", not a threat/win


# ---- config load --------------------------------------------------------------
def load_json(path):
    with open(path, 'r') as f:
        return json.load(f)


OIL_CLASSIFIER = load_json(os.path.join(HERE, 'oil_classifier.json'))
MATCH_MAP = load_json(os.path.join(HERE, 'competitor_match_map.json'))
MAPS = load_json(os.path.join(HERE, 'maps_to_jivo.json'))

# merged single-oil keyword table: oil_classifier is authoritative, augmented with
# the match-map's spelling variants (e.g. "til oil"/"til "). 'blend' stays a gate.
OIL_KW = {}
for _src in (OIL_CLASSIFIER.get('oil_type', {}), MATCH_MAP.get('oil_type_keywords', {})):
    for _tok, _kws in _src.items():
        if _tok == 'blend':
            continue
        OIL_KW.setdefault(_tok, set()).update(_kws)
BLEND_KW = list(OIL_CLASSIFIER['oil_type']['blend'])
SINGLE_PRIORITY = [t for t in OIL_CLASSIFIER['priority'] if t != 'blend']
GRADE_KW = OIL_CLASSIFIER['grade']            # canonical grade tokens (folds light/pure/classic -> extra_light)
GRADE_PRIORITY = OIL_CLASSIFIER['grade_priority']
ANCHORS = MATCH_MAP['anchors']                # the 9 JIVO anchors, authoritative order
DENSITY = MAPS.get('density_g_per_ml', {}).get('oil', 0.916)


def _norm(s):
    """Lowercase, fold kachchi/kachhi->kachi, hyphen/slash->space, collapse ws."""
    s = (s or '').lower()
    s = s.replace('kachchi', 'kachi').replace('kachhi', 'kachi')
    s = re.sub(r'[-/]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _kw_in(kw, text):
    """Word-boundary keyword test (so 'til' never matches 'lentil'/'until')."""
    kw = _norm(kw)
    if not kw:
        return False
    return re.search(r'(?<![a-z])' + re.escape(kw) + r'(?![a-z])', text) is not None


def oil_type_of(name):
    """Authoritative NAME-derived oil_type. Blend gate first; >1 distinct single
    oil also = blend; otherwise first match in priority order; else None."""
    n = _norm(name)
    if any(_kw_in(k, n) for k in BLEND_KW):
        return 'blend'
    hits = {tok for tok, kws in OIL_KW.items() if any(_kw_in(k, n) for k in kws)}
    if len(hits) > 1:
        return 'blend'
    for tok in SINGLE_PRIORITY:
        if tok in hits:
            return tok
    return next(iter(hits)) if hits else None


def grade_of(name):
    """Authoritative NAME-derived grade (most-specific first)."""
    n = _norm(name)
    for g in GRADE_PRIORITY:
        if any(_kw_in(k, n) for k in GRADE_KW.get(g, [])):
            return g
    return None


def modal(values):
    """Most common value, ties -> lowest (matches house modal_price)."""
    vals = [round(v) for v in values if v is not None]
    if not vals:
        return None
    cnt = Counter(vals)
    top = max(cnt.values())
    return min(v for v, n in cnt.items() if n == top)


def ist_today():
    return (datetime.datetime.utcnow() + datetime.timedelta(hours=5, minutes=30)).date().isoformat()


# ---- brand identity -----------------------------------------------------------
BRANDS_CFG = load_json(os.path.join(HERE, 'competitor_brands.json'))
OURS = [str(x).lower() for x in BRANDS_CFG.get('ours', ['jivo', 'sano'])]
ALIAS_TO_BRAND = {}
for _b in BRANDS_CFG.get('brands', []):
    ALIAS_TO_BRAND[_b['brand'].lower()] = _b['brand']
    for _a in _b.get('aliases', []) or []:
        if _a:
            ALIAS_TO_BRAND[_a.lower()] = _b['brand']


def is_jivo(r):
    raw = (r.get('brand') or '').lower()
    nm = (r.get('name') or '').lower()
    return any(o in raw or o in nm for o in OURS)


def brand_label(r):
    if is_jivo(r):
        return 'JIVO'
    raw = (r.get('brand') or '').strip().lower()
    if raw in ALIAS_TO_BRAND:
        return ALIAS_TO_BRAND[raw]
    nm = (r.get('name') or '').lower()
    for alias, brand in ALIAS_TO_BRAND.items():
        if alias and alias in nm:
            return brand
    return (r.get('brand') or 'Unknown').title()


def brand_allowed(r, allowed):
    rb = _norm(r.get('brand'))
    nm = _norm(r.get('name'))
    for a in allowed:
        al = _norm(a)
        if not al:
            continue
        if al == rb or _kw_in(al, nm) or (rb and _kw_in(al, rb)):
            return True
    return False


# ---- anchor matching ----------------------------------------------------------
def anchor_pack_label(a):
    v = a['vol_ml']
    return f"{v // 1000}L" if v % 1000 == 0 else f"{v}ml"


def jivo_matches(r, anchor):
    """A scraped JIVO row backs an anchor: oil_type + grade + pack (brand is ours,
    so brands_allowed / exclude_name are skipped)."""
    m = anchor['match']
    if r['_oil'] != m['oil_type']:
        return False
    if r['_grade'] not in m['grades_ok']:
        return False
    v = r.get('vol_ml')
    if v is None:
        return False
    return abs(v - m['pack_ml']) <= m.get('pack_tol', 0.15) * m['pack_ml']


def rival_matches(r, anchor):
    """A competitor row matches an anchor ONLY when every gate passes."""
    if r['_jivo']:
        return False
    m = anchor['match']
    if r['_oil'] is None or r['_oil'] != m['oil_type']:   # name-derived oil must equal anchor's
        return False
    if r['_grade'] not in m['grades_ok']:                  # grade never crossed
        return False
    v = r.get('vol_ml')
    if v is None or abs(v - m['pack_ml']) > m.get('pack_tol', 0.15) * m['pack_ml']:
        return False
    n = _norm(r.get('name'))
    if any(_kw_in(tok, n) for tok in m.get('exclude_name', [])):  # blends / wrong-grade tokens out
        return False
    return brand_allowed(r, m.get('brands_allowed', []))


def maps_bucket_for(anchor):
    """maps_to_jivo bucket that carries the JIVO BAU/MRP anchor per-litre."""
    for b in MAPS['buckets']:
        if b['category'] != anchor['oil_type'] or b['vol_ml'] != anchor['vol_ml']:
            continue
        bg = b.get('sub_grade')
        if bg is None or bg == anchor['grade']:
            return b
    return None


def aggregate(sub):
    """Modal roll-up of a set of rows for one SKU/anchor across pincodes."""
    sale = modal([r.get('sale') for r in sub])
    mrp = modal([r.get('mrp') for r in sub])
    pl = modal([r.get('per_litre') for r in sub])
    disc = round(100.0 * (mrp - sale) / mrp, 1) if (mrp and sale is not None) else None
    ranks = [r.get('rank') for r in sub if r.get('rank') is not None]
    name = Counter(r.get('name') for r in sub).most_common(1)[0][0]
    brand = Counter(r.get('_brand') for r in sub).most_common(1)[0][0]
    oil = Counter(r.get('_oil') for r in sub).most_common(1)[0][0]
    grade = Counter(r.get('_grade') for r in sub).most_common(1)[0][0]
    return {
        'sale': sale, 'mrp': mrp, 'pl': pl, 'disc': disc,
        'rank': modal(ranks) if ranks else None,
        'cities': {r.get('city') for r in sub},
        'pins': {r.get('pincode') for r in sub},
        'name': name, 'brand': brand, 'oil': oil, 'grade': grade,
        'vol': modal([r.get('vol_ml') for r in sub]),
        'instock': any(r.get('in_stock') for r in sub),
        'n': len(sub),
    }


def jivo_anchor_value(rows, anchor):
    """(agg_dict, source) for the JIVO side of an anchor. Prefer scraped JIVO rows;
    fall back to the maps_to_jivo BAU/MRP per-litre anchor."""
    sub = [r for r in rows if r['_jivo'] and jivo_matches(r, anchor)]
    if sub:
        return aggregate(sub), 'scraped'
    b = maps_bucket_for(anchor)
    if b is None:
        return None, 'none'
    vol = anchor['vol_ml']
    pl = b.get('jivo_bau_per_litre')
    mrp_pl = b.get('jivo_mrp_per_litre')
    sale = round(pl * vol / 1000.0) if pl is not None else None
    mrp = round(mrp_pl * vol / 1000.0) if mrp_pl is not None else None
    disc = round(100.0 * (mrp_pl - pl) / mrp_pl, 1) if (mrp_pl and pl is not None) else None
    return {
        'sale': sale, 'mrp': mrp, 'pl': round(pl) if pl is not None else None, 'disc': disc,
        'rank': None, 'cities': set(), 'pins': set(),
        'name': anchor.get('jivo_example', 'JIVO'), 'brand': 'JIVO',
        'oil': anchor['oil_type'], 'grade': anchor['grade'],
        'vol': vol, 'instock': True, 'n': 0,
    }, 'anchor (BAU)'


def rivals_for_anchor(rows, anchor):
    """Matched rival SKUs, one modal row per canonical, cheapest per-litre first."""
    matched = [r for r in rows if rival_matches(r, anchor)]
    by_canon = {}
    for r in matched:
        by_canon.setdefault(r.get('canonical') or (r.get('_brand'), r.get('name')), []).append(r)
    aggs = [aggregate(g) for g in by_canon.values()]
    aggs.sort(key=lambda a: (a['pl'] is None, a['pl'] if a['pl'] is not None else 0))
    return aggs


def verdict_for(rival_pl, jivo_pl):
    if rival_pl is None or jivo_pl is None:
        return "no rival data", None, None
    gap = rival_pl - jivo_pl
    gap_pct = round(100.0 * gap / jivo_pl, 1) if jivo_pl else None
    if gap < -EPS_PCT * jivo_pl:
        return "THREAT - rival cheaper", gap, gap_pct
    if gap > EPS_PCT * jivo_pl:
        return "JIVO cheaper", gap, gap_pct
    return "level", gap, gap_pct


# ---- xlsx helpers -------------------------------------------------------------
def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR
        cell.font = HDR_FONT
        cell.alignment = CEN
        cell.border = BORDER


def autosize(ws, maxw=46):
    for col in ws.columns:
        L = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[L].width = min(maxw, max(10, w + 2))


def pct_fraction(v):
    if v is None:
        return None
    return round(float(v) / 100.0, 4)


# ---- workbook builder ---------------------------------------------------------
def annotate(rows, platform):
    for r in rows:
        r['_platform'] = platform
        r['_jivo'] = is_jivo(r)
        r['_brand'] = brand_label(r)
        r['_oil'] = oil_type_of(r.get('name'))
        r['_grade'] = grade_of(r.get('name'))
        r['_anchor'] = ''
        if not r['_jivo']:
            for a in ANCHORS:
                if rival_matches(r, a):
                    r['_anchor'] = a['key']
                    break
    return rows


def build_workbook(captures, label, date, out_dir):
    """captures: list of (platform, rows, summary). Builds one workbook."""
    plats = [p for p, _, _ in captures]
    multi = len(plats) > 1
    all_rows = []
    for p, rows, _ in captures:
        all_rows.extend(annotate(rows, p))

    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    ws = wb.active
    ws.title = "Summary"
    disp = "All Quick-Commerce" if label == "AllQcomm" else label.replace('-', ' ').title()
    ws["A1"] = f"Competitor Price Watch - {disp}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    plat_list = ", ".join(sorted({p.replace('-', ' ').title() for p in plats}))
    ws["A2"] = (f"{date}  -  9 JIVO anchors vs name-matched rivals  -  "
                f"{len(all_rows)} datapoints across {plat_list}  -  "
                f"oil type derived from PRODUCT NAME (not the search query)  -  "
                f"modal per-litre (ties->lowest)")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:J2")

    comp_rows = [r for r in all_rows if not r['_jivo'] and r['_oil'] not in (None, 'blend')]
    kpis = [
        ("Platforms", len(set(plats))),
        ("Competitor brands", len({r['_brand'] for r in comp_rows})),
        ("Cities", len({r.get('city') for r in all_rows})),
        ("Pincodes", len({r.get('pincode') for r in all_rows})),
    ]
    for i, (k, v) in enumerate(kpis):
        c = 1 + i * 2
        ws.cell(row=4, column=c, value=k).font = Font(bold=True, size=10, color="555555")
        ws.cell(row=5, column=c, value=v).font = Font(bold=True, size=20, color=JIVO_GREEN)

    ws.cell(row=7, column=1, value="JIVO vs cheapest matched rival, per anchor").font = Font(bold=True, size=12)
    hdr = (["Platform"] if multi else []) + ["JIVO Anchor", "Oil / Grade / Pack", "JIVO Rs/L",
            "JIVO src", "# Rivals", "Cheapest Rival", "Rival Rs/L", "Gap Rs/L", "Gap %", "Verdict"]
    hr = 8
    for j, h in enumerate(hdr, 1):
        ws.cell(row=hr, column=j, value=h)
    style_header(ws, hr, len(hdr))
    rr = hr + 1
    for p, rows, _ in captures:
        prows = [r for r in all_rows if r['_platform'] == p]
        for a in ANCHORS:
            jv, src = jivo_anchor_value(prows, a)
            jpl = jv['pl'] if jv else None
            rivals = rivals_for_anchor(prows, a)
            if rivals:
                cheap = rivals[0]
                v, gap, gap_pct = verdict_for(cheap['pl'], jpl)
                cheap_br, cheap_pl = cheap['brand'], cheap['pl']
            else:
                cheap_br, cheap_pl, gap, gap_pct = "no direct competitor", None, None, None
                v = "no direct competitor"
            vals = ([p.replace('-', ' ').title()] if multi else []) + [
                a.get('jivo_example', a['key']),
                f"{a['oil_type']} / {a['grade']} / {anchor_pack_label(a)}",
                jpl, src, len(rivals), cheap_br, cheap_pl, gap, pct_fraction(gap_pct), v]
            for j, val in enumerate(vals, 1):
                cell = ws.cell(row=rr, column=j, value=val)
                cell.border = BORDER
                if j > 2 - (0 if multi else 0):
                    cell.alignment = CEN if j >= (4 if multi else 3) else LEFT
            base = 1 if multi else 0
            for col in (base + 3, base + 8):  # Rs/L, Gap Rs/L
                ws.cell(row=rr, column=col).number_format = '"Rs"#,##0'
            ws.cell(row=rr, column=base + 9).number_format = '0.0%'
            vc = ws.cell(row=rr, column=base + 10)
            gc = ws.cell(row=rr, column=base + 8)
            if "THREAT" in v:
                vc.fill = RED
                gc.fill = RED
            elif "JIVO cheaper" in v:
                vc.fill = GREEN
                gc.fill = GREEN
            elif v == "level":
                vc.fill = YEL
                gc.fill = YEL
            elif v == "no direct competitor":
                vc.fill = GREY
            rr += 1
    ws.freeze_panes = ws.cell(row=hr + 1, column=1).coordinate
    autosize(ws)

    # ---------- Sheet 2: Anchor Watch (one section per anchor) ----------
    ws = wb.create_sheet("Anchor Watch")
    cols = (["Platform"] if multi else []) + [
        "Brand", "Product", "Oil Type", "Grade", "Vol (ml)", "MRP Rs", "Sale Rs",
        "Discount %", "Rs/L", "Gap vs JIVO Rs/L", "Gap %", "Avg Rank", "Pins", "Verdict"]
    ncol = len(cols)
    for j, h in enumerate(cols, 1):
        ws.cell(row=1, column=j, value=h)
    style_header(ws, 1, ncol)
    ws.freeze_panes = "A2"
    base = 1 if multi else 0
    money_cols = [base + 5, base + 6, base + 8, base + 9]   # MRP, Sale, Rs/L, Gap Rs/L
    rr = 2

    def write_row(values, fill=None, bold=False):
        nonlocal rr
        for j, val in enumerate(values, 1):
            if j in (base + 7, base + 10) and isinstance(val, (int, float)):
                val = pct_fraction(val)
            cell = ws.cell(row=rr, column=j, value=val)
            cell.border = BORDER
            if j >= (3 if multi else 2):
                cell.alignment = CEN
            if bold:
                cell.font = Font(bold=True)
            if fill:
                cell.fill = fill
        for col in money_cols:
            ws.cell(row=rr, column=col).number_format = '"Rs"#,##0'
        ws.cell(row=rr, column=base + 7).number_format = '0.0%'   # Discount %
        ws.cell(row=rr, column=base + 10).number_format = '0.0%'  # Gap %
        rr += 1

    for idx, a in enumerate(ANCHORS, 1):
        # section band
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=ncol)
        band = ws.cell(row=rr, column=1,
                       value=f"{idx}. {a.get('jivo_example', a['key'])}   |   "
                             f"{a['oil_type']} / {a['grade']} / {anchor_pack_label(a)}")
        band.fill = SECTION_FILL
        band.font = SECTION_FONT
        band.alignment = LEFT
        rr += 1
        for p, rows, _ in captures:
            prows = [r for r in all_rows if r['_platform'] == p]
            jv, src = jivo_anchor_value(prows, a)
            jpl = jv['pl'] if jv else None
            ptag = [p.replace('-', ' ').title()] if multi else []
            # JIVO anchor row
            if jv:
                write_row(ptag + ["JIVO", f"{jv['name']}  [{src}]", jv['oil'], jv['grade'],
                                  jv['vol'], jv['mrp'], jv['sale'], jv['disc'], jpl,
                                  None, None, jv['rank'], len(jv['pins']) or None, "JIVO anchor"],
                          fill=JIVO_FILL, bold=True)
            # rivals
            rivals = rivals_for_anchor(prows, a)
            if rivals:
                for rv in rivals:
                    v, gap, gap_pct = verdict_for(rv['pl'], jpl)
                    write_row(ptag + [rv['brand'], rv['name'], rv['oil'], rv['grade'],
                                      rv['vol'], rv['mrp'], rv['sale'], rv['disc'], rv['pl'],
                                      gap, gap_pct, rv['rank'], len(rv['pins']) or None, v])
                    cell = ws.cell(row=rr - 1, column=base + 9)   # Rs/L cell
                    gcell = ws.cell(row=rr - 1, column=base + 10)  # Gap Rs/L
                    if "THREAT" in v:
                        cell.fill = RED
                        gcell.fill = RED
                    elif "JIVO cheaper" in v:
                        cell.fill = GREEN
                        gcell.fill = GREEN
                    elif v == "level":
                        cell.fill = YEL
                        gcell.fill = YEL
            else:
                note = "no direct competitor" + ("" if a['match'].get('allow_no_match', True)
                                                 else " (none found - expected one)")
                write_row(ptag + ["-", f"-- {note} --", a['oil_type'], a['grade'],
                                  None, None, None, None, None, None, None, None, None, ""],
                          fill=GREY)
        rr += 1   # blank spacer
    autosize(ws, maxw=52)

    # ---------- Sheet 3: Master Data (audit: NAME-derived oil_type + grade) ----------
    ws = wb.create_sheet("Master Data")
    cols = ["Platform", "Brand", "JIVO?", "Search Category (scraped)", "Oil Type (name)",
            "Grade (name)", "Matched Anchor", "Blend?", "City", "Pincode", "Store ID",
            "Name", "Pack", "Vol (ml)", "MRP Rs", "Sale Rs", "Rs/L", "Discount %",
            "In stock", "Rank", "Ad?", "Captured"]
    ws.append(cols)
    for x in sorted(all_rows, key=lambda r: (r['_platform'], r['_oil'] or 'zzz',
                                             r['_grade'] or '', r.get('_brand') or '',
                                             r.get('city') or '', r.get('pincode') or '')):
        ws.append([
            x['_platform'].replace('-', ' ').title(), x.get('_brand'),
            "Yes" if x['_jivo'] else "", x.get('category'), x.get('_oil') or '',
            x.get('_grade') or '', x.get('_anchor') or '',
            "Yes" if x.get('_oil') == 'blend' else "",
            x.get('city'), x.get('pincode'), x.get('store_id'),
            x.get('name'), x.get('pack'), x.get('vol_ml'), x.get('mrp'), x.get('sale'),
            x.get('per_litre'), pct_fraction(x.get('discount_pct')),
            "Yes" if x.get('in_stock') else "No", x.get('rank'),
            "Yes" if x.get('is_ad') else "", (x.get('captured_at') or '')[:16].replace('T', ' '),
        ])
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            if cell.column in (15, 16, 17):
                cell.number_format = '"Rs"#,##0'
            if cell.column == 18:
                cell.number_format = '0.0%'
        if row[2].value == "Yes":
            row[1].fill = JIVO_FILL
        if row[7].value == "Yes":
            row[4].fill = YEL              # blend flagged
        if row[18].value == "No":
            row[18].fill = RED
    autosize(ws)

    os.makedirs(out_dir, exist_ok=True)
    plat_file = ("AllQcomm" if label == "AllQcomm"
                 else label.replace('-', ' ').title().replace(' ', '-'))
    fname = os.path.join(out_dir, f"Competitor-Price-Watch-{plat_file}-{date}.xlsx")
    assert not os.path.basename(fname).startswith("Jivo-"), "filename must never start with Jivo-"
    wb.save(fname)
    return fname


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: build_competitor_report.py <platform> [date]")
    platform = sys.argv[1]
    date = sys.argv[2] if len(sys.argv) > 2 else ist_today()

    data_dir = os.environ.get('COMPETITOR_DATA_DIR', os.path.join(HERE, 'data'))
    out_dir = os.environ.get('COMPETITOR_OUT_DIR', os.path.join(ROOT, 'output'))
    in_path = os.environ.get('COMPETITOR_INPUT') or os.path.join(data_dir, f'{platform}_competitor_{date}.json')

    if not os.path.exists(in_path):
        sys.exit(f"competitor capture not found: {in_path}\n"
                 f"  run the scraper first: bash {HERE}/run_competitor.sh {platform}")

    cap = load_json(in_path)
    rows = cap.get('allRows', []) or []
    summary = cap.get('summary', {}) or {}

    produced = []
    # 1) per-platform workbook
    produced.append(build_workbook([(platform, rows, summary)], platform, date, out_dir))

    # 2) combined AllQcomm workbook: every *_competitor_<date>.json for THIS date.
    #    Globbing by <date> guarantees a TESTFIX run never reads a live YYYY-MM-DD capture.
    captures = []
    for path in sorted(glob.glob(os.path.join(data_dir, f'*_competitor_{date}.json'))):
        base = os.path.basename(path)
        p = base[:-len(f'_competitor_{date}.json')]
        c = load_json(path)
        captures.append((p, c.get('allRows', []) or [], c.get('summary', {}) or {}))
    if not captures:
        captures = [(platform, rows, summary)]
    produced.append(build_workbook(captures, "AllQcomm", date, out_dir))

    print("SAVED:")
    for f in produced:
        print("  ", f)
    print("Anchors:", len(ANCHORS), "| platforms in combined:", len({p for p, _, _ in captures}))


if __name__ == "__main__":
    main()
