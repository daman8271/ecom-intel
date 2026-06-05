#!/usr/bin/env python3
"""PRICE-MATCH dry-run: map every distinct platform listing -> master PRODUCT.
Prototype of the production matcher. Read-only; writes /tmp/pricematch/dryrun.json"""
import json, re, pathlib
from collections import Counter, defaultdict

BASE = pathlib.Path("/opt/ecom-intel/platforms")
OUT  = pathlib.Path("/tmp/pricematch")

# ---------------- master ----------------
import openpyxl
wb = openpyxl.load_workbook(OUT/"Amazon Price Match.xlsx", data_only=True)
ws = wb['url']; hdr=[c.value for c in ws[1]]; ix={h:i for i,h in enumerate(hdr)}
master=[]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[ix['PRODUCT']]: continue
    master.append(dict(
        product=r[ix['PRODUCT']].strip().upper(), asin=r[ix['ASIN']].strip(),
        title=r[ix['Title']], mrp=r[ix['MRP']],
        bau=r[ix['BAU Price']], svd=r[ix['SVD Price']], art=r[ix['ART PRICE']],
        fk_url=r[ix['FLIPKART URL']], bl_url=r[ix['BLINKIT URL']],
        zp_url=r[ix['ZEPTO URL']], sw_url=r[ix[' URL']], jm_url=r[ix['JIO MART URL']]))

def urlid(pat,u):
    m=re.search(pat,u or ''); return m.group(1) if m else None
zp_pvid2prod={urlid(r'/pvid/([a-f0-9-]+)',m['zp_url']):m['product'] for m in master if m['zp_url']}
fk_pid2prod ={urlid(r'pid=([A-Z0-9]+)',m['fk_url']):m['product'] for m in master if m['fk_url']}
asin2prod   ={m['asin']:m['product'] for m in master}
prodset     ={m['product'] for m in master}

# what is SO OLIVE / SANO? print titles for context
for m in master:
    if m['product'].startswith(('SO OLIVE','SANO CLASSIC 1','GOLD 1')):
        print(f"[master context] {m['product']:18s} -> {str(m['title'])[:90]}")

# ---------------- catalogs (internal names already maintained by ecom dept) ----------------
amz_cat=json.loads((BASE/'amazon/products.json').read_text())
asin2item={c['asin']:(c.get('item') or '').strip().upper() for c in amz_cat if c.get('asin')}

# ---------------- line/keyword ladder (ORDER MATTERS) ----------------
LADDER=[
 ('RICE BRAN',       r'rice\s*bran'),
 ('SUNFLOWER SEEDS', r'sunflower\s*seed'),
 ('PUMPKIN SEEDS',   r'pumpkin'),
 ('FLAX SEEDS',      r'flax'),
 ('CHIA SEEDS',      r'chia'),
 ('QUINOA SEEDS',    r'quinoa'),
 ('BASIL SEEDS',     r'basil'),
 ('BLACK OLIVE',     r'black\s*olive'),
 ('EXTRA LIGHT',     r'extra\s*li(?:ght|te)'),
 ('EXTRA VIRGIN',    r'extra\s*virgin'),
 ('POMACE',          r'pomace'),
 ('CANOLA',          r'canola'),
 ('GROUNDNUT',       r'groundnut|peanut'),
 ('COCONUT',         r'coconut'),
 ('YELLOW MUSTARD',  r'yellow\s*mustard'),
 ('MUSTARD',         r'mustard|sarso|kachi\s*ghani'),
 ('SESAME OIL',      r'sesame|gingelly|\btil\b'),
 ('SUNFLOWER',       r'sunflower'),
 ('SOYABEAN',        r'soya\s*bean|soyabean|soybean|\bsoya\b'),
 ('A2 GHEE',         r'ghee'),
 ('GOLD',            r'\bgold\b|blended'),
 ('CLASSIC',         r'classic'),
 ('BLACK CARDAMOM',  r'black\s*cardamom|badi\s*elaichi'),
 ('GREEN CARDAMOM',  r'green\s*cardamom|cardamom|elaichi'),
 ('BLACK PAPER',     r'black\s*pa?pper|black\s*paper|kali\s*mirch|black\s*pepper'),
 ('CINNAMON BARK',   r'cinnamon|dalchini'),
 ('CLOVE',           r'clove|laung'),
 ('ROSEMARY LEAVES', r'rosemary'),
 ('SAFFRON',         r'saffron|kesar'),
 ('SODA LEMON',      r'lemon\s*soda|soda\s*lemon|nimbu\s*soda'),
 ('SODA',            r'\bsoda\b'),
 ('SPRING WATER',    r'spring\s*water|\bwater\b'),
 ('COFFEE',          r'coffee|koffie'),
 ('RICE',            r'\brice\b'),
 ('JEERA',           r'jeera|cumin'),   # resolved to PUNJABI JEERA (ml) vs CUMIN SEEDS (g) below
 ('SO OLIVE',        r'olive'),         # generic olive fallback (after pomace/EL/EV/black-olive)
]
COMBO_RE=re.compile(r'\+|\bcombo\b|\bset of\s*[2-9]|\bpack of\s*([2-9])|buy\s*\d+\s*get', re.I)
POUCH_RE=re.compile(r'pouch', re.I)
SANO_RE =re.compile(r'\bdi\s*sano\b|\bdisano\b|\bsano\b', re.I)

def detect_line(text, uom_ml=None):
    t=' '+text.lower()+' '
    for line,pat in LADDER:
        if re.search(pat,t):
            if line=='JEERA':
                return 'PUNJABI JEERA' if (uom_ml and uom_ml>=100) else 'CUMIN SEEDS'
            return line
    return None

def magnitude(s):
    """trailing size on a master PRODUCT name -> magnitude in ml/g"""
    s=s.replace('GMS','G').replace('GM','G').replace('LTR','L').replace('LITRE','L')
    m=re.search(r'(\d+(?:\.\d+)?)\s*(ML|KG|G|L)\b', s)
    if not m: return None
    n=float(m.group(1)); u=m.group(2)
    return n*1000 if u in ('KG','L') else n

# parse master into signatures
def parse_master(m):
    p=m['product']
    brand='SANO' if p.startswith('SANO') else 'JIVO'
    combo=bool(re.search(r'\+',p))
    pouch=bool(POUCH_RE.search(p))
    line=detect_line(p, uom_ml=magnitude(p))
    mag=magnitude(p)
    if combo:  # e.g. 1+1L -> 2000
        parts=re.findall(r'(\d+)\s*\+\s*(\d+)\s*L',p)
        if parts: mag=(int(parts[0][0])+int(parts[0][1]))*1000
    return dict(brand=brand,line=line,mag=mag,combo=combo,pouch=pouch)
for m in master: m['sig']=parse_master(m)
sigidx=defaultdict(list)
for m in master:
    s=m['sig']; sigidx[(s['brand'],s['line'],s['mag'],s['combo'],s['pouch'])].append(m)

def name_match(raw, vol_ml, pack, mrp):
    text=f"{raw} {pack or ''}"
    brand='SANO' if SANO_RE.search(text) else 'JIVO'
    combo=bool(COMBO_RE.search(re.sub(r'\(pack of 1\)','',text,flags=re.I)))
    pouch=bool(POUCH_RE.search(text))
    line=detect_line(text, uom_ml=vol_ml)
    cands=sigidx.get((brand,line,vol_ml,combo,pouch),[])
    if not cands and pouch:  # pouch listing but master may only have bottle
        cands=sigidx.get((brand,line,vol_ml,combo,False),[])
    if len(cands)==1: return cands[0]['product'],'AUTO-NAME',[c['product'] for c in cands]
    if len(cands)>1:
        # MRP anchor tiebreak
        hit=[c for c in cands if mrp and c['mrp']==mrp]
        if len(hit)==1: return hit[0]['product'],'AUTO-NAME+MRP',[c['product'] for c in cands]
        return None,'AMBIGUOUS',[c['product'] for c in cands]
    return None,'NOT-IN-MASTER',[]

# ---------------- distinct listings per platform ----------------
def distinct(plat, keyf, idfields=()):
    rows=json.loads((BASE/plat/'result.json').read_text())['allRows']
    groups=defaultdict(list)
    for r in rows: groups[keyf(r)].append(r)
    out=[]
    for k,rs in groups.items():
        if k is None: continue
        raw=Counter(r.get('sku_raw') or '' for r in rs).most_common(1)[0][0]
        mrp=Counter(r.get('mrp') for r in rs if r.get('mrp')).most_common(1)
        out.append(dict(key=str(k), raw=raw,
            vol=Counter(r.get('vol_ml') for r in rs).most_common(1)[0][0],
            pack=Counter(r.get('pack') or '' for r in rs).most_common(1)[0][0],
            mrp=mrp[0][0] if mrp else None,
            extra={f:rs[0].get(f) for f in idfields}))
    return out

result={}
def classify(plat, listings, idjoin=None):
    res=[]
    for L in listings:
        prod=method=None; cands=[]
        if idjoin:
            prod,method=idjoin(L)
        if not prod:
            prod,method,cands=name_match(L['raw'],L['vol'],L['pack'],L['mrp'])
        combo=bool(COMBO_RE.search(re.sub(r'\(pack of 1\)','',L['raw'],flags=re.I)))
        res.append(dict(**{k:L[k] for k in ('key','raw','vol','mrp')},
                        match=prod,method=method,cands=cands,combo=combo))
    result[plat]=res
    c=Counter(r['method'] for r in res)
    print(f"\n{'='*18} {plat}  ({len(res)} distinct listings) {dict(c)}")
    for r in res:
        tag='COMBO' if r['combo'] else '     '
        if r['method'] in ('NOT-IN-MASTER','AMBIGUOUS'):
            print(f"  !! {r['method']:13s} {tag} {str(r['raw'])[:64]:64s} vol={r['vol']} mrp={r['mrp']} cands={r['cands']}")
        else:
            print(f"     {r['method']:13s} {tag} {str(r['raw'])[:64]:64s} -> {r['match']}")

# amazon family: asin -> master, else asin -> catalog item -> master, else name
def amz_join(L):
    a=L['extra']['asin']
    if a in asin2prod: return asin2prod[a],'ID-ASIN'
    item=asin2item.get(a)
    if item and item in prodset: return item,'ID-CATALOG'
    if item: return None,None  # has internal name but unpriced -> fall to name/NOT-IN-MASTER
    return None,None

classify('amazon',        distinct('amazon',        lambda r:r.get('asin'), ('asin',)), amz_join)
classify('amazon-fresh',  distinct('amazon-fresh',  lambda r:r.get('asin'), ('asin',)), amz_join)
classify('amazon-now',    distinct('amazon-now',    lambda r:r.get('asin'), ('asin',)), amz_join)

# flipkart: item name is the ecom dept's own internal name
def fk_join(L):
    item=(L['extra'].get('item') or '').strip().upper()
    fsn=L['extra'].get('fsn')
    if fsn in fk_pid2prod: return fk_pid2prod[fsn],'ID-FSN'
    if item in prodset: return item,'ID-CATALOG'
    return None,None
classify('flipkart', distinct('flipkart', lambda r:r.get('fsn'), ('fsn','item')), fk_join)

# zepto: variant_id -> master pvid
def zp_join(L):
    v=L['extra'].get('variant_id')
    if v in zp_pvid2prod: return zp_pvid2prod[v],'ID-PVID'
    return None,None
classify('zepto', distinct('zepto', lambda r:r.get('variant_id'), ('variant_id',)), zp_join)

classify('blinkit',          distinct('blinkit',          lambda r:r.get('canonical')))
classify('flipkart-minutes', distinct('flipkart-minutes', lambda r:r.get('canonical')))
classify('bigbasket',        distinct('bigbasket',        lambda r:r.get('sku_id'), ('sku_id','brand')))

# ---------------- summary ----------------
print("\n"+"="*70)
print("MASTER COVERAGE — which of the 96 priced SKUs we can see, per platform")
seen=defaultdict(set)
for plat,res in result.items():
    for r in res:
        if r['match']: seen[r['match']].add(plat)
mapped=set(seen)
print(f"master SKUs visible somewhere: {len(mapped)}/96; invisible: {sorted(prodset-mapped)[:50]}")
plat_count=Counter(p for s in seen.values() for p in s)
print("per-platform master-SKU coverage:", dict(plat_count))

# price-source duality checks
zr=json.loads((BASE/'zepto/result.json').read_text())['allRows']
print("\nzepto price_source values:", Counter(r.get('price_source') for r in zr))
br=json.loads((BASE/'bigbasket/result.json').read_text())['allRows']
print("bigbasket dual_pricing:", Counter(str(r.get('dual_pricing')) for r in br))
print("bigbasket sp_source:", Counter(r.get('sp_source') for r in br))

json.dump(result, open(OUT/'dryrun.json','w'), indent=1, default=str)
print("\nsaved", OUT/'dryrun.json')
