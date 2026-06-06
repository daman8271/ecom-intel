#!/usr/bin/env python3
"""master-v2: ingest team_sheet_2026-06-06.xlsx (113 SKUs, amazon_price_updated 17-col
format) -> master_v2.json, then extend sku_map.json in place:

  * renames merged as identity (ASIN-asserted): BLACK PAPER 100G -> BLACK PEPPER 100G,
    ROSEMARY LEAVES 150G -> ROSEMARY LEAVES 150 GMS
  * RICE 1KG marked retired (mapping kept)
  * regime prices (mrp/bau/svd/art) refreshed for every sheet SKU — the v2 sheet is THE master
  * 18 brand-new SKUs (12 same-oil 1L+1L combos + 6 beverages) mapped across platforms by
    REUSING frozen evidence only (no scraping, no LLM):
      - amazon: ASIN join vs platforms/amazon/result.json (sheet ASIN is identity-grade)
      - flipkart: exact INTERNAL catalog-name match in fragments/map-flipkart.json unpriced
        (catalog name = Jivo's own identity; listing MRP must not CONTRADICT the sheet —
        present-and-different => review, never silently mapped)
      - zepto: census-classified same-composition combo pvids in fragments/map-qcom.json,
        sheet-MRP anchored
      - bigbasket: EAN/sku_id join vs platforms/bigbasket/result.json (WG MANGO 500ML)
      - amazon-now: ASIN join vs platforms/amazon-now/result.json
      - flipkart-minutes: name+MRP-anchored water slug (id-fold fills the real pid)
  * owner-ratified rules enforced in code: combo -> same-composition only; name-match needs
    the sheet-MRP anchor else review; same-platform collision -> review; NEVER silently guess.
  * mapped listings are removed from sku_map['unpriced'] singles; the resolved
    B0CZP26VVN-duplicate review item is retired (it is now its own master row CANOLA 1L + 1L)
  * in-stock amazon listings whose live MRP diverges from the v2 sheet -> mrp_drift
  * finally re-applies apply_id_fold.py so folded ids/eans/urls survive any re-run

Deterministic, offline, idempotent — safe to re-run.
"""
import json, pathlib, subprocess, sys
from collections import Counter
from openpyxl import load_workbook

HERE = pathlib.Path(__file__).parent
ROOT = HERE.parent.parent
SHEET = HERE / "team_sheet_2026-06-06.xlsx"
V2_DATE = "2026-06-06"

RENAMES = {"BLACK PAPER 100G": "BLACK PEPPER 100G",
           "ROSEMARY LEAVES 150G": "ROSEMARY LEAVES 150 GMS"}
RETIRED = ["RICE 1KG"]

# ---------------- 1) ingest the sheet ----------------
ws = load_workbook(SHEET, read_only=True, data_only=True)["Sheet1"]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c else c for c in rows[0]]
ix = {h: i for i, h in enumerate(hdr)}
def cell(r, col):
    v = r[ix[col]]
    return v.strip() if isinstance(v, str) else v

master = {}
for r in rows[1:]:
    if not r[ix["PRODUCT"]]:
        continue
    p = cell(r, "PRODUCT")
    master[p] = {
        "product": p, "asin": cell(r, "ASIN"), "url": cell(r, "URL"),
        "mrp": cell(r, "MRP"), "asp": cell(r, "ASP"),
        "svd": cell(r, "SVD Price"), "bau": cell(r, "BAU Price"), "art": cell(r, "ART PRICE"),
        "url_price": cell(r, "URL PRICE"), "stock_status": cell(r, "STOCK STATUS"),
        "seller": cell(r, "SELLER"),
    }
assert len(master) == 113, f"expected 113 sheet SKUs, got {len(master)}"

(HERE / "master_v2.json").write_text(json.dumps({
    "version": V2_DATE,
    "source": f"team_sheet_{V2_DATE}.xlsx (e-com team master price sheet, amazon_price_updated 17-col format)",
    "renames": RENAMES, "dropped_from_sheet": RETIRED,
    "skus": {k: master[k] for k in sorted(master)},
}, indent=1, ensure_ascii=False))
print(f"master_v2.json: {len(master)} skus")

# ---------------- 2) load map + helpers ----------------
sm = json.loads((HERE / "sku_map.json").read_text())
skus = sm["skus"]

def load_json(p):
    return json.loads((ROOT / p).read_text())

amz = {r["asin"]: r for r in reversed(load_json("platforms/amazon/result.json")["allRows"]) if r.get("asin")}
fk_frag = load_json("tools/pricematch/fragments/map-flipkart.json")
qcom_frag = load_json("tools/pricematch/fragments/map-qcom.json")
fk_unpriced = {u["id"]: u for u in fk_frag["unpriced"] if u.get("id")}
qcom_unpriced = {u["id"]: u for u in qcom_frag["unpriced"] if u.get("id")}

# ---------------- 3) renames (identity merges, ASIN-asserted) ----------------
for old, new in RENAMES.items():
    if old in skus:
        old_asin = (skus[old].get("platforms", {}).get("amazon") or {}).get("id")
        assert old_asin == master[new]["asin"], \
            f"rename {old}->{new}: ASIN mismatch {old_asin} vs {master[new]['asin']} — NOT the same SKU, aborting"
        skus[new] = skus.pop(old)
        skus[new]["renamed_from"] = old
        print(f"renamed: {old} -> {new} (ASIN {old_asin} identical, merged)")

# ---------------- 4) retired ----------------
for p in RETIRED:
    if p in skus:
        skus[p]["retired"] = True
        skus[p]["retired_note"] = f"dropped from team master sheet v2 ({V2_DATE}); mapping kept for history"

# ---------------- 5) regimes refresh (the v2 sheet is THE master) ----------------
for p, m in master.items():
    skus.setdefault(p, {"regimes": {}, "platforms": {}})
    skus[p]["regimes"] = {"mrp": m["mrp"], "bau": m["bau"], "svd": m["svd"], "art": m["art"]}

# the 18 rows with no v1 entry (the 2 renames merge into existing entries instead).
# Pinned explicitly (not derived) so a re-run after the entries exist stays idempotent.
NEW_SKUS = sorted(['CANOLA 1L + 1L', 'EXTRA LIGHT 1+1L', 'EXTRA VIRGIN 1L + 1L', 'GINGER ALE 200ML',
                   'GOLD 1+1', 'JIVO POMACE 1L + 1L', 'JIVO PUNJABI SHIKANJI 160 MLS', 'JIVO WATER 1L',
                   'JIVO WATER 250ML', 'JIVO WATER 500 MLS', 'MUSTARD 1L + 1L', 'RICE BRAN 1L + 1L',
                   'SESAME 1L + 1L', 'SO OLIVE 1L + 1L', 'SOYABEAN 1L + 1L', 'SUNFLOWER 1L + 1L',
                   'WG MANGO JUICE 500ML', 'YELLOW MUSTARD 1L + 1L'])
assert all(p in master for p in NEW_SKUS)

reviews_add, drift_add, mapped_ids, summary = [], [], set(), {p: [] for p in NEW_SKUS}

def add_review(platform, id_, url, title, candidate, reason):
    reviews_add.append({"platform": platform, "id": id_, "url": url,
                        "title": title, "candidate": candidate, "reason": reason})
    summary[candidate].append(f"{platform}: -> REVIEW ({reason[:80]}…)" if len(reason) > 80
                              else f"{platform}: -> REVIEW ({reason})")

def set_entry(sku, platform, entry):
    skus[sku]["platforms"][platform] = entry
    mapped_ids.add((platform, str(entry["id"])))
    summary[sku].append(f"{platform}: {entry['method']} {entry['confidence']}")

# ---------------- 6) amazon: ASIN join (sheet ASIN = identity) ----------------
AMZ_NOTES = {
    "CANOLA 1L + 1L": "verifier-confirmed 2026-06-05 ('Pack of 2L' = 2x1L canola); was held as a known "
                      "DUPLICATE of CANOLA 1+1L (B0152TWWSQ) — team sheet v2 gives it its own master row, "
                      "old review item retired",
    "GINGER ALE 200ML": "wheatgrass-census identity: this ASIN is 'Wheatgrass Juice Sugar Free (Ginger Ale) 200ml'; "
                        "live amazon MRP 480 looks marketplace-inflated (census: BB 200ml flavours carry clean MRP 50)",
    "WG MANGO JUICE 500ML": "live amazon MRP 480 vs sheet 100 — census flags amazon WG MRPs as seller-inflated; "
                            "BB EAN row is the clean anchor",
    "JIVO WATER 1L": "live amazon MRP 1800 vs sheet 30 — census red flag: almost certainly a case-pack listing "
                     "sold per-case; see review",
    "SUNFLOWER 1L + 1L": "sheet ASIN is identity, but the scraped row for it reads '1 Litre' mrp 295 (a th=1 "
                         "variant capture?) vs sheet 1L+1L mrp 550; see review",
}
for p in NEW_SKUS:
    m = master[p]
    row = amz.get(m["asin"])
    if row:
        e = {"id": m["asin"], "url": f"https://www.amazon.in/dp/{m['asin']}",
             "title": row.get("sku_raw"), "mrp": row.get("mrp"),
             "sale_modal": row.get("sale"), "sale_min": row.get("sale"), "sale_max": row.get("sale"),
             "in_stock": bool(row.get("in_stock")), "method": "ID-ASIN", "confidence": "exact"}
        if p in AMZ_NOTES:
            e["note"] = AMZ_NOTES[p]
        set_entry(p, "amazon", e)
        if e["in_stock"] and e["mrp"] is not None and abs(e["mrp"] - m["mrp"]) > 1:
            drift_add.append({"product": p, "platform": "amazon", "platform_mrp": e["mrp"],
                              "official_mrp": m["mrp"],
                              "consensus": "v2 sheet vs live in-stock listing (new SKU, single platform)"})
    else:
        set_entry(p, "amazon", {
            "id": m["asin"], "url": m["url"], "title": None, "mrp": None,
            "sale_modal": None, "sale_min": None, "sale_max": None, "in_stock": None,
            "method": "ID-ASIN", "confidence": "exact",
            "note": f"identity from team master sheet v2 ({V2_DATE}); ASIN not in the 314-ASIN targeted "
                    f"scrape set yet (sheet says {m['stock_status']}"
                    + (f", seller {m['seller']}, url price {m['url_price']}" if m["seller"] else "") + ")"})

# variant-capture review for SUNFLOWER 1L + 1L
_sf = amz.get(master["SUNFLOWER 1L + 1L"]["asin"], {})
add_review("amazon", "B0992GRH3Z", "https://www.amazon.in/dp/B0992GRH3Z", _sf.get("sku_raw"),
           "SUNFLOWER 1L + 1L",
           "sheet ASIN mapped (identity), but scraped row reads '1 Litre' mrp 295 vs sheet 1L+1L mrp 550 — "
           "th=1 variant dimension suspected; confirm the ?th=1 variant is the 2-pack")
# case-pack review for JIVO WATER 1L
add_review("amazon", "B0CSPMN88F", "https://www.amazon.in/dp/B0CSPMN88F",
           amz.get("B0CSPMN88F", {}).get("sku_raw"), "JIVO WATER 1L",
           "sheet ASIN mapped (identity), but live MRP 1800 vs sheet 30 — census red flag: case-pack "
           "(12/24 bottles) sold per-case; per-bottle price-match must divide by case size once confirmed")

# ---------------- 7) flipkart: exact internal catalog-name matches ----------------
# (id, expected internal name) — MRP rule: equal->anchored exact; null->catalog identity stands;
# present-and-different => the code below throws the candidate to review instead.
FK_CANDIDATES = {
    "EXTRA LIGHT 1+1L":     ("QWRGGC46UKQKVCES", "EXTRA LIGHT 1+1L"),
    "EXTRA VIRGIN 1L + 1L": ("QWRGEMNS9ZPZGWJH", "Extra Virgin 1+1L"),
    "GOLD 1+1":             ("EDOGYJ7SYY4H2MDH", "GOLD 1+1L"),
    "SO OLIVE 1L + 1L":     ("EDOHBH2URWDMHPXC", "SO OLIVE 1+1L"),
    "SOYABEAN 1L + 1L":     ("EDOGGPFZ2FCJ6UWS", "SOYABEAN 1+1L"),
}
for sku, (pid, internal) in FK_CANDIDATES.items():
    u = fk_unpriced[pid]
    assert u.get("internal_name") == internal and u.get("combo"), (sku, pid)
    if u.get("mrp") is not None and abs(u["mrp"] - master[sku]["mrp"]) > 0:
        add_review("flipkart", pid, u.get("url"), u.get("title"), sku,
                   f"catalog internal name '{internal}' matches but listing mrp {u['mrp']} != sheet mrp "
                   f"{master[sku]['mrp']} — MRP conflict, not mapped")
        continue
    anchored = u.get("mrp") is not None
    set_entry(sku, "flipkart", {
        "id": pid, "url": u.get("url"), "title": u.get("title"), "item": internal,
        "sap_code": None, "mrp": u.get("mrp"), "sale_modal": None, "sale_min": None, "sale_max": None,
        "in_stock": None, "method": "ID-CATALOG", "confidence": "exact",
        "note": "exact Jivo-catalog internal-name match (frozen fragment evidence, no re-scrape); "
                + (f"listing mrp {u['mrp']} == sheet mrp (anchor confirmed)" if anchored
                   else "no live MRP on the listing — catalog identity stands, nothing contradicts")})

# flipkart REVIEW items (collisions / MRP conflicts / composition mismatches) — owner-ratified rules
add_review("flipkart", "EDOFRXTMJGAE9DCT", "https://www.flipkart.com/product/p/itme?pid=EDOFRXTMJGAE9DCT",
           "JIVO Cold Pressed Canola Oil 1+1 Ltr Canola Oil PET Bottle", "CANOLA 1L + 1L",
           "same-platform collision: flipkart's single 'CANOLA 1+1L' catalog FSN is already held by master "
           "CANOLA 1+1L (B0152TWWSQ family); the new CANOLA 1L + 1L (B0CZP26VVN) is the same composition — "
           "owner to confirm whether flipkart carries a second distinct FSN or the slot stays with v1")
for pid, mrp in [("EDOHYRRPJ3H7NHHN", 1899), ("QWRGEMQJ8TQRHASY", None)]:
    u = fk_unpriced[pid]
    add_review("flipkart", pid, u.get("url"), u.get("title"), "JIVO POMACE 1L + 1L",
               f"TWO FSNs share internal name 'JIVO POMACE 1+1L' (collision) and this one's mrp "
               f"{mrp} != sheet 2098 — owner pick" if mrp else
               "TWO FSNs share internal name 'JIVO POMACE 1+1L' (collision); this one has no live MRP — owner pick")
for pid, iname, mrp in [("QWRGEMQZXAYWVM4M", "MUSTARD 1+1L", 510),
                        ("EDOGE62HWA7UFXHW", "JIVO MUSTARD 1+1L", 510),
                        ("EDOHYSJ8BZYGCSBU", "JIVO MUSTARD 1+1L", None)]:
    u = fk_unpriced[pid]
    add_review("flipkart", pid, u.get("url"), u.get("title"), "MUSTARD 1L + 1L",
               f"3-way same-platform collision ('{iname}'); mrp {mrp} vs sheet 500"
               + (" — MRP conflict" if mrp else " — no live MRP") + "; amazon live mrp is 550 — owner pick")
u = fk_unpriced["EDOHNFQUDF32JHZR"]
add_review("flipkart", u["id"], u.get("url"), u.get("title"), "YELLOW MUSTARD 1L + 1L",
           "internal name 'YELLOW MUSTARD 1L + 1L' is an exact match but listing mrp 790 != sheet 750 — "
           "MRP conflict, not mapped")
u = fk_unpriced["EDOHF54ZZWHRJY2M"]
add_review("flipkart", u["id"], u.get("url"), u.get("title"), "SO OLIVE 1L + 1L",
           "second FSN with internal name 'SO OLIVE 1+1L' at mrp 699 vs sheet 650; the 650 twin "
           "EDOHBH2URWDMHPXC is mapped (anchor) — confirm this one as duplicate/secondary")
u = fk_unpriced["EDOHBH2FP9ES6GGA"]
add_review("flipkart", u["id"], u.get("url"), u.get("title"), "RICE BRAN 1L + 1L",
           "catalog 'RICE BRAN 2L' (single 2L bottle) carries mrp 570 == the new 1L+1L sheet mrp, but "
           "composition differs (2L bottle vs 2x1L combo) — owner-ratified rule forbids the map; confirm "
           "whether flipkart's '2L' is actually the twin-pack")
add_review("flipkart", "EDOG2NMPHJZH6GJC",
           "https://www.flipkart.com/jivo-extra-virgin-olive-oil-pet-bottle/p/itm6e4cfbc1f88ce?pid=EDOG2NMPHJZH6GJC",
           "EXTRA VIRGIN 1+1L (blessed as EXTRA VIRGIN 2L)", "EXTRA VIRGIN 1L + 1L",
           "FYI, mapping NOT changed: EXTRA VIRGIN 2L's owner-blessed flipkart slot is itself a 1+1L pack "
           "(mrp 3598 = 2x stale 1799). Now that a true EXTRA VIRGIN 1L + 1L master row exists (mapped to "
           "QWRGEMNS9ZPZGWJH @2998), owner may want to re-point the EV 2L slot")

# ---------------- 8) zepto: census-classified same-composition combos ----------------
u = qcom_unpriced["3d955a07-72db-4d1a-9d65-42c5a8dd0641"]
assert u["what"] == "Pomace Olive Oil Combo 1L x 2" and u["mrp"] == master["JIVO POMACE 1L + 1L"]["mrp"]
set_entry("JIVO POMACE 1L + 1L", "zepto", {
    "id": u["id"], "url": u["url"], "title": u["title"], "mrp": u["mrp"],
    "sale_modal": None, "sale_min": None, "sale_max": None, "in_stock": None,
    "method": "AUTO-NAME+MRP", "confidence": "anchored",
    "note": "census-classified combo 1L x 2 (same composition), mrp 2098 == v2 sheet exact; "
            "sale not captured in the frozen census"})
u = qcom_unpriced["597fe027-2e19-4a49-bc07-76ef57295ad6"]
add_review("zepto", u["id"], u["url"], u["title"], "EXTRA LIGHT 1+1L",
           "census-classified combo 1L x 2 (composition OK) but mrp 2998 != sheet 2798 — MRP conflict; "
           "NOTE amazon's ID-anchored live listing also says 2998, the v2 sheet MRP may be stale here")

# ---------------- 9) bigbasket: EAN join (WG MANGO 500ML) ----------------
bb_row = next(r for r in load_json("platforms/bigbasket/result.json")["allRows"]
              if r.get("sku_id") == "40335340")
assert bb_row["ean"] == "8905604001861" and bb_row["mrp"] == master["WG MANGO JUICE 500ML"]["mrp"]
set_entry("WG MANGO JUICE 500ML", "bigbasket", {
    "id": bb_row["sku_id"], "ean": bb_row["ean"],
    "url": "https://www.bigbasket.com" + bb_row["absolute_url"],
    "title": bb_row["sku_raw"], "mrp": bb_row["mrp"], "sale_modal": bb_row["sale"],
    "sale_min": bb_row["sale"], "sale_max": bb_row["sale"], "in_stock": bool(bb_row["in_stock"]),
    "method": "ID-EAN", "confidence": "exact",
    "note": "EAN 8905604001861 (member-session scrape), mrp 100 == v2 sheet exact"})

# ---------------- 10) amazon-now: ASIN join (WG MANGO 500ML) ----------------
now_rows = [r for r in load_json("platforms/amazon-now/result.json")["allRows"]
            if r.get("asin") == "B0DM2G4YCC"]
if now_rows:
    sales = [r["sale"] for r in now_rows if r.get("sale") is not None]
    mrps = [r["mrp"] for r in now_rows if r.get("mrp") is not None]
    set_entry("WG MANGO JUICE 500ML", "amazon-now", {
        "id": "B0DM2G4YCC", "url": "https://www.amazon.in/dp/B0DM2G4YCC",
        "title": now_rows[0].get("sku_raw"), "mrp": max(mrps) if mrps else None,
        "sale_modal": Counter(sales).most_common(1)[0][0] if sales else None,
        "sale_min": min(sales) if sales else None, "sale_max": max(sales) if sales else None,
        "in_stock": any(r.get("in_stock") for r in now_rows),
        "method": "ID-ASIN", "confidence": "exact",
        "note": f"genuine Now (ctnow) rows across {len(now_rows)} pincodes; sale 100 == v2 sheet mrp"})

# ---------------- 11) flipkart-minutes: water (name+MRP anchor; id-fold fills pid) ----------------
set_entry("JIVO WATER 1L", "flipkart-minutes", {
    "id": "jivo-mineral-water-1l", "url": None, "title": "JIVO Mineral Water",
    "item": None, "sap_code": None, "mrp": 30, "sale_modal": None, "sale_min": None, "sale_max": None,
    "in_stock": None, "method": "AUTO-NAME+MRP", "confidence": "anchored",
    "note": "1L water, modal mrp 30 == v2 sheet exact (fragment evidence; variants 25/36 noted there); "
            "id = canonical slug, real pid via id-fold"})

# ---------------- 12) bookkeeping: review/unpriced/drift/version ----------------
# retire the resolved duplicate-listing review item (B0CZP26VVN is now its own master row)
before = len(sm["review"])
sm["review"] = [r for r in sm["review"]
                if not (isinstance(r, dict) and r.get("id") == "B0CZP26VVN"
                        and r.get("candidate") == "CANOLA 1+1L")]
retired_reviews = before - len(sm["review"])

seen = {(r.get("platform"), r.get("id"), r.get("candidate")) for r in sm["review"] if isinstance(r, dict)}
for r in reviews_add:
    if (r["platform"], r["id"], r["candidate"]) not in seen:
        sm["review"].append(r)
        seen.add((r["platform"], r["id"], r["candidate"]))

sm["unpriced"] = [u for u in sm["unpriced"]
                  if u.get("family_census") or (u.get("platform"), str(u.get("id"))) not in mapped_ids]

seen_d = {(d.get("product"), d.get("platform")) for d in sm["mrp_drift"]}
for d in drift_add:
    if (d["product"], d["platform"]) not in seen_d:
        sm["mrp_drift"].append(d)
        seen_d.add((d["product"], d["platform"]))

sm["skus"] = {k: skus[k] for k in sorted(skus)}
sm["version"] = V2_DATE
sm["source"] = ("fleet pricemap 2026-06-05 (4 workers, lead-merged) + master v2 extension "
                f"{V2_DATE} (team sheet 113 SKUs: renames merged, RICE 1KG retired, regimes refreshed, "
                "20 new SKUs mapped from frozen evidence)")
(HERE / "sku_map.json").write_text(json.dumps(sm, indent=1, ensure_ascii=False))

# ---------------- 13) re-apply the id-fold (folded ids/eans/urls must survive) ----------------
subprocess.run([sys.executable, str(HERE / "apply_id_fold.py"),
                str(ROOT / "tools/cron/tests/id_fold_proposal.json"), str(HERE / "sku_map.json")],
               check=True)

# ---------------- 14) summary ----------------
n_map = sum(1 for s in sm["skus"].values() if s.get("platforms"))
print(f"\nsku_map.json: {len(sm['skus'])} skus ({n_map} with >=1 platform), "
      f"review {len(sm['review'])} (+{len(reviews_add)} new, -{retired_reviews} resolved), "
      f"unpriced {len(sm['unpriced'])}, mrp_drift {len(sm['mrp_drift'])}")
print(f"renames merged: {list(RENAMES.items())}; retired: {RETIRED}\n")
print("per new-SKU mapping summary:")
for p in NEW_SKUS:
    print(f"  {p}: " + ("; ".join(summary[p]) if summary[p] else "(no platform evidence)"))
