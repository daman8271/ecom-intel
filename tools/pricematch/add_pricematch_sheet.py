#!/usr/bin/env python3
"""
add_pricematch_sheet.py <platform> <xlsx_path> [--date YYYY-MM-DD]

Appends ONE worksheet named "Price Match" to an already-built per-platform
workbook: live price vs OUR reference price for today's pricing regime.

Color rule (owner's exact words, sacred):
    live price BELOW our reference  -> difference in RED
    live price ABOVE our reference  -> difference in GREEN
    within tolerance (MATCH)        -> plain

Data comes EXCLUSIVELY from tools/pricematch/pricematch_core.py (the
violations-engine core, frozen contract). If the core is missing or fails,
NO sheet is written — a sheet must never carry non-engine numbers/colors
(LEAD ruling). An embedded stub implementing the same contract exists for
DEVELOPMENT ONLY, behind PM_DEV_STUB=1 (default off).

Fail-safe like predict.py: ANY error -> print a warning, exit 0, workbook
untouched (all edits happen on a temp copy, atomically replaced on success).
Idempotent: re-running replaces the existing "Price Match" sheet, never
duplicates. A platform with zero mapped SKUs gets a sheet with a
"no mapped SKUs" note (still exit 0).

Python 3 stdlib + openpyxl only. NO LLM, NO network.
"""

import sys, os, json, shutil, datetime
from collections import Counter

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))          # /opt/ecom-intel

# ---------------------------------------------------------------- styling
# THE one red/green pair + canonical names come from tools/xlsx_dash (v2,
# fresh-eyes 2026-06-06). Guarded import: this script is fail-safe, so a
# missing helper degrades to the same hexes inline, never a crash.
try:
    sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(
        os.path.abspath(__file__))), ))  # /opt/ecom-intel/tools
    import xlsx_dash as _xd
    _RED_HEX, _RED_TXT = _xd.BAD_PAIR
    _GREEN_HEX, _GREEN_TXT = _xd.GOOD_PAIR
    _GLOSS = _xd.GLOSSARY
except Exception:
    _RED_HEX, _RED_TXT = "FFC7CE", "9C0006"
    _GREEN_HEX, _GREEN_TXT = "C6EFCE", "006100"
    _GLOSS = {"SVD": "Special Value Days (Fri–Sun agreed price list)"}

JIVO_GREEN = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO_GREEN)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=JIVO_GREEN, size=18)
SUB_FONT = Font(italic=True, color="555555", size=10)
NOTE_FONT = Font(italic=True, color="777777", size=9)
RED_FILL = PatternFill("solid", fgColor=_RED_HEX)
GREEN_FILL = PatternFill("solid", fgColor=_GREEN_HEX)
RED_FONT = Font(bold=True, color=_RED_TXT, size=11)
GREEN_FONT = Font(bold=True, color=_GREEN_TXT, size=11)
_thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

SHEET_NAME = "Price Match"
TOL = 1.0  # Rs.1 tolerance: |diff| <= 1 => MATCH (owner-ratified)

DISPLAY = {
    "blinkit": "Blinkit", "": "", "zepto": "Zepto",
    "flipkart-minutes": "Flipkart Minutes", "flipkart": "Flipkart",
    "amazon": "Amazon", "amazon-fresh": "Amazon Fresh",
    "amazon-now": "Amazon Now", "bigbasket": "BigBasket",
}


# ================================================================ stub core
# DEV-ONLY (PM_DEV_STUB=1; never reached in production — see _get_core).
# Implements the FROZEN pricematch_core contract (agent-1 brief) for working
# on this builder without the engine. Same record schema, same rules:
#   regime price = reference; tolerance Rs.1; BELOW = live_modal < ref-1;
#   live basis = modal of in-stock rows; stores_below = every row < ref-1;
#   only confidence exact/anchored participate; retired SKUs excluded;
#   no mapping on the platform => NOT_LISTED.
class _StubCore(object):
    name = "stub"

    # platform id field(s) in result.json rows, in lookup order
    ID_FIELDS = {
        "blinkit": ["prid"], "flipkart-minutes": ["fk_pid"],
        "flipkart": ["fsn"], "zepto": ["variant_id", "product_id"],
        "bigbasket": ["sku_id"], "amazon": ["asin"],
        "amazon-fresh": ["asin"], "amazon-now": ["asin"],
    }
    _DOW = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]

    def regime_for(self, date):
        if isinstance(date, str):
            date = datetime.date.fromisoformat(date)
        cfg = {}
        try:
            with open(os.path.join(HERE, "regime.json")) as f:
                cfg = json.load(f)
        except Exception:
            pass  # regime.json not created yet -> built-in defaults
        for ov in cfg.get("overrides") or []:
            if ov.get("date") == date.isoformat():
                return ov.get("regime", "BAU").upper()
        defaults = cfg.get("defaults") or {
            "mon": "BAU", "tue": "BAU", "wed": "BAU", "thu": "BAU",
            "fri": "SVD", "sat": "SVD", "sun": "SVD"}
        return defaults.get(self._DOW[date.weekday()], "BAU").upper()

    def load_context(self, date=None):
        if date is None:
            date = datetime.date.today()
        if isinstance(date, str):
            date = datetime.date.fromisoformat(date)
        with open(os.path.join(HERE, "sku_map.json")) as f:
            skus = json.load(f).get("skus") or {}
        return {"date": date, "regime": self.regime_for(date), "skus": skus}

    @staticmethod
    def _num(v):
        if isinstance(v, bool) or v is None:
            return None
        try:
            f = float(v)
            return f if f > 0 else None
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _modal(vals):
        if not vals:
            return None
        cnt = Counter(vals)
        top = max(cnt.values())
        return min(v for v, c in cnt.items() if c == top)  # tie -> lowest

    def _live_rows(self, platform):
        path = os.path.join(ROOT, "platforms", platform, "result.json")
        try:
            with open(path) as f:
                d = json.load(f)
        except Exception:
            return []
        return d.get("allRows") or [r for p in d.get("perPin", [])
                                    for r in p.get("rows", [])]

    def platform_comparison(self, ctx, platform):
        regime = ctx["regime"]
        idf = self.ID_FIELDS.get(platform, ["asin"])
        # index live rows by listing id
        by_id = {}
        for r in self._live_rows(platform):
            for f in idf:
                v = r.get(f)
                if v is not None and str(v).strip():
                    by_id.setdefault(str(v).strip().lower(), []).append(r)
                    break
        out = []
        for sku, e in ctx["skus"].items():
            if e.get("retired"):
                continue
            regs = e.get("regimes") or {}
            ref = self._num(regs.get(regime.lower()))
            pe = (e.get("platforms") or {}).get(platform)
            base = {"sku": sku, "platform": platform, "regime": regime,
                    "ref_price": ref, "live_modal": None, "live_min": None,
                    "live_max": None, "in_stock": False, "mrp_live": None,
                    "mrp_official": self._num(regs.get("mrp")),
                    "diff": None, "diff_pct": None, "stores_below": []}
            if not pe:
                out.append(dict(base, listing_id=None, url=None, title=None,
                                status="NOT_LISTED"))
                continue
            base.update(listing_id=pe.get("id"), url=pe.get("url"),
                        title=pe.get("title"))
            if pe.get("confidence") not in ("exact", "anchored"):
                out.append(dict(base, status="PENDING_REVIEW"))
                continue
            # join live rows on primary id + any alt listing ids
            ids = [pe.get("id")] + [a.get("id") for a in pe.get("alt") or []
                                    if isinstance(a, dict)]
            rows = []
            for i in ids:
                if i is not None:
                    rows.extend(by_id.get(str(i).strip().lower(), []))
            instock = [r for r in rows
                       if r.get("in_stock") and self._num(r.get("sale"))]
            prices = [self._num(r.get("sale")) for r in instock]
            prices = [p for p in prices if p]
            if ref is None:
                out.append(dict(base, status="NO_REF"))
                continue
            if not prices:
                out.append(dict(base, status="OOS"))
                continue
            modal = self._modal(prices)
            mrps = [self._num(r.get("mrp")) for r in instock]
            mrps = [m for m in mrps if m]
            diff = modal - ref
            below = sorted(({"pincode": r.get("pincode"),
                             "city": r.get("city"),
                             "price": self._num(r.get("sale"))}
                            for r in instock
                            if (self._num(r.get("sale")) or 0) < ref - TOL),
                           key=lambda s: s["price"])
            status = ("MATCH" if abs(diff) <= TOL
                      else "BELOW" if diff < -TOL else "ABOVE")
            out.append(dict(
                base, live_modal=modal, live_min=min(prices),
                live_max=max(prices), in_stock=True,
                mrp_live=self._modal(mrps), diff=round(diff, 2),
                diff_pct=round(diff / ref * 100.0, 2),
                stores_below=below, status=status))
        return out


def _get_core():
    """The real pricematch_core, ALWAYS, in production.

    LEAD RULING (PM_DEV_STUB gate): a core import/call failure must mean NO
    sheet — never stub numbers/colors diverging from engine truth. The stub
    survives only for development, behind an explicit PM_DEV_STUB=1 (default
    off): with it set, a missing/broken core falls back to the stub.
    """
    sys.path.insert(0, HERE)
    try:
        import pricematch_core as core
        for fn in ("load_context", "platform_comparison"):
            if not callable(getattr(core, fn, None)):
                raise AttributeError(f"pricematch_core.{fn} missing")
        return core, "pricematch_core"
    except Exception as e:
        if os.environ.get("PM_DEV_STUB") == "1":
            return _StubCore(), "stub"
        raise RuntimeError(
            f"pricematch_core unavailable ({e}); refusing to write a sheet "
            f"without the engine (dev-only: PM_DEV_STUB=1)") from e


# ================================================================ rendering
HEADERS = ["SKU", "Listing Title", None, "Live Price (₹)",
           "Diff (₹)", "Diff %", "Stock", "Status", "Stores Below Ref"]
COL_W = [22, 52, 14, 13, 11, 9, 8, 16, 30]

ORDER = {"BELOW": 0, "ABOVE": 1, "MATCH": 2, "OOS": 3,
         "PENDING_REVIEW": 4, "NO_REF": 5}


def _sort_key(r):
    o = ORDER.get(r.get("status"), 9)
    d = r.get("diff")
    if r.get("status") == "BELOW":
        return (o, d if d is not None else 0, r.get("sku") or "")
    if r.get("status") == "ABOVE":
        return (o, -(d if d is not None else 0), r.get("sku") or "")
    return (o, 0, r.get("sku") or "")


def render(ws, platform, records, date, regime):
    disp = DISPLAY.get(platform, platform.replace("-", " ").title())
    rows = sorted((r for r in records if r.get("status") != "NOT_LISTED"),
                  key=_sort_key)

    ws["A1"] = f"Jivo × {disp} — Price Match"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(HEADERS))
    # expand the regime once at first use (fresh-eyes: never bare "SVD")
    rlong = _GLOSS.get(regime)
    rtxt = f"{regime} — {rlong}" if rlong else f"{regime} day"
    ws["A2"] = (f"{date} · {rtxt} · live price vs our reference — "
                f"red = live BELOW our price, green = live ABOVE")
    ws["A2"].font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=len(HEADERS))
    # fresh-eyes MUST-5: a MATCH row can still show stores below — say why.
    ws["A3"] = ("Status compares the most-common live price across stores; "
                "'Stores Below Ref' counts individual stores selling below "
                "reference — a MATCH can still have stores below.")
    ws["A3"].font = NOTE_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3,
                   end_column=len(HEADERS))

    headers = list(HEADERS)
    headers[2] = f"Ref Price (₹, {regime})"
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = HDR; c.font = HDR_FONT; c.alignment = CEN; c.border = BORDER
    ws.freeze_panes = "A5"

    if not rows:
        ws["A5"] = "No mapped SKUs for this platform yet."
        ws["A5"].font = NOTE_FONT
        ws.merge_cells(start_row=5, start_column=1, end_row=5,
                       end_column=len(HEADERS))
        return Counter()

    counts = Counter()
    rr = 5
    for r in rows:
        st = r.get("status")
        counts[st] += 1
        sb = r.get("stores_below") or []
        if sb:
            w = sb[0]
            sb_txt = (f"{len(sb)} — worst: {w.get('city') or w.get('pincode')}"
                      f" @ ₹{w.get('price'):g}")
        else:
            sb_txt = "—" if st in ("BELOW", "ABOVE", "MATCH") else ""
        # diff_pct from the engine is ALREADY ×100 (-7.98 means -7.98%); Excel
        # percent formats scale ×100 again, so store the fraction + true "%"
        # format — renders -8.0% in every viewer (the -798% screenshot bug).
        pct = r.get("diff_pct")
        vals = [r.get("sku"), r.get("title") or "", r.get("ref_price"),
                r.get("live_modal"), r.get("diff"),
                (pct / 100.0 if pct is not None else None),
                ("Yes" if r.get("in_stock") else "No"), st, sb_txt]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=j, value=v)
            c.border = BORDER
            c.alignment = LEFT if j in (1, 2, 9) else CEN
        # signed whole-rupee diff (fresh-eyes: "-169.00" read as unitless)
        ws.cell(row=rr, column=5).number_format = "+₹#,##0;-₹#,##0"
        ws.cell(row=rr, column=6).number_format = "0.0%"
        # ---- THE color rule (sacred): RED = live below ref, GREEN = above
        dcell = ws.cell(row=rr, column=5)
        scell = ws.cell(row=rr, column=8)
        if st == "BELOW":
            dcell.fill = RED_FILL; dcell.font = RED_FONT
            scell.fill = RED_FILL
        elif st == "ABOVE":
            dcell.fill = GREEN_FILL; dcell.font = GREEN_FONT
            scell.fill = GREEN_FILL
        if not r.get("in_stock"):
            ws.cell(row=rr, column=7).fill = RED_FILL  # like the stock sheets
        rr += 1

    # ---- bottom summary row
    exposure = sum((r["ref_price"] - r["live_modal"]) for r in rows
                   if r.get("status") == "BELOW"
                   and r.get("ref_price") is not None
                   and r.get("live_modal") is not None)
    rr += 1
    parts = [f"BELOW: {counts.get('BELOW', 0)}",
             f"ABOVE: {counts.get('ABOVE', 0)}",
             f"MATCH: {counts.get('MATCH', 0)}",
             f"OOS: {counts.get('OOS', 0)}"]
    if counts.get("PENDING_REVIEW"):
        parts.append(f"PENDING REVIEW: {counts['PENDING_REVIEW']}")
    if counts.get("NO_REF"):
        parts.append(f"NO REF: {counts['NO_REF']}")
    parts.append(f"Below-ref exposure: ₹{exposure:,.0f}")
    sc = ws.cell(row=rr, column=1, value="   ·   ".join(parts))
    sc.font = Font(bold=True, size=11,
                   color="CC0000" if counts.get("BELOW") else JIVO_GREEN)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr,
                   end_column=len(HEADERS))

    for j, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return counts


# ================================================================ main
def _platform_unmapped(platform):
    """True iff NO sku in sku_map.json has a mapping on this platform
    (e.g. a future platform). Pure read of the map — no stub pricing."""
    try:
        with open(os.path.join(HERE, "sku_map.json")) as f:
            skus = json.load(f).get("skus") or {}
        return not any(platform in (e.get("platforms") or {})
                       for e in skus.values() if not e.get("retired"))
    except Exception:
        return False  # can't tell -> treat as a real failure (no sheet)


def build(platform, xlsx_path, date_str):
    core, core_name = _get_core()
    ctx = core.load_context(date_str) if date_str else core.load_context()
    try:
        records = core.platform_comparison(ctx, platform)
    except Exception:
        # LEAD RULING: a core failure means NO sheet (the except in main()
        # leaves the workbook untouched). Sole carve-out: a platform with
        # ZERO mapped SKUs gets the static "no mapped SKUs" note — records
        # stay empty, so no price/diff/color cell can ever come from here.
        if not _platform_unmapped(platform):
            raise
        records = []
    regime = (ctx.get("regime") if isinstance(ctx, dict) else None) \
        or core.regime_for(date_str or datetime.date.today())
    date = (ctx.get("date") if isinstance(ctx, dict) else None) \
        or date_str or datetime.date.today()

    # All edits on a temp copy; atomic replace only on full success.
    # Dotted hidden name: must end .xlsx for openpyxl, and must NOT match the
    # Jivo-*.xlsx glob run.sh uses to pick the newest report (a leftover tmp
    # after a hard kill must never be mistaken for the real workbook).
    tmp = os.path.join(os.path.dirname(os.path.abspath(xlsx_path)),
                       ".pm.tmp." + os.path.basename(xlsx_path))
    shutil.copy2(xlsx_path, tmp)
    try:
        wb = load_workbook(tmp)
        if SHEET_NAME in wb.sheetnames:          # idempotent: replace
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)         # appended after Predictions
        counts = render(ws, platform, records, date, regime)
        wb.save(tmp)
        os.replace(tmp, xlsx_path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass
    return counts, core_name, regime


def main(argv):
    args = [a for a in argv[1:] if not a.startswith("--")]
    date_str = None
    it = iter(argv[1:])
    for a in it:
        if a == "--date":
            date_str = next(it, None)
        elif a.startswith("--date="):
            date_str = a.split("=", 1)[1]
    if len(args) < 2:
        sys.stderr.write("usage: add_pricematch_sheet.py <platform> "
                         "<xlsx_path> [--date YYYY-MM-DD]\n")
        return 0  # fail-safe: never break the pipeline
    platform, xlsx_path = args[0], args[1]
    try:
        if not os.path.isfile(xlsx_path):
            raise FileNotFoundError(xlsx_path)
        counts, core_name, regime = build(platform, xlsx_path, date_str)
        sys.stderr.write(
            f"add_pricematch_sheet: '{SHEET_NAME}' written to {xlsx_path} "
            f"(core={core_name}, regime={regime}, "
            f"{sum(counts.values())} rows: "
            + ", ".join(f"{k}={v}" for k, v in sorted(counts.items()))
            + ")\n")
        return 0
    except Exception as e:
        # ANY error: warn, leave the workbook untouched, exit 0.
        import traceback
        sys.stderr.write(f"add_pricematch_sheet: FAILED (non-fatal, workbook "
                         f"untouched): {e}\n")
        traceback.print_exc(file=sys.stderr)
        return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
