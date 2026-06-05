#!/usr/bin/env python3
"""Merge fleet fragments -> sku_map.json (machine truth) + Jivo-SKU-Master-Map xlsx (human truth).

Inputs : tools/pricematch/fragments/map-{amazon,flipkart,qcom,bigbasket}.json
Outputs: tools/pricematch/sku_map.json
         tools/pricematch/Jivo-SKU-Master-Map-<date>.xlsx
Deterministic + idempotent. No network. No LLM.
"""
import json, sys, datetime, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).parent
FRAG = HERE / "fragments"
DATE = sys.argv[1] if len(sys.argv) > 1 else str(datetime.date.today())

PLATFORMS = ["amazon", "amazon-fresh", "amazon-now", "flipkart", "flipkart-minutes",
             "zepto", "blinkit", "bigbasket"]

# ---------------- load fragments ----------------
frags = {}
for name in ["map-amazon.json", "map-flipkart.json", "map-qcom.json", "map-bigbasket.json"]:
    frags[name] = json.loads((FRAG / name).read_text())

# master reference (for MRP + regime prices on the Excel)
vi = json.loads((HERE / "dryrun" / "verify_input.json").read_text())
master_info = {m["product"]: m for m in vi["master"]}

import openpyxl as _px
_wb = _px.load_workbook(HERE / "Amazon Price Match.xlsx", data_only=True)
_ws = _wb["url"]; _hdr = [c.value for c in _ws[1]]; _ix = {h: i for i, h in enumerate(_hdr)}
regimes = {}
for r in _ws.iter_rows(min_row=2, values_only=True):
    if r[_ix["PRODUCT"]]:
        regimes[r[_ix["PRODUCT"]].strip().upper()] = {
            "mrp": r[_ix["MRP"]], "bau": r[_ix["BAU Price"]],
            "svd": r[_ix["SVD Price"]], "art": r[_ix["ART PRICE"]]}

# ---------------- merge ----------------
sku_map = {"version": DATE, "source": "fleet pricemap 2026-06-05 (4 workers, lead-merged)",
           "skus": {}, "unpriced": [], "junk": [], "review": [],
           "master_mrp_stale": {}, "mrp_drift": []}

products = sorted(regimes.keys())
for p in products:
    entry = {"regimes": regimes[p], "platforms": {}}
    for f in frags.values():
        node = (f.get("mappings") or {}).get(p) or {}
        for plat, e in node.items():
            if e:
                entry["platforms"][plat] = e
    sku_map["skus"][p] = entry

for f in frags.values():
    sku_map["unpriced"] += f.get("unpriced", []) + [
        {"family_census": True, **fam} for fam in f.get("unpriced_census", [])]
    sku_map["junk"] += f.get("junk", [])
    sku_map["review"] += f.get("review", [])

# ---- stale-vs-drift computed deterministically from merged platform MRPs ----
# rule: cluster live MRPs (±1.5 = same, rupee rounding); a divergent cluster spanning >=2
# platforms AND outnumbering the sheet-agreeing platforms => master sheet is stale;
# any other divergent platform => per-platform mrp_drift flag.
worker_evidence = {}
for f in frags.values():
    for s in (f.get("master_mrp_stale") or []) + (f.get("mrp_drift") or []):
        if isinstance(s, dict):
            key = s.get("product") or s.get("master_product")
            if key:
                worker_evidence[key] = (worker_evidence.get(key, "") + " | " +
                                        str(s.get("evidence") or s.get("listing") or "")).strip(" |")
for p, entry in sku_map["skus"].items():
    sheet = entry["regimes"]["mrp"]
    # in-stock only: OOS marketplace rows carry 3P-gouge / case-pack MRPs (unverifiable)
    live = {plat: le.get("mrp") for plat, le in entry["platforms"].items()
            if le.get("mrp") is not None and le.get("in_stock")}
    if not live or sheet is None:
        continue
    agree = [plat for plat, v in live.items() if abs(v - sheet) <= 1]
    divergent = {plat: v for plat, v in live.items() if abs(v - sheet) > 1}
    clusters = []  # [(value, [plats])]
    for plat, v in sorted(divergent.items(), key=lambda x: x[1]):
        for c in clusters:
            if abs(c[0] - v) <= 1.5:
                c[1].append(plat); break
        else:
            clusters.append([v, [plat]])
    for v, plats in clusters:
        if len(plats) >= 2 and len(plats) > len(agree):
            sku_map["master_mrp_stale"][p] = {
                "product": p, "master_mrp": sheet, "live_mrp": v,
                "platforms": plats, "evidence": worker_evidence.get(p, "")}
        else:
            for plat in plats:
                sku_map["mrp_drift"].append({
                    "product": p, "platform": plat, "platform_mrp": v,
                    "official_mrp": sheet, "consensus": f"{len(agree)} platforms agree with sheet"})

(HERE / "sku_map.json").write_text(json.dumps(sku_map, indent=1, ensure_ascii=False))
print(f"sku_map.json: {len(sku_map['skus'])} skus, "
      f"{sum(1 for s in sku_map['skus'].values() if s['platforms'])} with >=1 platform, "
      f"review={len(sku_map['review'])}, unpriced={len(sku_map['unpriced'])}, "
      f"stale={len(sku_map['master_mrp_stale'])}")

# ---------------- Excel ----------------
JIVO = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO); HDRF = Font(bold=True, color="FFFFFF", size=10)
TITLE = Font(bold=True, color=JIVO, size=16); SUB = Font(italic=True, color="555555", size=9)
RED = PatternFill("solid", fgColor="F4CCCC"); GRN = PatternFill("solid", fgColor="D9EAD3")
YEL = PatternFill("solid", fgColor="FFF2CC"); BLU = Font(color="1155CC", underline="single", size=9)
thin = Side(style="thin", color="D0D0D0"); BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center"); SM = Font(size=9)

wb = Workbook()

def style_hdr(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = CEN; cell.border = BRD

def autosize(ws, maxw=46):
    for col in ws.columns:
        L = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[L].width = min(maxw, max(9, w + 2))

# Sheet 1: Master Map (wide) — one row per SKU, per-platform price + confidence
ws = wb.active; ws.title = "Master Map"
ws["A1"] = "Jivo SKU Master Map — every SKU pinned on every platform"; ws["A1"].font = TITLE
ws["A2"] = f"generated {DATE} from scraped snapshots + e-com master sheet; cell = live modal sale price (hyperlink = listing); blank = not listed"; ws["A2"].font = SUB
hdr = ["SKU", "MRP", "BAU", "SVD", "ART"] + [p for p in PLATFORMS]
ws.append([]); ws.append(hdr)
for p in products:
    e = sku_map["skus"][p]
    row = [p, e["regimes"]["mrp"], e["regimes"]["bau"], e["regimes"]["svd"], e["regimes"]["art"]]
    ws.append(row + [None] * len(PLATFORMS))
    r = ws.max_row
    for j, plat in enumerate(PLATFORMS):
        cell = ws.cell(row=r, column=6 + j)
        le = e["platforms"].get(plat)
        if not le: continue
        cell.value = le.get("sale_modal") if le.get("sale_modal") is not None else ("OOS" if le.get("in_stock") is False else "listed")
        if le.get("url"):
            cell.hyperlink = le["url"]; cell.font = BLU
        conf = le.get("confidence", "")
        if conf.startswith("exact"): cell.fill = GRN
        elif conf.startswith("anchored"): cell.fill = YEL
        cell.border = BRD; cell.alignment = CEN
style_hdr(ws, 4); ws.freeze_panes = "B5"; autosize(ws, 22)

# Sheet 2: Listing Details (long) — one row per SKU x platform listing
ws = wb.create_sheet("Listing Details")
ws.append(["SKU", "platform", "listing id", "url", "title", "live MRP", "sale modal",
           "sale min", "sale max", "in stock", "method", "confidence", "note"])
for p in products:
    for plat in PLATFORMS:
        le = sku_map["skus"][p]["platforms"].get(plat)
        if not le: continue
        ws.append([p, plat, le.get("id"), le.get("url"), (le.get("title") or le.get("item") or "")[:90],
                   le.get("mrp"), le.get("sale_modal"), le.get("sale_min"), le.get("sale_max"),
                   "Y" if le.get("in_stock") else "N", le.get("method"), le.get("confidence"),
                   le.get("note", "")])
        for c in ws[ws.max_row]: c.font = SM
style_hdr(ws); ws.freeze_panes = "A2"; autosize(ws)

# Sheet 3: Review — needs owner YES/NO
ws = wb.create_sheet("Review (confirm these)")
ws.append(["platform", "listing id", "url", "title", "candidate SKU", "reason", "OWNER: YES/NO"])
for it in sku_map["review"]:
    if isinstance(it, str):  # some fragments emit plain-string notes
        it = {"reason": it}
    ws.append([it.get("platform"), it.get("id"), it.get("url"), (it.get("title") or "")[:90],
               it.get("candidate"), (it.get("reason") or "")[:140], ""])
    ws.cell(row=ws.max_row, column=7).fill = YEL
style_hdr(ws); ws.freeze_panes = "A2"; autosize(ws)

# Sheet 4: MRP integrity — stale sheet rows + per-platform drift
ws = wb.create_sheet("MRP integrity")
ws.append(["type", "SKU", "sheet MRP", "live MRP", "where", "evidence / consensus"])
for k, s in sorted(sku_map["master_mrp_stale"].items()):
    ws.append(["MASTER SHEET STALE", k, s.get("master_mrp"), s.get("live_mrp"),
               ", ".join(s.get("platforms", [])), (s.get("evidence") or "")[:160]])
    ws.cell(row=ws.max_row, column=4).fill = RED
for s in sku_map["mrp_drift"]:
    ws.append(["PLATFORM DRIFT", s.get("product"), s.get("official_mrp"),
               s.get("platform_mrp"), s.get("platform"), s.get("consensus", "")])
    ws.cell(row=ws.max_row, column=4).fill = YEL
style_hdr(ws); ws.freeze_panes = "A2"; autosize(ws)

# Sheet 5: Unpriced census
ws = wb.create_sheet("Unpriced (no master row)")
ws.append(["family / listing", "platform(s)", "id", "mrp", "combo", "suggestion / title"])
fams = [u for u in sku_map["unpriced"] if u.get("family_census")]
singles = [u for u in sku_map["unpriced"] if not u.get("family_census")]
for f in fams:
    ws.append([f.get("family"), ", ".join(f.get("platforms", [])), "", "", "",
               f.get("suggestion", "")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10)
for u in singles:
    ws.append(["    " + (u.get("title") or "")[:80], u.get("platform"), u.get("id"),
               u.get("mrp"), "Y" if u.get("combo") else "", u.get("internal_name", "")])
    for c in ws[ws.max_row]: c.font = SM
style_hdr(ws); ws.freeze_panes = "A2"; autosize(ws)

out = HERE / f"Jivo-SKU-Master-Map-{DATE}.xlsx"
wb.save(out)
print("xlsx:", out)
