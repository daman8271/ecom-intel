#!/usr/bin/env python3
"""build_pricematch.py [--date YYYY-MM-DD] [--out DIR] — the master Price Match workbook.

Builds `Jivo-Price-Match-<date>.xlsx` — every platform x every SKU vs the day's
reference price (regime from tools/pricematch/regime.json via pricematch_core) — with
the **Ecom Head** executive view as the FIRST sheet. Color direction is sacred:
RED = live BELOW reference (the violation), GREEN = live ABOVE reference.

Sheets:
  1. Ecom Head          — date + regime badge, KPI cards, top-10 violations,
                          platform scoreboard, MRP-integrity flags, source footer.
  2. Matrix             — all SKUs x (MRP|BAU|SVD|ART + one col per platform);
                          live modal price, RED below / GREEN above, hyperlink to
                          the listing, cell comment carries the (±₹x) detail.
  3. Violations         — one row per (SKU, platform, store) below ref, by loss.
  4. Above reference    — the green list (platform-level).
  5. Coverage & pending — NOT_LISTED gaps, PENDING_REVIEW items, OOS list.

Also writes `<xlsx>.summary.json` (regime, KPI counts, a deterministic 6-line
Telegram markdown summary + caption) for the run_all.sh batch wiring, and prints
the absolute xlsx path as the LAST stdout line (run_all.sh reads it).

Consumes the FROZEN pricematch_core contract (load_context / all_comparisons /
summary / regime_for) — stdlib + openpyxl only, NO LLM, NO network.
"""
import argparse
import datetime
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
sys.path.insert(0, os.path.dirname(HERE))              # /opt/ecom-intel/tools

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import xlsx_dash as xd                                 # v2 house style (fresh-eyes)

# ---------------------------------------------------------------- price basis
# Basis history, all on 2026-06-11: modal → 560005 → 110095 → modal → a
# two-pin Matrix (560006+110092, one row per pin) live for about an hour →
# FINAL owner order: Matrix back to the pan-India modal. Exact-pin views
# live on in the PM Check sheets (core.PM_REF_PINCODES = 110092+560006) and
# the Pin Spread tabs. The two-pin Matrix renderer is preserved at commit
# c4fe6f14 if it's ever wanted again. A pincode string here restores a
# single-pin verdict basis for the whole builder; None = modal.
PM_MATRIX_BASIS_PINCODE = None

# ---------------------------------------------------------------- palette
# Single house palette from tools/xlsx_dash (fresh-eyes 2026-06-06: one brand
# green, ONE red/green compliance pair across every deliverable).
JIVO_GREEN = xd.JIVO_GREEN
BRAND      = xd.BRAND
BRAND_SOFT = xd.BRAND_SOFT
INK        = xd.INK
MUTED      = xd.MUTED
RULE       = xd.RULE
CANVAS     = xd.CANVAS
FAINT      = "D1D5DB"    # very light grey — the quiet "not sold here" dot (owner: kill the n/s wall)
POS        = xd.POS      # green text — ABOVE ref
NEG        = xd.NEG      # red text — BELOW ref (the violation)
WARN       = xd.WARN
RED_FILL   = PatternFill("solid", fgColor=xd.RED_FILL_HEX)
GREEN_FILL = PatternFill("solid", fgColor=xd.GREEN_FILL_HEX)
AMBER_FILL = PatternFill("solid", fgColor=xd.AMBER_FILL_HEX)
GREY_FILL  = PatternFill("solid", fgColor=CANVAS)
HDR_FILL   = PatternFill("solid", fgColor=JIVO_GREEN)
SOFT_FILL  = PatternFill("solid", fgColor=BRAND_SOFT)
RULE_FILL  = PatternFill("solid", fgColor=RULE)

F = xd.F

CANONICAL = ["flipkart-minutes", "flipkart", "zepto", "bigbasket",
             "amazon", "amazon-fresh", "amazon-now", "blinkit"]
PDISP = dict(xd.PLATFORM_DISPLAY)        # canonical full display names

# Price-match cluster band (owner ask 2026-06-08): when 2+ platforms price a SKU
# within ±₹PM_CLUSTER_BAND of each other they are "fighting for the same price" —
# price-match is ACTIVE for that SKU. Easy to retune later.
PM_CLUSTER_BAND = 5.0
# statuses that carry a REAL live price this run (exclude OOS/NOT_LISTED/PENDING/retired)
PM_PRICE_STATUSES = ("BELOW", "ABOVE", "MATCH", "NO_REF")

REGIME_BADGE = {"BAU": ("475569", "BAU day"),        # slate
                "SVD": (JIVO_GREEN, "SVD day"),      # Jivo green
                "ART": ("B45309", "ART day")}        # amber — announcement regime

def regime_long(regime):
    """'SVD' -> 'SVD — Special Value Days (…)' via the shared glossary."""
    full = xd.GLOSSARY.get(regime)
    return f"{regime} — {full}" if full else f"{regime} day"


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def esc_md(t):
    t = str(t)
    for ch in ("_", "*", "`", "["):
        t = t.replace(ch, "\\" + ch)
    return t


# ---------------------------------------------------------------- data
def gather(date_str):
    """Load everything through the frozen pricematch_core contract."""
    import pricematch_core as core

    ctx = core.load_context(date_str, basis_pincode=PM_MATRIX_BASIS_PINCODE)
    comps = core.all_comparisons(ctx)          # dict[platform, list[record]]
    try:
        summ = core.summary(ctx) or {}
    except Exception as e:                     # summary shape is the loosest part
        print(f"build_pricematch: core.summary failed (non-fatal): {e}", file=sys.stderr)
        summ = {}

    date = date_str or datetime.date.today().isoformat()
    regime = None
    for recs in comps.values():
        for r in recs:
            if r.get("regime"):
                regime = r["regime"]
                break
        if regime:
            break
    if not regime:
        try:
            regime = core.regime_for(date)
        except Exception:
            regime = core.regime_for(datetime.date.fromisoformat(date))

    with open(os.path.join(HERE, "sku_map.json"), encoding="utf-8") as fh:
        sku_map = json.load(fh)
    return date, regime, comps, summ, sku_map


def flatten(comps):
    """records flat + indexed by (sku, platform); first record wins on dupes."""
    flat, by_key = [], {}
    for p in CANONICAL:
        for r in comps.get(p, []):
            flat.append(r)
            by_key.setdefault((r.get("sku"), p), r)
    # any platform the canonical list doesn't know about (defensive)
    for p, recs in comps.items():
        if p in CANONICAL:
            continue
        for r in recs:
            flat.append(r)
            by_key.setdefault((r.get("sku"), p), r)
    return flat, by_key


def kpis_of(flat, summ, sku_map):
    skus = sku_map.get("skus", {})
    live_skus = [s for s, v in skus.items() if not v.get("retired")]
    counts = {}
    for r in flat:
        counts[r.get("status")] = counts.get(r.get("status"), 0) + 1

    # total below-ref exposure ₹: prefer the engine's global number if findable,
    # else Σ(ref − live_modal) over BELOW records (deterministic fallback).
    exposure = None
    candidates = []

    def walk(o, depth=0):
        if depth > 4:
            return
        if isinstance(o, dict):
            for k, v in o.items():
                if isinstance(v, (int, float)) and "exposure" in str(k).lower():
                    candidates.append((depth, float(v)))
                else:
                    walk(v, depth + 1)

    walk(summ)
    if candidates:
        candidates.sort()                      # shallowest key = the global one
        exposure = candidates[0][1]
    if exposure is None:
        exposure = sum((_f(r.get("ref_price")) or 0) - (_f(r.get("live_modal")) or 0)
                       for r in flat if r.get("status") == "BELOW"
                       and _f(r.get("ref_price")) and _f(r.get("live_modal")))
    # listing-unit counts (fresh-eyes: the cards must say which unit they count)
    listed_skus = {r.get("sku") for r in flat if r.get("status") != "NOT_LISTED"}
    return {
        "skus_tracked": len(live_skus),
        "platforms": len([p for p in CANONICAL]),
        "compliant": counts.get("MATCH", 0),
        "below": counts.get("BELOW", 0),
        "above": counts.get("ABOVE", 0),
        "oos": counts.get("OOS", 0),
        "pending": counts.get("PENDING_REVIEW", 0),
        "not_listed": counts.get("NOT_LISTED", 0),
        "exposure": round(exposure),
        "listings": len(flat),
        "mapped": sum(1 for r in flat if r.get("status") != "NOT_LISTED"),
        "listed_skus": len(listed_skus),
    }


def store_violations(flat):
    """Every (SKU, platform, store) below ref — the dark-store-level check."""
    rows = []
    for r in flat:
        ref = _f(r.get("ref_price"))
        stores = r.get("stores_below") or []
        for s in stores:
            price = _f(s.get("price"))
            if price is None or ref is None:
                continue
            rows.append({"sku": r.get("sku"), "platform": r.get("platform"),
                         "city": s.get("city") or "—", "pincode": s.get("pincode") or "—",
                         "price": price, "ref": ref, "loss": round(ref - price, 2),
                         "url": r.get("url")})
        if not stores and r.get("status") == "BELOW" and ref is not None \
                and _f(r.get("live_modal")) is not None:
            # platform-level BELOW with no store granularity (national platforms)
            rows.append({"sku": r.get("sku"), "platform": r.get("platform"),
                         "city": "(national)", "pincode": "—",
                         "price": _f(r.get("live_modal")), "ref": ref,
                         "loss": round(ref - _f(r.get("live_modal")), 2),
                         "url": r.get("url")})
    rows.sort(key=lambda x: -x["loss"])
    return rows


def mrp_flags_of(sku_map):
    """master_mrp_stale entries + the EL 1+1L 2798-vs-2998 drift case (owner brief)."""
    stale = sku_map.get("master_mrp_stale") or {}
    el = [d for d in (sku_map.get("mrp_drift") or [])
          if str(d.get("product", "")).upper().replace(" ", "") == "EXTRALIGHT1+1L"]
    return stale, el


# ---------------------------------------------------------------- sheet helpers
def title_cell(ws, row, text, size=20, color=INK, span=10):
    c = ws.cell(row, 1, text)
    c.font = Font(name=F, size=size, bold=True, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = max(22, size + 16)
    return c


def divider(ws, row, span=10):
    for c in range(1, span + 1):
        ws.cell(row, c).fill = RULE_FILL
    ws.row_dimensions[row].height = 3


def table_header(ws, row, cols, start_col=1):
    for i, h in enumerate(cols):
        c = ws.cell(row, start_col + i, h)
        c.font = Font(name=F, size=10, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def money(c, red=False, green=False, bold=False):
    c.number_format = "₹#,##0"
    color = NEG if red else (POS if green else INK)
    c.font = Font(name=F, size=10, bold=bold or red or green, color=color)


# ---------------------------------------------------------------- sheets
def sheet_ecom_head(wb, date, regime, flat, comps, kpi, sku_map, label=None):
    ws = wb.create_sheet("Ecom Head", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    for L in "ABCDEFGHIJ":
        ws.column_dimensions[L].width = 16

    # edition cover identity (fresh-eyes ed-MUST-1: a scoped edition must
    # never be mistakable for the cross-platform master)
    title = (f"Jivo — Price Match · {label} channels" if label
             else "Jivo — Price Match · Ecom Head")
    title_cell(ws, 1, title, span=8)
    # regime badge, top-right
    badge_color, badge_text = REGIME_BADGE.get(regime, (MUTED, f"{regime} day"))
    b = ws.cell(1, 9, f"{date}  ·  {badge_text}".upper())
    b.font = Font(name=F, size=11, bold=True, color="FFFFFF")
    b.fill = PatternFill("solid", fgColor=badge_color)
    b.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
    if label:
        xd.edition_badge(ws, 2, 9, f"{label} edition", span=2)

    # ONE reconciling caption: the cards below count LISTINGS (SKU×platform),
    # not SKUs — say so up front (fresh-eyes MUST-2).
    scope = (f"{', '.join(PDISP.get(p, p) for p in CANONICAL)} — the cross-"
             f"platform master covers all 8 platforms · " if label else "")
    h = ws.cell(2, 1, f"{kpi['skus_tracked']} master SKUs, {kpi['listed_skus']} listed here · "
                      f"{scope}of {kpi['listings']} SKU×platform listings: "
                      f"{kpi['compliant']} at agreed price · {kpi['below']} below · "
                      f"{kpi['above']} above · {kpi['oos']} OOS · "
                      f"{kpi['not_listed']} not listed · {kpi['pending']} pending · "
                      f"₹{kpi['exposure']:,} below-ref gap")
    h.font = Font(name=F, size=10, color=MUTED)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8 if label else 10)
    ws.row_dimensions[2].height = 22
    divider(ws, 3)

    # ---- KPI cards: 8 cards, 2 rows x 4, each spanning 2 cols x 3 rows ----
    cards = [
        ("MASTER SKUS", f"{kpi['skus_tracked']}", "tracked (retired excluded)", INK),
        ("PLATFORMS", f"{kpi['platforms']}", "in this report", INK),
        ("AT AGREED PRICE", f"{kpi['compliant']}", "listings matching reference (±₹1)", BRAND),
        ("BELOW REF", f"{kpi['below']}", "listing-level violations — under agreed price", NEG),
        ("ABOVE REF", f"{kpi['above']}", "listings priced over reference", POS),
        ("OOS LISTINGS", f"{kpi['oos']}", f"of {kpi['mapped']} mapped listings — no price verdict", MUTED),
        ("PENDING REVIEW", f"{kpi['pending']}", "mappings awaiting confirmation", WARN),
        ("BELOW-REF GAP ₹", f"₹{kpi['exposure']:,}", "Σ(agreed − live) over below-ref listings", NEG),
    ]
    r0 = 5
    for i, (label, value, sub, color) in enumerate(cards):
        rr = r0 + (i // 4) * 4
        cc = [1, 3, 6, 8][i % 4]                      # cols 1,3,6,8 (gutter mid-sheet)
        for dr in range(3):
            for dc in range(2):
                ws.cell(rr + dr, cc + dc).fill = GREY_FILL
        lc = ws.cell(rr, cc, label)
        lc.font = Font(name=F, size=8, bold=True, color=MUTED)
        lc.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
        vc = ws.cell(rr + 1, cc, value)
        vc.font = Font(name=F, size=20, bold=True, color=color)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sc = ws.cell(rr + 2, cc, sub)
        sc.font = Font(name=F, size=8, color=MUTED)
        sc.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    row = r0 + 8

    # ---- TOP-10 violations ----
    below = sorted([r for r in flat if r.get("status") == "BELOW"],
                   key=lambda r: _f(r.get("diff")) or 0)
    t = ws.cell(row, 1, "TOP VIOLATIONS — live BELOW the reference price")
    t.font = Font(name=F, size=12, bold=True, color=NEG)
    row += 1
    table_header(ws, row, ["SKU", "Platform", "Ref ₹", "Live ₹", "Diff ₹",
                           "Worst store (city @ ₹)"])
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
    row += 1
    if not below:
        c = ws.cell(row, 1, "None — every priced listing is at or above reference today.")
        c.font = Font(name=F, size=10, italic=True, color=POS)
        row += 1
    for r in below[:10]:
        ws.cell(row, 1, r.get("sku")).font = Font(name=F, size=10, color=INK)
        ws.cell(row, 2, PDISP.get(r.get("platform"), r.get("platform"))).font = \
            Font(name=F, size=10, color=MUTED)
        money(ws.cell(row, 3, _f(r.get("ref_price"))))
        money(ws.cell(row, 4, _f(r.get("live_modal"))), red=True)
        d = ws.cell(row, 5, _f(r.get("diff")))
        d.number_format = "₹#,##0;-₹#,##0"
        d.font = Font(name=F, size=10, bold=True, color=NEG)
        worst = "—"
        stores = [s for s in (r.get("stores_below") or []) if _f(s.get("price")) is not None]
        if stores:
            w = min(stores, key=lambda s: _f(s.get("price")))
            worst = f"{w.get('city') or '?'} @ ₹{_f(w.get('price')):,.0f}"
        wc = ws.cell(row, 6, worst)
        wc.font = Font(name=F, size=10, color=INK)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        row += 1
    row += 1

    # ---- platform scoreboard ----
    t = ws.cell(row, 1, "PLATFORM SCOREBOARD")
    t.font = Font(name=F, size=12, bold=True, color=INK)
    row += 1
    table_header(ws, row, ["Platform", "Mapped SKUs", "% ≥ ref", "Violations",
                           "Biggest offender"])
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    row += 1
    for p in CANONICAL:
        recs = comps.get(p, [])
        mapped = sum(1 for r in recs if r.get("status") != "NOT_LISTED")
        priced = [r for r in recs if r.get("status") in ("BELOW", "ABOVE", "MATCH")]
        ok = sum(1 for r in priced if r.get("status") in ("ABOVE", "MATCH"))
        pct = (100.0 * ok / len(priced)) if priced else None
        viols = [r for r in recs if r.get("status") == "BELOW"]
        offender = "—"
        if viols:
            w = min(viols, key=lambda r: _f(r.get("diff")) or 0)
            offender = f"{w.get('sku')} (−₹{abs(_f(w.get('diff')) or 0):,.0f})"
        ws.cell(row, 1, PDISP.get(p, p)).font = Font(name=F, size=10, bold=True, color=INK)
        ws.cell(row, 2, mapped).font = Font(name=F, size=10, color=INK)
        pc = ws.cell(row, 3, f"{pct:.0f}%" if pct is not None else "—")
        # fresh-eyes S9: threshold the colour or it stops meaning anything
        pc.font = Font(name=F, size=10, bold=True,
                       color=(POS if (pct or 0) >= 90 else
                              (WARN if (pct or 0) >= 70 else NEG)) if pct is not None else MUTED)
        vc = ws.cell(row, 4, len(viols))
        vc.font = Font(name=F, size=10, bold=True, color=NEG if viols else POS)
        oc = ws.cell(row, 5, offender)
        oc.font = Font(name=F, size=10, color=NEG if viols else MUTED)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        row += 1
    row += 1

    # ---- MRP-integrity flags ----
    stale, el = mrp_flags_of(sku_map)
    n_flags = len(stale) + len(el)
    t = ws.cell(row, 1, f"MRP-INTEGRITY FLAGS: {n_flags}")
    t.font = Font(name=F, size=12, bold=True, color=WARN if n_flags else POS)
    row += 1
    # plain language (fresh-eyes S10: no insider shorthand, state the action)
    detail = []
    if stale:
        detail.append(f"{len(stale)} master MRPs look stale vs live listings "
                      f"({', '.join(sorted(stale))}) — confirm with brand team")
    for d in el:
        detail.append(f"EXTRA LIGHT 1+1L: master MRP ₹{d.get('official_mrp')} vs "
                      f"{PDISP.get(d.get('platform'), d.get('platform'))} live "
                      f"₹{d.get('platform_mrp')} — confirm which is current")
    dc = ws.cell(row, 1, "; ".join(detail) if detail else "none on file")
    dc.font = Font(name=F, size=9, color=MUTED)
    dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 28
    row += 2

    # ---- footer ----
    divider(ws, row)
    row += 1
    # plain words, no internal file names (fresh-eyes MUST-4)
    fc = ws.cell(row, 1, f"Today is {regime_long(regime)}. Every reference price on "
                         f"this report is today's agreed {regime} price list.")
    fc.font = Font(name=F, size=9, italic=True, color=MUTED)
    fc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    # ---- APPENDED (2026-06-09): board-wide EXACT price-match section. Self-contained
    # fail-safe — a bug here skips ONLY this section; the KPI cards / violations /
    # scoreboard above are untouched and the workbook still builds. ----
    try:
        sheet_ecom_head_exact(ws, row, comps, sku_map)
    except Exception as e:
        import traceback
        print(f"build_pricematch: Ecom Head exact-match section SKIPPED (non-fatal): {e}",
              file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
    return ws


def pm_clusters(sku, by_key):
    """Price-match clusters for one SKU over the CURRENT CANONICAL platforms.

    Collect platforms that have a REAL live price this run (PM_PRICE_STATUSES with a
    numeric live_modal), sort by price, then greedily group so that within each group
    max(price) − min(price) ≤ PM_CLUSTER_BAND. A group of ≥2 platforms is a cluster.
    Scoped editions narrow naturally because CANONICAL is shrunk in place. Returns a
    list of clusters; each cluster = list of (platform, price) tuples, price-sorted.
    """
    priced = []
    for p in CANONICAL:
        r = by_key.get((sku, p))
        if not r or r.get("status") not in PM_PRICE_STATUSES:
            continue
        price = _f(r.get("live_modal"))
        if price is None:
            continue
        priced.append((p, price))
    return band_clusters(priced)


def band_clusters(priced):
    """Greedy ±PM_CLUSTER_BAND grouping of (platform, price) pairs.
    The Matrix feeds it per-pin prices; pm_clusters feeds it modal prices."""
    priced = sorted(priced, key=lambda x: x[1])
    clusters, group = [], []
    for p, price in priced:
        if group and price - group[0][1] > PM_CLUSTER_BAND:   # group[0] = min (sorted)
            if len(group) >= 2:
                clusters.append(group)
            group = []
        group.append((p, price))
    if len(group) >= 2:
        clusters.append(group)
    return clusters


def fmt_cluster(cluster):
    """'Blinkit · Zepto · Amazon @ ₹379–381' (single price when the band is flat)."""
    names = " · ".join(PDISP.get(p, p) for p, _ in cluster)
    lo = min(pr for _, pr in cluster)
    hi = max(pr for _, pr in cluster)
    band = f"₹{lo:,.0f}" if abs(hi - lo) < 0.005 else f"₹{lo:,.0f}–{hi:,.0f}"
    return f"{names} @ {band}"


def _board_exact_matches(sku_map, by_key):
    """Board-wide EXACT price-match: per master SKU, build {platform: live modal/national
    price} over the current CANONICAL platforms (SAME source the Matrix/Ecom Head already
    use — by_key + live_modal) and run the IDENTICAL-₹ grouping. Returns
    [{"sku", "groups"}] for SKUs with >=1 exact match, sorted by largest group desc, SKU.
    This is DISTINCT from pm_clusters (the ±₹5 'price-match active' band)."""
    import pricematch_core as core
    out = []
    for sku in _compete_skus(sku_map):
        pbp = {}
        for p in CANONICAL:
            r = by_key.get((sku, p))
            if not r or r.get("status") not in PM_PRICE_STATUSES:
                continue
            price = _f(r.get("live_modal"))
            if price is not None:
                pbp[p] = price
        groups = core.exact_price_match(pbp)
        if groups:
            out.append({"sku": sku, "groups": groups})
    out.sort(key=lambda m: (-max(len(g["platforms"]) for g in m["groups"]), m["sku"]))
    return out


def sheet_ecom_head_exact(ws, row, comps, sku_map):
    """APPENDED Ecom Head section: 'EXACT PRICE-MATCH ACROSS PLATFORMS' — a board-wide,
    all-platforms KPI + top table of SKUs where 2+ platforms sit at the IDENTICAL ₹.
    Fail-safe: wrapped by the caller; a bug skips only this section. Returns next row."""
    _, by_key = flatten(comps)
    matches = _board_exact_matches(sku_map, by_key)

    row = xd.section_title(ws, row, "EXACT PRICE-MATCH ACROSS PLATFORMS", c0=1, c1=10)
    sub = ws.cell(row, 1, "SKUs with 2+ platforms at the IDENTICAL ₹ today "
                          "(exact-equal — distinct from the Matrix ±₹5 cluster).")
    sub.font = Font(name=F, size=9, color=MUTED)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 24
    row += 1

    n = len(matches)
    noun = "SKU" if n == 1 else "SKUs"
    kc = ws.cell(row, 1, f"{n} {noun} matching exactly across platforms today")
    kc.font = Font(name=F, size=13, bold=True, color=BRAND if n else MUTED)
    kc.fill = SOFT_FILL if n else GREY_FILL
    kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 22
    row += 2

    if not matches:
        nc = ws.cell(row, 1, "No SKU has 2+ platforms at an identical price today.")
        nc.font = Font(name=F, size=10, italic=True, color=MUTED)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        return row + 1

    # flatten SKU→group(s); a SKU with two distinct identical-price clusters shows twice
    grows = [(m["sku"], g) for m in matches for g in m["groups"]]
    grows.sort(key=lambda x: (-len(x[1]["platforms"]), x[0], x[1]["price"]))

    # header (SKU | platforms | ₹) — house table style, merged spans
    def _hdr(cc0, cc1, text):
        ws.merge_cells(start_row=row, start_column=cc0, end_row=row, end_column=cc1)
        h = ws.cell(row, cc0, text)
        h.font = Font(name=F, size=10, bold=True, color="FFFFFF")
        h.fill = HDR_FILL
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _hdr(1, 3, "SKU")
    _hdr(4, 8, "Platforms matched at the identical ₹")
    _hdr(9, 10, "₹ price")
    ws.row_dimensions[row].height = 18
    row += 1

    TOP = 12
    for sku, g in grows[:TOP]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        sc = ws.cell(row, 1, sku)
        sc.font = Font(name=F, size=10, bold=True, color=INK)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        names = " = ".join(PDISP.get(p, p) for p in g["platforms"])
        pc = ws.cell(row, 4, f"⚡ {names}")
        pc.font = Font(name=F, size=10, bold=True, color=BRAND)
        pc.fill = SOFT_FILL
        pc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
        vc = ws.cell(row, 9, g["price"])
        vc.number_format = "₹#,##0"
        vc.font = Font(name=F, size=10, bold=True, color=INK)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        row += 1
    if len(grows) > TOP:
        mc = ws.cell(row, 1, f"… and {len(grows) - TOP} more exact-match group(s) — "
                             f"full per-pincode detail on the two PM Check sheets.")
        mc.font = Font(name=F, size=9, italic=True, color=MUTED)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1
    return row


def sheet_matrix(wb, date, regime, by_key, sku_map, listed_only=False):
    # Owner order 2026-06-06: show ONLY today's regime — MRP + ONE named
    # "Agreed price (<regime>)" column. The BAU/SVD/ART trio lives in the
    # team's own master sheet; our report shows the day's truth.
    # listed_only (scoped editions): keep only SKUs with >=1 listing on the
    # edition's platforms — the never-listed tail lives in Coverage & pending.
    ws = wb.create_sheet("Matrix")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 26
    regime_cols = ["MRP", f"Agreed price ({regime})"]
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 18
    for i in range(len(CANONICAL)):
        ws.column_dimensions[get_column_letter(2 + len(regime_cols) + i)].width = 14
    # NEW col at the very end: platforms whose live prices cluster within ±₹5
    pm_col = 2 + len(regime_cols) + len(CANONICAL)
    ws.column_dimensions[get_column_letter(pm_col)].width = 34

    headers = (["SKU"] + regime_cols + [PDISP.get(p, p) for p in CANONICAL]
               + [f"Price-Match active (±₹{PM_CLUSTER_BAND:.0f})"])
    table_header(ws, 1, headers)
    agreed_col = 3
    ws.cell(1, agreed_col).fill = PatternFill("solid", fgColor=INK)

    # legend at the TOP, inside the frozen pane, all states covered
    # (fresh-eyes MUST-5: it was the last row, below 115 data rows)
    basis_line = []
    if PM_MATRIX_BASIS_PINCODE:
        city = PM_COMPETE_CITY.get(PM_MATRIX_BASIS_PINCODE, "")
        basis_line = [(f"Live basis = price @ {PM_MATRIX_BASIS_PINCODE}"
                       + (f" ({city})" if city else "")
                       + " · national platforms = all-India price",
                       BRAND_SOFT, BRAND)]
    xd.legend(ws, 2, basis_line + [
        ("RED = live below agreed price (±₹1)", xd.RED_FILL_HEX, xd.RED_TEXT),
        ("GREEN = live above", xd.GREEN_FILL_HEX, xd.GREEN_TEXT),
        ("blue link = live listing · plain = at agreed price", None, "0563C1"),
        ("OOS = out of stock · ? = mapping under review · — = not listed", None, None),
        (f"⚡ Price-Match active = 2+ platforms within ₹{PM_CLUSTER_BAND:.0f} of "
         f"each other (last col)", BRAND_SOFT, BRAND),
        ("price here = modal (most-common) across pins · pin-by-pin detail = "
         "'Pin Spread · <platform>' sheets", None, None),
    ], span=2)

    skus = sku_map.get("skus", {})
    if listed_only:
        skus = {sku: e for sku, e in skus.items()
                if not e.get("retired") and any(
                    (by_key.get((sku, p)) or {}).get("status",
                     "NOT_LISTED") != "NOT_LISTED" for p in CANONICAL)}
    row = 3
    for sku, entry in skus.items():
        retired = bool(entry.get("retired"))
        name = ws.cell(row, 1, sku + ("  (retired)" if retired else ""))
        name.font = Font(name=F, size=10, bold=not retired,
                         color=MUTED if retired else INK)
        if retired:
            for c in range(1, len(headers) + 1):
                ws.cell(row, c).fill = GREY_FILL
        regs = entry.get("regimes") or {}
        for i, key in enumerate(("mrp", regime.lower())):
            c = ws.cell(row, 2 + i, _f(regs.get(key)))
            money(c)
            if retired:
                c.font = Font(name=F, size=10, color=MUTED)
                c.fill = GREY_FILL
            elif agreed_col == 2 + i:
                c.fill = SOFT_FILL
                c.font = Font(name=F, size=10, bold=True, color=BRAND)
        for j, p in enumerate(CANONICAL):
            cc = 2 + len(regime_cols) + j
            c = ws.cell(row, cc)
            if retired:
                c.value = "retired"
                c.font = Font(name=F, size=9, italic=True, color=MUTED)
                c.fill = GREY_FILL
                continue
            r = by_key.get((sku, p))
            status = (r or {}).get("status")
            if r is None or status == "NOT_LISTED":
                c.value = "—"
                c.font = Font(name=F, size=10, color=MUTED)
                c.fill = GREY_FILL
                c.alignment = Alignment(horizontal="center")
                continue
            if status == "OOS":
                c.value = "OOS"
                c.font = Font(name=F, size=9, bold=True, color=MUTED)
                c.alignment = Alignment(horizontal="center")
                if r.get("url"):
                    c.hyperlink = r["url"]
                continue
            if status == "PENDING_REVIEW":
                c.value = "?"
                c.font = Font(name=F, size=10, bold=True, color=WARN)
                c.alignment = Alignment(horizontal="center")
                c.comment = Comment("Mapping pending human review — not priced.",
                                    "pricematch", height=60, width=220)
                continue
            live = _f(r.get("live_modal"))
            ref = _f(r.get("ref_price"))
            diff = _f(r.get("diff"))
            c.value = live
            c.number_format = "₹#,##0"
            c.font = Font(name=F, size=10, color=INK)
            if r.get("url"):
                c.hyperlink = r["url"]
                c.font = Font(name=F, size=10, color="0563C1", underline="single")
            if status == "BELOW":                     # RED = below ref. Sacred.
                c.fill = RED_FILL
                c.font = Font(name=F, size=10, bold=True, color=NEG,
                              underline="single" if r.get("url") else None)
                c.comment = Comment(
                    f"−₹{abs(diff or 0):,.0f} BELOW {regime} ref ₹{ref:,.0f}"
                    + (f"\n{len(r.get('stores_below') or [])} store(s) below ref"
                       if r.get("stores_below") else ""),
                    "pricematch", height=70, width=240)
            elif status == "ABOVE":                   # GREEN = above ref.
                c.fill = GREEN_FILL
                c.font = Font(name=F, size=10, bold=True, color=POS,
                              underline="single" if r.get("url") else None)
                c.comment = Comment(f"+₹{abs(diff or 0):,.0f} ABOVE {regime} ref ₹{ref:,.0f}",
                                    "pricematch", height=60, width=240)
            elif status == "NO_REF":
                c.comment = Comment("No reference price for this SKU/regime.",
                                    "pricematch", height=50, width=220)
            # MATCH: value + hyperlink, no fill

        # ---- Price-Match active (±₹5) — the new cluster column ----
        pmc = ws.cell(row, pm_col)
        clusters = [] if retired else pm_clusters(sku, by_key)
        if clusters:
            pmc.value = "⚡ " + "  |  ".join(fmt_cluster(cl) for cl in clusters)
            pmc.fill = SOFT_FILL
            pmc.font = Font(name=F, size=10, bold=True, color=BRAND)
            n = sum(len(cl) for cl in clusters)
            pmc.comment = Comment(
                f"{len(clusters)} price-match cluster(s) · {n} platform listings "
                f"within ₹{PM_CLUSTER_BAND:.0f} of each other (live modal price).",
                "pricematch", height=70, width=260)
        else:
            pmc.value = "" if retired else "—"
            pmc.font = Font(name=F, size=10, color=MUTED)
            pmc.alignment = Alignment(horizontal="center")
            if retired:
                pmc.fill = GREY_FILL
        row += 1
    # (legend lives in the frozen top row — fresh-eyes: never the last row)
    # (pin spread lived BELOW this grid 2026-06-11 evening for a few hours;
    #  owner moved it the same night to one "Pin Spread · <platform>" sheet
    #  per QC platform — see add_pin_spread_sheets)
    return ws


# ---------------------------------------------------------------- pin spread
def add_pin_spread_sheets(wb, date, regime, by_key, sku_map):
    """One "Pin Spread · <platform>" sheet per QC platform in CANONICAL
    (owner order 2026-06-11 night; replaces the below-Matrix section).
    Per SKU: every price group — MODE row highlighted first, then deviating
    groups (worst SKUs first), FULL pin lists (no cap — the team actions
    pins). Uniform SKUs keep their full pin list in their own section.
    National platforms carry one all-India price -> no sheet."""
    import pricematch_core as core

    for plat in [p for p in CANONICAL if p not in core.NATIONAL_PLATFORMS]:
        disp = PDISP.get(plat, plat)
        ws = wb.create_sheet(f"Pin Spread · {disp}"[:31])
        ws.sheet_view.showGridLines = False
        for col, w in (("A", 46), ("B", 10), ("C", 7), ("D", 110)):
            ws.column_dimensions[col].width = w
        title_cell(ws, 1, f"Pincode spread — {disp}", size=14, span=4)
        s = ws.cell(2, 1,
                    f"{date} · {regime_long(regime)} · Matrix shows the modal "
                    "(most-common) price across pins; this sheet shows every "
                    "price group behind it. MODE row highlighted; a group below "
                    "the agreed price (±₹1) is RED. A pin with two store prices "
                    "appears in both groups. Worst SKUs (most price groups) first. "
                    "Amazon Core / Flipkart are national — one "
                    "all-India price, no spread (see Matrix).")
        s.font = Font(name=F, size=9, color=MUTED)
        s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        ws.row_dimensions[2].height = 38

        dev, uni = [], []
        for sku, entry in sku_map.get("skus", {}).items():
            if entry.get("retired"):
                continue
            r = by_key.get((sku, plat)) or {}
            pg = r.get("pin_groups")
            if pg:
                (dev if len(pg) > 1 else uni).append((sku, r, pg))
        dev.sort(key=lambda t: (-len(t[2]), t[0]))

        if not dev and not uni:
            c = ws.cell(4, 1, "No pinned in-stock rows for this platform today.")
            c.font = Font(name=F, size=10, italic=True, color=MUTED)
            continue

        table_header(ws, 4, ["SKU / price group", "₹", "Pins", "Pincodes"])
        ws.freeze_panes = "A5"
        row = 5

        def group_row(row, g, ref, with_mode_label=True):
            below_ref = ref is not None and g["price"] < ref - 1
            lab = ws.cell(row, 1, "MODE" if (g["is_mode"] and with_mode_label) else "")
            lab.font = Font(name=F, size=9, bold=True,
                            color=BRAND if g["is_mode"] else MUTED)
            lab.alignment = Alignment(horizontal="right")
            money(ws.cell(row, 2, g["price"]), red=below_ref,
                  bold=g["is_mode"] or below_ref)
            nc = ws.cell(row, 3, len(g["pincodes"]))
            nc.font = Font(name=F, size=10, color=MUTED)
            txt = ", ".join(g["pincodes"])
            pc = ws.cell(row, 4, txt)
            pc.font = Font(name=F, size=9, color=NEG if below_ref else INK)
            pc.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
            if below_ref:
                for cc in range(1, 5):
                    ws.cell(row, cc).fill = RED_FILL
            elif g["is_mode"]:
                for cc in range(1, 5):
                    ws.cell(row, cc).fill = SOFT_FILL
            # explicit height: Quick Look / cached viewers don't auto-fit wraps
            lines = max(1, -(-len(txt) // 100))
            if lines > 1:
                ws.row_dimensions[row].height = min(14 * lines, 150)
            return row + 1

        if dev:
            t = ws.cell(row, 1, f"PRICE SPREAD — {len(dev)} SKU(s) where pins disagree")
            t.font = Font(name=F, size=12, bold=True, color=INK)
            row += 2
            for sku, r, pg in dev:
                ref = _f(r.get("ref_price"))
                n_pins = len({pc for g in pg for pc in g["pincodes"]})
                hdr = ws.cell(row, 1, f"{sku} — mode ₹{pg[0]['price']:,.0f} · "
                                      f"{len(pg[0]['pincodes'])} of {n_pins} pins at mode"
                                      f" · {len(pg)} price groups")
                hdr.font = Font(name=F, size=10, bold=True, color=INK)
                row += 1
                for g in pg:
                    row = group_row(row, g, ref)
                row += 1

        if uni:
            t = ws.cell(row, 1, f"UNIFORM — every pin one price ({len(uni)} SKU(s))")
            t.font = Font(name=F, size=12, bold=True, color=POS)
            row += 2
            for sku, r, pg in sorted(uni, key=lambda t: t[0]):
                ref = _f(r.get("ref_price"))
                hdr = ws.cell(row, 1, f"{sku} — ₹{pg[0]['price']:,.0f} at all "
                                      f"{len(pg[0]['pincodes'])} pins")
                hdr.font = Font(name=F, size=10, bold=True, color=INK)
                row += 1
                row = group_row(row, pg[0], ref, with_mode_label=False)
                row += 1


def sheet_violations(wb, date, regime, vrows, below_listings):
    ws = wb.create_sheet("Violations")
    ws.sheet_view.showGridLines = False
    widths = [26, 16, 16, 10, 11, 11, 11]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    title_cell(ws, 1, "Below-reference violations — every store", size=14, color=NEG, span=7)
    tot = sum(v["loss"] for v in vrows)
    # reconcile the cover's listing-level count with this sheet's store rows
    # (fresh-eyes MUST-3) in plain words (S6: no "modal"/"dark store" jargon).
    h = ws.cell(2, 1, f"The cover's {below_listings} listing-level violations expand to "
                      f"{len(vrows)} store-level rows under the {regime} reference "
                      f"({date}) · Σ loss ₹{tot:,.0f}. The most-common price can hide "
                      f"one cheaper store; this sheet lists every store below reference.")
    h.font = Font(name=F, size=10, color=MUTED)
    h.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 28

    # ---- roll-up first (fresh-eyes S7: make the wall of rows navigable) ----
    groups = {}
    for v in vrows:
        g = groups.setdefault((v["sku"], v["platform"]),
                              {"n": 0, "loss": 0.0, "worst": None, "url": v.get("url")})
        g["n"] += 1
        g["loss"] += v["loss"]
        if g["worst"] is None or v["loss"] > g["worst"]["loss"]:
            g["worst"] = v
    ranked = sorted(groups.items(), key=lambda kv: -kv[1]["loss"])
    TOP = 15
    t = ws.cell(4, 1, "WORST FIRST — by SKU × platform"
                      + (f" (top {TOP} of {len(ranked)})" if len(ranked) > TOP else ""))
    t.font = Font(name=F, size=12, bold=True, color=NEG)
    table_header(ws, 5, ["SKU", "Platform", "Stores below", "Σ loss ₹",
                         "Worst store (city @ ₹)"])
    ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=7)
    row = 6
    if not ranked:
        c = ws.cell(row, 1, "None — no store anywhere is selling below reference today.")
        c.font = Font(name=F, size=10, italic=True, color=POS)
        row += 1
    for (sku, plat), g in ranked[:TOP]:
        sc = ws.cell(row, 1, sku)
        sc.font = Font(name=F, size=10, color=INK)
        if g.get("url"):
            sc.hyperlink = g["url"]
            sc.font = Font(name=F, size=10, color="0563C1", underline="single")
        ws.cell(row, 2, PDISP.get(plat, plat)).font = Font(name=F, size=10, color=MUTED)
        ws.cell(row, 3, g["n"]).font = Font(name=F, size=10, bold=True, color=INK)
        lc = ws.cell(row, 4, round(g["loss"]))
        lc.number_format = "₹#,##0"
        lc.font = Font(name=F, size=10, bold=True, color=NEG)
        lc.fill = RED_FILL
        w = g["worst"]
        wc = ws.cell(row, 5, f"{w['city']} @ ₹{w['price']:,.0f} (−₹{w['loss']:,.0f})")
        wc.font = Font(name=F, size=10, color=INK)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        row += 1
    row += 1

    # ---- full store-level detail (appendix), filterable ----
    t = ws.cell(row, 1, f"EVERY STORE — all {len(vrows)} rows, by loss")
    t.font = Font(name=F, size=12, bold=True, color=INK)
    row += 1
    hdr_row = row
    table_header(ws, hdr_row, ["SKU", "Platform", "City", "Pincode",
                               "Store ₹", "Ref ₹", "Loss ₹"])
    row += 1
    for v in vrows:
        sc = ws.cell(row, 1, v["sku"])
        sc.font = Font(name=F, size=10, color=INK)
        if v.get("url"):
            sc.hyperlink = v["url"]
            sc.font = Font(name=F, size=10, color="0563C1", underline="single")
        ws.cell(row, 2, PDISP.get(v["platform"], v["platform"])).font = \
            Font(name=F, size=10, color=MUTED)
        ws.cell(row, 3, v["city"]).font = Font(name=F, size=10, color=INK)
        ws.cell(row, 4, str(v["pincode"])).font = Font(name=F, size=10, color=MUTED)
        money(ws.cell(row, 5, v["price"]), red=True)
        money(ws.cell(row, 6, v["ref"]))
        lc = ws.cell(row, 7, v["loss"])
        lc.number_format = "₹#,##0"
        lc.font = Font(name=F, size=10, bold=True, color=NEG)
        lc.fill = RED_FILL
        row += 1
    if vrows:
        ws.auto_filter.ref = "A%d:G%d" % (hdr_row, row - 1)
    return ws


def sheet_above(wb, date, regime, flat):
    ws = wb.create_sheet("Above reference")
    ws.sheet_view.showGridLines = False
    widths = [26, 13, 11, 11, 11, 9, 60]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    title_cell(ws, 1, "Above-reference listings", size=14, color=POS, span=7)
    above = sorted([r for r in flat if r.get("status") == "ABOVE"],
                   key=lambda r: -(_f(r.get("diff")) or 0))
    over = sum(1 for r in above if (_f(r.get("diff_pct")) or 0) > 25)
    # fresh-eyes S8: heavy over-pricing is a sales risk, not a win — amber it
    h = ws.cell(2, 1, f"{len(above)} listings priced above the {regime} reference ({date})"
                      + (f" · AMBER = more than 25% over reference — overpriced, "
                         f"sales risk ({over})" if over else ""))
    h.font = Font(name=F, size=10, color=MUTED)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    table_header(ws, 4, ["SKU", "Platform", "Ref ₹", "Live ₹", "Diff ₹", "Diff %", "Listing"])
    ws.freeze_panes = "A5"
    row = 5
    if not above:
        c = ws.cell(row, 1, "None above reference today.")
        c.font = Font(name=F, size=10, italic=True, color=MUTED)
    for r in above:
        overpriced = (_f(r.get("diff_pct")) or 0) > 25
        ws.cell(row, 1, r.get("sku")).font = Font(name=F, size=10, color=INK)
        ws.cell(row, 2, PDISP.get(r.get("platform"), r.get("platform"))).font = \
            Font(name=F, size=10, color=MUTED)
        money(ws.cell(row, 3, _f(r.get("ref_price"))))
        lv = ws.cell(row, 4, _f(r.get("live_modal")))
        if overpriced:
            lv.number_format = "₹#,##0"
            lv.font = Font(name=F, size=10, bold=True, color=WARN)
            lv.fill = AMBER_FILL
        else:
            money(lv, green=True)
            lv.fill = GREEN_FILL
        dc = ws.cell(row, 5, _f(r.get("diff")))
        dc.number_format = "+₹#,##0;-₹#,##0"
        dc.font = Font(name=F, size=10, bold=True, color=WARN if overpriced else POS)
        pc = ws.cell(row, 6, (_f(r.get("diff_pct")) or 0) / 100.0)
        pc.number_format = "+0.0%;-0.0%"
        pc.font = Font(name=F, size=10, color=MUTED)
        u = ws.cell(row, 7, r.get("url") or "")
        u.font = Font(name=F, size=9, color="0563C1", underline="single")
        if r.get("url"):
            u.hyperlink = r["url"]
        row += 1
    return ws


def sheet_coverage(wb, date, flat, by_key, sku_map):
    ws = wb.create_sheet("Coverage & pending")
    ws.sheet_view.showGridLines = False
    widths = [26, 13, 70]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    title_cell(ws, 1, "Coverage gaps, pending mappings, out-of-stock", size=14, span=3)
    row = 3

    # NOT_LISTED gaps — one row per SKU listing its missing platforms
    skus = sku_map.get("skus", {})
    gaps = []
    for sku, entry in skus.items():
        if entry.get("retired"):
            continue
        missing = [p for p in CANONICAL
                   if (by_key.get((sku, p)) or {}).get("status", "NOT_LISTED") == "NOT_LISTED"]
        if missing:
            gaps.append((sku, missing))
    # zero-count sections are dropped entirely (fresh-eyes ed-S5)
    if gaps:
        t = ws.cell(row, 1, f"NOT LISTED — matrix gaps ({sum(len(m) for _, m in gaps)} "
                            f"SKU×platform cells across {len(gaps)} SKUs)")
        t.font = Font(name=F, size=12, bold=True, color=INK)
        row += 1
        table_header(ws, row, ["SKU", "Gaps", "Missing platforms"], start_col=1)
        row += 1
        all_str = ", ".join(PDISP.get(p, p) for p in CANONICAL)
        for sku, missing in sorted(gaps, key=lambda g: -len(g[1])):
            ws.cell(row, 1, sku).font = Font(name=F, size=10, color=INK)
            ws.cell(row, 2, len(missing)).font = Font(name=F, size=10, color=MUTED)
            miss_str = ", ".join(PDISP.get(p, p) for p in missing)
            if miss_str == all_str:
                miss_str = f"all {len(CANONICAL)} platforms in this report"
            ws.cell(row, 3, miss_str).font = Font(name=F, size=10, color=MUTED)
            row += 1
        row += 1

    # PENDING_REVIEW
    pend = [r for r in flat if r.get("status") == "PENDING_REVIEW"]
    if pend:
        t = ws.cell(row, 1, f"PENDING REVIEW — mapped but unconfirmed ({len(pend)})")
        t.font = Font(name=F, size=12, bold=True, color=WARN)
        row += 1
        table_header(ws, row, ["SKU", "Platform", "Listing"], start_col=1)
        row += 1
        for r in pend:
            ws.cell(row, 1, r.get("sku")).font = Font(name=F, size=10, color=INK)
            ws.cell(row, 2, PDISP.get(r.get("platform"), r.get("platform"))).font = \
                Font(name=F, size=10, color=MUTED)
            u = ws.cell(row, 3, r.get("url") or r.get("listing_id") or "")
            if r.get("url"):
                u.hyperlink = r["url"]
                u.font = Font(name=F, size=9, color="0563C1", underline="single")
            row += 1
        row += 1

    # OOS
    oos = [r for r in flat if r.get("status") == "OOS"]
    if oos:
        t = ws.cell(row, 1, f"OUT OF STOCK — no price verdict ({len(oos)})")
        t.font = Font(name=F, size=12, bold=True, color=MUTED)
        row += 1
        table_header(ws, row, ["SKU", "Platform", "Listing"], start_col=1)
        row += 1
        for r in oos:
            ws.cell(row, 1, r.get("sku")).font = Font(name=F, size=10, color=INK)
            ws.cell(row, 2, PDISP.get(r.get("platform"), r.get("platform"))).font = \
                Font(name=F, size=10, color=MUTED)
            u = ws.cell(row, 3, r.get("url") or r.get("listing_id") or "")
            if r.get("url"):
                u.hyperlink = r["url"]
                u.font = Font(name=F, size=9, color="0563C1", underline="single")
            row += 1
    if row == 3:
        ws.cell(3, 1, "No gaps — every master SKU is listed, confirmed and in stock."
                ).font = Font(name=F, size=10, italic=True, color=POS)
    return ws


# ---------------------------------------------------------------- competitor PM sheets (NEW 2026-06-09)
# Owner ask: two "PM Check" sheets that FLIP the question the agreed-price sheets answer.
# The agreed sheets ask "is OUR listing below the agreed price (a violation)". These ask
# "is a COMPETITOR undercutting our Amazon price". The COLOR POLARITY IS DELIBERATELY
# OPPOSITE and that is spelled out LOUDLY in each legend:
#     RED   = competitor BELOW our Amazon price  → they are UNDERCUTTING us
#     GREEN = competitor ABOVE our Amazon price
#     no fill = MATCH (within ±₹1)
# Pinned to the two owner reference pincodes — 110092 (Anand Vihar, Delhi) + 560006
# (J.C. Nagar, Bengaluru). Owner order 2026-06-11: 110095/560005 are RETIRED — removed
# from every platform's sweep list; they must never reappear in any sheet.
# NATIONAL platforms (bigbasket / flipkart MP /
# amazon core) carry ONE price at every pincode. Built on the FROZEN pricematch_core
# competitor_compare contract; the whole block is fail-safe in main() so an engine hiccup
# can never break the master workbook (and therefore never break tomorrow's batch).

PM_COMPETE_CITY = {"110095": "Delhi", "110092": "Delhi",
                   "560005": "Bengaluru", "560006": "Bengaluru"}


def _compete_skus(sku_map):
    """Master SKU names (retired excluded) in map order — the rows of both PM sheets."""
    return [s for s, e in (sku_map.get("skus") or {}).items() if not e.get("retired")]


def _index_compete(records):
    """competitor_compare() list → {(sku,pincode): record} + per-record cell index."""
    by_key = {}
    for r in records or []:
        key = (r.get("sku"), str(r.get("pincode")))
        by_key.setdefault(key, r)
        cells = {}
        for cell in r.get("cells") or []:
            cells.setdefault(cell.get("platform"), cell)
        r["_cell_by_platform"] = cells
    return by_key


def _pending_pincodes(records, perpincode_keys, ref_key, ref_is_perpincode):
    """Pincodes NOT yet swept = no per-pincode platform produced a real price there.

    560005 reads "pending" today and AUTO-CLEARS the moment tomorrow's sweep lands a
    real row (then the cell renders the live value). National platforms are ignored
    here — their one price applies everywhere and never counts as "swept" evidence."""
    all_pin = set()
    swept = set()
    for r in records or []:
        pin = str(r.get("pincode"))
        all_pin.add(pin)
        if ref_is_perpincode and _f(r.get("ref_price")) is not None:
            swept.add(pin)
        for cell in r.get("cells") or []:
            if cell.get("platform") in perpincode_keys and _f(cell.get("price")) is not None:
                swept.add(pin)
    return all_pin - swept


def _paint_compete_cell(c, cell, *, blank, national, pending, ref_name):
    """Render one competitor cell; return "RED" if it's a (counted) undercut, else None.

    OWNER POLARITY (opposite the agreed sheets): BELOW ref = competitor cheaper = RED.
    """
    if blank:                                    # / — no feed yet
        c.value = ""
        c.fill = GREY_FILL
        return None
    if pending:                                  # pincode not yet swept (560005 today)
        c.value = "pending"
        c.font = Font(name=F, size=9, italic=True, color=MUTED)
        c.alignment = Alignment(horizontal="center")
        c.fill = GREY_FILL
        return None
    status = (cell or {}).get("status")
    price = _f((cell or {}).get("price"))
    if cell is None or status in (None, "NOT_SERVICEABLE"):
        # owner: kill the n/s wall — q-comm carries ~9–25 of 113 SKUs, so most cells are
        # "not sold here". Render a quiet faint "·" instead of a loud "n/s" everywhere,
        # so real prices + OOS stand out. Legend explains the dot.
        c.value = "·"
        c.font = Font(name=F, size=9, color=FAINT)
        c.alignment = Alignment(horizontal="center")
        return None
    if status == "OOS":
        c.value = "OOS"
        c.font = Font(name=F, size=9, bold=True, color=MUTED)
        c.alignment = Alignment(horizontal="center")
        return None
    c.value = price
    c.number_format = "₹#,##0"
    out = None
    if status == "BELOW":                        # competitor cheaper → undercut → RED
        c.fill = RED_FILL
        c.font = Font(name=F, size=10, bold=True, color=NEG)
        out = "RED"
    elif status == "ABOVE":                      # competitor dearer → GREEN
        c.fill = GREEN_FILL
        c.font = Font(name=F, size=10, bold=True, color=POS)
    elif status == "MATCH":                       # within ±₹1 → no fill
        c.font = Font(name=F, size=10, color=INK)
    elif status == "NO_REF":                       # priced, but no Amazon ref to compare
        c.font = Font(name=F, size=10, color=MUTED)
    else:
        c.font = Font(name=F, size=10, color=INK)
    return out


def _paint_ref_cell(c, ref_price, *, pending, ref_is_perpincode):
    """The reference-platform price cell (our Amazon Now / Amazon Core price)."""
    if pending and ref_is_perpincode:
        c.value = "pending"
        c.font = Font(name=F, size=9, italic=True, color=MUTED)
        c.alignment = Alignment(horizontal="center")
        c.fill = GREY_FILL
        return
    rp = _f(ref_price)
    if rp is None:
        c.value = "·"                            # we don't sell it here — quiet, not a loud n/s
        c.font = Font(name=F, size=9, color=FAINT)
        c.alignment = Alignment(horizontal="center")
        return
    c.value = rp
    c.number_format = "₹#,##0"
    c.font = Font(name=F, size=10, bold=True, color=INK)
    c.fill = SOFT_FILL                            # the baseline everything is compared to


def _row_price_by_platform(rec, ref_name, comp_cols, *, include_ref):
    """Build {display_platform: price} for ONE pincode row from the cells already on it —
    the reference price + each competitor's live price at this pincode. Only platforms
    with a REAL numeric live price (priced statuses) are included; OOS/n/s/blank dropped.
    Keyed by DISPLAY name so exact_price_match's sorted group names read cleanly."""
    pbp = {}
    if include_ref:
        rp = _f(rec.get("ref_price"))
        if rp is not None:
            pbp[ref_name] = rp
    for _cc, spec in comp_cols:
        if spec.get("blank"):
            continue
        cell = (rec.get("_cell_by_platform") or {}).get(spec["key"]) or {}
        price = _f(cell.get("price"))
        if price is not None and cell.get("status") in PM_PRICE_STATUSES:
            pbp[spec["label"]] = price
    return pbp


def _paint_exact_cell(c, groups, *, pending):
    """The trailing 'Price match (same ₹)' cell. EXACT identical-₹ grouping (NOT the
    ±₹5 cluster). Soft sage fill (NEVER red/green — those mean undercut/above here).
    Returns True if it flagged a real exact match (so the caller can count the SKU)."""
    if pending:                                  # pincode not yet swept (560005 today)
        c.value = "pending"
        c.font = Font(name=F, size=9, italic=True, color=MUTED)
        c.alignment = Alignment(horizontal="center")
        c.fill = GREY_FILL
        return False
    if not groups:
        c.value = "—"
        c.font = Font(name=F, size=10, color=MUTED)
        c.alignment = Alignment(horizontal="center")
        return False
    parts = [f"{' = '.join(g['platforms'])} @ ₹{g['price']:,.0f}" for g in groups]
    c.value = "⚡ " + "  |  ".join(parts)
    c.fill = SOFT_FILL                           # BRAND_SOFT — explicitly NOT red/green
    c.font = Font(name=F, size=9, bold=True, color=BRAND)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    return True


def _write_exact_summary(ws, summary_row, last_c, n_skus):
    """The top summary line (row 4): count + definition of the EXACT-match signal.
    Sage-filled so it visually ties to the ⚡ cells below."""
    noun = "SKU" if n_skus == 1 else "SKUs"
    cell = ws.cell(summary_row, 1,
                   f"⚡ {n_skus} {noun} with an exact cross-platform price match today.")
    cell.font = Font(name=F, size=10, bold=True, color=BRAND)
    cell.fill = SOFT_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _compete_legend(ws, row, ref_name):
    """One tight legend: the essential opposite-polarity colour rule + a short marker key.
    (Owner 2026-06-09: strip the note clutter — keep only the red/green meaning + a minimal key.)"""
    xd.legend(ws, row, [
        (f"RED = below {ref_name} (UNDERCUTTING us)", xd.RED_FILL_HEX, xd.RED_TEXT),
        (f"GREEN = above {ref_name}", xd.GREEN_FILL_HEX, xd.GREEN_TEXT),
        ("no fill = match (±₹1)", None, INK),
        ("⚡ = exact same ₹", xd.BRAND_SOFT, INK),
        ("· not sold · OOS out of stock", None, None),
    ], span=2)


def _sku_has_compete_data(sku, by_key, pincodes, comp_specs):
    """True if at least one (non-blank) competitor actually carries this SKU (has a live
    price at any reference pincode). Quick-commerce platforms sell only ~9–25 of the 113
    master SKUs, so without this filter the sheet is a wall of 'n/s' for products nobody
    on these platforms sells. We show ONLY the SKUs with a real head-to-head."""
    for pin in pincodes:
        rec = by_key.get((sku, pin)) or {}
        cells = rec.get("_cell_by_platform") or {}
        for spec in comp_specs:
            if spec.get("blank"):
                continue
            cell = cells.get(spec["key"]) or {}
            if _f(cell.get("price")) is not None or \
               cell.get("status") in ("BELOW", "ABOVE", "MATCH", "NO_REF"):
                return True
    return False


def _render_compete_sheet(wb, sheet_name, title, date, regime, records, sku_map, *,
                          ref_key, ref_name, columns, ref_is_perpincode,
                          ref_national_label=None, show_seller=False):
    """Generic renderer for both PM Check sheets — SKU rows × per-pincode sub-blocks.

    columns = ordered list of {key, label, blank?, national?} for the competitor cols.
    Returns the count of RED (undercut) cells painted.
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    pincodes = [str(p) for p in PM_REF_PINCODES_LOCAL]
    by_key = _index_compete(records)
    perpincode_keys = {c["key"] for c in columns if not c.get("national")}
    pending = _pending_pincodes(records, perpincode_keys, ref_key, ref_is_perpincode)
    # Only SKUs a competitor actually sells (else 90+ all-'n/s' rows). Preserves map order.
    shown_skus = [s for s in _compete_skus(sku_map)
                  if _sku_has_compete_data(s, by_key, pincodes, columns)]

    # ---- columns layout ----
    col = 1
    SKU_C = col; col += 1
    SELLER_C = None
    if show_seller:
        SELLER_C = col; col += 1
    PIN_C = col; col += 1
    REF_C = col; col += 1
    comp_cols = []
    for spec in columns:
        comp_cols.append((col, spec)); col += 1
    PM_C = col; col += 1                          # NEW: "Price match (same ₹)" trailing col
    last_c = col - 1

    ws.column_dimensions[get_column_letter(SKU_C)].width = 30
    if SELLER_C:
        ws.column_dimensions[get_column_letter(SELLER_C)].width = 18
    ws.column_dimensions[get_column_letter(PIN_C)].width = 16
    ws.column_dimensions[get_column_letter(REF_C)].width = 16
    for cc, _ in comp_cols:
        ws.column_dimensions[get_column_letter(cc)].width = 15
    ws.column_dimensions[get_column_letter(PM_C)].width = 40

    # ---- title + regime badge (title spans all but the last col; badge sits in it) ----
    title_cell(ws, 1, title, size=16, span=max(1, last_c - 1))
    badge_color, badge_text = REGIME_BADGE.get(regime, (MUTED, f"{regime} day"))
    b = ws.cell(1, last_c, f"{date} · {badge_text}".upper())
    b.font = Font(name=F, size=10, bold=True, color="FFFFFF")
    b.fill = PatternFill("solid", fgColor=badge_color)
    b.alignment = Alignment(horizontal="center", vertical="center")

    # ---- tight sub-title: competitor live price vs our reference (colour is OPPOSITE the
    # agreed sheets — the one warning that must stay; everything else stripped per owner) ----
    sub = f"Competitor live price vs our {ref_name} — colour is OPPOSITE the agreed sheets (see legend)."
    hs = ws.cell(2, 1, sub)
    hs.font = Font(name=F, size=9, color=MUTED)
    hs.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_c)
    ws.row_dimensions[2].height = 24

    _compete_legend(ws, 3, ref_name)

    # ---- top summary line (row 4): defines + counts the EXACT-match signal. The
    # count is written AFTER the body loop (we don't know it yet); reserve the row now
    # so the header/freeze land below it. ----
    SUMMARY_ROW = 4
    ws.merge_cells(start_row=SUMMARY_ROW, start_column=1, end_row=SUMMARY_ROW, end_column=last_c)
    ws.row_dimensions[SUMMARY_ROW].height = 20

    # ---- header row ----
    hdr_row = 5
    ws.cell(hdr_row, SKU_C, "SKU")
    if SELLER_C:
        ws.cell(hdr_row, SELLER_C, "Buybox seller")
    ws.cell(hdr_row, PIN_C, "Pincode")
    ref_hdr = (ref_national_label or ref_name) + (" (national)" if not ref_is_perpincode else "")
    ws.cell(hdr_row, REF_C, ref_hdr + " — ref ₹")
    for cc, spec in comp_cols:
        lbl = spec["label"]
        if spec.get("national"):
            lbl += " (national)"
        elif spec.get("blank"):
            lbl += " — no feed yet"
        ws.cell(hdr_row, cc, lbl)
    ws.cell(hdr_row, PM_C, "Price match (same ₹)")
    for cc in range(1, last_c + 1):
        h = ws.cell(hdr_row, cc)
        h.font = Font(name=F, size=10, bold=True, color="FFFFFF")
        h.fill = HDR_FILL
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(hdr_row, REF_C).fill = PatternFill("solid", fgColor=INK)
    ws.row_dimensions[hdr_row].height = 28
    ws.freeze_panes = ws.cell(hdr_row + 1, PIN_C + 1).coordinate

    # ---- body: one sub-block (per pincode) per SKU (only SKUs a competitor sells) ----
    import pricematch_core as core
    red_count = 0
    exact_match_skus = set()                      # distinct SKUs with >=1 exact ₹ match
    row = hdr_row + 1
    if not shown_skus:
        nc = ws.cell(row, 1, "No competitor is selling any tracked SKU at these pincodes today.")
        nc.font = Font(name=F, size=10, italic=True, color=MUTED)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_c)
        _write_exact_summary(ws, SUMMARY_ROW, last_c, 0)
        return 0
    for sku in shown_skus:
        block_top = row
        seller_txt = None
        for pin in pincodes:
            rec = by_key.get((sku, pin)) or {}
            is_pending = pin in pending
            # SKU + seller written once (merged) at the block top
            if pin == pincodes[0]:
                sc = ws.cell(block_top, SKU_C, sku)
                sc.font = Font(name=F, size=10, bold=True, color=INK)
                sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                if show_seller:
                    seller_txt = (rec.get("ref_seller") or {}).get("seller") \
                        if isinstance(rec.get("ref_seller"), dict) else rec.get("ref_seller")
            pc = ws.cell(row, PIN_C, f"{pin} · {PM_COMPETE_CITY.get(pin, '')}".strip(" ·"))
            pc.font = Font(name=F, size=9, color=MUTED)
            pc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            _paint_ref_cell(ws.cell(row, REF_C), rec.get("ref_price"),
                            pending=is_pending, ref_is_perpincode=ref_is_perpincode)
            for cc, spec in comp_cols:
                cell = (rec.get("_cell_by_platform") or {}).get(spec["key"])
                # national columns are NEVER pending — their one price applies everywhere
                cell_pending = is_pending and not spec.get("national")
                res = _paint_compete_cell(
                    ws.cell(row, cc), cell, blank=bool(spec.get("blank")),
                    national=bool(spec.get("national")), pending=cell_pending,
                    ref_name=ref_name)
                if res == "RED":
                    red_count += 1
            # EXACT price-match indicator (trailing col): which platforms ON THIS ROW
            # sit at the identical ₹. Pending rows show "pending" (no data yet).
            # Inner guard (W3 note): a per-row failure degrades the cell to "—" rather
            # than dropping the whole sheet — matches the Ecom Head section's granularity.
            try:
                pbp = _row_price_by_platform(rec, ref_name, comp_cols, include_ref=True)
                groups = [] if is_pending else core.exact_price_match(pbp)
                if _paint_exact_cell(ws.cell(row, PM_C), groups, pending=is_pending):
                    exact_match_skus.add(sku)
            except Exception:
                ec = ws.cell(row, PM_C, "—")
                ec.font = Font(name=F, size=10, color=MUTED)
            # zebra the SKU/pincode gutter lightly on the 2nd pincode
            if pin != pincodes[0]:
                for cc in range(1, PIN_C + 1):
                    if ws.cell(row, cc).fill.fgColor.rgb in (None, "00000000"):
                        ws.cell(row, cc).fill = GREY_FILL
            row += 1
        # merge SKU (and seller) down the block
        if row - 1 > block_top:
            ws.merge_cells(start_row=block_top, start_column=SKU_C, end_row=row - 1, end_column=SKU_C)
            if SELLER_C:
                ws.merge_cells(start_row=block_top, start_column=SELLER_C, end_row=row - 1, end_column=SELLER_C)
        if SELLER_C:
            sl = ws.cell(block_top, SELLER_C, seller_txt or "—")
            sl.font = Font(name=F, size=9, color=INK if seller_txt else MUTED)
            sl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # thin divider under each SKU block
        for cc in range(1, last_c + 1):
            ws.cell(row, cc).fill = RULE_FILL
        ws.row_dimensions[row].height = 3
        row += 1
    _write_exact_summary(ws, SUMMARY_ROW, last_c, len(exact_match_skus))
    print(f"build_pricematch: {sheet_name} · exact cross-platform price-match SKUs = "
          f"{len(exact_match_skus)}", file=sys.stderr)
    return red_count


# the two owner reference pincodes (110092 Delhi, 560006 Bengaluru — 110095/560005
# retired by owner order 2026-06-11). Mirrors the frozen pricematch_core.PM_REF_PINCODES;
# bound from the engine at call time when present.
PM_REF_PINCODES_LOCAL = ["110092", "560006"]


def build_compete_sheets(wb, date, regime, sku_map):
    """Append the two competitor PM Check sheets; return {sheet: red_cell_count}.

    Binds to the FROZEN pricematch_core competitor_compare contract. Raises on a missing
    engine / bad data so the caller's fail-safe can skip JUST these sheets.
    """
    global PM_REF_PINCODES_LOCAL
    import pricematch_core as core
    PM_REF_PINCODES_LOCAL = [str(p) for p in getattr(core, "PM_REF_PINCODES", PM_REF_PINCODES_LOCAL)]
    skus = _compete_skus(sku_map)

    # ---- Sheet 1: Amazon Now PM Check ----
    now_cols = [
        {"key": "blinkit", "label": "Blinkit"},
        {"key": "zepto", "label": "Zepto"},
        {"key": "flipkart-minutes", "label": "Flipkart Minutes"},
        {"key": "bigbasket", "label": "BigBasket"},  # per-pincode now (QC pincode feed) — price at the ref pins
    ]
    now_recs = core.competitor_compare(
        "amazon-now", [c["key"] for c in now_cols], skus, pincodes=PM_REF_PINCODES_LOCAL)
    red_now = _render_compete_sheet(
        wb, "Amazon Now PM Check",
        "Amazon Now — Price-Match Check (quick-commerce)", date, regime, now_recs, sku_map,
        ref_key="amazon-now", ref_name="Amazon Now", columns=now_cols,
        ref_is_perpincode=True)

    # ---- Sheet 2: Amazon Core PM Check ----
    core_cols = [
        {"key": "blinkit", "label": "Blinkit"},
        {"key": "zepto", "label": "Zepto"},
        {"key": "flipkart-minutes", "label": "Flipkart Minutes"},
        {"key": "flipkart", "label": "Flipkart MP", "national": True},
    ]
    core_recs = core.competitor_compare(
        "amazon", [c["key"] for c in core_cols], skus, pincodes=PM_REF_PINCODES_LOCAL)
    red_core = _render_compete_sheet(
        wb, "Amazon Core PM Check",
        "Amazon Core — Price-Match Check (national buybox)", date, regime, core_recs, sku_map,
        ref_key="amazon", ref_name="Amazon Core", columns=core_cols,
        ref_is_perpincode=False, ref_national_label="Amazon Core", show_seller=True)

    return {"Amazon Now PM Check": red_now, "Amazon Core PM Check": red_core}


# ---------------------------------------------------------------- summary sidecar
def tg_summary(date, regime, kpi, flat):
    """Deterministic 6-line Telegram markdown (run_all.sh spools/sends this)."""
    below = sorted([r for r in flat if r.get("status") == "BELOW"],
                   key=lambda r: _f(r.get("diff")) or 0)
    top3 = " · ".join(
        f"{esc_md(r.get('sku'))} ({esc_md(PDISP.get(r.get('platform'), r.get('platform')))} "
        f"−₹{abs(_f(r.get('diff')) or 0):,.0f})"
        for r in below[:3]) or "none"
    lines = [
        "*Jivo Price Match — master* \U0001F3AF",
        f"{date} · {regime} day",
        f"\U0001F534 Below ref: {kpi['below']} listings · ₹{kpi['exposure']:,} exposure",
        f"Top offenders: {top3}",
        f"\U0001F7E2 Above: {kpi['above']} · Match: {kpi['compliant']} · "
        f"OOS: {kpi['oos']} · Pending: {kpi['pending']}",
        f"Coverage: {kpi['skus_tracked']} SKUs × {kpi['platforms']} platforms "
        f"({kpi['not_listed']} not listed)",
    ]
    return "\n".join(lines)


# ---------------------------------------------------------------- byte-stability
def stabilize_xlsx(path, stamp, props):
    """LEAD ruling 2026-06-06: same date ⇒ byte-identical workbook.

    openpyxl re-stamps properties.modified=now inside save() AND embeds the
    wall-clock in every zip entry's local header, so pinning wb.properties
    before save is not enough. Rewrite the saved zip deterministically:
    every entry's mtime pinned to the report date, and docProps/core.xml
    re-serialized (via openpyxl's own tostring) with created=modified=stamp.
    """
    import zipfile
    from openpyxl.xml.functions import tostring

    props.created = stamp
    props.modified = stamp
    core_xml = tostring(props.to_tree())
    with zipfile.ZipFile(path) as zin:
        items = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    dt = (stamp.year, stamp.month, stamp.day, 0, 0, 0)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as zout:
        for name, data in items:
            if name == "docProps/core.xml":
                data = core_xml
            zi = zipfile.ZipInfo(name, date_time=dt)
            zi.compress_type = zipfile.ZIP_DEFLATED
            zout.writestr(zi, data)


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Build the master Price Match workbook")
    ap.add_argument("--date", help="YYYY-MM-DD (default: today)")
    ap.add_argument("--out", help="output directory (default: this script's dir)")
    ap.add_argument("--platforms", help="CSV subset edition (e.g. amazon,amazon-now); default = all")
    ap.add_argument("--label", help="edition label for the filename (e.g. Amazon)")
    args = ap.parse_args()

    date, regime, comps, summ, sku_map = gather(args.date)
    if args.platforms:
        # Scoped edition: shrink the canonical column list IN PLACE (all sheet fns read it),
        # filter comps (flatten() has a defensive catch-all that would re-add others), and
        # drop the engine's GLOBAL summary so exposure falls back to the scoped sum.
        subset = [s.strip() for s in args.platforms.split(",") if s.strip()]
        CANONICAL[:] = [p for p in CANONICAL if p in subset]
        comps = {p: r for p, r in comps.items() if p in subset}
        summ = {}
    flat, by_key = flatten(comps)
    kpi = kpis_of(flat, summ, sku_map)
    vrows = store_violations(flat)

    wb = Workbook()
    wb.remove(wb.active)
    sheet_ecom_head(wb, date, regime, flat, comps, kpi, sku_map, label=args.label)
    sheet_matrix(wb, date, regime, by_key, sku_map, listed_only=bool(args.label))
    # Pin Spread tabs (owner order 2026-06-11 night: spread moved off the
    # Matrix into one sheet per QC platform). FAIL-SAFE like the PM Check
    # sheets: any error drops JUST these tabs, never the workbook.
    try:
        add_pin_spread_sheets(wb, date, regime, by_key, sku_map)
    except Exception as e:
        import traceback
        print(f"build_pricematch: pin-spread sheets SKIPPED (non-fatal): {e}",
              file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        for nm in [n for n in wb.sheetnames if n.startswith("Pin Spread")]:
            wb.remove(wb[nm])
    sheet_violations(wb, date, regime, vrows, kpi["below"])
    sheet_above(wb, date, regime, flat)
    sheet_coverage(wb, date, flat, by_key, sku_map)

    # NEW (2026-06-09): the two competitor "PM Check" sheets, appended AFTER the existing
    # sheets so those stay byte-identical. FAIL-SAFE — any error (engine absent, bad data,
    # render bug) skips JUST these two sheets and logs; the rest of the workbook still saves
    # so tomorrow's batch can never be broken by them. Only on the FULL national workbook
    # (scoped --platforms editions stay byte-identical / don't carry the cross-platform view).
    compete_reds = None
    if not args.platforms:
        try:
            compete_reds = build_compete_sheets(wb, date, regime, sku_map)
            print(f"build_pricematch: competitor PM sheets added · undercut(RED) cells "
                  f"{compete_reds}", file=sys.stderr)
        except Exception as e:
            import traceback
            print(f"build_pricematch: competitor PM sheets SKIPPED (non-fatal): {e}",
                  file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            # remove any half-built sheet so the workbook stays clean
            for nm in ("Amazon Now PM Check", "Amazon Core PM Check"):
                if nm in wb.sheetnames:
                    wb.remove(wb[nm])

    wb.active = 0

    out_dir = args.out or HERE
    tag = f"{args.label}-" if args.label else ""
    xlsx = os.path.join(out_dir, f"Jivo-Price-Match-{tag}{date}.xlsx")
    wb.save(xlsx)
    # Byte-stability (LEAD ruling 2026-06-06): pin doc props + zip mtimes to the
    # report date (midnight IST) so rebuilding the same day is byte-identical.
    stabilize_xlsx(xlsx, datetime.datetime.fromisoformat(date), wb.properties)

    sidecar = {
        "date": date, "regime": regime, "kpis": kpi,
        "store_violations": len(vrows),
        "summary_md": tg_summary(date, regime, kpi, flat),
        "caption": f"Jivo Price Match · {date} · {regime} day",
    }
    with open(xlsx + ".summary.json", "w", encoding="utf-8") as fh:
        json.dump(sidecar, fh, ensure_ascii=False, indent=1)

    # Best-effort: persist the machine price-match history (data/pricematch/*.csv) from
    # the comps/summ we already hold — AFTER the xlsx is saved + byte-stabilized, so a
    # history-write error can never fail the workbook build and can't touch the xlsx
    # bytes. Idempotent (replace-by-date). Skipped for scoped editions (--platforms): the
    # FULL national workbook owns the canonical daily history. (W1 — pm-history)
    # PM_SKIP_HISTORY=1 also skips it: test/diagnostic builds for a PAST --date would
    # otherwise overwrite that date's real history with TODAY's prices (caught
    # 2026-06-10 — check_safety's frozen --date 2026-06-06 had been silently
    # corrupting the 06-06 rows on every doctor/gate run).
    if not args.platforms and os.environ.get("PM_SKIP_HISTORY") != "1":
        try:
            import pricematch_history as pmh
            n, _ = pmh.record_run(date, regime, comps, store_violation_count=len(vrows))
            print(f"build_pricematch: pm-history captured {n} rows for {date}", file=sys.stderr)
        except Exception as e:  # never fail the build on a history hiccup
            print(f"build_pricematch: pm-history skipped (non-fatal): {e}", file=sys.stderr)

    print(f"build_pricematch: {regime} day · {kpi['below']} below-ref "
          f"(₹{kpi['exposure']:,} exposure) · {kpi['above']} above · "
          f"{kpi['compliant']} match · {kpi['oos']} OOS · {len(vrows)} store rows",
          file=sys.stderr)
    print(os.path.abspath(xlsx))               # LAST stdout line = the xlsx path
    return 0


if __name__ == "__main__":
    sys.exit(main())
