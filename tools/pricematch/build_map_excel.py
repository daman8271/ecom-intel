#!/usr/bin/env python3
"""Build the Jivo-SKU-Master-Map xlsx FROM sku_map.json (the enriched, folded, extended map).

Replaces the Excel half of build_master_map.py (which stays for fragment re-merges):
the workbook must always be generated from the final map, never from fragments —
otherwise late enrichment (id-fold urls, W1 extensions) never reaches the sheet.

Owner orders honoured here (2026-06-06):
  1. EVERY non-empty platform price cell carries a hyperlink — asserted programmatically
     after writing (workbook re-opened, cells counted vs hyperlinks, loud failure on
     mismatch). A map entry with no url renders the value WITH a visible "⚠ no-link"
     marker so it can never pass silently as a linked cell.
  2. EXACT, FULL listing titles everywhere — no truncation anywhere. Titles are taken
     from the freshest platforms/<p>/result.json via titles_overlay.json (built here):
       amazon                       -> sku_raw   (full PDP title, keyed by asin)
       amazon-fresh / amazon-now    -> amazon's PDP title for the same asin (their own
                                       sku_raw is the SHORTENED search-card title; the
                                       map url opens /dp/<asin>, i.e. the PDP), else
                                       their own sku_raw
       flipkart                     -> fk_name   (real on-page listing name; sku_raw /
                                       item / product_name are INTERNAL catalog names),
                                       keyed by fsn
       flipkart-minutes             -> sku_raw   (real listing name), keyed by fk_pid
       zepto                        -> sku_raw, keyed by variant_id (= the pvid in urls)
       blinkit                      -> sku_raw, keyed by prid
       bigbasket                    -> brand + sku_raw (page shows "Jivo <desc>"; the
                                       API desc alone omits the brand), keyed by sku_id
     Fallbacks: map entry's stored title -> master-sheet Amazon Title (by ASIN) ->
     internal name (last resort, counted + reported).

titles_overlay.json ({platform: {id: exact_title}}) is emitted as a standalone artifact
so W1 can fold it into sku_map.json later — this script never writes sku_map.json.

Usage: build_map_excel.py [DATE]          (default: today)
Deterministic + idempotent. No network. No LLM.
"""
import json, sys, datetime, pathlib, collections, re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).parent
ROOT = HERE.parent.parent
DATE = sys.argv[1] if len(sys.argv) > 1 else str(datetime.date.today())

PLATFORMS = ["amazon", "amazon-fresh", "amazon-now", "flipkart", "flipkart-minutes",
             "zepto", "blinkit", "bigbasket"]
AMZ_FAMILY = {"amazon", "amazon-fresh", "amazon-now"}
NO_LINK = "⚠ no-link"

# (id field, title field) per platform in platforms/<p>/result.json
RESULT_FIELDS = {
    "amazon":           ("asin",       "sku_raw"),
    "amazon-fresh":     ("asin",       "sku_raw"),
    "amazon-now":       ("asin",       "sku_raw"),
    "flipkart":         ("fsn",        "fk_name"),
    "flipkart-minutes": ("fk_pid",     "sku_raw"),
    "zepto":            ("variant_id", "sku_raw"),
    "blinkit":          ("prid",       "sku_raw"),
    "bigbasket":        ("sku_id",     "sku_raw"),
}


# ---------------- titles overlay ----------------
def build_overlay():
    """{platform: {id: exact_title}} from the freshest result.json per platform."""
    raw = {}  # platform -> id -> Counter(title)
    for plat, (idf, tf) in RESULT_FIELDS.items():
        f = ROOT / "platforms" / plat / "result.json"
        if not f.exists():
            print(f"overlay: platforms/{plat}/result.json MISSING — skipped", file=sys.stderr)
            continue
        d = json.loads(f.read_text())
        rows = d.get("allRows") or d.get("rows") or []
        node = raw.setdefault(plat, {})
        for r in rows:
            rid, title = r.get(idf), r.get(tf)
            if rid is None or not title:
                continue
            if plat == "bigbasket":
                brand = (r.get("brand") or "").strip()
                if brand and not title.lower().startswith(brand.lower()):
                    title = f"{brand} {title}"
            node.setdefault(str(rid), collections.Counter())[title.strip()] += 1

    def modal(counter):  # most frequent, then longest, then lexicographic — deterministic
        return sorted(counter.items(), key=lambda kv: (-kv[1], -len(kv[0]), kv[0]))[0][0]

    overlay = {plat: {rid: modal(c) for rid, c in node.items()} for plat, node in raw.items()}
    # amazon family: the map url opens the PDP — prefer amazon's full PDP title per asin
    pdp = overlay.get("amazon", {})
    for plat in ("amazon-fresh", "amazon-now"):
        for rid in list(overlay.get(plat, {})):
            if rid in pdp:
                overlay[plat][rid] = pdp[rid]
    return overlay


# master-sheet Amazon Title fallback (by ASIN)
def load_master_titles():
    titles = {}
    wb = openpyxl.load_workbook(HERE / "Amazon Price Match.xlsx", data_only=True)
    ws = wb["url"]; hdr = [c.value for c in ws[1]]; ix = {h: i for i, h in enumerate(hdr)}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ix["ASIN"]] and r[ix["Title"]]:
            titles[str(r[ix["ASIN"]]).strip()] = str(r[ix["Title"]]).strip()
    return titles


sku_map = json.loads((HERE / "sku_map.json").read_text())
overlay = build_overlay()
(HERE / "titles_overlay.json").write_text(
    json.dumps(overlay, indent=1, ensure_ascii=False, sort_keys=True))
print(f"titles_overlay.json: {sum(len(v) for v in overlay.values())} titles "
      f"across {len(overlay)} platforms")

# master-sheet titles by PRODUCT name (e-com sheet fallback for unresolved listings)
def load_master_titles_by_product():
    titles = {}
    wb = openpyxl.load_workbook(HERE / "Amazon Price Match.xlsx", data_only=True)
    ws = wb["url"]; hdr = [c.value for c in ws[1]]; ix = {h: i for i, h in enumerate(hdr)}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ix["PRODUCT"]] and r[ix["Title"]]:
            titles[str(r[ix["PRODUCT"]]).strip().upper()] = str(r[ix["Title"]]).strip()
    return titles


master_titles = load_master_titles()
master_titles_by_product = load_master_titles_by_product()
title_src = collections.Counter()

# OWNER ORDER (Punkirat feedback 2026-06-06): a title cell holds the exact on-page
# listing name ONLY — never internal names, never url-slugs, never annotations like
# "(blessed)" / "(unpriced)" / "(owner sheet URL)". Annotations move to the note column.
ANNOT_RE = re.compile(r"\s*\((?:blessed|unpriced|owner sheet)[^)]*\)", re.IGNORECASE)
SLUG_RE = re.compile(r"[a-z0-9]+(?:-[a-z0-9]+){2,}")


def team_lang(s):
    """Team-facing language for evidence/reason/note text (never title cells)."""
    if not s:
        return s
    return re.sub(r"(?:owner[- ])?blessed", "owner-verified", str(s), flags=re.IGNORECASE)


def exact_title(plat, le, sku=None):
    """Resolve the exact on-page listing title -> (title, displaced_note).

    title is "" when no on-page title exists offline — the explanation then lives in
    displaced_note, which call sites append to their note/reason column.
    """
    rid = le.get("id")
    t = overlay.get(plat, {}).get(str(rid)) if rid is not None else None
    if t:
        title_src["overlay"] += 1; return t, ""
    if plat in AMZ_FAMILY and rid is not None:
        t = overlay.get("amazon", {}).get(str(rid))
        if t:
            title_src["overlay-amazon"] += 1; return t, ""
    notes = []
    stored = le.get("title") or ""
    if stored:
        cleaned = ANNOT_RE.sub("", stored).strip()
        annot = stored.replace(cleaned, "").strip()
        if annot:
            notes.append(f"title annotation moved: {annot}")
        if cleaned and SLUG_RE.fullmatch(cleaned):
            notes.append(f"url-slug (not a title): {cleaned}")
            cleaned = ""
        if cleaned:
            title_src["map-title"] += 1
            return cleaned, "; ".join(notes)
    if plat in AMZ_FAMILY and rid is not None and str(rid) in master_titles:
        title_src["master-sheet"] += 1
        return master_titles[str(rid)], "; ".join(notes)
    prod = (sku or le.get("candidate") or "").strip().upper()
    if prod in master_titles_by_product:
        title_src["master-sheet-by-product"] += 1
        notes.append("title from e-com master sheet (listing not resolvable offline)")
        return master_titles_by_product[prod], "; ".join(notes)
    if rid is None and not le.get("url") and not stored:
        return "", ""  # a plain note row, not a listing — no title expected
    title_src["no-title"] += 1
    internal = le.get("item") or le.get("internal_name")
    notes.append(f"internal name: {internal} — listing not yet scraped, no on-page title"
                 if internal else "title not captured (listing not yet scraped)")
    return "", "; ".join(notes)


def slug_from(le):
    """The url-slug for the id column when a listing has no id (slug lives THERE only)."""
    stored = ANNOT_RE.sub("", le.get("title") or "").strip()
    if SLUG_RE.fullmatch(stored):
        return stored
    return None


# ---------------- Excel ----------------
JIVO = "008B3A"
HDR = PatternFill("solid", fgColor=JIVO); HDRF = Font(bold=True, color="FFFFFF", size=10)
TITLE = Font(bold=True, color=JIVO, size=16); SUB = Font(italic=True, color="555555", size=9)
RED = PatternFill("solid", fgColor="F4CCCC"); GRN = PatternFill("solid", fgColor="D9EAD3")
YEL = PatternFill("solid", fgColor="FFF2CC"); BLU = Font(color="1155CC", underline="single", size=9)
GRY = Font(italic=True, color="888888", size=10)
thin = Side(style="thin", color="D0D0D0"); BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center"); SM = Font(size=9)

wb = Workbook()
nolink = []   # (sku, platform, id) — expected EMPTY after the id-fold


def style_hdr(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = CEN; cell.border = BRD


def autosize(ws, maxw=46):
    for col in ws.columns:
        L = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[L].width = min(maxw, max(9, w + 2))


products = sorted(sku_map["skus"].keys())

# Sheet 1: Master Map (wide) — one row per SKU, per-platform price + confidence
ws = wb.active; ws.title = "Master Map"
ws["A1"] = "Jivo SKU Master Map — every SKU pinned on every platform"; ws["A1"].font = TITLE
ws["A2"] = f"generated {DATE} from sku_map.json (enriched) + e-com master sheet; cell = live modal sale price (every cell hyperlinks to the listing); blank = not listed"; ws["A2"].font = SUB
hdr = ["SKU", "MRP", "BAU", "SVD", "ART"] + [p for p in PLATFORMS]
ws.append([]); ws.append(hdr)
for p in products:
    e = sku_map["skus"][p]
    rg = e.get("regimes") or {}
    label = p + (" (retired)" if e.get("retired") else "")
    ws.append([label, rg.get("mrp"), rg.get("bau"), rg.get("svd"), rg.get("art")]
              + [None] * len(PLATFORMS))
    r = ws.max_row
    if e.get("retired"):
        ws.cell(row=r, column=1).font = GRY
    for j, plat in enumerate(PLATFORMS):
        cell = ws.cell(row=r, column=6 + j)
        le = (e.get("platforms") or {}).get(plat)
        if not le:
            continue
        if le.get("sale_modal") is not None:
            val = le.get("sale_modal")
        elif le.get("in_stock") is None:   # sheet-asserted listing, never scraped
            val = "listed"
        else:
            val = "listed" if le.get("in_stock") else "OOS"
        if le.get("url"):
            cell.value = val
            cell.hyperlink = le["url"]; cell.font = BLU
        else:  # post-fold this must never happen — make it impossible to miss
            cell.value = f"{val} {NO_LINK}"
            cell.fill = RED
            nolink.append((p, plat, le.get("id")))
        conf = le.get("confidence", "") or ""
        if conf.startswith("exact"): cell.fill = GRN
        elif conf.startswith("anchored"): cell.fill = YEL
        cell.border = BRD; cell.alignment = CEN
style_hdr(ws, 4); ws.freeze_panes = "B5"; autosize(ws, 22)

# Sheet 2: Listing Details (long) — one row per SKU x platform listing, EXACT full title
ws = wb.create_sheet("Listing Details")
ws.append(["SKU", "platform", "listing id", "url", "exact listing title", "live MRP",
           "sale modal", "sale min", "sale max", "in stock", "method", "confidence", "note"])
for p in products:
    for plat in PLATFORMS:
        le = (sku_map["skus"][p].get("platforms") or {}).get(plat)
        if not le:
            continue
        t, extra = exact_title(plat, le, sku=p)
        ws.append([p, plat, le.get("id") or slug_from(le), le.get("url"), t,
                   le.get("mrp"), le.get("sale_modal"), le.get("sale_min"), le.get("sale_max"),
                   "?" if le.get("in_stock") is None else ("Y" if le.get("in_stock") else "N"),
                   le.get("method"), le.get("confidence"),
                   team_lang("; ".join(x for x in (le.get("note", ""), extra) if x))])
        for c in ws[ws.max_row]: c.font = SM
        if le.get("url"):
            u = ws.cell(row=ws.max_row, column=4); u.hyperlink = le["url"]; u.font = BLU
style_hdr(ws); ws.freeze_panes = "A2"; autosize(ws)

# Sheet 3: Review — needs owner YES/NO, EXACT full title
ws = wb.create_sheet("Review (confirm these)")
ws.append(["platform", "listing id", "url", "exact listing title", "candidate SKU",
           "reason", "OWNER: YES/NO"])
for it in sku_map.get("review", []):
    if isinstance(it, str):  # some fragments emit plain-string notes
        it = {"reason": it}
    plat = it.get("platform") or ""
    t, extra = exact_title(plat, it)
    ws.append([plat, it.get("id") or slug_from(it), it.get("url"), t,
               it.get("candidate"),
               team_lang("; ".join(x for x in (it.get("reason") or "", extra) if x)), ""])
    if it.get("url"):
        u = ws.cell(row=ws.max_row, column=3); u.hyperlink = it["url"]; u.font = BLU
    ws.cell(row=ws.max_row, column=7).fill = YEL
style_hdr(ws); ws.freeze_panes = "A2"; autosize(ws)

# Sheet 4: MRP integrity — stale sheet rows + per-platform drift
ws = wb.create_sheet("MRP integrity")
ws.append(["type", "SKU", "sheet MRP", "live MRP", "where", "evidence / consensus"])
for k, s in sorted((sku_map.get("master_mrp_stale") or {}).items()):
    ws.append(["MASTER SHEET STALE", k, s.get("master_mrp"), s.get("live_mrp"),
               ", ".join(s.get("platforms", [])), team_lang(s.get("evidence") or "")])
    ws.cell(row=ws.max_row, column=4).fill = RED
for s in sku_map.get("mrp_drift", []):
    ws.append(["PLATFORM DRIFT", s.get("product"), s.get("official_mrp"),
               s.get("platform_mrp"), s.get("platform"), team_lang(s.get("consensus", ""))])
    ws.cell(row=ws.max_row, column=4).fill = YEL
style_hdr(ws); ws.freeze_panes = "A2"; autosize(ws)

# Sheet 5: Unpriced census — EXACT full title
ws = wb.create_sheet("Unpriced (no master row)")
ws.append(["family / listing", "platform(s)", "id", "mrp", "combo", "suggestion / internal name"])
fams = [u for u in sku_map.get("unpriced", []) if u.get("family_census")]
singles = [u for u in sku_map.get("unpriced", []) if not u.get("family_census")]
for f in fams:
    ws.append([f.get("family"), ", ".join(f.get("platforms", [])), "", "", "",
               f.get("suggestion", "")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10)
for u in singles:
    plat = u.get("platform") or ""
    t, extra = exact_title(plat, u)
    internal = u.get("internal_name", "")
    note6 = extra if internal and internal in extra else \
        "; ".join(x for x in (internal, extra) if x)
    ws.append(["    " + (t or "(title not captured)"), plat, u.get("id") or slug_from(u),
               u.get("mrp"), "Y" if u.get("combo") else "", team_lang(note6)])
    for c in ws[ws.max_row]: c.font = SM
    if u.get("url"):
        cell = ws.cell(row=ws.max_row, column=1); cell.hyperlink = u["url"]
        cell.font = Font(color="1155CC", underline="single", size=9)
style_hdr(ws); ws.freeze_panes = "A2"; autosize(ws)

out = HERE / f"Jivo-SKU-Master-Map-{DATE}.xlsx"
wb.save(out)


# ---------------- post-write hyperlink invariant (owner order #1) ----------------
def assert_hyperlinks(path):
    """Re-open the saved workbook: EVERY non-empty platform price cell on Master Map
    must carry a hyperlink (or wear the visible no-link marker). Fail loudly."""
    chk = openpyxl.load_workbook(path)
    ws = chk["Master Map"]
    filled = linked = marked = 0
    bad = []
    for r in range(5, ws.max_row + 1):
        for c in range(6, 6 + len(PLATFORMS)):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            filled += 1
            if cell.hyperlink is not None:
                linked += 1
            elif isinstance(cell.value, str) and NO_LINK in cell.value:
                marked += 1  # visible, listed on the bus — never a silent plain number
            else:
                bad.append((ws.cell(row=r, column=1).value,
                            ws.cell(row=4, column=c).value, cell.value))
    print(f"hyperlink invariant: {filled} non-empty price cells, {linked} hyperlinked, "
          f"{marked} marked '{NO_LINK}'")
    if bad:
        for b in bad:
            print(f"  UNLINKED+UNMARKED: sku={b[0]!r} platform={b[1]!r} value={b[2]!r}",
                  file=sys.stderr)
        sys.exit(f"FAIL: {len(bad)} price cell(s) violate the hyperlink invariant")
    return filled, linked, marked


filled, linked, marked = assert_hyperlinks(out)


def assert_title_hygiene(path):
    """Punkirat checks: no slug-shaped titles, no (blessed|unpriced|owner sheet)
    annotations in ANY title cell, no 'blessed' jargon anywhere in the workbook."""
    chk = openpyxl.load_workbook(path)
    title_cols = {"Listing Details": 5, "Review (confirm these)": 4,
                  "Unpriced (no master row)": 1}
    bad = []
    for sheet, col in title_cols.items():
        for r in range(2, chk[sheet].max_row + 1):
            v = chk[sheet].cell(row=r, column=col).value
            if not isinstance(v, str):
                continue
            t = v.strip()
            if SLUG_RE.fullmatch(t):
                bad.append((sheet, r, "slug-shaped title", t))
            if ANNOT_RE.search(t):
                bad.append((sheet, r, "annotation in title", t))
    for sheet in chk.sheetnames:
        for row in chk[sheet].iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and re.search(r"\bblessed\b", v, re.IGNORECASE):
                    bad.append((sheet, "-", "'blessed' jargon", v[:80]))
    if bad:
        for b in bad:
            print(f"  TITLE-LEAK: sheet={b[0]} row={b[1]} {b[2]}: {b[3]!r}", file=sys.stderr)
        sys.exit(f"FAIL: {len(bad)} title-hygiene violation(s)")
    print("title hygiene: 0 slug titles, 0 annotations in title cells, 0 'blessed' jargon")


assert_title_hygiene(out)

print(f"xlsx: {out}")
print(f"skus: {len(products)} | title sources: {dict(title_src)}")
if nolink:
    print(f"NO-LINK entries ({len(nolink)} — post these on the bus): {nolink}", file=sys.stderr)
else:
    print("no-link entries: 0 (expected)")
