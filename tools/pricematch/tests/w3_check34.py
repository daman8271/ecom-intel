#!/usr/bin/env python3
"""W3 adversarial gate — Check 3 (rebuilt xlsx: new sheets + existing byte-identical)
and Check 4 (build_pricematch new-sheet fail-safe). Run from tools/pricematch/.

Builds today's workbook with the WORKING-TREE build_pricematch.py (W2) and compares
the EXISTING sheets against a workbook built from the git-HEAD build_pricematch.py
(pre-W2 baseline) — both using the same working-tree core, so only W2's additive
change is isolated. Existing sheets must be cell-for-cell identical.
"""
import hashlib, json, os, subprocess, sys, tempfile, shutil
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
PM = os.path.abspath(os.path.join(HERE, ".."))
ROOT = os.path.abspath(os.path.join(PM, "..", ".."))
DATE = "2026-06-09"
NEW_SHEETS = ("Amazon Now PM Check", "Amazon Core PM Check")

results = []
def check(name, ok, detail=""):
    results.append((name, ok, detail))
    print(("PASS " if ok else "FAIL ") + name + (("  -- " + detail) if detail and not ok else ""))

def build(build_script_path, out_dir):
    """Run a given build_pricematch.py against the live repo for DATE."""
    env = dict(os.environ)
    p = subprocess.run([sys.executable, build_script_path, "--date", DATE, "--out", out_dir],
                       cwd=PM, capture_output=True, text=True, timeout=600, env=env)
    return p

def sheet_fingerprint(ws):
    """Stable hash of a worksheet's cell values + fill colors + dims."""
    h = hashlib.sha256()
    h.update(f"DIMS:{ws.max_row}x{ws.max_column}\n".encode())
    for row in ws.iter_rows():
        for c in row:
            if c.value is None and (c.fill is None or c.fill.fgColor is None):
                continue
            fg = ""
            try:
                if c.fill and c.fill.patternType:
                    fg = str(c.fill.fgColor.rgb)
            except Exception:
                fg = "?"
            h.update(f"{c.coordinate}|{c.value!r}|{fg}\n".encode())
    # merged cells + column widths affect layout identity
    h.update(("MERGES:" + ",".join(sorted(str(m) for m in ws.merged_cells.ranges))).encode())
    return h.hexdigest()

def main():
    tmp = tempfile.mkdtemp(prefix="w3-check34-")
    try:
        after_dir = os.path.join(tmp, "after"); os.makedirs(after_dir)
        before_dir = os.path.join(tmp, "before"); os.makedirs(before_dir)

        # ---- AFTER: working-tree (W2) build script ----
        pa = build(os.path.join(PM, "build_pricematch.py"), after_dir)
        after_xlsx = os.path.join(after_dir, f"Jivo-Price-Match-{DATE}.xlsx")
        check("workbook builds with W2 build script (rc=0)", pa.returncode == 0 and os.path.exists(after_xlsx),
              f"rc={pa.returncode} stderr_tail={pa.stderr[-800:]}")
        if not os.path.exists(after_xlsx):
            return finish()
        # surface the fail-safe stderr line for the record
        for ln in pa.stderr.splitlines():
            if "competitor PM sheets" in ln:
                print("    [build stderr] " + ln)

        wb = load_workbook(after_xlsx)
        names = wb.sheetnames

        # ---- C3: both new sheets present, appended AFTER existing ----
        check("Sheet 'Amazon Now PM Check' present", "Amazon Now PM Check" in names, str(names))
        check("Sheet 'Amazon Core PM Check' present", "Amazon Core PM Check" in names, str(names))
        existing = [n for n in names if n not in NEW_SHEETS]
        if all(n in names for n in NEW_SHEETS):
            idx_new = min(names.index(n) for n in NEW_SHEETS)
            check("new sheets appended AFTER existing", idx_new == len(existing),
                  f"new sheets start at {idx_new}, existing count {len(existing)}")

        # ---- C3: legend states the RED = cheaper-than-Amazon rule (loud) ----
        for sh, refname in [("Amazon Now PM Check", "Amazon Now"), ("Amazon Core PM Check", "Amazon Core")]:
            if sh not in names: continue
            ws = wb[sh]
            txt = " \n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
            low = txt.lower()
            has_red = ("red" in low and "cheaper than" in low and refname.lower() in low and "undercut" in low)
            check(f"[{sh}] legend states RED=cheaper-than-{refname} (undercut)", has_red,
                  "legend text missing the loud RED rule")
            #  present but blank
            has_ = "" in low
            check(f"[{sh}]  column present (blank placeholder)", has_, "no  header found")
            # 560005 pending
            has_pending = ("560005" in txt) and ("pending" in low)
            check(f"[{sh}] 560005 shown as pending", has_pending,
                  "expected 560005 block + 'pending' marker")

        # ---- C3: BB tagged national on sheet 1 ----
        if "Amazon Now PM Check" in names:
            ws = wb["Amazon Now PM Check"]
            txt = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value).lower()
            check("[Sheet1] BigBasket tagged national", ("bigbasket" in txt or "big basket" in txt) and "national" in txt,
                  "expected BigBasket + 'national' tag")

        # ---- C3: buybox seller shown on sheet 2 ----
        if "Amazon Core PM Check" in names:
            ws = wb["Amazon Core PM Check"]
            txt = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value).lower()
            check("[Sheet2] buybox seller surfaced", "seller" in txt or "buybox" in txt or "jivo mart" in txt or "rk world" in txt,
                  "expected a buybox-seller column/note")

        # ---- BEFORE: git-HEAD build script (pre-W2) for byte-identical existing sheets ----
        head_script = os.path.join(tmp, "build_HEAD.py")
        gp = subprocess.run(["git", "show", "HEAD:tools/pricematch/build_pricematch.py"],
                            cwd=ROOT, capture_output=True, text=True)
        before_ok = gp.returncode == 0
        if before_ok:
            with open(head_script, "w") as f: f.write(gp.stdout)
            # HEAD script must import its sibling modules from PM dir; run it from PM via a shim copy
            shim = os.path.join(PM, ".build_HEAD_shim.py")
            shutil.copy(head_script, shim)
            try:
                pb = build(shim, before_dir)
            finally:
                os.remove(shim)
            before_xlsx = os.path.join(before_dir, f"Jivo-Price-Match-{DATE}.xlsx")
            if pb.returncode == 0 and os.path.exists(before_xlsx):
                wbb = load_workbook(before_xlsx)
                # the named existing sheets the task calls out, plus all others
                target = [n for n in ("Ecom Head", "Matrix", "Violations") if n in wbb.sheetnames]
                allexist = [n for n in wbb.sheetnames if n not in NEW_SHEETS]
                mismatches = []
                for n in allexist:
                    if n not in wb.sheetnames:
                        mismatches.append(n + " (missing in W2)"); continue
                    fa = sheet_fingerprint(wb[n]); fb = sheet_fingerprint(wbb[n])
                    if fa != fb:
                        mismatches.append(n)
                check("EXISTING sheets byte-identical (Ecom Head/Matrix/Violations + all)",
                      not mismatches, "differing sheets: %s" % mismatches)
                # explicit per-sheet line for the named three
                for n in target:
                    same = sheet_fingerprint(wb[n]) == sheet_fingerprint(wbb[n])
                    check(f"  existing sheet '{n}' identical", same)
            else:
                check("HEAD baseline workbook builds", False,
                      f"rc={pb.returncode} stderr={pb.stderr[-600:]}")
        else:
            check("git HEAD build_pricematch.py retrievable", False, gp.stderr[-300:])

        # ================= CHECK 4: fail-safe =================
        # Inject a fault into build_compete_sheets via a monkeypatch wrapper module and
        # confirm: (a) the workbook STILL builds & saves, (b) the 2 new sheets are absent,
        # (c) all existing sheets remain present.
        fault_dir = os.path.join(tmp, "fault"); os.makedirs(fault_dir)
        driver = os.path.join(tmp, "fault_driver.py")
        with open(driver, "w") as f:
            f.write(
                "import sys, runpy\n"
                f"sys.path.insert(0, {PM!r})\n"
                "import build_pricematch as B\n"
                "def boom(*a, **k):\n"
                "    raise RuntimeError('INJECTED competitor-sheet fault')\n"
                "B.build_compete_sheets = boom\n"
                f"sys.argv = ['build_pricematch.py','--date','{DATE}','--out',{fault_dir!r}]\n"
                "B.main()\n"
            )
        pf = subprocess.run([sys.executable, driver], cwd=PM, capture_output=True, text=True, timeout=600)
        fault_xlsx = os.path.join(fault_dir, f"Jivo-Price-Match-{DATE}.xlsx")
        check("C4 fail-safe: workbook still builds when compete sheets throw",
              pf.returncode == 0 and os.path.exists(fault_xlsx),
              f"rc={pf.returncode} stderr={pf.stderr[-600:]}")
        if os.path.exists(fault_xlsx):
            wbf = load_workbook(fault_xlsx)
            check("C4 fail-safe: faulted run OMITS the 2 new sheets",
                  not any(n in wbf.sheetnames for n in NEW_SHEETS), str(wbf.sheetnames))
            check("C4 fail-safe: all existing sheets still present",
                  all(n in wbf.sheetnames for n in ("Ecom Head", "Matrix", "Violations")),
                  str(wbf.sheetnames))
            check("C4 fail-safe: skip is logged (non-fatal)",
                  "SKIPPED" in pf.stderr or "non-fatal" in pf.stderr, pf.stderr[-300:])

        return finish()
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

def finish():
    nfail = sum(1 for _, ok, _ in results if not ok)
    print(f"\n== w3_check34: {len(results)} checks, {nfail} FAIL ==")
    sys.exit(1 if nfail else 0)

if __name__ == "__main__":
    main()
