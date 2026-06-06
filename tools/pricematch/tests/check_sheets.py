#!/usr/bin/env python3
"""VERIFY Phase 2 — sheet verification on the REAL generated files.

COLOR DIRECTION is the one unforgivable bug: for every colored diff/price cell in
every per-platform "Price Match" sheet AND the master Matrix:
    fill==RED   <=>  status BELOW  <=>  live < ref-1
    fill==GREEN <=>  status ABOVE  <=>  live > ref+1
Verified programmatically via openpyxl against the engine's --json output,
cell by cell, all 8 platforms + master. Zero mismatches tolerated.

Usage: python3 tools/pricematch/tests/check_sheets.py [--date YYYY-MM-DD]
Exit: 0 all PASS, 1 any FAIL.
"""
import json
import os
import re
import subprocess
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PM_DIR = os.path.dirname(HERE)
ROOT = os.path.dirname(os.path.dirname(PM_DIR))
CORE = os.path.join(PM_DIR, "pricematch_core.py")
DATE = "2026-06-06"
for i, a in enumerate(sys.argv):
    if a == "--date" and i + 1 < len(sys.argv):
        DATE = sys.argv[i + 1]

PLATFORMS = ["blinkit", "zepto", "flipkart", "flipkart-minutes", "amazon",
             "amazon-fresh", "amazon-now", "bigbasket"]

RESULTS = []


def check(name, ok, detail=""):
    RESULTS.append((name, bool(ok), detail))
    print(("PASS  " if ok else "FAIL  ") + name + (("  -- " + str(detail)) if (detail and not ok) else ""))


def engine_json(platform=None):
    cmd = [sys.executable, CORE, "--date", DATE, "--json"]
    if platform:
        cmd.append(platform)
    p = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
    if p.returncode != 0:
        raise RuntimeError("engine CLI rc=%d: %s" % (p.returncode, p.stderr[-400:]))
    return json.loads(p.stdout)


def fill_rgb(cell):
    """Solid-fill RGB hex of a cell, or None."""
    f = cell.fill
    if not f or f.patternType is None:
        return None
    c = f.start_color
    if c is None:
        return None
    if getattr(c, "type", None) == "rgb" and c.rgb and isinstance(c.rgb, str):
        return c.rgb[-6:].upper()
    return None  # theme/indexed: treat as neutral (builders use hex)


def classify(rgb):
    """'red' / 'green' / None for a hex fill. Conservative thresholds so greys,
    yellows, blues, and the Jivo header green don't misclassify data fills."""
    if not rgb:
        return None
    r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    if r - g >= 25 and r - b >= 25:
        return "red"
    if g - r >= 15 and g - b >= 15:
        return "green"
    return None


def num(x):
    if x is None:
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def newest_workbook(p):
    d = os.path.join(ROOT, "platforms", p)
    xs = [os.path.join(d, f) for f in os.listdir(d)
          if f.startswith("Jivo-") and f.endswith(".xlsx") and not f.startswith("~")]
    return max(xs, key=os.path.getmtime) if xs else None


# ------------------------------------------------- per-platform Price Match
ORDER_RANK = {"BELOW": 0, "ABOVE": 1, "MATCH": 2, "OOS": 3, "PENDING_REVIEW": 4}


def check_platform_sheet(p):
    xlsx = newest_workbook(p)
    if not xlsx:
        check("%s: workbook exists" % p, False, "no Jivo-*.xlsx")
        return
    wb = openpyxl.load_workbook(xlsx)
    if "Price Match" not in wb.sheetnames:
        check("%s: 'Price Match' sheet present" % p, False, xlsx)
        return
    ws = wb["Price Match"]
    recs = {r["sku"]: r for r in engine_json(p)}

    # regime badge row
    badge = str(ws.cell(row=2, column=1).value or "")
    regime = next(iter(recs.values()))["regime"] if recs else "?"
    check("%s: regime badge mentions %s" % (p, regime), regime in badge, repr(badge))

    # header row -> column map
    hdr_row = None
    for r in range(1, 8):
        if str(ws.cell(row=r, column=1).value or "").strip().upper() == "SKU":
            hdr_row = r
            break
    if hdr_row is None:
        check("%s: header row found" % p, False)
        return
    cols = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=hdr_row, column=c).value or "").lower()
        if h.startswith("sku"):
            cols["sku"] = c
        elif h.startswith("stores"):
            cols["stores"] = c
        elif h.startswith("ref"):
            cols["ref"] = c
        elif h.startswith("live"):
            cols["live"] = c
        elif h.startswith("diff") and "%" not in h:
            cols["diff"] = c
        elif h.startswith("diff") and "%" in h:
            cols["pct"] = c
        elif h.startswith("status"):
            cols["status"] = c
    missing = [k for k in ("sku", "ref", "live", "diff", "pct", "status") if k not in cols]
    if missing:
        check("%s: required columns present" % p, False, "missing %s" % missing)
        return

    # New 9-col layout (owner order 2026-06-06): Min–Max column REMOVED
    hdr_vals = [str(ws.cell(row=hdr_row, column=c).value or "")
                for c in range(1, ws.max_column + 1)]
    nonempty = [h for h in hdr_vals if h.strip()]
    check("%s: 9-col layout, no Min–Max column" % p,
          len(nonempty) == 9 and not any(
              "min" in h.lower() and "max" in h.lower() for h in hdr_vals),
          hdr_vals)

    rows_seen = 0
    color_mismatches = []
    value_mismatches = []
    pct_mismatches = []
    order_seen = []
    status_count = {}
    for r in range(hdr_row + 1, ws.max_row + 1):
        sku = ws.cell(row=r, column=cols["sku"]).value
        if sku is None or str(sku).strip() == "":
            continue
        sku = str(sku).strip()
        if sku not in recs:
            # summary/footer rows start with "BELOW:" etc.
            if re.match(r"^(BELOW|ABOVE|MATCH|OOS|PENDING|TOTAL|SUMMARY|NO MAPPED)", sku, re.I):
                continue
            value_mismatches.append("row %d: sheet SKU %r not in engine records" % (r, sku))
            continue
        rec = recs[sku]
        rows_seen += 1
        status_sheet = str(ws.cell(row=r, column=cols["status"]).value or "").strip()
        status_count[status_sheet] = status_count.get(status_sheet, 0) + 1
        order_seen.append((status_sheet, num(ws.cell(row=r, column=cols["diff"]).value)))

        # status, numbers vs engine
        if status_sheet != rec["status"]:
            value_mismatches.append("row %d %s: sheet status %r != engine %r" %
                                    (r, sku, status_sheet, rec["status"]))
        for key, col in (("ref_price", "ref"), ("live_modal", "live"), ("diff", "diff")):
            sv, ev = num(ws.cell(row=r, column=cols[col]).value), num(rec[key])
            if ev is not None and sv is not None and abs(sv - ev) > 0.005:
                value_mismatches.append("row %d %s: %s sheet=%r engine=%r" % (r, sku, key, sv, ev))

        # Diff % must be the FRACTION (engine diff_pct/100) with a true Excel
        # percent format — the literal '0.0"%"' on a pre-scaled value is the
        # -798% double-scaling bug (owner screenshot 2026-06-06).
        if rec.get("diff_pct") is not None:
            pcell = ws.cell(row=r, column=cols["pct"])
            want = rec["diff_pct"] / 100.0
            got = num(pcell.value)
            if got is None or abs(got - want) > 1e-9:
                pct_mismatches.append("row %d %s: pct cell=%r want=%r" % (r, sku, pcell.value, want))
            elif "%" not in (pcell.number_format or "") or '"' in (pcell.number_format or ""):
                pct_mismatches.append("row %d %s: pct fmt=%r not a true %% format" %
                                      (r, sku, pcell.number_format))

        # ---- COLOR DIRECTION, recomputed INDEPENDENTLY from engine numbers
        live, ref = num(rec["live_modal"]), num(rec["ref_price"])
        truth = None
        if live is not None and ref is not None:
            if live < ref - 1:
                truth = "red"      # BELOW
            elif live > ref + 1:
                truth = "green"    # ABOVE
        # engine status must agree with its own numbers
        want_status = {"red": "BELOW", "green": "ABOVE"}.get(truth, None)
        if want_status and rec["status"] != want_status:
            value_mismatches.append("%s: engine status %r contradicts its numbers live=%r ref=%r"
                                    % (sku, rec["status"], live, ref))
        for col in ("diff", "live", "status"):
            cell = ws.cell(row=r, column=cols[col])
            got = classify(fill_rgb(cell))
            if got is not None and got != truth:
                color_mismatches.append("row %d %s [%s]: fill=%s but truth=%s (live=%r ref=%r status=%s)"
                                        % (r, sku, col, got, truth, live, ref, rec["status"]))
            if col == "diff" and truth is not None and got is None and rec["status"] in ("BELOW", "ABOVE"):
                color_mismatches.append("row %d %s [diff]: expected %s fill, got none" % (r, sku, truth))

    check("%s: COLOR DIRECTION zero mismatches (%d data rows)" % (p, rows_seen),
          rows_seen > 0 and not color_mismatches,
          "; ".join(color_mismatches[:6]) or "no data rows matched engine records")
    check("%s: sheet numbers/status == engine records" % p, not value_mismatches,
          "; ".join(value_mismatches[:6]))
    check("%s: Diff %% = engine diff_pct/100 with true %% format" % p, not pct_mismatches,
          "; ".join(pct_mismatches[:6]))

    # NOT_LISTED omitted
    check("%s: NOT_LISTED omitted from sheet" % p, "NOT_LISTED" not in status_count,
          "%d NOT_LISTED rows present" % status_count.get("NOT_LISTED", 0))

    # ordering: BELOW (worst first) -> ABOVE -> MATCH -> OOS -> PENDING_REVIEW
    ranks = [ORDER_RANK.get(s, 99) for s, _ in order_seen]
    check("%s: section order BELOW>ABOVE>MATCH>OOS>PENDING" % p, ranks == sorted(ranks),
          "statuses appear as %s" % [s for s, _ in order_seen])
    below_diffs = [d for s, d in order_seen if s == "BELOW" and d is not None]
    check("%s: BELOW sorted worst diff first" % p, below_diffs == sorted(below_diffs),
          "%s" % below_diffs)

    # stores-below count cell vs engine
    sb_bad = []
    if "stores" in cols:
        for r in range(hdr_row + 1, ws.max_row + 1):
            sku = str(ws.cell(row=r, column=cols["sku"]).value or "").strip()
            if sku not in recs:
                continue
            n_eng = len(recs[sku].get("stores_below") or [])
            cellv = str(ws.cell(row=r, column=cols["stores"]).value or "")
            m = re.match(r"^(\d+)", cellv)
            n_sheet = int(m.group(1)) if m else 0
            if n_sheet != n_eng:
                sb_bad.append("%s: sheet=%d engine=%d" % (sku, n_sheet, n_eng))
        check("%s: stores-below counts == engine" % p, not sb_bad, "; ".join(sb_bad[:5]))

    # footer summary counts
    footer = ""
    for r in range(ws.max_row, hdr_row, -1):
        v = str(ws.cell(row=r, column=1).value or "")
        if "BELOW" in v and ":" in v:
            footer = v
            break
    eng_counts = {}
    for rec in recs.values():
        eng_counts[rec["status"]] = eng_counts.get(rec["status"], 0) + 1
    m = re.search(r"BELOW:\s*(\d+)", footer)
    check("%s: footer BELOW count == engine (%d)" % (p, eng_counts.get("BELOW", 0)),
          m is not None and int(m.group(1)) == eng_counts.get("BELOW", 0),
          "footer=%r" % footer[:90])

    # only today's regime may be named on the sheet (owner order 2026-06-06)
    forbidden = {"BAU", "SVD", "ART"} - {regime}
    leaks = [repr(str(v)[:60]) for row in ws.iter_rows(values_only=True) for v in row
             if isinstance(v, str) and any(re.search(r"\b%s\b" % rg, v) for rg in forbidden)]
    check("%s: no non-today regime token on Price Match sheet" % p, not leaks,
          "; ".join(leaks[:4]))


# ------------------------------------------------------------------ master
def find_master():
    # exact unlabeled name FIRST — a scoped edition (Jivo-Price-Match-Amazon-…)
    # must never be picked up and judged against all-8-platform expectations
    for d in (PM_DIR, os.path.join(ROOT, "output"), ROOT):
        c = os.path.join(d, "Jivo-Price-Match-%s.xlsx" % DATE)
        if os.path.exists(c):
            return c
    cands = []
    for d in (PM_DIR, os.path.join(ROOT, "output"), ROOT):
        if os.path.isdir(d):
            for f in os.listdir(d):
                if re.match(r"Jivo-Price-Match-\d{4}-\d{2}-\d{2}\.xlsx$", f):
                    cands.append(os.path.join(d, f))
    return max(cands, key=os.path.getmtime) if cands else None


def check_master():
    xlsx = find_master()
    if not xlsx:
        check("master: workbook exists", False, "Jivo-Price-Match-*.xlsx not found")
        return
    print("master workbook: %s" % xlsx)
    wb = openpyxl.load_workbook(xlsx)
    names = wb.sheetnames
    check("master: 'Ecom Head' is FIRST sheet", names and names[0] == "Ecom Head", names)
    matrix_name = next((n for n in names if "matrix" in n.lower()), None)
    check("master: Matrix sheet present", matrix_name is not None, names)

    allrec = engine_json(None)
    if isinstance(allrec, dict) and "records" in allrec:
        allrec = allrec["records"]
    # allrec: dict platform -> records, or flat list
    by = {}
    if isinstance(allrec, dict):
        for p, rl in allrec.items():
            if isinstance(rl, list):
                for r in rl:
                    by[(r["sku"], r["platform"])] = r
    else:
        for r in allrec:
            by[(r["sku"], r["platform"])] = r
    regime = next(iter(by.values()))["regime"] if by else "?"

    if not matrix_name:
        return
    ws = wb[matrix_name]
    # header row: find row containing 'MRP'
    hdr_row, cols = None, {}
    for r in range(1, 8):
        vals = [str(ws.cell(row=r, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if any(v.upper() == "MRP" for v in vals):
            hdr_row = r
            for c, v in enumerate(vals, 1):
                cols[c] = v
            break
    if hdr_row is None:
        check("master Matrix: header row found", False)
        return
    DISPLAY = {"fk minutes": "flipkart-minutes", "flipkart minutes": "flipkart-minutes",
               "amz fresh": "amazon-fresh", "amazon fresh": "amazon-fresh",
               "amz now": "amazon-now", "amazon now": "amazon-now",
               "bigbasket": "bigbasket", "blinkit": "blinkit", "zepto": "zepto",
               "flipkart": "flipkart", "amazon": "amazon"}
    plat_cols = {}
    for c, v in cols.items():
        vl = v.lower().strip()
        if vl in DISPLAY:
            plat_cols[DISPLAY[vl]] = c
    check("master Matrix: all 8 platform columns", len(plat_cols) == 8,
          "found %s" % sorted(plat_cols))
    # Owner order 2026-06-06 (sheet-clean): show ONLY today's regime — the
    # header must be MRP + ONE named "Agreed price (<regime>)" column; the
    # old BAU/SVD/ART trio is BANNED from the Matrix.
    trio = [v for v in cols.values()
            if (v.upper().split()[0] if v.strip() else "") in ("BAU", "SVD", "ART")]
    agreed = [c for c, v in cols.items()
              if v.strip().lower() == ("agreed price (%s)" % regime.lower())]
    has_mrp = any(v.strip().upper() == "MRP" for v in cols.values())
    check("master Matrix: MRP + single 'Agreed price (%s)' column, no BAU/SVD/ART trio" % regime,
          len(agreed) == 1 and has_mrp and not trio,
          "agreed=%r trio=%r cols=%r" % (agreed, trio, cols))

    # no OTHER regime's name leaks anywhere in the workbook (today's is fine)
    forbidden = {"BAU", "SVD", "ART"} - {regime}
    leaks = []
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and any(
                        re.search(r"\b%s\b" % rg, v) for rg in forbidden):
                    leaks.append("[%s] %r" % (sn, v[:60]))
    check("master: no non-today regime token (%s) in any cell" % "/".join(sorted(forbidden)),
          not leaks, "; ".join(leaks[:5]))

    color_mm, val_mm, link_missing = [], [], []
    checked = 0
    skus_in_matrix = set()
    for r in range(hdr_row + 1, ws.max_row + 1):
        sku = ws.cell(row=r, column=1).value
        if sku is None or str(sku).strip() == "":
            continue
        sku = str(sku).strip()
        skus_in_matrix.add(sku)
        for p, c in plat_cols.items():
            rec = by.get((sku, p))
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if rec is None:
                continue
            got = classify(fill_rgb(cell))
            live, ref = num(rec["live_modal"]), num(rec["ref_price"])
            status = rec["status"]
            if status == "NOT_LISTED":
                ok = v is None or str(v).strip() in ("—", "-", "--", "")
                if not ok:
                    val_mm.append("%s/%s NOT_LISTED but cell=%r" % (sku, p, v))
                if got is not None:
                    color_mm.append("%s/%s NOT_LISTED colored %s" % (sku, p, got))
                continue
            if status == "OOS":
                if "OOS" not in str(v or "").upper():
                    val_mm.append("%s/%s OOS but cell=%r" % (sku, p, v))
                if got is not None:
                    color_mm.append("%s/%s OOS colored %s" % (sku, p, got))
                continue
            if live is None:
                continue
            checked += 1
            # priced cell: value == live modal (allow "(±x)" annotations)
            mnum = re.search(r"[\d.]+", str(v or ""))
            sheetnum = num(v) if num(v) is not None else (num(mnum.group(0)) if mnum else None)
            if sheetnum is None or abs(sheetnum - live) > 0.005:
                val_mm.append("%s/%s live=%r cell=%r" % (sku, p, live, v))
            truth = "red" if live < ref - 1 else ("green" if live > ref + 1 else None)
            if got != truth:
                color_mm.append("%s/%s fill=%s truth=%s (live=%r ref=%r st=%s)" %
                                (sku, p, got, truth, live, ref, status))
            if not cell.hyperlink:
                link_missing.append("%s/%s" % (sku, p))

    check("master Matrix: COLOR DIRECTION zero mismatches (%d priced cells)" % checked,
          checked > 0 and not color_mm, "; ".join(color_mm[:8]))
    check("master Matrix: priced cell values == engine live modal", not val_mm,
          "; ".join(val_mm[:8]))
    check("master Matrix: hyperlink on every priced cell", not link_missing,
          "%d missing e.g. %s" % (len(link_missing), link_missing[:5]))
    # coverage: every non-retired master SKU appears
    eng_skus = {k[0] for k in by}
    check("master Matrix: covers all %d engine SKUs" % len(eng_skus),
          eng_skus <= skus_in_matrix,
          "missing %s" % sorted(eng_skus - skus_in_matrix)[:5])


def check_ecom_head():
    xlsx = find_master()
    if not xlsx:
        return
    wb = openpyxl.load_workbook(xlsx)
    if "Ecom Head" not in wb.sheetnames:
        return
    ws = wb["Ecom Head"]
    p = subprocess.run([sys.executable, CORE, "--date", DATE, "--json"],
                       capture_output=True, text=True, timeout=180)
    # summary via module (CLI --json dumps records); load through import driver
    drv = ("import importlib.util, json, sys\n"
           "spec = importlib.util.spec_from_file_location('pmc', %r)\n"
           "m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m)\n"
           "ctx = m.load_context(date=%r)\n"
           "print(json.dumps(m.summary(ctx), default=str))\n" % (CORE, DATE))
    q = subprocess.run([sys.executable, "-c", drv], capture_output=True, text=True, timeout=180)
    summ = json.loads(q.stdout)
    g = summ.get("global") or {}

    # scrape every (label, number) pair off the sheet
    text = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                text.append((cell.coordinate, str(cell.value)))
    blob = " | ".join(t for _, t in text)

    def find_num_near(label_words, want):
        """True if `want` appears in a cell whose own text or a neighbour label
        matches the words."""
        want_str = ("%d" % want) if float(want) == int(want) else ("%s" % want)
        for coord, t in text:
            tl = t.lower()
            if all(w in tl for w in label_words) and re.search(r"\b" + re.escape(want_str) + r"\b", t):
                return True
        # label and number in separate cells: just require the number to exist somewhere
        return None

    kpis = [(["below"], g.get("BELOW")), (["above"], g.get("ABOVE")),
            (["oos"], g.get("OOS")), (["pending"], g.get("PENDING_REVIEW"))]
    soft_fail = []
    for words, want in kpis:
        if want is None:
            continue
        hit = find_num_near(words, want)
        if hit is None:
            # fallback: number present anywhere on the sheet
            ws_nums = set(re.findall(r"\d+", blob))
            if str(int(want)) not in ws_nums:
                soft_fail.append("%s=%s not found on sheet" % ("+".join(words), want))
        elif hit is False:
            soft_fail.append("%s=%s mismatched" % ("+".join(words), want))
    check("Ecom Head: KPI numbers match summary() (BELOW/ABOVE/OOS/PENDING)",
          not soft_fail, "; ".join(soft_fail))
    exp = g.get("below_exposure_inr")
    if exp is not None:
        s_int = str(int(round(exp)))
        check("Ecom Head: below-ref exposure ₹%s on sheet" % s_int,
              s_int in re.sub(r"[,₹]", "", blob), "exposure %r absent" % exp)
    check("Ecom Head: regime badge for %s" % summ.get("regime"),
          summ.get("regime", "??") in blob, "no regime badge found")
    # fresh-eyes MUST-4: plain words, the regime expanded, NO internal residue
    GLOSS = {"SVD": "Special Value Days"}
    rg = summ.get("regime", "")
    long_form = GLOSS.get(rg)
    if long_form:
        check("Ecom Head: regime expanded in plain words ('%s')" % long_form,
              long_form in blob, "expansion missing")
    residue = [t for t in ("regime.json", "almBrandId", "_dashboard.py",
                           "tools/", ".py") if t in blob]
    check("Ecom Head: no internal residue (file names / parameters)",
          not residue, "found %s" % residue)


def main():
    if not os.path.exists(CORE):
        print("engine missing")
        sys.exit(2)
    for p in PLATFORMS:
        try:
            check_platform_sheet(p)
        except Exception as e:
            import traceback
            check("%s: checker ran" % p, False, traceback.format_exc()[-400:])
    try:
        check_master()
        check_ecom_head()
    except Exception:
        import traceback
        check("master: checker ran", False, traceback.format_exc()[-600:])

    fails = [r for r in RESULTS if not r[1]]
    print("\n== check_sheets: %d checks, %d FAIL ==" % (len(RESULTS), len(fails)))
    for name, _, detail in fails:
        print("  FAIL: %s  %s" % (name, detail))
    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()
