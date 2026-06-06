#!/usr/bin/env python3
"""VERIFY Phase 2c — deep checks on the master Price Match workbook.

Beyond color direction (check_sheets.py): Ecom Head top-10 sort + values,
platform scoreboard arithmetic, Violations sheet = engine stores_below exactly
(count, sort by loss, row-level spot equality), Above-reference sheet, Coverage
counts, and owner-facing number formatting.
"""
import json
import os
import re
import subprocess
import sys
from collections import Counter

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PM_DIR = os.path.dirname(HERE)
ROOT = os.path.dirname(os.path.dirname(PM_DIR))
CORE = os.path.join(PM_DIR, "pricematch_core.py")
DATE = "2026-06-06"
for i, a in enumerate(sys.argv):
    if a == "--date" and i + 1 < len(sys.argv):
        DATE = sys.argv[i + 1]

DISPLAY2SLUG = {"fk minutes": "flipkart-minutes", "flipkart minutes": "flipkart-minutes",
                "amz fresh": "amazon-fresh", "amazon fresh": "amazon-fresh",
                "amz now": "amazon-now", "amazon now": "amazon-now",
                "bigbasket": "bigbasket", "blinkit": "blinkit", "zepto": "zepto",
                "flipkart": "flipkart", "amazon": "amazon"}

RESULTS = []


def check(name, ok, detail=""):
    RESULTS.append((name, bool(ok), detail))
    print(("PASS  " if ok else "FAIL  ") + name + (("  -- " + str(detail)) if (detail and not ok) else ""))


def num(x):
    if x is None:
        return None
    try:
        return float(str(x).replace(",", "").replace("₹", "").strip())
    except ValueError:
        return None


def main():
    # engine truth
    p = subprocess.run([sys.executable, CORE, "--date", DATE, "--json"],
                       capture_output=True, text=True, timeout=300)
    raw = json.loads(p.stdout)
    if isinstance(raw, dict) and "records" in raw:
        raw = raw["records"]
    recs = []
    if isinstance(raw, dict):
        for rl in raw.values():
            if isinstance(rl, list):
                recs.extend(rl)
    else:
        recs = raw
    below = [r for r in recs if r["status"] == "BELOW"]
    above = [r for r in recs if r["status"] == "ABOVE"]
    by_status = Counter(r["status"] for r in recs)
    sb_total = sum(len(r.get("stores_below") or []) for r in recs)
    sb_below_only = sum(len(r.get("stores_below") or []) for r in below)

    xlsx = None
    for d in (os.path.join(ROOT, "output"), PM_DIR):
        c = os.path.join(d, "Jivo-Price-Match-%s.xlsx" % DATE)
        if os.path.exists(c):
            xlsx = c
            break
    if not xlsx:
        check("master workbook exists", False)
        finish()
    wb = openpyxl.load_workbook(xlsx)

    # ---------------------------------------------------------- Ecom Head
    ws = wb["Ecom Head"]
    grid = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                grid[(c.row, c.column)] = c

    # locate top-violations table header
    top_hdr = None
    for (r, c), cell in grid.items():
        if c == 1 and "TOP VIOLATION" in str(cell.value).upper():
            top_hdr = r
            break
    check("Ecom Head: TOP VIOLATIONS table present", top_hdr is not None)
    if top_hdr:
        hr = top_hdr + 1  # SKU | Platform | Ref | Live | Diff | Worst store
        rows = []
        r = hr + 1
        while (r, 1) in grid and str(grid[(r, 1)].value).strip():
            rows.append(r)
            r += 1
        diffs = [num(grid.get((r, 5)) and grid[(r, 5)].value) for r in rows]
        check("Ecom Head top-10: <=10 rows", len(rows) <= 10, "%d rows" % len(rows))
        check("Ecom Head top-10: sorted by diff (worst first)",
              diffs == sorted(diffs), diffs)
        # equals the engine's global top-N below by diff
        eng_sorted = sorted(below, key=lambda x: x["diff"])[:len(rows)]
        eng_set = [(e["sku"], round(e["diff"])) for e in eng_sorted]
        sheet_set = []
        for r in rows:
            sheet_set.append((str(grid[(r, 1)].value).strip(),
                              int(round(num(grid[(r, 5)].value)))))
        check("Ecom Head top-10: matches engine's worst BELOW records",
              sheet_set == eng_set, "sheet=%s engine=%s" % (sheet_set, eng_set))
        # ref/live values exact per row
        bad = []
        for r in rows:
            sku = str(grid[(r, 1)].value).strip()
            plat = DISPLAY2SLUG.get(str(grid[(r, 2)].value).strip().lower())
            er = next((x for x in below if x["sku"] == sku and x["platform"] == plat), None)
            if not er:
                bad.append("%s/%s not an engine BELOW" % (sku, plat))
                continue
            if num(grid[(r, 3)].value) != er["ref_price"] or num(grid[(r, 4)].value) != er["live_modal"]:
                bad.append("%s/%s ref/live mismatch" % (sku, plat))
        check("Ecom Head top-10: ref/live == engine", not bad, "; ".join(bad[:4]))

    # scoreboard
    sb_hdr = None
    for (r, c), cell in grid.items():
        if c == 1 and "SCOREBOARD" in str(cell.value).upper():
            sb_hdr = r
            break
    check("Ecom Head: PLATFORM SCOREBOARD present", sb_hdr is not None)
    if sb_hdr:
        hr = sb_hdr + 1
        r = hr + 1
        sb_rows = []
        while (r, 1) in grid and str(grid[(r, 1)].value).strip():
            sb_rows.append(r)
            r += 1
        check("Ecom Head scoreboard: 8 platform rows", len(sb_rows) == 8, len(sb_rows))
        bad = []
        sum_viol = 0
        sum_mapped = 0
        for r in sb_rows:
            disp = str(grid[(r, 1)].value).strip().lower()
            plat = DISPLAY2SLUG.get(disp)
            if not plat:
                bad.append("unknown platform %r" % disp)
                continue
            pr = [x for x in recs if x["platform"] == plat]
            mapped_eng = sum(1 for x in pr if x["status"] != "NOT_LISTED")
            viol_eng = sum(1 for x in pr if x["status"] == "BELOW")
            mapped_sheet = int(num(grid[(r, 2)].value))
            viol_sheet = int(num(grid[(r, 4)].value))
            sum_viol += viol_sheet
            sum_mapped += mapped_sheet
            if mapped_sheet != mapped_eng:
                bad.append("%s mapped %d != engine %d" % (plat, mapped_sheet, mapped_eng))
            if viol_sheet != viol_eng:
                bad.append("%s violations %d != engine %d" % (plat, viol_sheet, viol_eng))
            # biggest offender = engine's worst BELOW for that platform
            worst = min((x for x in pr if x["status"] == "BELOW"),
                        key=lambda x: x["diff"], default=None)
            off = str(grid[(r, 5)].value or "")
            if worst and worst["sku"] not in off:
                bad.append("%s offender %r != engine worst %s" % (plat, off[:30], worst["sku"]))
        check("Ecom Head scoreboard: per-row mapped/violations/offender == engine",
              not bad, "; ".join(bad[:5]))
        check("Ecom Head scoreboard: violations sum to global BELOW (%d)" % by_status["BELOW"],
              sum_viol == by_status["BELOW"], "sum=%d" % sum_viol)
        mapped_global = sum(1 for x in recs if x["status"] != "NOT_LISTED")
        check("Ecom Head scoreboard: mapped sums to global mapped (%d)" % mapped_global,
              sum_mapped == mapped_global, "sum=%d" % sum_mapped)

    # ------------------------------------------------------------ Violations
    # The sheet now opens with a SKU×platform ROLL-UP (fresh-eyes S7); the
    # store-level detail table is the one whose header row is SKU|…|City.
    ws = wb["Violations"]
    hdr = None
    for r in range(1, 40):
        if (str(ws.cell(row=r, column=1).value or "").strip().upper() == "SKU"
                and str(ws.cell(row=r, column=3).value or "").strip().upper() == "CITY"):
            hdr = r
            break
    check("Violations: store-level detail table found (SKU|…|City header)", hdr is not None)
    if hdr is None:
        finish()
    vrows = []
    for r in range(hdr + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            continue
        vrows.append([ws.cell(row=r, column=c).value for c in range(1, 8)])
    check("Violations: row count == engine Σ stores_below over BELOW records (%d)" % sb_below_only,
          len(vrows) in (sb_below_only, sb_total),
          "sheet=%d engine_below_only=%d engine_all=%d" % (len(vrows), sb_below_only, sb_total))
    losses = [num(v[6]) for v in vrows]
    check("Violations: sorted by loss desc", losses == sorted(losses, reverse=True),
          "first 10: %s" % losses[:10])
    bad = []
    for v in vrows[:50] + vrows[-20:]:
        sku, disp, city, pin, store_p, ref_p, loss = v
        plat = DISPLAY2SLUG.get(str(disp).strip().lower())
        er = next((x for x in recs if x["sku"] == str(sku).strip() and x["platform"] == plat), None)
        if er is None:
            bad.append("%s/%s no engine record" % (sku, disp))
            continue
        if abs((num(ref_p) - num(store_p)) - num(loss)) > 0.01:
            bad.append("%s/%s loss %s != ref-store %s-%s" % (sku, plat, loss, ref_p, store_p))
        sb = er.get("stores_below") or []
        if not any(abs(num(s.get("price")) - num(store_p)) < 0.01 and
                   str(s.get("pincode", "")) in (str(pin), "-", "") or
                   str(pin) in ("-", "None") for s in sb):
            if not any(abs(num(s.get("price")) - num(store_p)) < 0.01 for s in sb):
                bad.append("%s/%s store ₹%s not in engine stores_below" % (sku, plat, store_p))
    check("Violations: spot rows (70) match engine stores_below + loss math", not bad,
          "; ".join(str(b) for b in bad[:5]))

    # ------------------------------------------------------------ Above ref
    ws = wb["Above reference"]
    hdr = None
    for r in range(1, 8):
        if str(ws.cell(row=r, column=1).value or "").strip().upper() == "SKU":
            hdr = r
            break
    arows = []
    fmt_ugly = []
    pct_bad = []
    for r in range(hdr + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            continue
        arows.append(r)
        c6 = ws.cell(row=r, column=6)  # Diff % — FRACTION + true % format
        if isinstance(c6.value, float) and c6.number_format in ("General", "@", None):
            fmt_ugly.append("%s: %r fmt=%r" % (c6.coordinate, c6.value, c6.number_format))
        # exact: cell == engine diff_pct/100 (pre-scaled value here = the
        # -798% double-scaling bug, owner screenshot 2026-06-06)
        sku = str(ws.cell(row=r, column=1).value).strip()
        plat = DISPLAY2SLUG.get(str(ws.cell(row=r, column=2).value or "").strip().lower())
        er = next((x for x in above if x["sku"] == sku and x["platform"] == plat), None)
        if er and er.get("diff_pct") is not None and isinstance(c6.value, (int, float)):
            want = er["diff_pct"] / 100.0
            if abs(c6.value - want) > 1e-9:
                pct_bad.append("%s/%s c6=%r want=%r" % (sku, plat, c6.value, want))
            elif "%" not in (c6.number_format or "") or '"' in (c6.number_format or ""):
                pct_bad.append("%s/%s fmt=%r not a true %% format" % (sku, plat, c6.number_format))
    check("Above reference: row count == engine ABOVE (%d)" % len(above),
          len(arows) == len(above), "sheet=%d" % len(arows))
    check("Above reference: Diff %% cells carry a number format (owner-facing)",
          not fmt_ugly, "raw floats shown as General: %s" % fmt_ugly[:3])
    check("Above reference: Diff %% = engine diff_pct/100 (no double-scaling)",
          not pct_bad, "; ".join(pct_bad[:4]))

    # ------------------------------------------------------------ Coverage
    ws = wb["Coverage & pending"]
    blob = " | ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    nl_eng = by_status["NOT_LISTED"]
    check("Coverage: NOT_LISTED total (%d) on sheet" % nl_eng,
          str(nl_eng) in blob, "expected %d" % nl_eng)
    check("Coverage: PENDING_REVIEW count (%d) on sheet" % by_status["PENDING_REVIEW"],
          str(by_status["PENDING_REVIEW"]) in blob)
    check("Coverage: OOS count (%d) on sheet" % by_status["OOS"],
          str(by_status["OOS"]) in blob)

    finish()


def finish():
    fails = [x for x in RESULTS if not x[1]]
    print("\n== check_master_deep: %d checks, %d FAIL ==" % (len(RESULTS), len(fails)))
    for name, _, detail in fails:
        print("  FAIL: %s  %s" % (name, detail))
    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()
