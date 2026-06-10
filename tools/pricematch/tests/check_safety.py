#!/usr/bin/env python3
"""VERIFY Phase 2b/3a — idempotency + originals-untouched + appender fail-safety.

All mutation tests run on TEMP COPIES of real workbooks; nothing in the repo is
modified. Bare python3.

Checks:
  1. Appender idempotency: re-running add_pricematch_sheet.py on a workbook does
     NOT duplicate the "Price Match" sheet (same sheet count, one Price Match).
  2. Existing sheets untouched: content-hash every original sheet pre/post append
     (values + fills + fonts + merged ranges) — zero drift.
  3. Master builder determinism: build twice → byte-stable output.
  4. Appender crash-safety: engine import broken → appender exits 0, workbook
     byte-identical (atomic, untouched).
  5. Unknown platform (zero mapped SKUs) → exit 0, sheet with a note, no crash.
"""
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PM_DIR = os.path.dirname(HERE)
ROOT = os.path.dirname(os.path.dirname(PM_DIR))
APPENDER = os.path.join(PM_DIR, "add_pricematch_sheet.py")
BUILDER = os.path.join(PM_DIR, "build_pricematch.py")
DATE = "2026-06-06"

RESULTS = []


def check(name, ok, detail=""):
    RESULTS.append((name, bool(ok), detail))
    print(("PASS  " if ok else "FAIL  ") + name + (("  -- " + str(detail)) if (detail and not ok) else ""))


def sheet_content_hash(ws):
    h = hashlib.sha256()
    h.update(str(sorted(str(m) for m in ws.merged_cells.ranges)).encode())
    for row in ws.iter_rows():
        for c in row:
            if c.value is None and (not c.fill or c.fill.patternType is None):
                continue
            fill = c.fill.start_color.rgb if (c.fill and c.fill.patternType) else ""
            font = "%s|%s|%s" % (c.font.bold, c.font.size,
                                 c.font.color.rgb if (c.font.color and c.font.color.type == "rgb") else "")
            h.update(("%s=%r;%s;%s\n" % (c.coordinate, c.value, fill, font)).encode())
    return h.hexdigest()


def wb_hashes(path, skip=("Price Match",)):
    wb = openpyxl.load_workbook(path)
    return {n: sheet_content_hash(wb[n]) for n in wb.sheetnames if n not in skip}


def newest_workbook(p):
    d = os.path.join(ROOT, "platforms", p)
    xs = [os.path.join(d, f) for f in os.listdir(d)
          if f.startswith("Jivo-") and f.endswith(".xlsx")]
    return max(xs, key=os.path.getmtime) if xs else None


def run_appender(platform, xlsx, env=None, date=DATE):
    e = dict(os.environ)
    if env:
        e.update(env)
    return subprocess.run([sys.executable, APPENDER, platform, xlsx, "--date", date],
                          capture_output=True, text=True, timeout=300, env=e, cwd=ROOT)


def main():
    if not os.path.exists(APPENDER):
        print("appender missing — SHEETS-A not landed")
        sys.exit(2)

    base = tempfile.mkdtemp(prefix="pmsafety-")
    try:
        src = newest_workbook("amazon")  # Amazon-family: the append path (owner 2026-06-10: tab is Amazon-only)
        tmp_x = os.path.join(base, os.path.basename(src))
        shutil.copy2(src, tmp_x)

        # --- 1+2: idempotency + originals untouched
        pre = wb_hashes(tmp_x)
        wb = openpyxl.load_workbook(tmp_x, read_only=True)
        n_sheets_pre = len(wb.sheetnames)
        had_pm = "Price Match" in wb.sheetnames
        wb.close()
        p1 = run_appender("amazon", tmp_x)
        check("appender re-run: exit 0", p1.returncode == 0,
              "rc=%d stderr=%s" % (p1.returncode, p1.stderr[-300:]))
        wb = openpyxl.load_workbook(tmp_x, read_only=True)
        names = wb.sheetnames
        wb.close()
        check("appender re-run: exactly ONE 'Price Match' sheet",
              names.count("Price Match") == 1, names)
        expect_n = n_sheets_pre if had_pm else n_sheets_pre + 1
        check("appender re-run: sheet count stable (%d)" % expect_n,
              len(names) == expect_n, "pre=%d post=%d %s" % (n_sheets_pre, len(names), names))
        post = wb_hashes(tmp_x)
        drift = [n for n in pre if pre[n] != post.get(n)]
        gone = [n for n in pre if n not in post]
        check("original sheets content-identical pre/post append",
              not drift and not gone, "drift=%s gone=%s" % (drift, gone))

        # run twice more: still one sheet (true idempotency)
        run_appender("amazon", tmp_x)
        wb = openpyxl.load_workbook(tmp_x, read_only=True)
        check("appender 3x: still one 'Price Match', count stable",
              wb.sheetnames.count("Price Match") == 1 and len(wb.sheetnames) == expect_n,
              wb.sheetnames)
        wb.close()

        # --- 4: crash-safety — break the engine import inside a sandbox copy of tools/pricematch
        sandbox_pm = os.path.join(base, "tools", "pricematch")
        os.makedirs(sandbox_pm)
        for f in os.listdir(PM_DIR):
            fp = os.path.join(PM_DIR, f)
            if os.path.isfile(fp) and (f.endswith(".py") or f.endswith(".json")):
                shutil.copy2(fp, os.path.join(sandbox_pm, f))
        with open(os.path.join(sandbox_pm, "pricematch_core.py"), "w") as f:
            f.write("raise RuntimeError('VERIFY: simulated engine crash')\n")
        crash_x = os.path.join(base, "crash-" + os.path.basename(src))
        shutil.copy2(src, crash_x)
        with open(crash_x, "rb") as f:
            pre_bytes = hashlib.sha256(f.read()).hexdigest()
        p4 = subprocess.run([sys.executable, os.path.join(sandbox_pm, "add_pricematch_sheet.py"),
                             "amazon", crash_x, "--date", DATE],
                            capture_output=True, text=True, timeout=120, cwd=base)
        with open(crash_x, "rb") as f:
            post_bytes = hashlib.sha256(f.read()).hexdigest()
        check("crashing engine: appender exits 0 (fail-safe)", p4.returncode == 0,
              "rc=%d stderr=%s" % (p4.returncode, p4.stderr[-300:]))
        check("crashing engine: workbook byte-identical", pre_bytes == post_bytes)
        check("crashing engine: warning printed", bool((p4.stdout + p4.stderr).strip()),
              "silent failure — no warning at all")

        # --- 5: non-Amazon platform => tab stripped (owner 2026-06-10), exit 0
        unk_x = os.path.join(base, "unk-" + os.path.basename(src))
        shutil.copy2(src, unk_x)
        p5 = run_appender("not-a-platform", unk_x)
        check("unknown platform: exit 0", p5.returncode == 0,
              "rc=%d stderr=%s" % (p5.returncode, p5.stderr[-300:]))
        try:
            wb = openpyxl.load_workbook(unk_x, read_only=True)
            unk_names = wb.sheetnames
            wb.close()
            ok_open = True
        except Exception as e:
            ok_open = False
            unk_names = []
        check("unknown platform: workbook still opens", ok_open)
        check("non-Amazon platform: 'Price Match' tab stripped",
              ok_open and "Price Match" not in unk_names, unk_names)

        # --- 3: master builder determinism (byte-stable)
        if os.path.exists(BUILDER):
            d1, d2 = os.path.join(base, "b1"), os.path.join(base, "b2")
            os.makedirs(d1)
            os.makedirs(d2)
            # PM_SKIP_HISTORY: a test build for the frozen past DATE must NEVER
            # touch data/pricematch/*.csv (it was overwriting the real 2026-06-06
            # history with today's prices on every run — caught 2026-06-10).
            env_nohist = dict(os.environ, PM_SKIP_HISTORY="1")
            outs = []
            for d in (d1, d2):
                p = subprocess.run([sys.executable, BUILDER, "--date", DATE, "--out", d],
                                   capture_output=True, text=True, timeout=600, cwd=ROOT,
                                   env=env_nohist)
                if p.returncode != 0:
                    # builder may not support --out; try cwd-based
                    p = subprocess.run([sys.executable, BUILDER, "--date", DATE],
                                       capture_output=True, text=True, timeout=600, cwd=d,
                                       env=env_nohist)
                xs = [os.path.join(d, f) for f in os.listdir(d) if f.endswith(".xlsx")]
                outs.append((p, max(xs, key=os.path.getmtime) if xs else None))
            (pa, fa), (pb, fb) = outs
            if fa and fb:
                ha = hashlib.sha256(open(fa, "rb").read()).hexdigest()
                hb = hashlib.sha256(open(fb, "rb").read()).hexdigest()
                check("master builder: byte-stable across two builds", ha == hb,
                      "%s != %s" % (ha[:12], hb[:12]))
            else:
                check("master builder: produced output in sandbox dir", False,
                      "rc=%s/%s — could not build into a temp dir (stderr: %s)" %
                      (pa.returncode, pb.returncode, (pa.stderr or "")[-200:]))
        else:
            check("master builder exists", False, "build_pricematch.py missing (SHEETS-B pending)")
    finally:
        shutil.rmtree(base, ignore_errors=True)

    fails = [r for r in RESULTS if not r[1]]
    print("\n== check_safety: %d checks, %d FAIL ==" % (len(RESULTS), len(fails)))
    for name, _, detail in fails:
        print("  FAIL: %s  %s" % (name, detail))
    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()
