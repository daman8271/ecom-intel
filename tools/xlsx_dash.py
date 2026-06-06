#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_dash.py — shared Excel dashboard helpers for ecom-intel (pure openpyxl).

One small, dependency-free library so every workbook (platform Leadership Views,
pricematch dailies, the Ecom Head master) shares ONE visual language:
the ink + sage palette already used by tools/predict.py + tools/pricematch/,
KPI cards, banded tables, conditional-format data bars / color scales,
Indian ₹ lakh/crore number formats, and text sparklines.

THE DURABILITY CONTRACT
-----------------------
Every element this module draws survives an openpyxl load()+save() round-trip
unchanged AND renders in cache-only preview viewers (macOS Quick Look, Google
Drive, Office mobile). Our workbooks are re-opened and appended to by later
pipeline stages, and the owner reads them in preview apps. Native openpyxl
charts carry ZERO cached values, so they render BLANK in every preview viewer
(the 2026-06-06 empty-Leadership-View root cause) — and Excel-AUTHORED
charts/images/shapes are destroyed by an openpyxl load+save. So:

  * NO native charts here — ever. Trend/proportion visuals are done with
    conditional-format data bars, 3-color scales, icon sets and unicode
    text sparklines/meters, all of which round-trip losslessly and render
    everywhere.
  * Native charts are allowed ONLY in the final stage that touches a workbook,
    accepting they will be blank in preview apps.

Self-test (builds demo.xlsx exercising every helper, round-trips it, asserts
nothing was lost):    python3 tools/xlsx_dash.py --selftest
"""

from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- palette
# The ink + sage system (tools/predict.py Leadership View + tools/pricematch).
JIVO_GREEN = "008B3A"   # THE one brand green (fresh-eyes 2026-06-06: single green)
BRAND      = JIVO_GREEN  # alias — section titles, data bars (was 1F8A4C; unified)
BRAND_SOFT = "D6F0E0"   # soft sage — subtle emphasis fills
INK        = "111827"   # near-black — primary text
MUTED      = "6B7280"   # secondary text, card labels
RULE       = "E5E7EB"   # hairline dividers, card borders
CANVAS     = "F9FAFB"   # zebra stripe / page canvas
POS        = "047857"   # success green (text)
WARN       = "B45309"   # amber (text)
NEG        = "B91C1C"   # risk red (text)
ACCENT     = "0F766E"   # neutral teal — competitor / "other" series
NEUTRAL    = "9CA3AF"

# Classic Excel compliance fills (the red/green every stakeholder already reads).
# THE one red pair and THE one green pair (fill, text) — fresh-eyes 2026-06-06
# found F4CCCC/CC0000 vs FFC7CE/B91C1C drift across files; use ONLY these.
RED_FILL_HEX,   RED_TEXT   = "FFC7CE", "9C0006"
GREEN_FILL_HEX, GREEN_TEXT = "C6EFCE", "006100"
AMBER_FILL_HEX, AMBER_TEXT = "FFEB9C", "9C6500"
BAD_PAIR   = (RED_FILL_HEX, RED_TEXT)     # below agreed price / violation
GOOD_PAIR  = (GREEN_FILL_HEX, GREEN_TEXT)  # compliant / at-or-above
AMBER_PAIR = (AMBER_FILL_HEX, AMBER_TEXT)  # overpriced — sales risk / caution

F = "Calibri"           # ships with every Excel; never substitute-fonted

# ------------------------------------------------- canonical display names
# ONE set of platform names across every deliverable (fresh-eyes: "FK Minutes"
# vs "Flipkart Minutes" vs "Amz Now" drift). Use PLATFORM_DISPLAY everywhere;
# PLATFORM_SHORT only where a column is genuinely too narrow.
PLATFORM_DISPLAY = {
    "blinkit":          "Blinkit",
    "zepto":            "Zepto",
    "":        "",
    "flipkart-minutes": "Flipkart Minutes",
    "flipkart":         "Flipkart",
    "bigbasket":        "BigBasket",
    "amazon":           "Amazon",
    "amazon-fresh":     "Amazon Fresh",
    "amazon-now":       "Amazon Now",
}
PLATFORM_SHORT = {
    "flipkart-minutes": "FK Minutes",
    "amazon-fresh":     "Amz Fresh",
    "amazon-now":       "Amz Now",
}


def platform_name(key, short=False):
    """Canonical display name for a platform key ('amazon-now' → 'Amazon Now')."""
    if short and key in PLATFORM_SHORT:
        return PLATFORM_SHORT[key]
    return PLATFORM_DISPLAY.get(key, str(key).replace("-", " ").title())


# ------------------------------------------------------------- glossary
# Plain-language replacements/expansions for the jargon fresh-eyes flagged.
# Rule: expand each term ONCE per sheet at first use (or use the plain term).
GLOSSARY = {
    "SVD":        "Special Value Days (Fri–Sun agreed price list)",
    "ART":        "festival/event agreed price list, applied when announced",
    "regime":     "price plan",
    "modal":      "most-common price",
    "exposure":   "₹ gap below agreed price",
    "dark store": "delivery store",
}


def gloss(term):
    """First-use expansion: gloss('SVD') → 'SVD — Special Value Days (…)'."""
    full = GLOSSARY.get(term)
    return "%s — %s" % (term, full) if full else str(term)

TONES = {                # kpi_card / big_number / chip semantic tones
    "good":    POS,
    "bad":     NEG,
    "warn":    WARN,
    "brand":   BRAND,
    "neutral": INK,
    "muted":   MUTED,
}

# ------------------------------------------------------- number formats
# Indian grouping via conditional sections: crore / lakh / thousands.
# (Excel allows max 2 conditions + fallback. NOTE: negatives land in the
# fallback section UNSIGNED — fine for prices/counts, don't use for deltas.)
FMT_INR      = '[>=10000000]"₹"##\\,##\\,##\\,##0;[>=100000]"₹"##\\,##\\,##0;"₹"#,##0'
FMT_INR_DEC  = '[>=10000000]"₹"##\\,##\\,##\\,##0.00;[>=100000]"₹"##\\,##\\,##0.00;"₹"#,##0.00'
FMT_INT_IN   = '[>=10000000]##\\,##\\,##\\,##0;[>=100000]##\\,##\\,##0;#,##0'
FMT_PCT      = '0%'
FMT_PCT_DEC  = '0.0%'
FMT_INT      = '#,##0'
FMT_DEC      = '#,##0.00'

_THIN = Side(style="thin", color=RULE)
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")


def _tone(tone):
    """Accept a semantic tone name or a raw hex string."""
    return TONES.get(tone, tone or INK)


def _argb(hex6):
    """CF rule colors should be 8-digit ARGB; openpyxl 00-prefixes 6-digit hex
    (alpha 0 — Excel happens to ignore it, but FF is the documented form)."""
    h = str(hex6).lstrip("#").upper()
    return h if len(h) == 8 else "FF" + h


# ---------------------------------------------------------------- canvas
def no_gridlines(ws, zoom=110):
    """Hide gridlines + set zoom — the single biggest visual upgrade."""
    ws.sheet_view.showGridLines = False
    if zoom:
        ws.sheet_view.zoomScale = zoom


def col_grid(ws, n=10, width=16, first=1):
    """Uniform column rhythm: n columns of equal width starting at `first`."""
    for i in range(first, first + n):
        ws.column_dimensions[get_column_letter(i)].width = width


def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def fit_to_width(ws, print_area=None):
    """One-page-wide print setup (executives print these)."""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    if print_area:
        ws.print_area = print_area


# ---------------------------------------------------------------- chrome
def title_block(ws, row, title, subtitle="", c0=1, c1=10):
    """20pt title + muted subtitle + 3pt hairline. Returns next free row."""
    ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
    t = ws.cell(row, c0, title)
    t.font = Font(name=F, size=20, bold=True, color=INK)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 36
    r = row + 1
    if subtitle:
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
        h = ws.cell(r, c0, subtitle)
        h.font = Font(name=F, size=11, color=MUTED)
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 22
        r += 1
    hairline(ws, r, c0, c1)
    return r + 1


def hairline(ws, row, c0=1, c1=10):
    """3pt-high solid rule-colored divider row."""
    for c in range(c0, c1 + 1):
        ws.cell(row, c).fill = PatternFill("solid", fgColor=RULE)
    ws.row_dimensions[row].height = 3


def section_title(ws, row, text, c0=1, c1=10, fill=BRAND):
    """Full-width section header bar (white caps on brand green). Returns row+1."""
    ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
    c = ws.cell(row, c0, str(text).upper())
    c.font = Font(name=F, size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1


# ---------------------------------------------------------------- KPI cards
def kpi_card(ws, r, c, title, value, sub="", tone="neutral", span=2):
    """
    One KPI card: caps label / 22pt value / muted sub-line, hairline border,
    white fill, spanning `span` columns x 4 rows (r..r+3). The card grid rows
    get fixed heights (set once per row index, idempotent across a strip).
    Returns (next_row, next_col) = (r+4, c+span) so cards tile naturally.
    """
    c0, c1 = c, c + span - 1
    ws.row_dimensions[r].height = 20      # label
    ws.row_dimensions[r + 1].height = 38  # value
    ws.row_dimensions[r + 2].height = 16  # sub
    ws.row_dimensions[r + 3].height = 8   # breathing room inside the border
    for rr in range(r, r + 4):
        for cc in range(c0, c1 + 1):
            cell = ws.cell(rr, cc)
            cell.fill = _WHITE_FILL
            cell.border = Border(
                left=_THIN if cc == c0 else None,
                right=_THIN if cc == c1 else None,
                top=_THIN if rr == r else None,
                bottom=_THIN if rr == r + 3 else None)
    ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
    lc = ws.cell(r, c0, str(title).upper())
    lc.font = Font(name=F, size=9, bold=True, color=MUTED)
    lc.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
    ws.merge_cells(start_row=r + 1, start_column=c0, end_row=r + 1, end_column=c1)
    vc = ws.cell(r + 1, c0, value)
    vc.font = Font(name=F, size=22, bold=True, color=_tone(tone))
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=r + 2, start_column=c0, end_row=r + 2, end_column=c1)
    sc = ws.cell(r + 2, c0, sub or "")
    sc.font = Font(name=F, size=9, color=MUTED)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return r + 4, c + span


def big_number(ws, r, c, value, tone="neutral", size=28, span=2, num_fmt=None):
    """A lone oversized figure (hero stat). Returns the value cell."""
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    cell = ws.cell(r, c, value)
    cell.font = Font(name=F, size=size, bold=True, color=_tone(tone))
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if num_fmt:
        cell.number_format = num_fmt
    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, size + 14)
    return cell


def chip(ws, r, c, text, fill=BRAND_SOFT, color=INK):
    """Small badge cell (e.g. regime BAU/SVD/ART, status tags)."""
    cell = ws.cell(r, c, str(text))
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=F, size=9, bold=True, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def edition_badge(ws, r, c, text, span=2, fill=JIVO_GREEN, color="FFFFFF"):
    """
    Prominent cover badge ("AMAZON EDITION", "SVD DAY") — bold white caps on
    brand green, merged across `span` columns, hairline-bordered. Fresh-eyes
    2026-06-06: every scoped/edition workbook MUST carry one on its first
    sheet so it can never be mistaken for the cross-platform master.
    """
    c1 = c + span - 1
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)
    for cc in range(c, c1 + 1):
        ws.cell(r, cc).border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    cell = ws.cell(r, c, str(text).upper())
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=F, size=11, bold=True, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 20)
    return cell


def legend(ws, r, items, c=1, span=2):
    """
    Compact self-demonstrating legend row, meant to sit at the TOP of a sheet
    (under the header, inside the frozen pane — fresh-eyes: never row 117).
    items = [(text, fill_hex_or_None, text_hex_or_None)] — each item renders
    as one merged cell STYLED LIKE THE STATE IT EXPLAINS, e.g.
      legend(ws, 2, [("RED = below agreed price", *BAD_PAIR),
                     ("GREEN = at/above", *GOOD_PAIR),
                     ("blue = exact match", None, "1D4ED8"),
                     ("OOS = out of stock", None, None),
                     ("? = mapping pending", None, None)])
    Returns r + 1.
    """
    cc = c
    for text, fill, color in items:
        c1 = cc + span - 1
        if span > 1:
            ws.merge_cells(start_row=r, start_column=cc, end_row=r, end_column=c1)
        cell = ws.cell(r, cc, str(text))
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name=F, size=9, bold=bool(fill), color=color or MUTED)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cc = c1 + 1
    ws.row_dimensions[r].height = 16
    return r + 1


# ---------------------------------------------------------------- tables
def banded_table(ws, r, c, headers, rows, widths=None, num_fmts=None,
                 aligns=None, zebra=CANVAS, header_fill=None):
    """
    Manual zebra table (NOT a native Excel table — manual fills are the most
    round-trip-durable striping there is, and they never fight later CF rules).
      headers   list[str]
      rows      list[list]      (values; None renders as empty)
      widths    list[int|None]  per-column widths (optional)
      num_fmts  list[str|None]  per-column number formats (optional)
      aligns    list[str|None]  per-column horizontal alignment (optional)
      header_fill  hex or None  None = muted caps on hairline underline (house
                                default); a hex = white caps on that fill.
    Returns the last written row index.
    """
    n = len(headers)
    if widths:
        for j, w in enumerate(widths):
            if w:
                ws.column_dimensions[get_column_letter(c + j)].width = w
    for j, h in enumerate(headers):
        hc = ws.cell(r, c + j, str(h).upper() if not header_fill else str(h))
        if header_fill:
            hc.font = Font(name=F, size=10, bold=True, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor=header_fill)
        else:
            hc.font = Font(name=F, size=9, bold=True, color=MUTED)
            hc.border = Border(bottom=Side(style="thin", color=NEUTRAL))
        hc.alignment = Alignment(
            horizontal=(aligns[j] if aligns and aligns[j] else "left"),
            vertical="center", indent=1)
    ws.row_dimensions[r].height = 18
    for i, rw in enumerate(rows):
        rr = r + 1 + i
        stripe = PatternFill("solid", fgColor=zebra) if i % 2 == 0 else _WHITE_FILL
        for j in range(n):
            v = rw[j] if j < len(rw) else None
            cell = ws.cell(rr, c + j, v)
            cell.fill = stripe
            cell.font = Font(name=F, size=11, color=INK)
            cell.alignment = Alignment(
                horizontal=(aligns[j] if aligns and aligns[j] else "left"),
                vertical="center", indent=1)
            if num_fmts and j < len(num_fmts) and num_fmts[j]:
                cell.number_format = num_fmts[j]
    return r + len(rows)


# ------------------------------------------- conditional formats (durable)
def data_bar(ws, ref, color=BRAND, vmin=0, vmax=None):
    """
    In-cell data bars over `ref` ('E5:E25'). vmax=None scales to the range max.
    Solid-ish bar in brand green by default — our no-charts proportional visual.
    """
    rule = DataBarRule(
        start_type="num", start_value=vmin,
        end_type=("max" if vmax is None else "num"),
        end_value=vmax,
        color=_argb(color), showValue=True, minLength=None, maxLength=None)
    ws.conditional_formatting.add(ref, rule)


def color_scale(ws, ref, low=RED_FILL_HEX, mid="FFFFFF", high=GREEN_FILL_HEX):
    """3-color scale (default red→white→green, classic compliance read)."""
    ws.conditional_formatting.add(ref, ColorScaleRule(
        start_type="min", start_color=_argb(low),
        mid_type="percentile", mid_value=50, mid_color=_argb(mid),
        end_type="max", end_color=_argb(high)))


def icon_set(ws, ref, style="3TrafficLights1", values=(0, 33, 67), pct=True,
             reverse=False, show_value=True):
    """Icon set CF (traffic lights / arrows). Thresholds default to terciles."""
    ws.conditional_formatting.add(ref, IconSetRule(
        icon_style=style, type=("percent" if pct else "num"),
        values=list(values), showValue=show_value, reverse=reverse))


# ----------------------------------------------------- text visuals (no charts)
_SPARK = "▁▂▃▄▅▆▇█"


def spark(values, lo=None, hi=None):
    """Unicode sparkline string for a numeric series (None → gap). Durable
    everywhere a chart isn't. Style the cell MUTED/small via spark_cell()."""
    nums = [v for v in values if v is not None]
    if not nums:
        return ""
    lo = min(nums) if lo is None else lo
    hi = max(nums) if hi is None else hi
    span = (hi - lo) or 1.0
    out = []
    for v in values:
        if v is None:
            out.append(" ")
        else:
            idx = int(round((v - lo) / span * (len(_SPARK) - 1)))
            out.append(_SPARK[max(0, min(len(_SPARK) - 1, idx))])
    return "".join(out)


def meter(value, vmax, width=12):
    """Text progress bar: '▮▮▮▮▮▯▯▯▯▯▯▯' for value/vmax."""
    if not vmax or vmax <= 0:
        return "▯" * width
    filled = int(round(max(0.0, min(1.0, float(value) / float(vmax))) * width))
    return "▮" * filled + "▯" * (width - filled)


def spark_cell(ws, r, c, text, tone="brand", size=11):
    """Place a spark()/meter() string with the house look."""
    cell = ws.cell(r, c, text)
    cell.font = Font(name=F, size=size, color=_tone(tone))
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return cell


# ---------------------------------------------------------------- extras
def excel_app_workaround():
    """
    openpyxl >=3.1.4 stamps docProps/app.xml with
    'Microsoft Excel Compatible / Openpyxl 3.1.5'; Excel then mis-renders
    openpyxl-DRAWN native charts (tiny overlapping titles, missing axes).
    Call ONCE, before save, in the one stage allowed to draw charts
    (predict.py Leadership View) — it patches the stamp to 'Microsoft Excel'
    for every subsequent Workbook save in the process. No-op for chart-free
    workbooks; harmless to call twice.
    Ref: https://groups.google.com/g/openpyxl-users/c/khC6BTqaH3Y
    """
    from openpyxl.packaging.extended import ExtendedProperties
    if getattr(ExtendedProperties, "_xd_app_patched", False):
        return
    orig = ExtendedProperties.__init__

    def patched(self, *a, **kw):
        orig(self, *a, **kw)
        self.Application = "Microsoft Excel"

    ExtendedProperties.__init__ = patched
    ExtendedProperties._xd_app_patched = True


def note(ws, r, c, text, author="ecom-intel"):
    """Cell comment for drill-down detail (hover in Excel). Round-trip safe."""
    ws.cell(r, c).comment = Comment(text, author)


def footnote(ws, r, text, c0=1, c1=10):
    """Italic muted footer line (data source / generated-at)."""
    ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
    cell = ws.cell(r, c0, text)
    cell.font = Font(name=F, size=9, italic=True, color=MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return cell


# ================================================================ self-test
def _build_demo(path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Demo Dashboard"
    no_gridlines(ws)
    col_grid(ws, n=10, width=16)
    r = title_block(ws, 1, "Jivo × %s — Leadership View" % platform_name("amazon-now"),
                    "82% catalog live · avg 18% off · " + gloss("SVD"))
    edition_badge(ws, r, 9, "Amazon edition")
    r += 1
    r = legend(ws, r, [("RED = below agreed price",) + BAD_PAIR,
                       ("GREEN = at/above",) + GOOD_PAIR,
                       ("AMBER = overpriced",) + AMBER_PAIR,
                       ("OOS = out of stock", None, None),
                       ("? = mapping pending", None, None)])
    r += 1
    # KPI strip: 5 cards
    cards = [("CATALOG LIVE", "82%", "51 of 62 SKUs", "good"),
             ("AVG DISCOUNT", "18%", "across 48 priced", "good"),
             ("DEEP DISC ≥50%", "3", "6% of priced", "warn"),
             ("STOCK-OUT RISK", "9", "SKU(s)", "bad"),
             ("COVERAGE", "312", "pincodes live", "brand")]
    cc = 1
    for t, v, s, tone in cards:
        nr, cc = kpi_card(ws, r, cc, t, v, s, tone)
    r = nr + 1
    r = section_title(ws, r, "Top SKUs by availability")
    last = banded_table(
        ws, r, 1,
        ["SKU", "City", "Sale ₹", "MRP ₹", "Live %", "Trend"],
        [["Jivo Canola 1L", "Delhi", 399, 650, 0.92, spark([3, 4, 4, 5, 6])],
         ["Jivo Olive Pomace 1L", "Mumbai", 549, 925, 0.81, spark([6, 5, 5, 4, 4])],
         ["Jivo Canola 5L", "Pune", 1849, 3050, 0.64, spark([2, 3, 2, 4, 3])],
         ["Jivo Mustard 1L", "Jaipur", 189, 240, 0.55, spark([4, 4, 3, 3, 2])]],
        widths=[24, 14, 12, 12, 10, 14],
        num_fmts=[None, None, FMT_INR, FMT_INR, FMT_PCT, None],
        aligns=[None, None, "right", "right", "right", None])
    # conditional formats over the table body
    body_live = "E%d:E%d" % (r + 1, last)
    data_bar(ws, body_live, color=BRAND, vmin=0, vmax=1)
    color_scale(ws, "C%d:C%d" % (r + 1, last))
    icon_set(ws, "D%d:D%d" % (r + 1, last))
    r = last + 2
    big_number(ws, r, 1, 1234567.0, tone="brand", num_fmt=FMT_INR)
    chip(ws, r, 4, "SVD day", BRAND_SOFT, POS)
    spark_cell(ws, r, 5, meter(7, 10))
    note(ws, r, 1, "GMV proxy = Σ sale_price across live datapoints")
    footnote(ws, r + 2, "xlsx_dash selftest · every element here must survive a round-trip")
    freeze(ws, "A2")
    fit_to_width(ws)
    wb.save(path)


def _snapshot(path):
    """Capture every feature the durability contract covers."""
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Demo Dashboard"]
    cf = {}
    for rng in ws.conditional_formatting:
        key = str(rng.sqref)
        cf[key] = sorted(r.type for r in rng.rules)
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.comment:
                continue
            cells[cell.coordinate] = (
                cell.value, cell.number_format,
                cell.font.b, cell.font.sz, cell.font.color.rgb if cell.font.color else None,
                cell.fill.fgColor.rgb if cell.fill and cell.fill.fill_type else None,
                cell.comment.text if cell.comment else None)
    return {
        "gridlines": ws.sheet_view.showGridLines,
        "zoom": ws.sheet_view.zoomScale,
        "freeze": ws.freeze_panes,
        "merges": sorted(str(m) for m in ws.merged_cells.ranges),
        "cf": cf,
        "fit": ws.page_setup.fitToWidth,
        "cells": cells,
    }


def _selftest():
    import os
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        p1 = os.path.join(td, "demo.xlsx")
        p2 = os.path.join(td, "demo.rt.xlsx")
        _build_demo(p1)
        before = _snapshot(p1)
        # the round trip a later pipeline stage performs
        from openpyxl import load_workbook
        load_workbook(p1).save(p2)
        after = _snapshot(p2)
        failures = []
        for k in before:
            if before[k] != after[k]:
                failures.append(k)
        assert before["cf"], "no conditional formats captured — selftest is broken"
        assert before["merges"], "no merges captured — selftest is broken"
        assert any(v[6] for v in before["cells"].values()), "no comment captured"
        if failures:
            for k in failures:
                print("LOST IN ROUND-TRIP: %s\n  before=%r\n  after =%r"
                      % (k, before[k], after[k]))
            raise SystemExit("xlsx_dash selftest FAILED: %s" % ", ".join(failures))
        print("xlsx_dash selftest OK — %d cells, %d merges, %d CF ranges "
              "survived the round-trip intact"
              % (len(before["cells"]), len(before["merges"]), len(before["cf"])))
        # leave a copy for eyeballing if requested
        keep = os.environ.get("XLSX_DASH_DEMO")
        if keep:
            import shutil
            shutil.copy(p1, keep)
            print("demo written to %s" % keep)


if __name__ == "__main__":
    import sys
    if "--selftest" in sys.argv:
        _selftest()
    else:
        print(__doc__)
