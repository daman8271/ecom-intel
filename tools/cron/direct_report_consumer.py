#!/usr/bin/env python3
"""Validate and promote final Mac-built reports; never accepts shard data."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
from pathlib import Path
import re
import shlex
import shutil
import subprocess
import sys
from typing import Any

from openpyxl import load_workbook


IST = dt.timezone(dt.timedelta(hours=5, minutes=30))
RUN_ID_PATTERN = re.compile(r"^[0-9]{8}-[0-9]{6}-[a-z0-9][a-z0-9-]*$")
REQUIRED_SHEETS = {
    "Summary",
    "Master Data",
    "Pricing Matrix",
    "Stock Status",
    "Discount Analysis",
    "Coverage & Gaps",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def atomic_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.tmp.{os.getpid()}")
    with temp.open("w", encoding="ascii") as handle:
        json.dump(value, handle, indent=2, sort_keys=True, ensure_ascii=True)
        handle.write("\n")
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp, path)


def expected_names(platform: str, date_ist: str) -> set[str]:
    main = f"Jivo-{platform.title()}-Live-Report-{date_ist}.xlsx"
    if platform == "blinkit":
        return {main, f"Jivo-Blinkit-Not-Listed-Pincodes-{date_ist}.xlsx"}
    return {main}


def validate_workbook(path: Path, main: bool) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if main:
            missing = REQUIRED_SHEETS - set(workbook.sheetnames)
            if missing:
                raise ValueError(f"missing workbook sheets: {sorted(missing)}")
            if workbook["Master Data"].max_row < 2:
                raise ValueError("Master Data is empty")
    finally:
        workbook.close()


def copy_fsync(source: Path, target: Path) -> None:
    with source.open("rb") as incoming, target.open("wb") as outgoing:
        shutil.copyfileobj(incoming, outgoing, 1024 * 1024)
        outgoing.flush()
        os.fsync(outgoing.fileno())


def acknowledge(receipt_path: Path, run_id: str, host: str, root: str) -> str:
    if not host:
        return "disabled"
    remote = f"{root.rstrip('/')}/{run_id}.json"
    quoted_root = shlex.quote(root)
    quoted_part = shlex.quote(f"{remote}.part")
    quoted_remote = shlex.quote(remote)
    try:
        subprocess.run(
            ["ssh", "-o", "BatchMode=yes", "-o", "ConnectTimeout=15", host,
             f"mkdir -p -- {quoted_root}"],
            check=True,
            timeout=30,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        subprocess.run(
            ["scp", "-q", "-o", "BatchMode=yes", str(receipt_path), f"{host}:{quoted_part}"],
            check=True,
            timeout=60,
        )
        subprocess.run(
            ["ssh", "-o", "BatchMode=yes", host, f"mv -- {quoted_part} {quoted_remote}"],
            check=True,
            timeout=30,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return "accepted"
    except (OSError, subprocess.SubprocessError):
        return "pending"


def acknowledge_once(receipt_path: Path, run_id: str, host: str, root: str) -> str:
    marker = receipt_path.with_suffix(".promotion-acked.json")
    if marker.is_file():
        return "accepted"
    status = acknowledge(receipt_path, run_id, host, root)
    if status == "accepted":
        atomic_json(marker, {"status": status, "run_id": run_id})
    return status


def consume_run(
    run_dir: Path,
    date_ist: str,
    output_dir: Path,
    receipt_root: Path,
    stable_age: int,
    ack_host: str,
    ack_root: str,
) -> tuple[str, str]:
    source_receipt = run_dir / "report.ready.json"
    age = dt.datetime.now().timestamp() - source_receipt.stat().st_mtime
    if age < stable_age:
        return "waiting", f"receipt is only {int(age)}s old"

    receipt_bytes = source_receipt.read_bytes()
    receipt_sha = hashlib.sha256(receipt_bytes).hexdigest()
    receipt = json.loads(receipt_bytes.decode("utf-8-sig"))
    run_id = run_dir.name
    platform = str(receipt.get("platform", "")).lower()
    if receipt.get("schema") != "jivo-direct-report-receipt-v1":
        raise ValueError("unsupported receipt schema")
    if receipt.get("status") != "ready" or receipt.get("review_verdict") != "OK":
        raise ValueError("report is not quality-approved")
    if receipt.get("run_id") != run_id:
        raise ValueError("receipt run_id does not match directory")
    if not RUN_ID_PATTERN.fullmatch(run_id):
        raise ValueError("unsafe run_id")
    if receipt.get("date_ist") != date_ist or not run_id.startswith(date_ist.replace("-", "")):
        raise ValueError("receipt date/run does not match requested IST date")
    if platform not in {"blinkit", "zepto"}:
        raise ValueError("unsupported platform")
    if not str(receipt.get("attempt_id", "")):
        raise ValueError("attempt_id is missing")
    for field in ("plan_sha256", "source_sha256", "scraper_sha256", "merged_sha256", "merge_receipt_sha256"):
        value = receipt.get(field)
        if not isinstance(value, str) or not re.fullmatch(r"[0-9a-f]{64}", value):
            raise ValueError(f"{field} is missing or invalid")
    input_hashes = receipt.get("input_result_sha256")
    if not isinstance(input_hashes, dict) or set(input_hashes) != {"macpro", "windows"}:
        raise ValueError("input result hashes are incomplete")
    if any(not isinstance(value, str) or not re.fullmatch(r"[0-9a-f]{64}", value) for value in input_hashes.values()):
        raise ValueError("input result hash is invalid")
    if int(receipt.get("merged_bytes") or 0) <= 0 \
       or int(receipt.get("pincodes_total") or 0) <= 0 \
       or int(receipt.get("total_rows") or 0) <= 0:
        raise ValueError("merged byte/pincode/row counts are invalid")
    policy = receipt.get("quality_policy")
    if not isinstance(policy, dict):
        raise ValueError("quality policy is missing")
    if platform == "blinkit":
        required_policy = {
            "require_auth": True,
            "require_oos_probe": True,
            "require_pdp_oos_probe": True,
            "require_pdp_price_probe": True,
            "max_pdp_price_probe_failed": 0,
            "max_unverified_oos": 0,
        }
        if any(policy.get(key) != value for key, value in required_policy.items()):
            raise ValueError("Blinkit quality policy is incomplete")
    elif float(policy.get("min_serviceable_pct") or 0) <= 0 \
         or float(policy.get("min_rows_per_source_pincode") or 0) <= 0:
        raise ValueError("Zepto quality policy is incomplete")

    accepted_path = receipt_root / date_ist / f"{run_id}.json"
    if accepted_path.exists():
        accepted = load_json(accepted_path)
        if accepted.get("source_receipt_sha256") != receipt_sha:
            raise ValueError("accepted receipt conflicts with current source receipt")
        verified = accepted.get("workbooks")
        if not isinstance(verified, list) or not verified:
            raise ValueError("accepted workbook manifest is invalid")
        accepted_names = [str(entry.get("name", "")) for entry in verified if isinstance(entry, dict)]
        required_names = expected_names(platform, date_ist)
        if len(accepted_names) != len(required_names) or set(accepted_names) != required_names:
            raise ValueError("accepted workbook manifest does not match platform")
        restored = 0
        for entry in verified:
            name = str(entry.get("name", ""))
            if Path(name).name != name or name not in expected_names(platform, date_ist):
                raise ValueError("accepted workbook name is invalid")
            destination = output_dir / name
            if destination.exists():
                if sha256_file(destination) != entry.get("sha256"):
                    raise ValueError(f"accepted output hash changed: {name}")
                continue
            source = run_dir / name
            if not source.is_file() or sha256_file(source) != entry.get("sha256"):
                raise ValueError(f"accepted output is missing and source cannot restore it: {name}")
            output_dir.mkdir(parents=True, exist_ok=True)
            temp = output_dir / f".{name}.restore-{run_id}.part"
            try:
                copy_fsync(source, temp)
                if sha256_file(temp) != entry["sha256"]:
                    raise ValueError(f"restored hash mismatch: {name}")
                os.replace(temp, destination)
                restored += 1
            finally:
                temp.unlink(missing_ok=True)
        ack = acknowledge_once(accepted_path, run_id, ack_host, ack_root)
        return "existing", f"already accepted; restored={restored}; Mac promotion acknowledgement={ack}"

    entries = receipt.get("workbooks")
    if not isinstance(entries, list):
        raise ValueError("workbooks manifest is missing")
    by_name = {str(entry.get("name", "")): entry for entry in entries if isinstance(entry, dict)}
    required = expected_names(platform, date_ist)
    if len(entries) != len(required) or len(by_name) != len(entries) or set(by_name) != required:
        raise ValueError(f"workbook manifest mismatch: expected {sorted(required)}")

    verified: list[dict[str, Any]] = []
    for name in sorted(required):
        if Path(name).name != name:
            raise ValueError("unsafe workbook name")
        source = run_dir / name
        entry = by_name[name]
        if not source.is_file() or source.stat().st_size < 10_000:
            raise ValueError(f"workbook missing or too small: {name}")
        actual_bytes = source.stat().st_size
        actual_sha = sha256_file(source)
        if entry.get("bytes") != actual_bytes or entry.get("sha256") != actual_sha:
            raise ValueError(f"workbook hash/size mismatch: {name}")
        validate_workbook(source, "-Live-Report-" in name)
        verified.append({"name": name, "bytes": actual_bytes, "sha256": actual_sha})

    output_dir.mkdir(parents=True, exist_ok=True)
    for entry in verified:
        destination = output_dir / entry["name"]
        if destination.exists() and sha256_file(destination) != entry["sha256"]:
            raise ValueError(f"refusing to overwrite different report: {entry['name']}")

    staged: list[tuple[Path, Path, str]] = []
    try:
        for entry in verified:
            source = run_dir / entry["name"]
            destination = output_dir / entry["name"]
            temp = output_dir / f".{entry['name']}.direct-{run_id}.part"
            copy_fsync(source, temp)
            if sha256_file(temp) != entry["sha256"]:
                raise ValueError(f"staged hash mismatch: {entry['name']}")
            staged.append((temp, destination, entry["sha256"]))
        for temp, destination, expected_sha in staged:
            if destination.exists() and sha256_file(destination) != expected_sha:
                raise ValueError(f"destination changed during promotion: {destination.name}")
            os.replace(temp, destination)
    finally:
        for temp, _, _ in staged:
            temp.unlink(missing_ok=True)

    accepted = {
        "schema": "jivo-direct-promotion-receipt-v1",
        "status": "accepted",
        "platform": platform,
        "date_ist": date_ist,
        "run_id": run_id,
        "attempt_id": receipt["attempt_id"],
        "merged_sha256": receipt["merged_sha256"],
        "source_receipt_sha256": receipt_sha,
        "workbooks": verified,
        "accepted_at": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
    }
    atomic_json(accepted_path, accepted)
    ack = acknowledge_once(accepted_path, run_id, ack_host, ack_root)
    return "accepted", f"promoted {len(verified)} workbook(s); Mac promotion acknowledgement={ack}"


def pending_delivery_count(receipt_root: Path, date_ist: str, delivery_root: Path) -> int:
    pending: set[str] = set()
    delivered: set[str] = set()
    for path in (receipt_root / date_ist).glob("*.json"):
        try:
            promotion = load_json(path)
            if promotion.get("status") != "accepted":
                continue
            for entry in promotion.get("workbooks", []):
                name = str(entry["name"])
                if "-Not-Listed-Pincodes-" in name:
                    continue
                receipt = delivery_root / date_ist / f"{name}.json"
                if receipt.is_file():
                    confirmation = load_json(receipt)
                    if confirmation.get("sha256") == entry.get("sha256") and confirmation.get("messageId"):
                        delivered.add(name)
                        continue
                pending.add(name)
        except (OSError, ValueError, KeyError, json.JSONDecodeError):
            continue
    return len(pending - delivered)


def consume_failure(source: Path, date_ist: str, failure_root: Path) -> tuple[str, dict[str, str]]:
    run_dir = source.parent
    run_id = run_dir.name
    if run_dir.is_symlink() or source.is_symlink() or not RUN_ID_PATTERN.fullmatch(run_id):
        raise ValueError("unsafe endpoint failure path")
    receipt = load_json(source)
    expected = {
        "schema": "jivo-direct-failure-receipt-v1",
        "date_ist": date_ist,
        "run_id": run_id,
        "status": "failed",
    }
    for key, value in expected.items():
        if receipt.get(key) != value:
            raise ValueError(f"endpoint failure receipt mismatch: {key}")
    if receipt.get("platform") not in {"blinkit", "zepto"}:
        raise ValueError("endpoint failure platform is invalid")
    if not str(receipt.get("attempt_id", "")) or not str(receipt.get("phase", "")):
        raise ValueError("endpoint failure attempt/phase is missing")
    if not isinstance(receipt.get("plan_sha256"), str) or len(receipt["plan_sha256"]) != 64:
        raise ValueError("endpoint failure plan hash is invalid")
    source_sha = sha256_file(source)
    accepted = failure_root / date_ist / f"{run_id}.json"
    detail = {
        "run_id": run_id,
        "platform": str(receipt["platform"]),
        "phase": str(receipt["phase"]),
        "reason": str(receipt.get("reason", "unspecified")),
    }
    if accepted.is_file():
        previous = load_json(accepted)
        if previous.get("source_receipt_sha256") != source_sha:
            raise ValueError("endpoint failure receipt changed after acceptance")
        return "existing", detail
    atomic_json(
        accepted,
        {
            "schema": "jivo-direct-failure-accepted-v1",
            "source_receipt_sha256": source_sha,
            **detail,
        },
    )
    return "new", detail


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inbox", default="shards/mac-direct-ready")
    parser.add_argument("--output", default="output")
    parser.add_argument("--receipts", default="logs/direct-report-receipts")
    parser.add_argument("--date", default=dt.datetime.now(IST).date().isoformat())
    parser.add_argument("--stable-age", type=int, default=30)
    parser.add_argument("--ack-host", default="macpro")
    parser.add_argument("--ack-root", default="/Users/danny./ecom-direct/receipts/promotion")
    parser.add_argument("--delivery-receipts", default="logs/delivery-receipts")
    parser.add_argument("--failure-receipts", default="logs/direct-report-failures")
    args = parser.parse_args()

    inbox = Path(args.inbox)
    summary: dict[str, Any] = {
        "date": args.date,
        "new": 0,
        "existing": 0,
        "waiting": 0,
        "rejected": 0,
        "pending_delivery": 0,
        "endpoint_failures": [],
        "errors": [],
    }
    if inbox.exists():
        for source_receipt in sorted(inbox.glob("*/report.ready.json")):
            run_dir = source_receipt.parent
            if not run_dir.name.startswith(args.date.replace("-", "")):
                continue
            rejection_path = (
                Path(args.receipts)
                / args.date
                / "rejected"
                / f"{hashlib.sha256(run_dir.name.encode()).hexdigest()}.json"
            )
            try:
                source_sha = sha256_file(source_receipt)
            except OSError:
                source_sha = "unreadable"
            if rejection_path.is_file():
                try:
                    rejected = load_json(rejection_path)
                    if rejected.get("source_receipt_sha256") == source_sha:
                        summary["rejected"] += 1
                        continue
                except (OSError, ValueError):
                    pass
            try:
                if run_dir.is_symlink() or source_receipt.is_symlink():
                    raise ValueError("symlinked inbox paths are prohibited")
                status, message = consume_run(
                    run_dir,
                    args.date,
                    Path(args.output),
                    Path(args.receipts),
                    args.stable_age,
                    args.ack_host,
                    args.ack_root,
                )
                if status == "accepted":
                    summary["new"] += 1
                else:
                    summary[status] += 1
                print(f"{run_dir.name}: {status}: {message}", file=sys.stderr)
            except Exception as exc:
                summary["errors"].append({"run_id": run_dir.name, "error": str(exc)})
                print(f"{run_dir.name}: rejected: {exc}", file=sys.stderr)
                atomic_json(
                    rejection_path,
                    {
                        "schema": "jivo-direct-rejection-v1",
                        "run_id": run_dir.name,
                        "source_receipt_sha256": source_sha,
                        "error": str(exc),
                    },
                )
    if inbox.exists():
        for failure in sorted(inbox.glob("*/failure.json")):
            if not failure.parent.name.startswith(args.date.replace("-", "")):
                continue
            try:
                status, detail = consume_failure(failure, args.date, Path(args.failure_receipts))
                if status == "new":
                    summary["endpoint_failures"].append(detail)
            except Exception as exc:
                summary["errors"].append({"run_id": failure.parent.name, "error": str(exc)})
                print(f"{failure.parent.name}: rejected failure receipt: {exc}", file=sys.stderr)
    for accepted_path in (Path(args.receipts) / args.date).glob("*.json"):
        try:
            accepted = load_json(accepted_path)
            run_id = str(accepted.get("run_id", ""))
            if accepted.get("status") == "accepted" and RUN_ID_PATTERN.fullmatch(run_id):
                acknowledge_once(accepted_path, run_id, args.ack_host, args.ack_root)
        except (OSError, ValueError, KeyError, json.JSONDecodeError):
            continue
    summary["pending_delivery"] = pending_delivery_count(
        Path(args.receipts), args.date, Path(args.delivery_receipts)
    )
    print(json.dumps(summary, sort_keys=True))
    return 1 if summary["errors"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
