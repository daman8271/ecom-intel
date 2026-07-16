#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
from pathlib import Path
import shutil
import subprocess
import tempfile
import unittest

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[3]
CONSUMER = ROOT / "tools/cron/direct_competitor_consumer.py"
GATE = ROOT / "tools/cron/direct_competitor_is_accepted.py"
DATE = "2026-07-15"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def brand_hash(brands: list[str]) -> str:
    return hashlib.sha256(json.dumps(brands, separators=(",", ":")).encode("ascii")).hexdigest()


class DirectCompetitorConsumerTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.base = Path(self.temp.name)
        self.inbox = self.base / "inbox"
        self.output = self.base / "output"
        self.data = self.base / "data"
        self.logs = self.base / "logs"
        self.receipts = self.base / "receipts"
        self.failures = self.base / "failures"

    def tearDown(self) -> None:
        self.temp.cleanup()

    def build_package(self, platform: str = "zepto", seconds: str = "000000") -> tuple[Path, dict, Path, Path]:
        workflow = "blinkit-top8" if platform == "blinkit" else "zepto-competitor"
        run_id = f"20260715-{seconds}-{platform}-competitor-direct-a01"
        run = self.inbox / run_id
        run.mkdir(parents=True)
        label = platform.title()
        book = run / f"Competitor-Price-Watch-{label}-{DATE}.xlsx"
        sheets = (
            ["Summary", "City-Pin-SKU Prices", "Run Scope", "Anchor Watch", "Master Data"]
            if platform == "blinkit" else ["Summary", "Anchor Watch", "Master Data"]
        )
        pins = 75 if platform == "blinkit" else 25
        brands_display = [
            "Borges", "Del Monte", "Figaro", "Fortune", "Gulab",
            "Hudson", "Oreal", "Saffola", "Sundrop", "Tata",
        ]
        brands = sorted(brand.casefold() for brand in brands_display)
        captured_at = "2026-07-14T19:00:00Z"
        rows = []
        per_pin = []
        for index in range(pins):
            pincode = f"{100000 + index:06d}"
            brand = brands_display[index % len(brands_display)]
            row = {
                "platform": platform,
                "city": f"City-{index // 3:02d}",
                "pincode": pincode,
                "store_id": f"store-{index}",
                "brand": brand,
                "name": f"{brand} Test Oil {index}",
                "canonical": f"oil-{index}",
                "category": "cooking oil",
                "pack": "1 L",
                "vol_ml": 1000,
                "per_litre": 100,
                "sale": 100,
                "mrp": 110,
                "discount_pct": 9.0909,
                "in_stock": 1,
                "rank": index + 1,
                "is_ad": 0,
                "captured_at": captured_at,
            }
            rows.append(row)
            per_pin.append({
                "city": row["city"],
                "locality": f"Locality-{index}",
                "lat": 20.0 + index / 1000,
                "lon": 70.0 + index / 1000,
                "pincode": pincode,
                "resolved": True,
                "auth_accepted": 1,
                "serviceable": True,
                "blocked": None,
                "rows": [row],
            })
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name in sheets:
            sheet = workbook.create_sheet(name)
            if name == "Master Data":
                sheet.append([
                    "Platform", "Brand", "JIVO?", "Search Category (scraped)", "Oil Type (name)",
                    "Grade (name)", "Matched Anchor", "Blend?", "City", "Pincode", "Store ID",
                    "Name", "Pack", "Vol (ml)", "MRP Rs", "Sale Rs", "Rs/L", "Discount %",
                    "In stock", "Rank", "Ad?", "Captured",
                ])
                for row in rows:
                    sheet.append([
                        platform.title(), row["brand"], "", row["category"], "", "", "", "",
                        row["city"], row["pincode"], row["store_id"], row["name"], row["pack"],
                        row["vol_ml"], row["mrp"], row["sale"], row["per_litre"],
                        round(row["discount_pct"] / 100, 4), "Yes", row["rank"], "",
                        row["captured_at"][:16].replace("T", " "),
                    ])
            elif name == "Run Scope":
                for row_number in range(1, 83):
                    sheet.cell(row=row_number, column=1, value=f"meta-{row_number}")
                for row_number, item in enumerate(per_pin, 8):
                    sheet.cell(row=row_number, column=1, value=item["city"])
                    sheet.cell(row=row_number, column=2, value=item["pincode"])
                    sheet.cell(row=row_number, column=3, value=item["locality"])
                    sheet.cell(row=row_number, column=4, value="Mac Pro" if row_number % 2 else "Windows")
                    sheet.cell(row=row_number, column=5, value="Yes")
                    sheet.cell(row=row_number, column=6, value="Yes")
                    sheet.cell(row=row_number, column=7, value=1)
                    sheet.cell(row=row_number, column=8, value=item["rows"][0]["brand"])
            else:
                sheet.append(["header", "value"])
                for index in range(120):
                    sheet.append([f"row-{index}", hashlib.sha256(f"{name}-{index}".encode()).hexdigest()])
        workbook.save(book)

        summary = {
            "mode": "competitor",
            "platform": platform,
            "date_ist": DATE,
            "run_id": run_id,
            "pincodes_total": pins,
            "pincodes_blocked": 0,
            "pincodes_with_rows": pins,
            "total_rows": pins,
            "unique_brands": len(brands),
            "partial": False,
            "captured_at": captured_at,
            "scope": {
                "anchors": ["Jivo", "Sano"],
                "competitors": brands_display,
                "capture_brands": ["Jivo", "Sano", *brands_display],
            },
        }
        if platform == "blinkit":
            summary.update({
                "pincodes_resolved": 75,
                "auth_verified": 1,
                "auth_verified_pincodes": 75,
            })
        else:
            summary["pincodes_serviceable"] = 25
        capture = run / f"{platform}_competitor_{DATE}.json"
        capture.write_text(
            json.dumps({"summary": summary, "perPin": per_pin, "allRows": rows, "partial": False}),
            encoding="utf-8",
        )

        audit_name = f"blinkit-top8-{DATE}.audit.json" if platform == "blinkit" else f"zepto-competitor-{DATE}.audit.json"
        audit = run / audit_name
        audit_value = {
            "schema": "jivo-direct-competitor-delivery-audit-v1",
            "platform": platform,
            "workflow_kind": workflow,
            "date_ist": DATE,
            "run_id": run_id,
            "attempt_id": "01",
            "status": "OK",
            "brand_set": brands,
            "brand_set_count": len(brands),
            "brand_set_sha256": brand_hash(brands),
            "pincodes_total": pins,
            "total_rows": pins,
            "merged_sha256": sha256(capture),
            "workbook_sha256": sha256(book),
        }
        audit.write_text(json.dumps(audit_value), encoding="utf-8")

        support = [
            {"name": name, "sha256": sha256(ROOT / "tools/competitor" / name)}
            for name in (
                "category_queries.json", "competitor_brands.json", "competitor_match_map.json",
                "maps_to_jivo.json", "oil_classifier.json",
            )
        ]
        code_names = ["build_competitor_report.py"]
        if platform == "blinkit":
            code_names += ["select_blinkit_top8_pincodes.py", "build_blinkit_top8_daily.py"]
        code = [{"name": name, "sha256": sha256(ROOT / "tools/competitor" / name)} for name in code_names]
        policy = {
            "min_rows_per_source_pincode": 1.0,
            "min_unique_brands": 8,
            "baseline_min_row_fraction": 0.4,
            "baseline_min_brand_fraction": 0.6,
        }
        if platform == "blinkit":
            policy.update({"require_resolved": True, "require_auth": True, "require_rows_each_pincode": True})
        else:
            policy["min_serviceable_pct"] = 20.0
        receipt = {
            "schema": "jivo-direct-competitor-report-receipt-v1",
            "platform": platform,
            "workflow_kind": workflow,
            "date_ist": DATE,
            "run_id": run_id,
            "attempt_id": "01",
            "status": "ready",
            "review_verdict": "OK",
            "plan_sha256": "a" * 64,
            "source_sha256": "b" * 64,
            "scraper_sha256": "c" * 64,
            "merge_receipt_sha256": "d" * 64,
            "merged_sha256": sha256(capture),
            "merged_bytes": capture.stat().st_size,
            "input_result_sha256": {"macpro": "e" * 64, "windows": "f" * 64},
            "input_progress_sha256": {"macpro": "1" * 64, "windows": "2" * 64},
            "input_terminal_sha256": {"macpro": "3" * 64, "windows": "4" * 64},
            "support_files": support,
            "code_files": code,
            "quality_policy": policy,
            "baseline": {
                "name": "baseline.json",
                "capture_sha256": "9" * 64,
                "capture_bytes": 100,
                "total_rows": pins,
                "pincodes_with_rows": pins,
                "unique_brands": len(brands),
            },
            "anchor_brands": ["Jivo", "Sano"],
            "competitor_brands": brands_display,
            "capture_brands": ["Jivo", "Sano", *brands_display],
            "brand_aliases": {"oriel": "oreal"},
            "pincodes_total": pins,
            "total_rows": pins,
            "brand_set": brands,
            "brand_set_count": len(brands),
            "brand_set_sha256": brand_hash(brands),
            "merged_capture": {"name": capture.name, "bytes": capture.stat().st_size, "sha256": sha256(capture)},
            "audits": [],
            "workbooks": [{"name": book.name, "bytes": book.stat().st_size, "sha256": sha256(book)}],
        }
        receipt["scraper_sha256"] = sha256(ROOT / "platforms" / platform / "scrape.js")
        quality = run / "quality-audit.json"
        quality_value = {
            "schema": "jivo-direct-competitor-quality-audit-v1",
            "plan_sha256": receipt["plan_sha256"],
            "merge_receipt_sha256": receipt["merge_receipt_sha256"],
            "merged_sha256": receipt["merged_sha256"],
            "workbook_sha256": receipt["workbooks"][0]["sha256"],
            "platform": platform,
            "workflow_kind": workflow,
            "date_ist": DATE,
            "run_id": run_id,
            "attempt_id": "01",
            "verdict": "OK",
            "brand_set": brands,
            "brand_set_count": len(brands),
            "brand_set_sha256": brand_hash(brands),
            "pincodes_total": pins,
            "total_rows": pins,
            "input_result_sha256": receipt["input_result_sha256"],
            "input_progress_sha256": receipt["input_progress_sha256"],
            "input_terminal_sha256": receipt["input_terminal_sha256"],
            "support_files": support,
            "code_files": code,
            "quality_policy": policy,
            "anchor_brands": receipt["anchor_brands"],
            "competitor_brands": receipt["competitor_brands"],
            "capture_brands": receipt["capture_brands"],
        }
        quality.write_text(json.dumps(quality_value), encoding="utf-8")
        receipt["audits"] = [
            {"name": quality.name, "bytes": quality.stat().st_size, "sha256": sha256(quality)},
            {"name": audit.name, "bytes": audit.stat().st_size, "sha256": sha256(audit)},
        ]
        (run / "report.ready.json").write_text(json.dumps(receipt), encoding="utf-8")
        return run, receipt, book, capture

    def refresh_package(self, run: Path, receipt: dict, book: Path, capture: Path) -> None:
        receipt["merged_sha256"] = sha256(capture)
        receipt["merged_bytes"] = capture.stat().st_size
        receipt["merged_capture"] = {
            "name": capture.name, "bytes": capture.stat().st_size, "sha256": sha256(capture),
        }
        receipt["workbooks"] = [
            {"name": book.name, "bytes": book.stat().st_size, "sha256": sha256(book)}
        ]
        for name in ("quality-audit.json", f"{receipt['workflow_kind']}-{DATE}.audit.json"):
            path = run / name
            if not path.exists():
                continue
            value = json.loads(path.read_text())
            value["merged_sha256"] = sha256(capture)
            value["workbook_sha256"] = sha256(book)
            path.write_text(json.dumps(value))
        receipt["audits"] = [
            {"name": name, "bytes": (run / name).stat().st_size, "sha256": sha256(run / name)}
            for name in ("quality-audit.json", f"{receipt['workflow_kind']}-{DATE}.audit.json")
        ]
        (run / "report.ready.json").write_text(json.dumps(receipt))

    def consume(self) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [
                "python3", str(CONSUMER), "--inbox", str(self.inbox), "--output", str(self.output),
                "--data", str(self.data), "--audit-dir", str(self.logs), "--receipts", str(self.receipts),
                "--failure-receipts", str(self.failures), "--date", DATE, "--stable-age", "0", "--ack-host", "",
            ],
            text=True, capture_output=True,
        )

    def gate(self, file: Path, platform: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            ["python3", str(GATE), "--file", str(file), "--date", DATE, "--platform", platform,
             "--receipts", str(self.receipts)],
            text=True, capture_output=True,
        )

    def test_valid_zepto_package_promotes_once_and_gates(self) -> None:
        run, _, book, capture = self.build_package()
        self.assertNotEqual(self.gate(book, "zepto").returncode, 0)
        first = self.consume()
        self.assertEqual(first.returncode, 0, first.stderr)
        self.assertEqual(json.loads(first.stdout)["new"], 1)
        promoted_book = self.output / book.name
        promoted_capture = self.data / capture.name
        promoted_audit = self.logs / f"zepto-competitor-{DATE}.audit.json"
        self.assertEqual(sha256(promoted_book), sha256(book))
        self.assertEqual(sha256(promoted_capture), sha256(capture))
        self.assertTrue(promoted_audit.is_file())
        self.assertEqual(self.gate(promoted_book, "zepto").returncode, 0)
        second = self.consume()
        self.assertEqual(second.returncode, 0, second.stderr)
        self.assertEqual(json.loads(second.stdout)["existing"], 1)
        self.assertTrue((self.receipts / DATE / f"{run.name}.json").is_file())

    def test_valid_blinkit_package_promotes_delivery_audit(self) -> None:
        _, _, book, capture = self.build_package("blinkit")
        result = self.consume()
        self.assertEqual(result.returncode, 0, result.stderr)
        promoted_audit = self.logs / f"blinkit-top8-{DATE}.audit.json"
        self.assertTrue(promoted_audit.is_file())
        self.assertEqual(self.gate(self.output / book.name, "blinkit").returncode, 0)
        self.assertEqual(sha256(self.data / capture.name), sha256(capture))

    def test_tampered_capture_is_rejected_without_promotion(self) -> None:
        _, _, book, capture = self.build_package()
        capture.write_bytes(capture.read_bytes() + b"tamper")
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("hash/size mismatch", result.stderr)
        self.assertFalse((self.output / book.name).exists())

    def test_missing_endpoint_hash_manifest_is_rejected(self) -> None:
        run, receipt, book, _ = self.build_package()
        del receipt["input_progress_sha256"]
        (run / "report.ready.json").write_text(json.dumps(receipt))
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("input_progress_sha256", result.stderr)
        self.assertFalse((self.output / book.name).exists())

    def test_missing_code_manifest_entry_is_rejected(self) -> None:
        run, receipt, _, _ = self.build_package("blinkit")
        receipt["code_files"] = receipt["code_files"][:1]
        (run / "report.ready.json").write_text(json.dumps(receipt))
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("code_files is missing", result.stderr)

    def test_ours_in_rival_brand_set_is_rejected(self) -> None:
        run, receipt, _, _ = self.build_package()
        receipt["brand_set"].append("jivo")
        receipt["brand_set"] = sorted(receipt["brand_set"])
        receipt["brand_set_count"] = len(receipt["brand_set"])
        receipt["brand_set_sha256"] = brand_hash(receipt["brand_set"])
        (run / "report.ready.json").write_text(json.dumps(receipt))
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("rival-only", result.stderr)

    def test_tampered_accepted_destination_fails_gate_and_consumer(self) -> None:
        _, _, book, _ = self.build_package()
        self.assertEqual(self.consume().returncode, 0)
        promoted = self.output / book.name
        promoted.write_bytes(promoted.read_bytes() + b"tamper")
        self.assertNotEqual(self.gate(promoted, "zepto").returncode, 0)
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("accepted destination hash changed", result.stderr)

    def test_missing_accepted_destination_is_restored(self) -> None:
        _, _, book, _ = self.build_package()
        self.assertEqual(self.consume().returncode, 0)
        promoted = self.output / book.name
        promoted.unlink()
        result = self.consume()
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(sha256(promoted), sha256(book))

    def test_tampered_promoted_audit_revokes_delivery_gate(self) -> None:
        _, _, book, _ = self.build_package()
        self.assertEqual(self.consume().returncode, 0)
        promoted_book = self.output / book.name
        promoted_audit = self.logs / f"zepto-competitor-{DATE}.audit.json"
        promoted_audit.write_bytes(promoted_audit.read_bytes() + b"tamper")
        self.assertNotEqual(self.gate(promoted_book, "zepto").returncode, 0)
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("accepted destination hash changed", result.stderr)

    def test_tampered_audit_is_rejected(self) -> None:
        run, receipt, _, _ = self.build_package()
        audit = run / receipt["audits"][0]["name"]
        audit.write_bytes(audit.read_bytes() + b"tamper")
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("hash/size mismatch", result.stderr)

    def test_endpoint_failure_is_accepted_once(self) -> None:
        run_id = "20260715-000010-zepto-competitor-direct-a01"
        run = self.inbox / run_id
        run.mkdir(parents=True)
        failure = {
            "schema": "jivo-direct-competitor-failure-receipt-v1",
            "platform": "zepto",
            "workflow_kind": "zepto-competitor",
            "date_ist": DATE,
            "run_id": run_id,
            "attempt_id": "01",
            "status": "failed",
            "plan_sha256": "a" * 64,
            "phase": "windows-terminal",
            "reason": "runner_rc=1",
        }
        (run / "failure.json").write_text(json.dumps(failure))
        first = self.consume()
        self.assertEqual(first.returncode, 0, first.stderr)
        detail = json.loads(first.stdout)["endpoint_failures"][0]
        second = self.consume()
        self.assertEqual(second.returncode, 0, second.stderr)
        self.assertEqual(len(json.loads(second.stdout)["endpoint_failures"]), 1)
        Path(detail["alert_marker"]).write_text(json.dumps({
            "status": "alerted",
            "source_receipt_sha256": detail["source_receipt_sha256"],
        }))
        third = self.consume()
        self.assertEqual(third.returncode, 0, third.stderr)
        self.assertEqual(len(json.loads(third.stdout)["endpoint_failures"]), 0)

    def test_actual_one_brand_capture_cannot_claim_ten(self) -> None:
        run, receipt, book, capture = self.build_package()
        value = json.loads(capture.read_text())
        for row in value["allRows"]:
            row["brand"] = "Fortune"
        for item in value["perPin"]:
            for row in item["rows"]:
                row["brand"] = "Fortune"
        capture.write_text(json.dumps(value))
        self.refresh_package(run, receipt, book, capture)
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("brand coverage", result.stderr)

    def test_rows_concentrated_on_one_pin_are_rejected(self) -> None:
        run, receipt, book, capture = self.build_package()
        value = json.loads(capture.read_text())
        first_pin = value["perPin"][0]["pincode"]
        for row in value["allRows"]:
            row["pincode"] = first_pin
        for item in value["perPin"]:
            for row in item["rows"]:
                row["pincode"] = first_pin
        capture.write_text(json.dumps(value))
        self.refresh_package(run, receipt, book, capture)
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("geo identity differs", result.stderr)

    def test_per_pin_flattened_rows_and_counters_are_authoritative(self) -> None:
        mutations = {
            "duplicate pin": lambda value: value["perPin"][1].update(
                {"pincode": value["perPin"][0]["pincode"]}
            ),
            "flattened mismatch": lambda value: value["perPin"][0]["rows"][0].update(
                {"sale": 99, "discount_pct": 10}
            ),
            "blocked mismatch": lambda value: value["perPin"][0].update({"blocked": "HTTP 429"}),
            "metadata mismatch": lambda value: value["perPin"][0].update({"city": "Wrong City"}),
        }
        for index, (label, mutate) in enumerate(mutations.items(), 1):
            with self.subTest(label=label):
                run, receipt, book, capture = self.build_package(seconds=f"0001{index:02d}")
                value = json.loads(capture.read_text())
                mutate(value)
                capture.write_text(json.dumps(value))
                self.refresh_package(run, receipt, book, capture)
                result = self.consume()
                self.assertNotEqual(result.returncode, 0)
                self.assertFalse((self.output / book.name).exists())

    def test_workbook_same_count_with_changed_pin_is_rejected(self) -> None:
        run, receipt, book, capture = self.build_package()
        workbook = load_workbook(book)
        workbook["Master Data"]["J2"] = "999999"
        workbook.save(book)
        self.refresh_package(run, receipt, book, capture)
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("Master Data content differs", result.stderr)

    def test_attempt_must_match_run_suffix(self) -> None:
        run, receipt, _, _ = self.build_package()
        receipt["attempt_id"] = "99"
        (run / "report.ready.json").write_text(json.dumps(receipt))
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("attempt identity", result.stderr)

    def test_receipt_first_then_exact_artifact_arrival_recovers(self) -> None:
        run, _, book, capture = self.build_package()
        saved = capture.read_bytes()
        capture.unlink()
        first = self.consume()
        self.assertNotEqual(first.returncode, 0)
        capture.write_bytes(saved)
        second = self.consume()
        self.assertEqual(second.returncode, 0, second.stderr)
        self.assertEqual(json.loads(second.stdout)["new"], 1)
        self.assertTrue((self.output / book.name).is_file())

    def test_rejection_alert_remains_pending_until_exact_ack(self) -> None:
        run, receipt, _, _ = self.build_package()
        receipt["schema"] = "bad"
        (run / "report.ready.json").write_text(json.dumps(receipt))
        first = self.consume()
        first_summary = json.loads(first.stdout)
        self.assertEqual(len(first_summary["rejection_alerts"]), 1)
        second = self.consume()
        second_summary = json.loads(second.stdout)
        self.assertEqual(len(second_summary["rejection_alerts"]), 1)
        alert = second_summary["rejection_alerts"][0]
        marker = Path(alert["alert_marker"])
        marker.write_text(json.dumps({
            "status": "alerted",
            "source_receipt_sha256": alert["source_receipt_sha256"],
            "package_state_sha256": alert["package_state_sha256"],
        }))
        third = self.consume()
        self.assertEqual(json.loads(third.stdout)["rejection_alerts"], [])

    def test_poison_package_does_not_block_valid_package(self) -> None:
        bad, receipt, _, _ = self.build_package(seconds="000001")
        receipt["schema"] = "bad"
        (bad / "report.ready.json").write_text(json.dumps(receipt))
        _, _, good_book, _ = self.build_package(seconds="000002")
        result = self.consume()
        self.assertNotEqual(result.returncode, 0)
        self.assertEqual(json.loads(result.stdout)["new"], 1)
        self.assertTrue((self.output / good_book.name).is_file())

    def test_wrapper_has_no_scrape_merge_or_main_mailer(self) -> None:
        wrapper = (ROOT / "tools/cron/consume_direct_competitor_reports.sh").read_text()
        self.assertNotIn("mail_price_data.sh", wrapper)
        self.assertNotIn("run_direct_competitor", wrapper)
        self.assertNotIn("scrape.js", wrapper)


if __name__ == "__main__":
    unittest.main()
