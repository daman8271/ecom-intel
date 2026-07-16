#!/usr/bin/env python3
"""Validate and promote final Mac-built competitor packages only.

This consumer never accepts endpoint shards and never performs capture, merge, or
workbook construction.  Its accepted receipt is the delivery authorization.
"""

from __future__ import annotations

import argparse
from collections import Counter
import datetime as dt
import functools
import hashlib
import json
import math
import os
from pathlib import Path
import re
import shlex
import shutil
import subprocess
import sys
from typing import Any

from openpyxl import load_workbook


IST = dt.timezone(dt.timedelta(hours=5, minutes=30))
HEX64 = re.compile(r"^[0-9a-f]{64}$")
RUN_ID = re.compile(
    r"^(?P<day>[0-9]{8})-[0-9]{6}-(?P<platform>blinkit|zepto)-competitor-direct-a(?P<attempt>[0-9]{2})$"
)
SCHEMA = "jivo-direct-competitor-report-receipt-v1"
PROMOTION_SCHEMA = "jivo-direct-competitor-promotion-receipt-v1"
FAILURE_SCHEMA = "jivo-direct-competitor-failure-receipt-v1"
FAILURE_ACCEPTED_SCHEMA = "jivo-direct-competitor-failure-accepted-v1"
REPO_ROOT = Path(__file__).resolve().parents[2]
COMPETITOR_ROOT = REPO_ROOT / "tools/competitor"
ANCHOR_BRANDS = {"jivo", "sano"}
BRAND_ALIASES = {"oriel": "oreal"}
MASTER_DATA_HEADER = [
    "Platform", "Brand", "JIVO?", "Search Category (scraped)", "Oil Type (name)",
    "Grade (name)", "Matched Anchor", "Blend?", "City", "Pincode", "Store ID",
    "Name", "Pack", "Vol (ml)", "MRP Rs", "Sale Rs", "Rs/L", "Discount %",
    "In stock", "Rank", "Ad?", "Captured",
]

PLATFORM = {
    "blinkit": {
        "workflow": "blinkit-top8",
        "pins": 75,
        "label": "Blinkit",
        "sheets": ["Summary", "City-Pin-SKU Prices", "Run Scope", "Anchor Watch", "Master Data"],
        "code": {
            "select_blinkit_top8_pincodes.py",
            "build_blinkit_top8_daily.py",
            "build_competitor_report.py",
        },
        "brands": {
            "borges", "del monte", "figaro", "fortune", "gulab",
            "hudson", "oreal", "saffola", "sundrop", "tata",
        },
    },
    "zepto": {
        "workflow": "zepto-competitor",
        "pins": 25,
        "label": "Zepto",
        "sheets": ["Summary", "Anchor Watch", "Master Data"],
        "code": {"build_competitor_report.py"},
    },
}

REQUIRED_SUPPORT = {
    "category_queries.json",
    "competitor_brands.json",
    "competitor_match_map.json",
    "maps_to_jivo.json",
    "oil_classifier.json",
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def atomic_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.tmp.{os.getpid()}")
    with temp.open("w", encoding="ascii") as handle:
        json.dump(value, handle, indent=2, sort_keys=True, ensure_ascii=True)
        handle.write("\n")
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp, path)


def copy_fsync(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with source.open("rb") as incoming, target.open("wb") as outgoing:
        shutil.copyfileobj(incoming, outgoing, 1024 * 1024)
        outgoing.flush()
        os.fsync(outgoing.fileno())


def normalized_brands(value: Any) -> list[str]:
    if not isinstance(value, list):
        raise ValueError("brand_set must be a list")
    brands = sorted({" ".join(str(item).split()).casefold() for item in value if str(item).strip()})
    if len(brands) != len(value):
        raise ValueError("brand_set contains blank or duplicate names")
    return brands


def brand_hash(brands: list[str]) -> str:
    raw = json.dumps(brands, ensure_ascii=True, separators=(",", ":")).encode("ascii")
    return hashlib.sha256(raw).hexdigest()


def valid_hash(value: Any) -> bool:
    return isinstance(value, str) and HEX64.fullmatch(value) is not None


def validate_hash_map(receipt: dict[str, Any], name: str) -> dict[str, str]:
    values = receipt.get(name)
    if not isinstance(values, dict) or set(values) != {"macpro", "windows"}:
        raise ValueError(f"{name} must bind macpro and windows")
    if any(not valid_hash(value) for value in values.values()):
        raise ValueError(f"{name} contains an invalid hash")
    return values


def validate_named_hashes(value: Any, field: str, required: set[str]) -> list[dict[str, str]]:
    if not isinstance(value, list) or not value:
        raise ValueError(f"{field} manifest is missing")
    output: list[dict[str, str]] = []
    names: set[str] = set()
    for item in value:
        if not isinstance(item, dict):
            raise ValueError(f"{field} entry is invalid")
        name = item.get("name")
        digest = item.get("sha256")
        if not isinstance(name, str) or Path(name).name != name or name in names:
            raise ValueError(f"{field} name is unsafe or duplicated")
        if not valid_hash(digest):
            raise ValueError(f"{field} hash is invalid: {name}")
        names.add(name)
        output.append({"name": name, "sha256": digest})
    if not required.issubset(names):
        raise ValueError(f"{field} is missing required files: {sorted(required - names)}")
    return output


def validate_local_manifest(
    value: Any, field: str, required: set[str], root: Path
) -> list[dict[str, str]]:
    manifest = validate_named_hashes(value, field, required)
    for item in manifest:
        path = root / item["name"]
        if path.is_symlink() or not path.is_file() or sha256_file(path) != item["sha256"]:
            raise ValueError(f"{field} does not match trusted control file: {item['name']}")
    return manifest


def normalize_brand(value: Any) -> str:
    brand = " ".join(str(value or "").split()).casefold()
    return BRAND_ALIASES.get(brand, brand)


def capture_scope(receipt: dict[str, Any]) -> tuple[set[str], set[str], dict[str, str]]:
    anchors = set(normalized_brands(receipt.get("anchor_brands")))
    competitors = set(normalized_brands(receipt.get("competitor_brands")))
    capture = set(normalized_brands(receipt.get("capture_brands")))
    if anchors != ANCHOR_BRANDS:
        raise ValueError("anchor_brands must be exactly Jivo and Sano")
    if competitors != set(receipt["brand_set"]):
        raise ValueError("competitor_brands differs from the reviewed brand_set")
    if capture != anchors | competitors:
        raise ValueError("capture_brands must equal anchors plus reviewed competitors")
    aliases = receipt.get("brand_aliases")
    if not isinstance(aliases, dict) or any(
        not isinstance(key, str) or not isinstance(value, str)
        for key, value in aliases.items()
    ):
        raise ValueError("brand_aliases is invalid")
    normalized_aliases = {
        " ".join(key.split()).casefold(): " ".join(value.split()).casefold()
        for key, value in aliases.items()
    }
    if any(target not in capture for target in normalized_aliases.values()):
        raise ValueError("brand_aliases targets a brand outside capture_brands")
    return competitors, capture, normalized_aliases


def normalize_row_brand(value: Any, aliases: dict[str, str]) -> str:
    brand = normalize_brand(value)
    return aliases.get(brand, brand)


def row_digest(row: Any) -> str:
    return json.dumps(row, ensure_ascii=True, sort_keys=True, separators=(",", ":"))


def finite_number(value: Any) -> float:
    if isinstance(value, bool):
        raise ValueError("boolean is not a numeric price")
    number = float(value)
    if not math.isfinite(number):
        raise ValueError("non-finite numeric value")
    return number


def captured_on_date(value: Any, date_ist: str) -> bool:
    try:
        parsed = dt.datetime.fromisoformat(str(value or "").replace("Z", "+00:00"))
    except ValueError:
        return False
    if parsed.tzinfo is None:
        return False
    return parsed.astimezone(IST).date().isoformat() == date_ist


def package_state_sha(run_dir: Path) -> str:
    state: list[dict[str, Any]] = []
    for path in sorted(run_dir.iterdir(), key=lambda item: item.name):
        item: dict[str, Any] = {"name": path.name}
        if path.is_symlink():
            item["kind"] = "symlink"
        elif path.is_file():
            try:
                item.update({"kind": "file", "bytes": path.stat().st_size, "sha256": sha256_file(path)})
            except OSError:
                item["kind"] = "unreadable"
        else:
            item["kind"] = "other"
        state.append(item)
    payload = json.dumps(state, ensure_ascii=True, sort_keys=True, separators=(",", ":")).encode("ascii")
    return hashlib.sha256(payload).hexdigest()


def alert_marker_matches(marker: Path, source_sha: str, package_sha: str | None = None) -> bool:
    if marker.is_symlink() or not marker.is_file():
        return False
    try:
        value = load_json(marker)
    except (OSError, ValueError, json.JSONDecodeError):
        return False
    return value.get("status") == "alerted" \
        and value.get("source_receipt_sha256") == source_sha \
        and (package_sha is None or value.get("package_state_sha256") == package_sha)


def expected_names(platform: str, date_ist: str) -> tuple[str, str]:
    spec = PLATFORM[platform]
    workbook = f"Competitor-Price-Watch-{spec['label']}-{date_ist}.xlsx"
    capture = f"{platform}_competitor_{date_ist}.json"
    return workbook, capture


def delivery_audit_name(platform: str, date_ist: str) -> str:
    if platform == "blinkit":
        return f"blinkit-top8-{date_ist}.audit.json"
    return f"zepto-competitor-{date_ist}.audit.json"


def validate_manifest_entry(value: Any, expected_name: str, field: str) -> dict[str, Any]:
    if not isinstance(value, dict) or value.get("name") != expected_name:
        raise ValueError(f"{field} name mismatch")
    if Path(expected_name).name != expected_name:
        raise ValueError(f"unsafe {field} name")
    if not isinstance(value.get("bytes"), int) or value["bytes"] <= 0 or not valid_hash(value.get("sha256")):
        raise ValueError(f"{field} size/hash is invalid")
    return {"name": expected_name, "bytes": value["bytes"], "sha256": value["sha256"]}


def verify_file(run_dir: Path, entry: dict[str, Any], minimum: int = 1) -> Path:
    path = run_dir / entry["name"]
    if path.is_symlink() or not path.is_file() or path.stat().st_size < minimum:
        raise ValueError(f"artifact missing or too small: {entry['name']}")
    if path.stat().st_size != entry["bytes"] or sha256_file(path) != entry["sha256"]:
        raise ValueError(f"artifact hash/size mismatch: {entry['name']}")
    return path


def validate_quality_policy(platform: str, policy: Any) -> dict[str, Any]:
    if not isinstance(policy, dict):
        raise ValueError("quality_policy is missing")
    if platform == "blinkit":
        exact = {
            "require_resolved": True,
            "require_auth": True,
            "require_rows_each_pincode": True,
        }
        if any(policy.get(key) != value for key, value in exact.items()):
            raise ValueError("Blinkit quality_policy is incomplete")
    elif float(policy.get("min_serviceable_pct") or 0) <= 0:
        raise ValueError("Zepto quality_policy has no serviceability floor")
    for key in ("min_rows_per_source_pincode", "min_unique_brands", "baseline_min_row_fraction", "baseline_min_brand_fraction"):
        if float(policy.get(key) or 0) <= 0:
            raise ValueError(f"quality_policy has no positive {key}")
    return policy


def validate_capture(path: Path, receipt: dict[str, Any], platform: str) -> dict[str, Any]:
    capture = load_json(path)
    if not isinstance(capture, dict):
        raise ValueError("merged capture is not an object")
    summary = capture.get("summary")
    per_pin = capture.get("perPin")
    rows = capture.get("allRows")
    if not isinstance(summary, dict) or not isinstance(per_pin, list) \
       or not isinstance(rows, list) or not rows:
        raise ValueError("merged capture summary/perPin/allRows is invalid")
    competitors, capture_brands, aliases = capture_scope(receipt)
    pins = [str(item.get("pincode") or "").strip() for item in per_pin if isinstance(item, dict)]
    expected_pins = set(pins)
    if len(per_pin) != PLATFORM[platform]["pins"] or len(expected_pins) != len(per_pin) or "" in expected_pins:
        raise ValueError("merged capture perPin membership is not exact")
    flattened: list[dict[str, Any]] = []
    blocked = resolved = authenticated = serviceable = with_rows = 0
    metadata: dict[str, dict[str, Any]] = {}
    for item in per_pin:
        if not isinstance(item, dict) or not isinstance(item.get("rows"), list):
            raise ValueError("merged capture perPin entry is invalid")
        pin = str(item.get("pincode") or "").strip()
        if any(field not in item for field in ("city", "locality", "lat", "lon")):
            raise ValueError(f"merged capture perPin metadata is incomplete: {pin}")
        try:
            finite_number(item["lat"])
            finite_number(item["lon"])
        except (TypeError, ValueError) as exc:
            raise ValueError(f"merged capture perPin coordinates are invalid: {pin}") from exc
        metadata[pin] = item
        pin_rows = item["rows"]
        flattened.extend(pin_rows)
        blocked += int(bool(item.get("blocked") or item.get("partial_block")))
        resolved += int(bool(item.get("resolved")))
        authenticated += int(item.get("auth_accepted") == 1)
        serviceable += int(bool(item.get("serviceable")))
        with_rows += int(bool(pin_rows))
    checks = {
        "mode": summary.get("mode") == "competitor",
        "platform": summary.get("platform") == platform,
        "date": summary.get("date_ist") == receipt["date_ist"],
        "run": summary.get("run_id") == receipt["run_id"],
        "pins": summary.get("pincodes_total") == PLATFORM[platform]["pins"],
        "rows": summary.get("total_rows") == len(rows) == receipt["total_rows"],
        "with_rows": int(summary.get("pincodes_with_rows") or 0) == with_rows,
        "blocked": int(summary.get("pincodes_blocked") or 0) == blocked == 0,
        "partial": summary.get("partial") is False and capture.get("partial") is False,
        "captured_at": captured_on_date(summary.get("captured_at"), receipt["date_ist"]),
    }
    if not all(checks.values()):
        raise ValueError(f"merged capture identity/quality failed: {checks}")
    if Counter(map(row_digest, flattened)) != Counter(map(row_digest, rows)):
        raise ValueError("merged capture perPin rows do not equal allRows")
    summary_scope = summary.get("scope")
    if not isinstance(summary_scope, dict) \
       or set(normalized_brands(summary_scope.get("anchors"))) != ANCHOR_BRANDS \
       or set(normalized_brands(summary_scope.get("competitors"))) != competitors \
       or set(normalized_brands(summary_scope.get("capture_brands"))) != capture_brands:
        raise ValueError("merged capture summary brand scope differs from receipt")

    actual_competitors: set[str] = set()
    actual_brands: set[str] = set()
    listing_keys: list[tuple[str, str, str, str]] = []
    for index, row in enumerate(rows):
        if not isinstance(row, dict):
            raise ValueError(f"merged capture row is not an object: {index}")
        pin = str(row.get("pincode") or "").strip()
        item = metadata.get(pin)
        brand = normalize_row_brand(row.get("brand"), aliases)
        required = ("platform", "city", "pincode", "brand", "canonical", "sale", "mrp", "discount_pct", "in_stock")
        if item is None or any(field not in row for field in required):
            raise ValueError(f"merged capture row membership/fields are invalid: {index}")
        if str(row.get("platform") or "").casefold() != platform \
           or str(row.get("city") or "").strip() != str(item.get("city") or "").strip():
            raise ValueError(f"merged capture row geo identity differs from perPin: {index}")
        if brand not in capture_brands:
            raise ValueError(f"merged capture row brand is outside capture scope: {index}")
        actual_brands.add(brand)
        if brand in competitors:
            actual_competitors.add(brand)
        if not captured_on_date(row.get("captured_at"), receipt["date_ist"]):
            raise ValueError(f"merged capture row date is invalid: {index}")
        try:
            sale = finite_number(row.get("sale"))
            mrp = finite_number(row.get("mrp"))
            discount = finite_number(row.get("discount_pct"))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"merged capture row price is invalid: {index}") from exc
        expected_discount = max(0.0, (mrp - sale) * 100 / mrp) if mrp > 0 else -1
        if sale <= 0 or mrp <= 0 or mrp < sale - 0.5 or not 0 <= discount <= 100 \
           or abs(discount - expected_discount) > 0.25:
            raise ValueError(f"merged capture row price arithmetic failed: {index}")
        if row.get("in_stock") not in (0, 1, False, True):
            raise ValueError(f"merged capture row stock value is invalid: {index}")
        listing_id = str(row.get("prid") or row.get("product_id") or row.get("variant_id") \
                         or row.get("canonical") or "").strip()
        listing_keys.append((pin, str(row.get("canonical") or ""), listing_id, brand))
    if len(listing_keys) != len(set(listing_keys)):
        raise ValueError("merged capture contains duplicate listing rows")
    minimum_brands = int(receipt["quality_policy"]["min_unique_brands"])
    if len(actual_competitors) < minimum_brands:
        raise ValueError(
            f"merged capture competitor brand coverage {len(actual_competitors)} is below {minimum_brands}"
        )
    if int(summary.get("unique_brands") or -1) != len(actual_brands):
        raise ValueError("merged capture unique brand count differs from rows")
    baseline = receipt["baseline"]
    if len(rows) < math.ceil(
        int(baseline["total_rows"]) * float(receipt["quality_policy"]["baseline_min_row_fraction"])
    ):
        raise ValueError("merged capture rows collapsed against baseline")
    if len(actual_brands) < math.ceil(
        int(baseline["unique_brands"]) * float(receipt["quality_policy"]["baseline_min_brand_fraction"])
    ):
        raise ValueError("merged capture brands collapsed against baseline")
    if platform == "blinkit":
        if with_rows != 75 or resolved != 75 or authenticated != 75 \
           or summary.get("pincodes_resolved") != 75 \
           or summary.get("auth_verified") != 1 or summary.get("auth_verified_pincodes") != 75:
            raise ValueError("Blinkit capture is not 75/75 resolved and authenticated")
    else:
        minimum = math.ceil(PLATFORM[platform]["pins"] * float(receipt["quality_policy"]["min_serviceable_pct"]) / 100)
        row_floor = math.ceil(
            PLATFORM[platform]["pins"] * float(receipt["quality_policy"]["min_rows_per_source_pincode"])
        )
        if int(summary.get("pincodes_serviceable") or 0) != serviceable or serviceable < minimum:
            raise ValueError("Zepto capture is below the serviceability floor")
        if len(rows) < row_floor:
            raise ValueError("Zepto capture is below the total-row floor")
    return capture


@functools.lru_cache(maxsize=1)
def workbook_brand_config() -> tuple[list[str], dict[str, str]]:
    config = load_json(COMPETITOR_ROOT / "competitor_brands.json")
    ours = [str(item).casefold() for item in config.get("ours", ["jivo", "sano"])]
    aliases: dict[str, str] = {}
    for item in config.get("brands") or []:
        canonical = str(item.get("brand") or "")
        aliases[canonical.casefold()] = canonical
        for alias in item.get("aliases") or []:
            if alias:
                aliases[str(alias).casefold()] = canonical
    return ours, aliases


def workbook_brand_label(row: dict[str, Any]) -> str:
    ours, aliases = workbook_brand_config()
    raw = str(row.get("brand") or "").strip()
    raw_key = raw.casefold()
    name = str(row.get("name") or "").casefold()
    if any(item in raw_key or item in name for item in ours):
        return "JIVO"
    if raw_key in aliases:
        return aliases[raw_key]
    for alias, canonical in aliases.items():
        if alias and alias in name:
            return canonical
    return raw.title() if raw else "Unknown"


def workbook_value(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return round(float(value), 6)
    return str(value)


def workbook_projection(row: dict[str, Any], platform: str) -> tuple[Any, ...]:
    discount = row.get("discount_pct")
    return tuple(map(workbook_value, (
        platform.replace("-", " ").title(),
        workbook_brand_label(row),
        row.get("city"),
        str(row.get("pincode") or ""),
        row.get("store_id"),
        row.get("name"),
        row.get("pack"),
        row.get("vol_ml"),
        row.get("mrp"),
        row.get("sale"),
        row.get("per_litre"),
        round(float(discount) / 100, 4) if discount is not None else None,
        "Yes" if row.get("in_stock") else "No",
        row.get("rank"),
        "Yes" if row.get("is_ad") else "",
        str(row.get("captured_at") or "")[:16].replace("T", " "),
    )))


def validate_workbook(
    path: Path, platform: str, receipt: dict[str, Any], capture: dict[str, Any]
) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        expected = PLATFORM[platform]["sheets"]
        if workbook.sheetnames != expected:
            raise ValueError(f"workbook sheets mismatch: {workbook.sheetnames}")
        master = workbook["Master Data"]
        header = [cell.value for cell in next(master.iter_rows(min_row=1, max_row=1))]
        if header != MASTER_DATA_HEADER:
            raise ValueError("workbook Master Data header mismatch")
        if master.max_row - 1 != receipt["total_rows"]:
            raise ValueError("workbook Master Data row count mismatch")
        actual = Counter(
            tuple(map(workbook_value, (
                values[0], values[1], values[8], str(values[9] or ""), values[10], values[11],
                values[12], values[13], values[14], values[15], values[16], values[17],
                values[18], values[19], values[20], values[21],
            )))
            for values in master.iter_rows(min_row=2, values_only=True)
        )
        expected_rows = Counter(workbook_projection(row, platform) for row in capture["allRows"])
        if actual != expected_rows:
            raise ValueError("workbook Master Data content differs from merged capture")
        if platform == "blinkit":
            scope = workbook["Run Scope"]
            if scope.max_row != 82:
                raise ValueError("Blinkit Run Scope row count mismatch")
            pin_map = {str(item["pincode"]): item for item in capture["perPin"]}
            seen: set[str] = set()
            for values in scope.iter_rows(min_row=8, max_row=82, values_only=True):
                pin = str(values[1] or "")
                item = pin_map.get(pin)
                if item is None or pin in seen:
                    raise ValueError("Blinkit Run Scope pincode membership mismatch")
                seen.add(pin)
                if str(values[0] or "") != str(item.get("city") or "") \
                   or str(values[2] or "") != str(item.get("locality") or "") \
                   or values[4] != "Yes" or values[5] != "Yes" \
                   or int(values[6] or -1) != len(item.get("rows") or []):
                    raise ValueError(f"Blinkit Run Scope row differs from capture: {pin}")
            if seen != set(pin_map):
                raise ValueError("Blinkit Run Scope coverage mismatch")
    finally:
        workbook.close()


def validate_audits(
    run_dir: Path,
    receipt: dict[str, Any],
    platform: str,
    workbook: dict[str, Any],
    capture: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    values = receipt.get("audits")
    if not isinstance(values, list) or not values:
        raise ValueError("audits manifest is missing")
    verified: list[dict[str, Any]] = []
    names: set[str] = set()
    parsed: dict[str, dict[str, Any]] = {}
    for value in values:
        if not isinstance(value, dict) or not isinstance(value.get("name"), str):
            raise ValueError("audit manifest entry is invalid")
        name = value["name"]
        if Path(name).name != name or name in names:
            raise ValueError("audit name is unsafe or duplicated")
        entry = validate_manifest_entry(value, name, "audit")
        path = verify_file(run_dir, entry)
        try:
            audit = load_json(path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            raise ValueError(f"audit is not valid JSON: {name}") from exc
        if not isinstance(audit, dict):
            raise ValueError(f"audit is not an object: {name}")
        names.add(name)
        parsed[name] = audit
        verified.append(entry)
    required_name = delivery_audit_name(platform, receipt["date_ist"])
    if required_name not in parsed or "quality-audit.json" not in parsed:
        raise ValueError(f"required delivery audit is missing: {required_name}")
    audit = parsed[required_name]
    quality = parsed["quality-audit.json"]
    brands = receipt["brand_set"]
    common = {
        "schema": audit.get("schema") == "jivo-direct-competitor-delivery-audit-v1",
        "platform": audit.get("platform") == platform,
        "workflow": audit.get("workflow_kind") == receipt["workflow_kind"],
        "date": audit.get("date_ist") == receipt["date_ist"],
        "run": audit.get("run_id") == receipt["run_id"],
        "attempt": audit.get("attempt_id") == receipt["attempt_id"],
        "status": audit.get("status") == "OK",
        "brands": audit.get("brand_set") == brands,
        "brand_count": audit.get("brand_set_count") == len(brands),
        "brand_hash": audit.get("brand_set_sha256") == receipt["brand_set_sha256"],
        "pins": audit.get("pincodes_total") == receipt["pincodes_total"],
        "rows": audit.get("total_rows") == receipt["total_rows"],
        "capture": audit.get("merged_sha256") == receipt["merged_capture"]["sha256"],
        "workbook": audit.get("workbook_sha256") == workbook["sha256"],
    }
    if not all(common.values()):
        raise ValueError(f"delivery audit provenance/quality failed: {common}")
    quality_checks = {
        "schema": quality.get("schema") == "jivo-direct-competitor-quality-audit-v1",
        "plan": quality.get("plan_sha256") == receipt["plan_sha256"],
        "merge_receipt": quality.get("merge_receipt_sha256") == receipt["merge_receipt_sha256"],
        "capture": quality.get("merged_sha256") == receipt["merged_capture"]["sha256"],
        "workbook": quality.get("workbook_sha256") == workbook["sha256"],
        "platform": quality.get("platform") == platform,
        "workflow": quality.get("workflow_kind") == receipt["workflow_kind"],
        "date": quality.get("date_ist") == receipt["date_ist"],
        "run": quality.get("run_id") == receipt["run_id"],
        "attempt": quality.get("attempt_id") == receipt["attempt_id"],
        "verdict": quality.get("verdict") == "OK",
        "brands": quality.get("brand_set") == brands,
        "brand_count": quality.get("brand_set_count") == len(brands),
        "brand_hash": quality.get("brand_set_sha256") == receipt["brand_set_sha256"],
        "pins": quality.get("pincodes_total") == receipt["pincodes_total"],
        "rows": quality.get("total_rows") == receipt["total_rows"],
        "results": quality.get("input_result_sha256") == receipt["input_result_sha256"],
        "progress": quality.get("input_progress_sha256") == receipt["input_progress_sha256"],
        "terminals": quality.get("input_terminal_sha256") == receipt["input_terminal_sha256"],
        "support": quality.get("support_files") == receipt["support_files"],
        "code": quality.get("code_files") == receipt["code_files"],
        "policy": quality.get("quality_policy") == receipt["quality_policy"],
        "anchors": quality.get("anchor_brands") == receipt["anchor_brands"],
        "competitors": quality.get("competitor_brands") == receipt["competitor_brands"],
        "capture_brands": quality.get("capture_brands") == receipt["capture_brands"],
    }
    if not all(quality_checks.values()):
        raise ValueError(f"quality audit provenance/quality failed: {quality_checks}")
    if platform == "blinkit":
        if capture["summary"].get("pincodes_resolved") != 75 \
           or capture["summary"].get("auth_verified_pincodes") != 75:
            raise ValueError("Blinkit delivery audit capture coverage failed")
    return verified, next(item for item in verified if item["name"] == required_name)


def acknowledge(receipt_path: Path, run_id: str, host: str, root: str) -> str:
    if not host:
        return "disabled"
    remote = f"{root.rstrip('/')}/{run_id}.json"
    root_q, part_q, remote_q = map(shlex.quote, (root, f"{remote}.part", remote))
    try:
        subprocess.run(
            ["ssh", "-o", "BatchMode=yes", "-o", "ConnectTimeout=15", host, f"mkdir -p -- {root_q}"],
            check=True, timeout=30, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        subprocess.run(
            ["scp", "-q", "-o", "BatchMode=yes", str(receipt_path), f"{host}:{part_q}"],
            check=True, timeout=60, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        subprocess.run(
            ["ssh", "-o", "BatchMode=yes", host, f"mv -- {part_q} {remote_q}"],
            check=True, timeout=30, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        return "accepted"
    except (OSError, subprocess.SubprocessError):
        return "pending"


def acknowledge_once(receipt_path: Path, run_id: str, host: str, root: str) -> str:
    marker = receipt_path.with_suffix(".promotion-acked.json")
    if marker.is_file():
        return "accepted"
    status = acknowledge(receipt_path, run_id, host, root)
    if status == "accepted":
        atomic_json(marker, {"status": status, "run_id": run_id})
    return status


def destination_manifest(
    output_dir: Path,
    data_dir: Path,
    audit_dir: Path,
    platform: str,
    date_ist: str,
    workbook: dict[str, Any],
    capture: dict[str, Any],
    delivery_audit: dict[str, Any],
) -> list[dict[str, Any]]:
    values = [
        {**workbook, "kind": "workbook", "destination": str((output_dir / workbook["name"]).resolve())},
        {**capture, "kind": "merged_capture", "destination": str((data_dir / capture["name"]).resolve())},
    ]
    target = audit_dir / delivery_audit_name(platform, date_ist)
    values.append({**delivery_audit, "kind": "delivery_audit", "destination": str(target.resolve())})
    return values


def restore_or_verify(run_dir: Path, artifacts: list[dict[str, Any]]) -> int:
    restored = 0
    for item in artifacts:
        destination = Path(item["destination"])
        if destination.exists():
            if destination.is_symlink() or sha256_file(destination) != item["sha256"]:
                raise ValueError(f"accepted destination hash changed: {destination}")
            continue
        source = run_dir / item["name"]
        if not source.is_file() or sha256_file(source) != item["sha256"]:
            raise ValueError(f"accepted artifact cannot be restored: {item['name']}")
        temp = destination.with_name(f".{destination.name}.restore.part")
        try:
            copy_fsync(source, temp)
            if sha256_file(temp) != item["sha256"]:
                raise ValueError(f"restored hash mismatch: {item['name']}")
            os.replace(temp, destination)
            restored += 1
        finally:
            temp.unlink(missing_ok=True)
    return restored


def persist_source_receipt(
    source: Path, receipt_root: Path, date_ist: str, run_id: str, expected_sha: str
) -> dict[str, Any]:
    target = receipt_root / date_ist / "source" / f"{run_id}.report.ready.json"
    target.parent.mkdir(parents=True, exist_ok=True)
    expected_size = source.stat().st_size
    if target.exists():
        if target.is_symlink() or target.stat().st_size != expected_size \
           or sha256_file(target) != expected_sha:
            raise ValueError("persisted competitor source receipt changed")
    else:
        temp = target.with_name(f".{target.name}.part.{os.getpid()}")
        try:
            copy_fsync(source, temp)
            if temp.stat().st_size != expected_size or sha256_file(temp) != expected_sha:
                raise ValueError("persisted competitor source receipt copy mismatch")
            os.replace(temp, target)
        finally:
            temp.unlink(missing_ok=True)
    return {
        "destination": str(target.resolve()),
        "bytes": expected_size,
        "sha256": expected_sha,
    }


def consume_run(
    run_dir: Path,
    date_ist: str,
    output_dir: Path,
    data_dir: Path,
    audit_dir: Path,
    receipt_root: Path,
    stable_age: int,
    ack_host: str,
    ack_root: str,
) -> tuple[str, str]:
    source_receipt = run_dir / "report.ready.json"
    age = dt.datetime.now().timestamp() - source_receipt.stat().st_mtime
    if age < stable_age:
        return "waiting", f"receipt is only {int(age)}s old"
    if run_dir.is_symlink() or source_receipt.is_symlink():
        raise ValueError("symlinked inbox paths are prohibited")
    match = RUN_ID.fullmatch(run_dir.name)
    if not match:
        raise ValueError("unsafe competitor run_id")
    receipt_bytes = source_receipt.read_bytes()
    receipt_sha = hashlib.sha256(receipt_bytes).hexdigest()
    receipt = json.loads(receipt_bytes.decode("utf-8-sig"))
    platform = match.group("platform")
    spec = PLATFORM[platform]
    expected = {
        "schema": SCHEMA,
        "platform": platform,
        "workflow_kind": spec["workflow"],
        "date_ist": date_ist,
        "run_id": run_dir.name,
        "status": "ready",
        "review_verdict": "OK",
        "pincodes_total": spec["pins"],
    }
    for key, value in expected.items():
        if receipt.get(key) != value:
            raise ValueError(f"competitor receipt mismatch: {key}")
    if match.group("day") != date_ist.replace("-", "") \
       or receipt.get("attempt_id") != match.group("attempt"):
        raise ValueError("competitor run date/attempt identity is invalid")
    for field in ("plan_sha256", "source_sha256", "scraper_sha256", "merge_receipt_sha256"):
        if not valid_hash(receipt.get(field)):
            raise ValueError(f"{field} is missing or invalid")
    if not isinstance(receipt.get("total_rows"), int) or receipt["total_rows"] <= 0:
        raise ValueError("total_rows is invalid")
    receipt["quality_policy"] = validate_quality_policy(platform, receipt.get("quality_policy"))
    validate_hash_map(receipt, "input_result_sha256")
    validate_hash_map(receipt, "input_progress_sha256")
    validate_hash_map(receipt, "input_terminal_sha256")
    receipt["support_files"] = validate_local_manifest(
        receipt.get("support_files"), "support_files", REQUIRED_SUPPORT, COMPETITOR_ROOT
    )
    receipt["code_files"] = validate_local_manifest(
        receipt.get("code_files"), "code_files", spec["code"], COMPETITOR_ROOT
    )
    trusted_scraper = REPO_ROOT / f"platforms/{platform}/scrape.js"
    if trusted_scraper.is_symlink() or not trusted_scraper.is_file() \
       or sha256_file(trusted_scraper) != receipt["scraper_sha256"]:
        raise ValueError("scraper_sha256 differs from the trusted control revision")

    brands = normalized_brands(receipt.get("brand_set"))
    if len(brands) != receipt.get("brand_set_count") or brand_hash(brands) != receipt.get("brand_set_sha256"):
        raise ValueError("brand_set count/hash mismatch")
    if {"jivo", "sano"} & set(brands):
        raise ValueError("brand_set must be rival-only")
    if platform == "blinkit" and set(brands) != spec["brands"]:
        raise ValueError("Blinkit approved rival brand set mismatch")
    if platform == "zepto" and len(brands) < 8:
        raise ValueError("Zepto rival brand set is unexpectedly small")
    receipt["brand_set"] = brands
    capture_scope(receipt)
    baseline = receipt.get("baseline")
    if not isinstance(baseline, dict) or not valid_hash(baseline.get("capture_sha256")) \
       or not isinstance(baseline.get("capture_bytes"), int) or baseline["capture_bytes"] <= 0 \
       or int(baseline.get("total_rows") or 0) <= 0 \
       or int(baseline.get("unique_brands") or 0) <= 0:
        raise ValueError("baseline provenance is missing or invalid")

    workbook_name, capture_name = expected_names(platform, date_ist)
    workbooks = receipt.get("workbooks")
    if not isinstance(workbooks, list) or len(workbooks) != 1:
        raise ValueError("workbooks manifest must contain exactly one workbook")
    workbook = validate_manifest_entry(workbooks[0], workbook_name, "workbook")
    capture_entry = validate_manifest_entry(receipt.get("merged_capture"), capture_name, "merged_capture")
    if receipt.get("merged_sha256") != capture_entry["sha256"] or receipt.get("merged_bytes") != capture_entry["bytes"]:
        raise ValueError("legacy merged hash/size fields disagree with merged_capture")
    workbook_path = verify_file(run_dir, workbook, minimum=10_000)
    capture_path = verify_file(run_dir, capture_entry)
    capture = validate_capture(capture_path, receipt, platform)
    validate_workbook(workbook_path, platform, receipt, capture)
    audits, delivery_audit = validate_audits(run_dir, receipt, platform, workbook, capture)
    artifacts = destination_manifest(
        output_dir, data_dir, audit_dir, platform, date_ist,
        workbook, capture_entry, delivery_audit,
    )
    promotion_provenance = {
        "platform": platform,
        "workflow_kind": spec["workflow"],
        "date_ist": date_ist,
        "run_id": run_dir.name,
        "attempt_id": receipt["attempt_id"],
        "plan_sha256": receipt["plan_sha256"],
        "source_sha256": receipt["source_sha256"],
        "scraper_sha256": receipt["scraper_sha256"],
        "merge_receipt_sha256": receipt["merge_receipt_sha256"],
        "source_receipt_sha256": receipt_sha,
        "merged_sha256": capture_entry["sha256"],
        "merged_bytes": capture_entry["bytes"],
        "input_result_sha256": receipt["input_result_sha256"],
        "input_progress_sha256": receipt["input_progress_sha256"],
        "input_terminal_sha256": receipt["input_terminal_sha256"],
        "support_files": receipt["support_files"],
        "code_files": receipt["code_files"],
        "quality_policy": receipt["quality_policy"],
        "baseline": receipt["baseline"],
        "anchor_brands": receipt["anchor_brands"],
        "competitor_brands": receipt["competitor_brands"],
        "capture_brands": receipt["capture_brands"],
        "brand_aliases": receipt["brand_aliases"],
        "pincodes_total": receipt["pincodes_total"],
        "total_rows": receipt["total_rows"],
        "brand_set": receipt["brand_set"],
        "brand_set_count": receipt["brand_set_count"],
        "brand_set_sha256": receipt["brand_set_sha256"],
    }

    accepted_path = receipt_root / date_ist / f"{run_dir.name}.json"
    if accepted_path.is_file():
        if accepted_path.is_symlink():
            raise ValueError("symlinked accepted promotion is prohibited")
        accepted = load_json(accepted_path)
        if accepted.get("schema") != PROMOTION_SCHEMA or accepted.get("source_receipt_sha256") != receipt_sha:
            raise ValueError("accepted promotion conflicts with source receipt")
        if any(accepted.get(key) != value for key, value in promotion_provenance.items()):
            raise ValueError("accepted promotion provenance changed")
        if accepted.get("artifacts") != artifacts:
            raise ValueError("accepted promotion artifact manifest changed")
        evidence = accepted.get("source_receipt")
        if not isinstance(evidence, dict):
            raise ValueError("accepted promotion source receipt evidence is missing")
        evidence_path = Path(str(evidence.get("destination") or ""))
        if evidence_path.is_symlink() or not evidence_path.is_file() \
           or evidence_path.stat().st_size != evidence.get("bytes") \
           or sha256_file(evidence_path) != evidence.get("sha256") \
           or evidence.get("sha256") != receipt_sha:
            raise ValueError("accepted promotion source receipt evidence changed")
        restored = restore_or_verify(run_dir, artifacts)
        ack = acknowledge_once(accepted_path, run_dir.name, ack_host, ack_root)
        return "existing", f"already accepted; restored={restored}; Mac acknowledgement={ack}"

    for item in artifacts:
        destination = Path(item["destination"])
        if destination.exists() and (destination.is_symlink() or sha256_file(destination) != item["sha256"]):
            raise ValueError(f"refusing to overwrite different competitor artifact: {destination}")

    staged: list[tuple[Path, Path, str]] = []
    try:
        for item in artifacts:
            source = run_dir / item["name"]
            destination = Path(item["destination"])
            temp = destination.with_name(f".{destination.name}.direct-{run_dir.name}.part")
            copy_fsync(source, temp)
            if sha256_file(temp) != item["sha256"]:
                raise ValueError(f"staged hash mismatch: {item['name']}")
            staged.append((temp, destination, item["sha256"]))
        for temp, destination, expected_sha in staged:
            if destination.exists() and sha256_file(destination) != expected_sha:
                raise ValueError(f"destination changed during promotion: {destination}")
            os.replace(temp, destination)
    finally:
        for temp, _, _ in staged:
            temp.unlink(missing_ok=True)

    source_evidence = persist_source_receipt(
        source_receipt, receipt_root, date_ist, run_dir.name, receipt_sha
    )
    accepted = {
        "schema": PROMOTION_SCHEMA,
        "status": "accepted",
        **promotion_provenance,
        "source_receipt": source_evidence,
        "artifacts": artifacts,
        "audits": audits,
        "accepted_at": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
    }
    atomic_json(accepted_path, accepted)
    ack = acknowledge_once(accepted_path, run_dir.name, ack_host, ack_root)
    return "accepted", f"promoted {len(artifacts)} artifact(s); Mac acknowledgement={ack}"


def consume_failure(source: Path, date_ist: str, failure_root: Path) -> tuple[str, dict[str, str]]:
    run_dir = source.parent
    match = RUN_ID.fullmatch(run_dir.name)
    if run_dir.is_symlink() or source.is_symlink() or not match:
        raise ValueError("unsafe competitor failure path")
    receipt = load_json(source)
    platform = match.group("platform")
    expected = {
        "schema": FAILURE_SCHEMA,
        "platform": platform,
        "workflow_kind": PLATFORM[platform]["workflow"],
        "date_ist": date_ist,
        "run_id": run_dir.name,
        "status": "failed",
    }
    for key, value in expected.items():
        if receipt.get(key) != value:
            raise ValueError(f"competitor failure receipt mismatch: {key}")
    if match.group("day") != date_ist.replace("-", "") \
       or receipt.get("attempt_id") != match.group("attempt"):
        raise ValueError("competitor failure date/attempt is invalid")
    if not valid_hash(receipt.get("plan_sha256")) or not str(receipt.get("phase") or "") \
       or not str(receipt.get("reason") or ""):
        raise ValueError("competitor failure provenance is incomplete")
    source_sha = sha256_file(source)
    target = failure_root / date_ist / f"{run_dir.name}.json"
    alert_marker = target.with_suffix(".alerted.json")
    detail = {
        "run_id": run_dir.name,
        "platform": platform,
        "workflow_kind": PLATFORM[platform]["workflow"],
        "attempt_id": receipt["attempt_id"],
        "plan_sha256": receipt["plan_sha256"],
        "source_receipt_sha256": source_sha,
        "phase": str(receipt["phase"]),
        "reason": str(receipt["reason"]),
        "alert_marker": str(alert_marker),
    }
    if target.is_file():
        if load_json(target).get("source_receipt_sha256") != source_sha:
            raise ValueError("accepted competitor failure receipt changed")
        if alert_marker.is_file():
            alerted = load_json(alert_marker)
            if alerted.get("source_receipt_sha256") == source_sha and alerted.get("status") == "alerted":
                return "existing", detail
        return "pending_alert", detail
    atomic_json(target, {
        "schema": FAILURE_ACCEPTED_SCHEMA,
        "status": "accepted",
        "date_ist": date_ist,
        "source_receipt_sha256": source_sha,
        **detail,
    })
    return "new", detail


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inbox", default="shards/mac-direct-competitor-ready")
    parser.add_argument("--output", default="output")
    parser.add_argument("--data", default="tools/competitor/data")
    parser.add_argument("--audit-dir", default="logs")
    parser.add_argument("--receipts", default="logs/direct-competitor-report-receipts")
    parser.add_argument("--failure-receipts", default="logs/direct-competitor-report-failures")
    parser.add_argument("--date", default=dt.datetime.now(IST).date().isoformat())
    parser.add_argument("--stable-age", type=int, default=30)
    parser.add_argument("--ack-host", default="macpro")
    parser.add_argument("--ack-root", default="/Users/danny./ecom-direct/receipts/competitor-promotion")
    args = parser.parse_args()

    inbox = Path(args.inbox)
    receipt_root = Path(args.receipts)
    summary: dict[str, Any] = {
        "date": args.date, "new": 0, "existing": 0, "waiting": 0,
        "rejected": 0, "endpoint_failures": [], "rejection_alerts": [], "errors": [],
    }
    if inbox.exists():
        for source in sorted(inbox.glob("*/report.ready.json")):
            run_dir = source.parent
            if not run_dir.name.startswith(args.date.replace("-", "")):
                continue
            rejection = receipt_root / args.date / "rejected" / f"{hashlib.sha256(run_dir.name.encode()).hexdigest()}.json"
            try:
                source_sha = sha256_file(source)
            except OSError:
                source_sha = ""
            state_sha = package_state_sha(run_dir)
            if not source_sha:
                source_sha = hashlib.sha256(
                    f"unreadable:{run_dir.name}:{state_sha}".encode("utf-8")
                ).hexdigest()
            if rejection.is_file():
                try:
                    rejected = load_json(rejection)
                    if rejected.get("source_receipt_sha256") == source_sha \
                       and rejected.get("package_state_sha256") == state_sha:
                        summary["rejected"] += 1
                        marker = rejection.with_suffix(".alerted.json")
                        if not alert_marker_matches(marker, source_sha, state_sha):
                            summary["rejection_alerts"].append({
                                "kind": "report-package",
                                "run_id": run_dir.name,
                                "error": str(rejected.get("error") or "rejected"),
                                "source_receipt_sha256": source_sha,
                                "package_state_sha256": state_sha,
                                "alert_marker": str(marker),
                            })
                        continue
                except (OSError, ValueError, json.JSONDecodeError):
                    pass
            try:
                status, message = consume_run(
                    run_dir, args.date, Path(args.output), Path(args.data), Path(args.audit_dir),
                    receipt_root, args.stable_age, args.ack_host, args.ack_root,
                )
                summary["new" if status == "accepted" else status] += 1
                print(f"{run_dir.name}: {status}: {message}", file=sys.stderr)
            except Exception as exc:
                summary["rejected"] += 1
                summary["errors"].append({"run_id": run_dir.name, "error": str(exc)})
                print(f"{run_dir.name}: rejected: {exc}", file=sys.stderr)
                atomic_json(rejection, {
                    "schema": "jivo-direct-competitor-rejection-v1",
                    "run_id": run_dir.name,
                    "source_receipt_sha256": source_sha,
                    "package_state_sha256": state_sha,
                    "error": str(exc),
                })
                summary["rejection_alerts"].append({
                    "kind": "report-package",
                    "run_id": run_dir.name,
                    "error": str(exc),
                    "source_receipt_sha256": source_sha,
                    "package_state_sha256": state_sha,
                    "alert_marker": str(rejection.with_suffix(".alerted.json")),
                })
        for source in sorted(inbox.glob("*/failure.json")):
            if not source.parent.name.startswith(args.date.replace("-", "")):
                continue
            failure_rejection = Path(args.failure_receipts) / args.date / "rejected" / \
                f"{hashlib.sha256(source.parent.name.encode()).hexdigest()}.json"
            try:
                failure_source_sha = sha256_file(source)
            except OSError:
                failure_source_sha = ""
            failure_state_sha = package_state_sha(source.parent)
            if not failure_source_sha:
                failure_source_sha = hashlib.sha256(
                    f"unreadable:{source.parent.name}:{failure_state_sha}".encode("utf-8")
                ).hexdigest()
            if failure_rejection.is_file():
                try:
                    rejected = load_json(failure_rejection)
                    if rejected.get("source_receipt_sha256") == failure_source_sha \
                       and rejected.get("package_state_sha256") == failure_state_sha:
                        marker = failure_rejection.with_suffix(".alerted.json")
                        if not alert_marker_matches(marker, failure_source_sha, failure_state_sha):
                            summary["rejection_alerts"].append({
                                "kind": "failure-receipt",
                                "run_id": source.parent.name,
                                "error": str(rejected.get("error") or "rejected"),
                                "source_receipt_sha256": failure_source_sha,
                                "package_state_sha256": failure_state_sha,
                                "alert_marker": str(marker),
                            })
                        continue
                except (OSError, ValueError, json.JSONDecodeError):
                    pass
            try:
                status, detail = consume_failure(source, args.date, Path(args.failure_receipts))
                if status in {"new", "pending_alert"}:
                    summary["endpoint_failures"].append(detail)
            except Exception as exc:
                summary["rejected"] += 1
                summary["errors"].append({"run_id": source.parent.name, "error": str(exc)})
                print(f"{source.parent.name}: rejected failure receipt: {exc}", file=sys.stderr)
                atomic_json(failure_rejection, {
                    "schema": "jivo-direct-competitor-failure-rejection-v1",
                    "run_id": source.parent.name,
                    "source_receipt_sha256": failure_source_sha,
                    "package_state_sha256": failure_state_sha,
                    "error": str(exc),
                })
                summary["rejection_alerts"].append({
                    "kind": "failure-receipt",
                    "run_id": source.parent.name,
                    "error": str(exc),
                    "source_receipt_sha256": failure_source_sha,
                    "package_state_sha256": failure_state_sha,
                    "alert_marker": str(failure_rejection.with_suffix(".alerted.json")),
                })

    for accepted in (receipt_root / args.date).glob("*.json"):
        try:
            value = load_json(accepted)
            run_id = str(value.get("run_id") or "")
            if value.get("schema") == PROMOTION_SCHEMA and RUN_ID.fullmatch(run_id):
                acknowledge_once(accepted, run_id, args.ack_host, args.ack_root)
        except (OSError, ValueError, KeyError, json.JSONDecodeError):
            continue
    print(json.dumps(summary, sort_keys=True))
    return 1 if summary["errors"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
