#!/usr/bin/env python3
"""
check_layout.py — post-batch layout gate (alert-only, never breaks the chain).

Owner rule 2026-06-10: agreed-price (BAU/SVD/ART) content is Amazon-only.
After each batch this verifies, for today's books:
  - competitor platforms (zepto/blinkit/flipkart/flipkart-minutes/bigbasket):
      NO "Price Match" tab, NO agreed-price block on the Leadership View
  - Amazon family (amazon/amazon-fresh/amazon-now):
      "Price Match" tab present

Telegram: FAILs always ping. PASS pings exactly once (first run after install,
state file logs/.layout_gate_bootstrapped) so the owner gets one positive
confirmation, then silence-unless-broken. --dry-run = check + print only.
Exit 0 always (alert-only, mirrors the doctor philosophy).
"""

import datetime
import glob
import os
import sys
import urllib.parse
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))                 # /opt/ecom-intel
STATE = os.path.join(ROOT, "logs", ".layout_gate_bootstrapped")
TG_ENV = "/root/.config/tg/env"

COMPETITORS = ["zepto", "blinkit", "flipkart", "flipkart-minutes", "bigbasket"]
AMAZON_FAMILY = ["amazon", "amazon-fresh", "amazon-now"]
BAU_MARKERS = ("price plan", "Below reference", "BELOW reference",
               "below reference", "agreed price")

# BUG#1 PERMANENT GUARD (2026-07-04, goal #61): a cell with a percent number_format must
# hold a FRACTION (0.363 → "36.3%"), never a raw percent (36.3 → "3630%"). This scanner
# catches a regression before it ships: (a) quoted-percent formats like 0.0"%" (the old
# buggy style), and (b) a %-formatted cell whose value is >= PCT_INFLATION_CEIL, i.e.
# >= 2000% — unambiguous inflation. Real gaps/ratios (e.g. a 161% price gap = 1.61) stay
# far below the ceiling, so this does not false-alarm on legitimate large percentages.
PCT_INFLATION_CEIL = 20.0


def newest_book(platform, date):
    pat = os.path.join(ROOT, "platforms", platform, f"Jivo-*{date}*.xlsx")
    books = sorted(glob.glob(pat), key=os.path.getmtime)
    return books[-1] if books else None


def check_book(platform, path):
    """Return a list of problem strings (empty = compliant)."""
    import openpyxl
    problems = []
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        has_tab = "Price Match" in wb.sheetnames
        if platform in AMAZON_FAMILY:
            if not has_tab:
                problems.append(f"{platform}: 'Price Match' tab MISSING")
            return problems
        if has_tab:
            problems.append(f"{platform}: 'Price Match' tab present (Amazon-only)")
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows():
            for c in row:
                v = str(c.value or "")
                if any(m in v for m in BAU_MARKERS):
                    problems.append(f"{platform}: agreed-price text on "
                                    f"{ws.title} r{c.row}: {v[:60]!r}")
                    return problems          # one example is enough
        return problems
    finally:
        wb.close()


def check_percent_inflation(date, books_dir=None):
    """Scan today's delivered workbooks (<books_dir>/*<date>*.xlsx — platform reports AND the
    competitor/price-watch books) for the percent-inflation bug. Returns a list of problem
    strings (empty = clean). Alert-only, like the rest of this gate. books_dir defaults to
    output/ (production); tests pass a temp dir."""
    import openpyxl
    problems = []
    books_dir = books_dir or os.path.join(ROOT, "output")
    books = sorted(glob.glob(os.path.join(books_dir, f"*{date}*.xlsx")),
                   key=os.path.getmtime)
    for path in books:
        base = os.path.basename(path)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            for ws in wb.worksheets:
                hits = 0
                for row in ws.iter_rows():
                    for c in row:
                        nf = c.number_format or ""
                        if "%" not in nf:
                            continue
                        if '"%"' in nf or "'%'" in nf:
                            problems.append(f"{base}[{ws.title}] {c.coordinate}: quoted-percent "
                                            f"format {nf!r} — use 0.0% on a fraction")
                            hits += 1
                        else:
                            v = c.value
                            if isinstance(v, (int, float)) and abs(v) >= PCT_INFLATION_CEIL:
                                problems.append(f"{base}[{ws.title}] {c.coordinate}: value {v} in "
                                                f"{nf!r} renders {v * 100:.0f}% — raw value, "
                                                f"should be a fraction")
                                hits += 1
                        if hits >= 3:
                            break
                    if hits >= 3:
                        break               # a few examples per sheet is enough
        finally:
            wb.close()
    return problems


def telegram(text):
    tok = chat = None
    try:
        for line in open(TG_ENV):
            line = line.strip()
            if line.startswith("TELEGRAM_BOT_TOKEN="):
                tok = line.split("=", 1)[1]
            elif line.startswith("TELEGRAM_CHAT_ID="):
                chat = line.split("=", 1)[1]
        if not (tok and chat):
            raise RuntimeError("tg env incomplete")
        urllib.request.urlopen(
            f"https://api.telegram.org/bot{tok}/sendMessage",
            data=urllib.parse.urlencode({"chat_id": chat, "text": text}).encode(),
            timeout=30)
    except Exception as e:                    # alert-only: never raise
        sys.stderr.write(f"check_layout: telegram failed (non-fatal): {e}\n")


def main(argv):
    dry = "--dry-run" in argv
    date = datetime.date.today().isoformat()
    problems, missing, checked = [], [], 0
    for p in COMPETITORS + AMAZON_FAMILY:
        path = newest_book(p, date)
        if not path:
            missing.append(p)                 # batch may have held a report
            continue
        checked += 1
        problems.extend(check_book(p, path))

    pct_problems = check_percent_inflation(date)

    if problems or pct_problems:
        parts = ["🚨 BATCH GATE FAIL — today's books have issues:"]
        if problems:
            parts.append("Amazon-only rule:\n"
                         + "\n".join("• " + q for q in problems[:8]))
        if pct_problems:
            parts.append("Percent inflation (bug#1 guard):\n"
                         + "\n".join("• " + q for q in pct_problems[:8]))
        if missing:
            parts.append(f"(books missing: {', '.join(missing)})")
        msg = "\n".join(parts)
        print(msg)
        if not dry:
            telegram(msg)
        return 0
    note = (f"layout+percent gate PASS — {checked} book(s) compliant, no percent inflation"
            + (f" ({', '.join(missing)} not built yet)" if missing else ""))
    print(note)
    # bootstrap ping only on a FULL pass (all 8 books present) — a late chain
    # at 12:10 holds the confirmation for the 15:10 run instead
    if not dry and not missing and not os.path.exists(STATE):
        telegram(f"✅ Batch verified ({date}) — all {checked} reports checked: "
                 "competitor books carry no agreed-price content / no Price "
                 "Match tab, Amazon books intact. The layout gate now runs "
                 "after every batch and pings again only if one ever breaks "
                 "the rule.")
        open(STATE, "w").write(date + "\n")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
