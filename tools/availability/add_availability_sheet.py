#!/usr/bin/env python3
"""add_availability_sheet.py — append an "Availability" sheet to a platform's report.

The daily scrape set IS the serviceable footprint (where the platform delivers). This
sheet lists EVERY one of those pincodes and marks whether Jivo is actually on sale —
so "delivers but Jivo NOT available" pincodes are visible, not silently dropped. That
is the whole point of the availability tracker.

Source: platforms/<platform>/result.json -> perPin (each has rows = Jivo products found,
and a serviceability signal: 'serviceable' (amazon/zepto) or 'resolved' (blinkit)).

usage: python3 add_availability_sheet.py <platform> <report.xlsx>
Best-effort: never raises in a way that fails the run.
"""
import json
import os
import sys

BASE = "/opt/ecom-intel/platforms"


def is_serviceable(x):
    if "serviceable" in x:
        return bool(x.get("serviceable"))
    if "resolved" in x:
        return bool(x.get("resolved"))
    return bool(x.get("rows"))  # fallback: had Jivo rows -> definitely serviceable


def first_price(rows):
    for r in rows or []:
        v = str(r.get("sale") or r.get("price") or r.get("mrp") or "").strip()
        if v:
            return v
    return ""


def main():
    if len(sys.argv) < 3:
        print("usage: add_availability_sheet.py <platform> <report.xlsx>")
        return
    plat, xlsx = sys.argv[1], sys.argv[2]
    rp = os.path.join(BASE, plat, "result.json")
    if not os.path.exists(rp) or not os.path.exists(xlsx):
        print(f"[availability] {plat}: missing result.json or report; skip")
        return
    try:
        d = json.load(open(rp))
    except Exception as e:
        print(f"[availability] {plat}: result.json unreadable ({e}); skip")
        return

    rows = []
    nserv = njivo = 0
    for x in d.get("perPin", []):
        if not isinstance(x, dict) or not is_serviceable(x):
            continue
        nserv += 1
        jr = x.get("rows") or []
        jivo = bool(jr)
        if jivo:
            njivo += 1
        rows.append((x.get("city", ""), str(x.get("pincode", "")),
                     "Yes" if jivo else "No", first_price(jr)))
    rows.sort(key=lambda r: (r[0], r[1]))

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as e:
        print(f"[availability] openpyxl missing ({e}); skip")
        return

    wb = load_workbook(xlsx)
    if "Availability" in wb.sheetnames:
        del wb["Availability"]
    ws = wb.create_sheet("Availability", 1)
    ws.append([f"JIVO Availability — {plat} (every serviceable pincode)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Serviceable: {nserv}", f"Jivo on sale: {njivo}",
               f"Delivers but Jivo NOT available: {nserv - njivo}"])
    ws.append(["City", "Pincode", "Jivo available?", "Price (INR)"])
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1B7A3D")
    for c in ws[3]:
        c.font, c.fill = hf, hfill
    no_fill = PatternFill("solid", fgColor="FDE8E8")
    for city, pin, jivo, price in rows:
        ws.append([city, pin, jivo, price])
        if jivo == "No":
            for c in ws[ws.max_row]:
                c.fill = no_fill
    for col, w in zip("ABCD", [22, 12, 16, 12]):
        ws.column_dimensions[col].width = w
    wb.save(xlsx)
    print(f"[availability] {plat}: {nserv} serviceable / {njivo} Jivo / "
          f"{nserv - njivo} deliver-but-no-Jivo -> 'Availability' sheet added")


if __name__ == "__main__":
    main()
