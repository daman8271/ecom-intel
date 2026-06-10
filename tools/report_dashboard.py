#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
report_dashboard.py <platform> <xlsx_path> [--date YYYY-MM-DD] [--force]

(Re)generates the FIRST sheet "Leadership View" of a platform workbook as a
chart-free, DURABLE dashboard — the LAST step of the run.sh chain (after the
Predictions and Price Match appends).

WHY THIS EXISTS (root cause, 2026-06-06)
-----------------------------------------
The old Leadership View (tools/predict.py:build_dashboard) draws NATIVE
openpyxl charts. Those charts carry ZERO cached values (no numCache/strCache),
so every cache-only viewer — macOS Quick Look (the owner's screenshot),
Google Drive preview, Office mobile, WPS — renders the chart bands EMPTY.
Only desktop Excel re-resolves the cell references. On top of that, any
build_excel-only rebuild of an archive workbook wiped the sheet entirely.

THE DURABILITY CONTRACT (mirrors tools/xlsx_dash.py)
----------------------------------------------------
Everything drawn here is styled cells, conditional-format data bars / color
scales, and unicode text sparklines. All of it renders in EVERY viewer and
survives any number of openpyxl load()+save() round-trips. NO native charts.
Because this runs LAST, anything that re-opens the workbook earlier in the
chain can no longer kill the first page.

FAIL-SAFE: any error → warning on stderr, exit 0, workbook untouched
(all edits happen on a temp copy, atomic os.replace on full success).
Idempotent: an existing "Leadership View" sheet is replaced in full.
"""

import csv
import datetime
import json
import os
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)

import xlsx_dash as xd  # noqa: E402  (shared dashboard helpers)

SHEET_NAME = "Leadership View"
IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))

# PERCENT RULE (memory: xlsx-percent-format-double-scaling): literal '0.0"%"'
# formats get ×100-scaled by the owner's preview app. ALWAYS store the
# FRACTION (value/100) with a true percent format — renders right everywhere.
FMT_PCT_TRUE = "0.0%"


def pct(v):
    """0-100-scale number -> fraction for a true-% formatted cell (rounded so
    float artifacts like -0.08130000000000001 never reach a formula bar)."""
    return None if v is None else round(v / 100.0, 4)


# Plain-words glossary (fresh-eyes 2026-06-06: SVD/regime/exposure were never
# expanded anywhere). Prefer the shared xlsx_dash GLOSSARY once it lands.
_GLOSSARY_FALLBACK = {
    "SVD": "Special Value Days (Fri–Sun agreed price list)",
    "BAU": "Business As Usual (weekday agreed price list)",
    # owner-confirmed 2026-06-06 (bus): ART = festival pricing
    "ART": "festival/event agreed price list, applied when announced",
}


def regime_expansion(regime):
    g = dict(_GLOSSARY_FALLBACK)
    g.update(getattr(xd, "GLOSSARY", None) or {})   # shared entries win
    return g.get(str(regime), None)


# Owner 2026-06-10: the agreed price list (BAU/SVD/ART) is an Amazon agreement —
# only the Amazon-family reports carry the reference-compliance block. Every
# other platform has NO agreed price; those platforms are compared against
# Amazon in the Price Match workbook's PM Check sheets instead.
AGREED_PRICE_PLATFORMS = {"amazon", "amazon-now", "amazon-fresh"}


# ------------------------------------------------------------- data helpers
def _f(v):
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return None


def in_stock_of(row):
    """True / False / None from the cross-platform in_stock field."""
    v = row.get("in_stock")
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v > 0
    s = str(v).strip().lower()
    if s in ("1", "true", "yes", "in stock", "instock", "available"):
        return True
    if s in ("0", "false", "no", "out of stock", "oos", "unavailable"):
        return False
    return None


def display_name(row, platform):
    """Real on-page listing name (title-hygiene rules): flipkart real title is
    fk_name; bigbasket card titles drop the brand prefix."""
    name = row.get("fk_name") or row.get("sku_raw") or row.get("canonical") or "?"
    name = str(name).strip()
    if platform == "bigbasket" and not name.lower().startswith("jivo"):
        name = "Jivo " + name
    return name


def load_result(platform):
    path = os.path.join(ROOT, "platforms", platform, "result.json")
    with open(path, encoding="utf-8") as f:
        d = json.load(f)
    rows = d.get("allRows") or [r for p in d.get("perPin", [])
                                for r in (p.get("rows") or [])]
    # cities probed this run where NOT ONE pincode returned a Jivo row
    # (named in the action list — "5 cities have zero Jivo" begs "which?")
    zero_cities = []
    per_pin = d.get("perPin") or []
    if per_pin:
        by_city = {}
        for p in per_pin:
            c = (str(p.get("city") or "").strip()) or None
            if not c:
                continue
            by_city.setdefault(c, 0)
            by_city[c] += len(p.get("rows") or [])
        zero_cities = sorted(c for c, n in by_city.items() if n == 0)
    return d.get("summary") or {}, rows, zero_cities


def captured_ist(summary):
    cap = summary.get("captured_at")
    try:
        dt = datetime.datetime.fromisoformat(str(cap).replace("Z", "+00:00"))
        return dt.astimezone(IST)
    except Exception:
        return None


def load_history(platform, last_n=10):
    """Per-run aggregates from data/<p>/history.csv (append-ordered).
    Returns list[(run_id, rows, instock_pct, avg_disc, n_pincodes)] for the
    last `last_n` runs."""
    path = os.path.join(ROOT, "data", platform, "history.csv")
    if not os.path.isfile(path):
        return []
    agg = {}      # run_id -> [rows, in, discs_sum, discs_n, set(pincodes)]
    order = []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rid = row.get("run_id") or "?"
            if rid not in agg:
                agg[rid] = [0, 0, 0.0, 0, set()]
                order.append(rid)
            a = agg[rid]
            a[0] += 1
            ins = str(row.get("in_stock", "")).strip()
            if ins in ("1", "True", "true"):
                a[1] += 1
                d = _f(row.get("discount_pct"))
                if d is not None:
                    a[2] += d
                    a[3] += 1
            pc = (row.get("pincode") or "").strip()
            if pc:
                a[4].add(pc)
    # comparable-sweep filter (fresh-eyes 2026-06-06): drop off-hours debug/
    # heal runs (a handful of rows) that read as fake collapse-and-recovery
    # in the momentum table. Same rule as predict.py: keep runs with row
    # count >=25% of the median; the newest run is always kept.
    if order:
        counts = [agg[x][0] for x in order]
        med = sorted(counts)[len(counts) // 2]
        floor = 0.25 * med
        keep = [x for x in order if agg[x][0] >= floor]
        if order[-1] not in keep:
            keep.append(order[-1])
        order = keep
    out = []
    for rid in order[-last_n:]:
        rows_n, n_in, dsum, dn, pins = agg[rid]
        out.append((rid, rows_n,
                    (100.0 * n_in / rows_n) if rows_n else None,
                    (dsum / dn) if dn else None,
                    len(pins)))
    return out


def pricematch_block(platform, date_str):
    """Compliance numbers via tools/pricematch/pricematch_core. Returns dict
    or None when price-match is unavailable (no sheet lies, just omitted)."""
    try:
        sys.path.insert(0, os.path.join(HERE, "pricematch"))
        import pricematch_core as core
        ctx = core.load_context(date_str) if date_str else core.load_context()
        recs = core.platform_comparison(ctx, platform)
        regime = (ctx.get("regime") if isinstance(ctx, dict) else None) \
            or core.regime_for(date_str or datetime.date.today())
    except Exception as e:
        sys.stderr.write("report_dashboard: price-match block skipped "
                         "(non-fatal): %s\n" % e)
        return None
    counts = {}
    for r in recs:
        counts[r.get("status") or "?"] = counts.get(r.get("status") or "?", 0) + 1
    below = [r for r in recs if r.get("status") == "BELOW"]
    # house definition (pricematch_core.summary): Σ(ref − live_modal) in ₹
    exposure = sum((r.get("ref_price") or 0) - (r.get("live_modal") or 0)
                   for r in below
                   if r.get("ref_price") is not None
                   and r.get("live_modal") is not None)
    violations = sum(len(r.get("stores_below") or []) for r in below)
    below.sort(key=lambda r: (r.get("diff_pct") if r.get("diff_pct")
                              is not None else 0))
    return {"counts": counts, "exposure": round(exposure),
            "violations": violations, "regime": regime, "below": below}


# ------------------------------------------------------------- the dashboard
def build_sheet(wb, platform, date_str):
    summary, rows, zero_cities = load_result(platform)
    if not rows:
        raise RuntimeError("result.json has no rows — refusing to draw an "
                           "empty dashboard over the existing sheet")

    pname = (getattr(xd, "PLATFORM_DISPLAY", None) or {}).get(
        platform, platform.replace("-", " ").title())
    cap = captured_ist(summary)
    when = cap.strftime("%Y-%m-%d %H:%M IST") if cap else (date_str or "")

    # ---- core splits ----
    n_in = n_oos = n_nf = 0
    for r in rows:
        at = str(r.get("availability_text") or "").upper()
        if at in ("NOT FOUND", "BLOCKED"):
            n_nf += 1
            continue
        st = in_stock_of(r)
        if st is True:
            n_in += 1
        elif st is False:
            n_oos += 1
        else:
            n_nf += 1
    stocked = n_in + n_oos
    instock_pct = (100.0 * n_in / stocked) if stocked else 0.0

    bands = [("0-20%", 0, 20), ("20-40%", 20, 40),
             ("40-60%", 40, 60), ("60%+", 60, 1e9)]
    band_counts = [0] * len(bands)
    discs = []
    for r in rows:
        if in_stock_of(r) is not True:
            continue
        d = _f(r.get("discount_pct"))
        if d is None:
            continue
        discs.append(d)
        for i, (_, lo, hi) in enumerate(bands):
            if lo <= d < hi:
                band_counts[i] += 1
                break
    avg_disc = (sum(discs) / len(discs)) if discs else 0.0
    deep = sum(1 for d in discs if d >= 50)

    # per-SKU stock-out risk: SKUs that are OOS at EVERY location they appear
    per_sku = {}
    for r in rows:
        k = r.get("canonical") or r.get("sku_raw") or "?"
        e = per_sku.setdefault(k, {"in": 0, "oos": 0, "name": display_name(r, platform)})
        st = in_stock_of(r)
        if st is True:
            e["in"] += 1
        elif st is False:
            e["oos"] += 1
    risk_skus = [e["name"] for e in per_sku.values()
                 if e["oos"] > 0 and e["in"] == 0]

    # city spread (in-stock datapoints)
    cities = {}
    for r in rows:
        if in_stock_of(r) is not True:
            continue
        c = (str(r.get("city") or "?").strip()) or "?"
        cities[c] = cities.get(c, 0) + 1
    city_rows = sorted(cities.items(), key=lambda kv: -kv[1])[:8]

    # cheapest SKUs by ₹/L + top discounts (in-stock, real prices only)
    best_pl = {}
    best_disc = {}
    for r in rows:
        if in_stock_of(r) is not True:
            continue
        k = r.get("canonical") or r.get("sku_raw") or "?"
        pl = _f(r.get("per_litre"))
        if pl and pl > 0 and (k not in best_pl or pl < best_pl[k]["pl"]):
            best_pl[k] = {"pl": pl, "name": display_name(r, platform),
                          "city": r.get("city") or "—",
                          "sale": _f(r.get("sale")), "mrp": _f(r.get("mrp"))}
        d = _f(r.get("discount_pct"))
        if d and d > 0 and (k not in best_disc or d > best_disc[k]["d"]):
            best_disc[k] = {"d": d, "name": display_name(r, platform),
                            "city": r.get("city") or "—",
                            "sale": _f(r.get("sale")), "mrp": _f(r.get("mrp"))}
    cheapest = sorted(best_pl.values(), key=lambda e: e["pl"])[:8]
    discounts = sorted(best_disc.values(), key=lambda e: -e["d"])[:8]

    pm = (pricematch_block(platform, date_str)
          if platform in AGREED_PRICE_PLATFORMS else None)
    hist = load_history(platform, last_n=10)

    pin_tot = summary.get("pincodes_total")
    pin_with = summary.get("pincodes_with_jivo")
    national = not pin_tot or int(pin_tot or 0) <= 1
    skus = summary.get("unique_skus") or len(per_sku)
    nrows = summary.get("total_rows") or len(rows)

    # ============================ draw ============================
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, 0)
    wb.active = 0
    xd.no_gridlines(ws, zoom=110)
    xd.col_grid(ws, n=10, width=16)
    xd.fit_to_width(ws)

    headline = "%.0f%% in stock   ·   avg %.0f%% off   ·   %d SKU(s) at stock-out risk" \
        % (instock_pct, avg_disc, len(risk_skus))
    if pm:
        headline += "   ·   %d below reference" % pm["counts"].get("BELOW", 0)
    r = xd.title_block(ws, 1, "Jivo × %s — Leadership View" % pname,
                       headline + "   ·   " + when)

    # ---- KPI strip (plain cells: render EVERYWHERE) ----
    cov_val = "National" if national else "%s/%s" % (pin_with, pin_tot)
    cov_sub = "single national listing" if national else "pincodes with Jivo live"
    # honest tone: 31% coverage must not glow green (fresh-eyes 2026-06-06)
    if national:
        cov_tone = "brand"
    else:
        try:
            ratio = float(pin_with) / float(pin_tot)
        except (TypeError, ValueError, ZeroDivisionError):
            ratio = None
        cov_tone = ("good" if ratio is not None and ratio >= 0.7 else
                    "warn" if ratio is not None and ratio >= 0.4 else
                    "bad" if ratio is not None else "neutral")
    cards = [
        ("COVERAGE", cov_val, cov_sub, cov_tone),
        ("CATALOG SKUs", "%d" % int(skus), "%s datapoints this run" % nrows, "neutral"),
        ("IN STOCK", "%.0f%%" % instock_pct,
         "%d of %d datapoints" % (n_in, stocked),
         "good" if instock_pct >= 80 else ("warn" if instock_pct >= 60 else "bad")),
        ("AVG DISCOUNT", "%.0f%%" % avg_disc, "%d priced in-stock rows" % len(discs),
         "good" if avg_disc > 0 else "neutral"),
    ]
    if pm:
        nb = pm["counts"].get("BELOW", 0)
        cards.append(("BELOW REFERENCE", "%d" % nb,
                      "%d match · %d above · vs %s plan"
                      % (pm["counts"].get("MATCH", 0),
                         pm["counts"].get("ABOVE", 0), pm["regime"]),
                      "bad" if nb else "good"))
    else:
        cards.append(("DEEP DISC ≥50%", "%d" % deep, "of %d priced" % len(discs),
                      "warn" if deep else "good"))
    cc = 1
    for t, v, s, tone in cards:
        nr, cc = xd.kpi_card(ws, r, cc, t, v, s, tone)
    r = nr + 1

    # ================= PORTFOLIO HEALTH (durable content) =================
    r = xd.section_title(ws, r, "Portfolio health")
    r += 1
    top = r

    # left: stock split (counts + text meters — no chart needed)
    split_rows = [["In stock", n_in, xd.meter(n_in, max(1, len(rows)))],
                  ["Out of stock", n_oos, xd.meter(n_oos, max(1, len(rows)))]]
    if n_nf:
        split_rows.append(["Not found / blocked", n_nf,
                           xd.meter(n_nf, max(1, len(rows)))])
    xd.banded_table(
        ws, r, 1, ["Stock split", "Rows", "% of rows"],
        split_rows,
        widths=[18, 10, 18], num_fmts=[None, xd.FMT_INT, None],
        aligns=[None, "right", None])
    left_end = r + (3 if n_nf else 2)
    xd.data_bar(ws, "B%d:B%d" % (r + 1, left_end), color=xd.BRAND)

    # left, below: discount mix
    r2 = left_end + 2
    xd.banded_table(
        ws, r2, 1, ["Discount band", "Rows", "% of priced rows"],
        [[label, band_counts[i],
          xd.meter(band_counts[i], max(1, sum(band_counts)))]
         for i, (label, _, _) in enumerate(bands)],
        widths=[18, 10, 18], num_fmts=[None, xd.FMT_INT, None],
        aligns=[None, "right", None])
    left_end2 = r2 + len(bands)
    xd.data_bar(ws, "B%d:B%d" % (r2 + 1, left_end2), color=xd.ACCENT)

    # right: where Jivo is live by city (quick-comm) / availability note (national)
    if city_rows and not national:
        xd.banded_table(
            ws, top, 5,
            ["Where Jivo is live (city)", "In-stock rows", "vs top city"],
            [[c, n, xd.meter(n, city_rows[0][1])] for c, n in city_rows],
            widths=[26, 14, 18], num_fmts=[None, xd.FMT_INT, None],
            aligns=[None, "right", None])
        right_end = top + len(city_rows)
        xd.data_bar(ws, "F%d:F%d" % (top + 1, right_end), color=xd.BRAND)
    else:
        # national platforms: top categories by live rows when present
        cats = {}
        for rr in rows:
            cat = (str(rr.get("category") or "").strip() or "(uncategorised)").title()
            e = cats.setdefault(cat, [0, 0])
            e[1] += 1
            if in_stock_of(rr) is True:
                e[0] += 1
        cat_rows = sorted(((c, i, t, 100.0 * i / t) for c, (i, t) in cats.items()
                           if t >= 2), key=lambda x: x[3])[:8]
        if cat_rows:
            xd.banded_table(
                ws, top, 5,
                ["Category (lowest live %)", "Live", "Listed", "Live %"],
                [[c, i, t, pct(p)] for c, i, t, p in cat_rows],
                widths=[26, 10, 10, 12],
                num_fmts=[None, xd.FMT_INT, xd.FMT_INT, FMT_PCT_TRUE],
                aligns=[None, "right", "right", "right"])
            right_end = top + len(cat_rows)
            xd.data_bar(ws, "H%d:H%d" % (top + 1, right_end),
                        color=xd.BRAND, vmin=0, vmax=1)
        else:
            right_end = top
    r = max(left_end2, right_end) + 2

    # ================= PRICING & COMPETITION (durable content) =============
    r = xd.section_title(ws, r, "Pricing & competition")
    r += 1
    if pm:
        xd.chip(ws, r, 1, "Today's price plan: %s" % pm["regime"],
                xd.BRAND_SOFT, xd.GREEN_TEXT)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        msg = ("%d SKU(s) priced BELOW reference · exposure ₹%s · "
               "%d store-pincode violation(s)"
               % (pm["counts"].get("BELOW", 0), "{:,}".format(pm["exposure"]),
                  pm["violations"])
               if pm["counts"].get("BELOW")
               else "No SKU priced below reference this run")
        cell = ws.cell(r, 3, msg)
        cell.font = xd.Font(name=xd.F, size=11, bold=True,
                            color=xd.NEG if pm["counts"].get("BELOW") else xd.POS)
        cell.alignment = xd.Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
        r += 1
        # expand the regime code once, in plain words (fresh-eyes 2026-06-06)
        exp = regime_expansion(pm["regime"])
        gloss = ("%s = %s. " % (pm["regime"], exp)) if exp else ""
        xd.footnote(ws, r, gloss + '"Reference" = today\'s agreed price per '
                          'SKU; "exposure" = total ₹ gap below it.')
        r += 2
        if pm["below"]:
            body = []
            for rec in pm["below"][:8]:
                title = rec.get("title") or rec.get("sku") or "?"
                # engine diff_pct is ×100 (-16.69 == -16.69%) -> store fraction
                body.append([title, rec.get("ref_price"), rec.get("live_modal"),
                             pct(rec.get("diff_pct")),
                             len(rec.get("stores_below") or [])])
            last = xd.banded_table(
                ws, r, 1,
                ["Below reference (worst first)", "Ref ₹", "Live ₹",
                 "Diff %", "Stores below"],
                body, widths=[40, 11, 11, 10, 13],
                num_fmts=[None, xd.FMT_INR, xd.FMT_INR, FMT_PCT_TRUE, xd.FMT_INT],
                aligns=[None, "right", "right", "right", "right"])
            for rr in range(r + 1, last + 1):
                ws.cell(rr, 4).font = xd.Font(name=xd.F, size=11, bold=True,
                                              color=xd.NEG)
            xd.data_bar(ws, "E%d:E%d" % (r + 1, last), color=xd.NEG)
            r = last + 2
    elif platform in AGREED_PRICE_PLATFORMS:
        xd.footnote(ws, r, "Price-match reference unavailable for this run — "
                           "compliance block omitted.")
        r += 2

    # cheapest ₹/L (left) + top discounts (right), side by side
    top = r
    if cheapest:
        last = xd.banded_table(
            ws, top, 1, ["Cheapest ₹/L (in stock)", "City", "₹/L", "Sale ₹"],
            [[e["name"], e["city"], round(e["pl"], 1), e["sale"]] for e in cheapest],
            widths=[34, 12, 10, 10],
            num_fmts=[None, None, xd.FMT_INR, xd.FMT_INR],
            aligns=[None, None, "right", "right"])
        xd.data_bar(ws, "C%d:C%d" % (top + 1, last), color=xd.ACCENT)
        left_end = last
    else:
        left_end = top
    if discounts:
        last = xd.banded_table(
            ws, top, 6, ["Top discounts (in stock)", "City", "Disc %", "Sale ₹"],
            [[e["name"], e["city"], pct(e["d"]), e["sale"]] for e in discounts],
            widths=[34, 12, 10, 10],
            num_fmts=[None, None, FMT_PCT_TRUE, xd.FMT_INR],
            aligns=[None, None, "right", "right"])
        xd.data_bar(ws, "H%d:H%d" % (top + 1, last), color=xd.BRAND,
                    vmin=0, vmax=1)
        right_end = last
    else:
        right_end = top
    r = max(left_end, right_end) + 2

    # ================= MOMENTUM (text sparks + table — no charts) ==========
    if len(hist) >= 2:
        r = xd.section_title(ws, r, "Momentum (last %d runs)" % len(hist))
        r += 1
        ip_series = [h[2] for h in hist]
        ad_series = [h[3] for h in hist]
        xd.spark_cell(ws, r, 1, "In-stock %%  %s"
                      % xd.spark(ip_series, lo=0, hi=100), tone="brand")
        xd.spark_cell(ws, r, 4, "Avg disc %%  %s" % xd.spark(ad_series), tone="warn")
        xd.spark_cell(ws, r, 7, "Rows/run  %s" % xd.spark([h[1] for h in hist]),
                      tone="muted")
        r += 2
        body = []
        for rid, rows_n, ip, ad, npins in hist:
            label = rid
            try:
                label = "%s %s:%s" % (rid[5:10], rid[11:13], rid[13:15])
            except Exception:
                pass
            body.append([label, rows_n, npins, pct(ip), pct(ad)])
        last = xd.banded_table(
            ws, r, 1, ["Run", "Rows", "Pincodes", "In-stock %", "Avg disc %"],
            body, widths=[16, 10, 10, 12, 12],
            num_fmts=[None, xd.FMT_INT, xd.FMT_INT, FMT_PCT_TRUE, FMT_PCT_TRUE],
            aligns=[None, "right", "right", "right", "right"])
        xd.data_bar(ws, "D%d:D%d" % (r + 1, last), color=xd.BRAND,
                    vmin=0, vmax=1)
        xd.data_bar(ws, "E%d:E%d" % (r + 1, last), color=xd.ACCENT,
                    vmin=0, vmax=1)
        r = last + 2

    # ================= WHAT TO ACT ON =======================================
    r = xd.section_title(ws, r, "What to act on")
    r += 1
    actions = []
    if pm and pm["counts"].get("BELOW"):
        actions.append("%d SKU(s) priced below the %s reference — exposure "
                       "₹%s across %d store-pincode(s); see the Price Match "
                       "sheet." % (pm["counts"]["BELOW"], pm["regime"],
                                   "{:,}".format(pm["exposure"]),
                                   pm["violations"]))
    if risk_skus:
        sample = "; ".join(risk_skus[:3])
        actions.append("%d SKU(s) out of stock everywhere they are listed: %s%s"
                       % (len(risk_skus), sample,
                          " …" if len(risk_skus) > 3 else ""))
    if deep:
        actions.append("%d in-stock row(s) at ≥50%% discount — margin check."
                       % deep)
    if zero_cities:
        sample = ", ".join(zero_cities[:5])
        actions.append("%d city/cities have ZERO Jivo this run: %s%s"
                       % (len(zero_cities), sample,
                          " …" if len(zero_cities) > 5 else ""))
    if not national and pin_tot and pin_with is not None:
        gap = int(pin_tot) - int(pin_with)
        if gap > 0:
            actions.append("%d of %s probed pincode(s) have no Jivo presence "
                           "this run." % (gap, pin_tot))
    if len(hist) >= 2 and hist[-1][4] and hist[-2][4]:
        dpin = hist[-1][4] - hist[-2][4]
        if dpin:
            actions.append("Coverage %s: %+d pincode(s) vs previous run."
                           % ("grew" if dpin > 0 else "shrank", dpin))
    if not actions:
        actions.append("No red flags this run — stock, pricing and coverage "
                       "all within expected ranges.")
    for a in actions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        cell = ws.cell(r, 1, "▸  " + a)
        cell.font = xd.Font(name=xd.F, size=11, color=xd.INK)
        cell.alignment = xd.Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1
    # ONE timestamp (the scrape's captured_at), no internal tool paths
    # (fresh-eyes 2026-06-06): execs forward this; keep it clean.
    xd.footnote(ws, r, "Data captured %s · %s datapoints · source: %s live "
                       "storefront" % (when, nrows, pname))
    xd.freeze(ws, "A4")
    return {"cards": len(cards), "actions": len(actions),
            "pm": bool(pm), "hist_runs": len(hist)}


def build(platform, xlsx_path, date_str, force=False):
    from openpyxl import load_workbook

    # date guard: never paint TODAY's result.json onto an archive workbook
    base = os.path.basename(xlsx_path)
    m = base.rsplit("-", 3)
    fdate = None
    if base.endswith(".xlsx") and len(m) == 4:
        fdate = "-".join(m[-3:]).replace(".xlsx", "")
    summary, _, _ = load_result(platform)
    cap = captured_ist(summary)
    cap_date = cap.strftime("%Y-%m-%d") if cap else None
    if fdate and cap_date and fdate != cap_date and not force:
        raise RuntimeError(
            "date mismatch: workbook is %s but result.json was captured %s "
            "(--force to override)" % (fdate, cap_date))

    # all edits on a temp copy; atomic replace only on full success.
    # dotted hidden name (must NOT match the Jivo-*.xlsx glob run.sh uses).
    tmp = os.path.join(os.path.dirname(os.path.abspath(xlsx_path)),
                       ".lv.tmp." + base)
    shutil.copy2(xlsx_path, tmp)
    try:
        wb = load_workbook(tmp)
        stats = build_sheet(wb, platform, date_str or cap_date)
        wb.save(tmp)
        os.replace(tmp, xlsx_path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass
    return stats


def main(argv):
    args = [a for a in argv[1:] if not a.startswith("--")]
    force = "--force" in argv
    date_str = None
    it = iter(argv[1:])
    for a in it:
        if a == "--date":
            date_str = next(it, None)
        elif a.startswith("--date="):
            date_str = a.split("=", 1)[1]
    if len(args) < 2:
        sys.stderr.write("usage: report_dashboard.py <platform> <xlsx_path> "
                         "[--date YYYY-MM-DD] [--force]\n")
        return 0  # fail-safe: never break the pipeline
    platform, xlsx_path = args[0], args[1]
    try:
        if not os.path.isfile(xlsx_path):
            raise FileNotFoundError(xlsx_path)
        stats = build(platform, xlsx_path, date_str, force=force)
        sys.stderr.write(
            "report_dashboard: '%s' regenerated as sheet 1 of %s "
            "(%d KPI cards, %d action items, price-match=%s, %d history runs)\n"
            % (SHEET_NAME, xlsx_path, stats["cards"], stats["actions"],
               "yes" if stats["pm"] else "no", stats["hist_runs"]))
        return 0
    except Exception as e:
        # ANY error: warn, leave the workbook untouched, exit 0.
        import traceback
        sys.stderr.write("report_dashboard: FAILED (non-fatal, workbook "
                         "untouched): %s\n" % e)
        traceback.print_exc(file=sys.stderr)
        return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
