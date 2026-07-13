#!/usr/bin/env python3
import hashlib
import json
import os
import shutil
import subprocess
from collections import Counter

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = "/opt/ecom-intel"
RUN = os.path.join(ROOT, "shards/runs/20260713-blinkit-competitor-top8-3pcity")
CAPTURE_OUT = os.path.join(
    ROOT, "tools/competitor/data/blinkit_competitor_2026-07-13.json"
)
WORKBOOK = os.path.join(
    ROOT, "output/Competitor-Price-Watch-Blinkit-2026-07-13.xlsx"
)
TARGET_BRANDS = [
    "Fortune", "Saffola", "Borges", "Tata", "Del Monte", "Figaro",
    "Sundrop", "Gulab",
]
ALLOWED_RAW = {x.casefold() for x in TARGET_BRANDS} | {"jivo", "sano"}
DEVICES = (
    ("Windows", "windows", "pincodes.shard0.json"),
    ("Mac Pro", "mac", "pincodes.shard1.json"),
)


def read_json(path):
    with open(path, encoding="utf-8") as handle:
        return json.load(handle)


def norm_brand(value):
    return " ".join(str(value or "").split()).casefold()


def merge_capture():
    expected = read_json(os.path.join(RUN, "pincodes.all.json"))
    expected_pins = {str(row["pincode"]) for row in expected}
    if len(expected) != 75 or len(expected_pins) != 75:
        raise SystemExit("scope gate failed: expected 75 unique pincodes")

    progress = {}
    pin_device = {}
    summaries = []
    rows = []
    for device_label, prefix, shard_name in DEVICES:
        shard = read_json(os.path.join(RUN, shard_name))
        for item in shard:
            pin_device[str(item["pincode"])] = device_label

        device_progress = read_json(os.path.join(RUN, f"{prefix}.progress.json"))
        for pin, item in device_progress.items():
            if pin in progress:
                raise SystemExit(f"duplicate progress pincode: {pin}")
            progress[pin] = item

        capture = read_json(os.path.join(RUN, f"{prefix}.capture.json"))
        summaries.append(capture.get("summary") or {})
        rows.extend(capture.get("allRows") or [])

        progress_rows = sum(len(item.get("rows") or []) for item in device_progress.values())
        if progress_rows != len(capture.get("allRows") or []):
            raise SystemExit(
                f"{device_label} row gate failed: progress={progress_rows}, "
                f"capture={len(capture.get('allRows') or [])}"
            )

    if set(progress) != expected_pins:
        missing = sorted(expected_pins - set(progress))
        extra = sorted(set(progress) - expected_pins)
        raise SystemExit(f"coverage gate failed: missing={missing}, extra={extra}")

    unresolved = sorted(pin for pin, item in progress.items() if not item.get("resolved"))
    blocked = sorted(
        pin for pin, item in progress.items()
        if item.get("blocked") or item.get("partial_block")
    )
    auth_rejected = sorted(
        pin for pin, item in progress.items() if item.get("auth_accepted") != 1
    )
    if unresolved or blocked or auth_rejected:
        raise SystemExit(
            f"live gate failed: unresolved={unresolved}, blocked={blocked}, "
            f"auth_rejected={auth_rejected}"
        )
    empty_pins = sorted(
        pin for pin, item in progress.items() if not (item.get("rows") or [])
    )
    if empty_pins:
        raise SystemExit(f"row coverage gate failed: empty pincodes={empty_pins}")

    unexpected = sorted({norm_brand(row.get("brand")) for row in rows} - ALLOWED_RAW)
    if unexpected:
        raise SystemExit(f"brand gate failed: unexpected={unexpected}")
    bad_rows = []
    for index, row in enumerate(rows):
        pin = str(row.get("pincode") or "")
        if (
            pin not in expected_pins
            or row.get("platform") != "blinkit"
            or not norm_brand(row.get("brand"))
            or not isinstance(row.get("sale"), (int, float))
            or row.get("sale") <= 0
            or row.get("in_stock") not in (0, 1)
        ):
            bad_rows.append(index)
    if bad_rows:
        raise SystemExit(f"row quality gate failed: bad row indexes={bad_rows[:20]}")
    found = {norm_brand(row.get("brand")) for row in rows}
    missing_brands = [b for b in TARGET_BRANDS if b.casefold() not in found]
    if missing_brands:
        raise SystemExit(f"brand gate failed: no rows captured for {missing_brands}")

    deduped = {}
    for row in rows:
        key = (
            str(row.get("pincode")), str(row.get("store_id")),
            norm_brand(row.get("brand")), str(row.get("canonical")),
        )
        deduped.setdefault(key, row)
    rows = list(deduped.values())

    city_counts = Counter(str(row["city"]) for row in expected)
    if len(city_counts) != 25 or set(city_counts.values()) != {3}:
        raise SystemExit(f"city scope gate failed: {dict(city_counts)}")

    summary = {
        "mode": "competitor",
        "platform": "blinkit",
        "pincodes_total": 75,
        "pincodes_resolved": 75,
        "pincodes_unresolved": 0,
        "pincodes_blocked": 0,
        "pincodes_with_rows": sum(bool(item.get("rows")) for item in progress.values()),
        "total_rows": len(rows),
        "unique_skus": len({row.get("canonical") for row in rows}),
        "wall_s": max(int(item.get("wall_s") or 0) for item in summaries),
        "partial": False,
        "auth_session": 1,
        "auth_required": 1,
        "auth_verified": 1,
        "auth_verified_pincodes": 75,
        "captured_at": max(str(item.get("captured_at") or "") for item in summaries),
        "scope": {
            "cities": 25,
            "pincodes_per_city": 3,
            "devices": ["Windows", "Mac Pro"],
            "competitors": TARGET_BRANDS,
            "selection_source": "platforms/blinkit/pincodes.daily.json",
        },
    }
    merged = {"summary": summary, "allRows": rows}
    if os.path.exists(CAPTURE_OUT):
        backup = os.path.join(RUN, "preexisting-blinkit-competitor-2026-07-13.json")
        if not os.path.exists(backup):
            shutil.copy2(CAPTURE_OUT, backup)
    with open(CAPTURE_OUT, "w", encoding="utf-8") as handle:
        json.dump(merged, handle, indent=2, ensure_ascii=False)

    audit = {
        "summary": summary,
        "capture_sha256": hashlib.sha256(open(CAPTURE_OUT, "rb").read()).hexdigest(),
        "rows_by_brand": dict(sorted(Counter(
            norm_brand(row.get("brand")) for row in rows
        ).items())),
        "rows_by_city": dict(sorted(Counter(str(row.get("city")) for row in rows).items())),
        "pin_device": pin_device,
    }
    with open(os.path.join(RUN, "merge-audit.json"), "w", encoding="utf-8") as handle:
        json.dump(audit, handle, indent=2, ensure_ascii=False)
    return expected, progress, pin_device, summary


def add_scope_sheet(expected, progress, pin_device, summary):
    workbook = load_workbook(WORKBOOK)
    if "Run Scope" in workbook.sheetnames:
        del workbook["Run Scope"]
    sheet = workbook.create_sheet("Run Scope", 1)
    green = "008B3A"
    header_fill = PatternFill("solid", fgColor=green)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet["A1"] = "Blinkit Competitor Run Scope"
    sheet["A1"].font = Font(size=18, bold=True, color=green)
    sheet.merge_cells("A1:H1")
    sheet["A2"] = (
        f"13 Jul 2026 | 25 cities | 3 pincodes per city | 75/75 resolved | "
        f"{summary['total_rows']} datapoints"
    )
    sheet["A2"].font = Font(italic=True, color="555555")
    sheet.merge_cells("A2:H2")
    sheet["A4"] = "Top 8 competitors"
    sheet["B4"] = ", ".join(TARGET_BRANDS)
    sheet.merge_cells("B4:H4")
    sheet["A5"] = "Devices"
    sheet["B5"] = "Windows laptop + Mac Pro (2 workers each)"
    sheet.merge_cells("B5:H5")

    headers = [
        "City", "Pincode", "Locality", "Device", "Resolved", "Authenticated",
        "Rows", "Brands captured",
    ]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(row=7, column=column, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    ordered = sorted(expected, key=lambda row: (str(row["city"]), int(row["competitor_sample_index"])))
    for row_number, item in enumerate(ordered, 8):
        pin = str(item["pincode"])
        result = progress[pin]
        brands = sorted({
            " ".join(str(row.get("brand") or "").split())
            for row in result.get("rows") or [] if row.get("brand")
        }, key=str.casefold)
        values = [
            item.get("city"), pin, item.get("locality"), pin_device[pin],
            "Yes" if result.get("resolved") else "No",
            "Yes" if result.get("auth_accepted") == 1 else "No",
            len(result.get("rows") or []), ", ".join(brands),
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A7:H{sheet.max_row}"
    widths = [20, 12, 32, 12, 12, 15, 10, 70]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    workbook.save(WORKBOOK)


if __name__ == "__main__":
    expected_rows, run_progress, devices, merged_summary = merge_capture()
    subprocess.run([
        "python3",
        os.path.join(ROOT, "tools/competitor/build_competitor_report.py"),
        "blinkit",
        "2026-07-13",
    ], check=True)
    add_scope_sheet(expected_rows, run_progress, devices, merged_summary)
    print(json.dumps(merged_summary, sort_keys=True))
